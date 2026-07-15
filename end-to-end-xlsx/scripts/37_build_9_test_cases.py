#!/usr/bin/env python3
"""Build the nine-case 1C/Fitbase review workbook for the 2026-06-30 backup.

The workbook is deliberately built from the current full-cutoff membership staging
CSV.  Case metadata below contains only the additional 1C diagnostics that are not
part of the 22-column Fitbase import row.
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension


ROOT = Path(__file__).resolve().parents[1]
RUN_DIR = ROOT / "work" / "20260630_full_cutoff"
STATUS_PATH = RUN_DIR / "status.json"
STAGING_PATH = RUN_DIR / "imports" / "staging" / "membership_import_rows.csv"
DEFAULT_OUTPUT = ROOT / "9-test-cases" / "9_test_cases_20260630.xlsx"

CUTOFF_AT = "2026-06-30 23:27:03"
CUTOFF_DATE = "2026-06-30"
DATE_STAMP = "20260630"

CLIENT_HEADERS = [
    "tag",
    "contract_id",
    "client_id",
    "phone",
    "client_fio",
    "contract_name",
    "card",
    "duration",
    "duration_type",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "freeze",
    "guests",
    "visits_left",
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "manager",
    "филиал",
]

CLIENT_RUS_HEADERS = [
    "Тег",
    "Внутренний номер абонемента",
    "Внутренний номер клиента",
    "Телефон клиента",
    "ФИО клиента  *",
    "Название абонемента *",
    "Номер карты",
    "Продолжительность",
    "Тип продолжительности",
    "Дата  добавления *",
    "Дата оплаты *",
    "Дата активации ",
    "Дата окончания",
    "Осталось дней для заморозки",
    "Осталось гостевых визитов",
    "Осталось посещений *",
    "Стоимость *",
    "Оплачено *",
    "Осталось оплатить *",
    "Тип оплаты",
    "Менеджер ",
    "Филиал продажи",
]

PROBLEM_NAMES = {
    1: "1 — Цена есть, платёж не найден",
    2: "2 — Цена = 0, direct-платёж есть",
    3: "3 — Долг без «рассрочки» в названии",
}

PROBLEM_FILLS = {
    1: "FCE8E6",
    2: "FFF2CC",
    3: "DDEBF7",
}

CASE_DEFINITIONS: list[dict[str, Any]] = [
    {
        "problem": 1,
        "contract_id": "00000152421",
        "client_id": "000036065",
        "reason": "Активация в день продажи: бронью или далёким стартом объяснить нельзя.",
        "subscription_ref": "9ADD2FC1387179134A06A5C4EA7636AC",
        "product_ref": "B1B2000C29D830FD11F109745F404A62",
        "sale_number": "00000035172",
        "sale_line": 10990,
        "rg_price": 10990,
        "rg_paid_candidate": 10990,
        "rg_fld5963": 3,
        "direct_documents": "нет",
        "direct_dates": "",
        "direct_amounts": "",
        "direct_sum": 0,
        "accum_charge": 10990,
        "accum_paid": 0,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "Регистр и строка продажи говорят о полной оплате, но нет ни direct-платежа, ни закрывающего RecordKind=0.",
        "question": "Если оплата была, где она хранится в 1С? Если оплаты не было, нужно ли исправить _Fld3072?",
    },
    {
        "problem": 1,
        "contract_id": "00000152393",
        "client_id": "000024933",
        "reason": "Активация на 90 дней позже продажи. Нужно проверить, связано ли это с бронью или отложенным стартом.",
        "subscription_ref": "92709296F0C260784FD1A0A761E427D6",
        "product_ref": "B1B2000C29D830FD11F109745F404A62",
        "sale_number": "00000035131",
        "sale_line": 10990,
        "rg_price": 10990,
        "rg_paid_candidate": 10990,
        "rg_fld5963": 2,
        "direct_documents": "нет",
        "direct_dates": "",
        "direct_amounts": "",
        "direct_sum": 0,
        "accum_charge": 10990,
        "accum_paid": 0,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "При будущей активации денежные поля уже выглядят как полная оплата, но платёж/закрытие отсутствует.",
        "question": "Меняется ли смысл _Fld3072 при отложенном старте? Как отличить оплату от брони?",
    },
    {
        "problem": 1,
        "contract_id": "00000152418",
        "client_id": "000047380",
        "reason": "У одного клиента два одинаковых абонемента. Нужно понять, это две продажи или дубль пакетной операции.",
        "subscription_ref": "9BE006D4B00FEF344C4B7B459E7A0231",
        "product_ref": "B1B2000C29D830FD11F109745F404A62",
        "sale_number": "00000035169",
        "sale_line": 10990,
        "rg_price": 10990,
        "rg_paid_candidate": 10990,
        "rg_fld5963": 4,
        "direct_documents": "нет",
        "direct_dates": "",
        "direct_amounts": "",
        "direct_sum": 0,
        "accum_charge": 10990,
        "accum_paid": 0,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "00000152419",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "Две продажи одного продукта имеют одинаковый разрыв: _Fld3072=10990, direct-платежа нет.",
        "question": "Договоры 00000152418/00000152419 — две реальные продажи или дубль?",
    },
    {
        "problem": 2,
        "contract_id": "00000150339",
        "client_id": "000073892",
        "reason": "Один платёж полностью закрывает продажу; дополнительно есть смена владельца.",
        "subscription_ref": "AA5629D3CAE33BB74B4780BBB795A571",
        "product_ref": "B21C000C29D830FD11F156779FB32F41",
        "sale_number": "00000028625",
        "sale_line": 11990,
        "rg_price": 0,
        "rg_paid_candidate": 0,
        "rg_fld5963": 1,
        "direct_documents": "00000029276",
        "direct_dates": "2026-05-15 17:53:30",
        "direct_amounts": "11990",
        "direct_sum": 11990,
        "accum_charge": 11990,
        "accum_paid": 11990,
        "refund_count": 0,
        "owner_change": "00000056684: 000073894 → 000073892; 2026-06-14 18:53:02",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "Строка продажи, direct-платёж и _AccumRg3305 все дают 11990, но _Fld3070/_Fld3072 остались нулевыми.",
        "question": "Почему переоформление не сохранило/не восстановило цену в _InfoRg3060?",
    },
    {
        "problem": 2,
        "contract_id": "00000134111",
        "client_id": "000027270",
        "reason": "Полная оплата двумя документами 16000+990; один TOP-1 payment недостаточен.",
        "subscription_ref": "9A8B633F84B3FADC4D617BAEB3CB57CD",
        "product_ref": "B069000C29D830FD11F0152EE144CBEF",
        "sale_number": "00000023896",
        "sale_line": 16990,
        "rg_price": 0,
        "rg_paid_candidate": 0,
        "rg_fld5963": 1,
        "direct_documents": "00000023076; 00000023077",
        "direct_dates": "2025-04-12 11:34:27; 2025-04-12 11:34:27",
        "direct_amounts": "16000; 990",
        "direct_sum": 16990,
        "accum_charge": 16990,
        "accum_paid": 16990,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Промышленная)",
        "conclusion": "Два уникальных _Document152 и движения закрывают строку 16990, а регистр выдаёт price=0.",
        "question": "Можно ли массово пересчитать _Fld3070/_Fld3072 из строки продажи и суммы распределённых оплат?",
    },
    {
        "problem": 2,
        "contract_id": "00000140456",
        "client_id": "000001070",
        "reason": "Полная оплата тремя траншами за 3,5 месяца; нулевая цена осталась и после закрытия.",
        "subscription_ref": "A5981F3CADD9643E4DFAFD0B602A7CDD",
        "product_ref": "AA9EA4BF01266AD311E8D82589232D62",
        "sale_number": "00000054176",
        "sale_line": 16990,
        "rg_price": 0,
        "rg_paid_candidate": 0,
        "rg_fld5963": 4,
        "direct_documents": "00000054495; 00000060790; 00000077639",
        "direct_dates": "2025-09-16 10:59:07; 2025-10-17 09:58:51; 2025-12-30 09:55:39",
        "direct_amounts": "8495; 4248; 4247",
        "direct_sum": 16990,
        "accum_charge": 16990,
        "accum_paid": 16990,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "Три direct-платежа и _AccumRg3305 дают полное закрытие 16990; _InfoRg3060 остался нулевым.",
        "question": "Какое событие 1С должно обновлять регистр после каждого транша или полного закрытия?",
    },
    {
        "problem": 3,
        "contract_id": "00000133458",
        "client_id": "000026977",
        "reason": "direct_sum=11993 равен текущему payment_left. Такое же совпадение есть у 197 из 203 договоров группы.",
        "subscription_ref": "97BD7C157B40C06845D4E3F01F687315",
        "product_ref": "B051000C29D830FD11F003F365FABD18",
        "sale_number": "00000019988",
        "sale_line": 15990,
        "rg_price": 15990,
        "rg_paid_candidate": 3997,
        "rg_fld5963": 8,
        "direct_documents": "00000019280; 00000031556",
        "direct_dates": "2025-03-27 15:30:46; 2025-05-23 10:15:45",
        "direct_amounts": "7995; 3998",
        "direct_sum": 11993,
        "accum_charge": 15990,
        "accum_paid": 11993,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Промышленная)",
        "conclusion": "Сейчас paid=3997/left=11993, но direct и RecordKind=0 дают уже оплаченные 11993; вероятно, поля перевёрнуты.",
        "question": "Подтвердите, что paid=11993 и left=3997. Что именно означает _Fld3072?",
    },
    {
        "problem": 3,
        "contract_id": "00000137554",
        "client_id": "000012851",
        "reason": "Оплачена ровно половина. Direct-платёж, регистр и движение дают одинаковую сумму.",
        "subscription_ref": "B1267CB90D0571FC41FD5447D919C8B0",
        "product_ref": "AFC6000C29D830FD11EF9824DEDCBC71",
        "sale_number": "00000041801",
        "sale_line": 10990,
        "rg_price": 10990,
        "rg_paid_candidate": 5495,
        "rg_fld5963": 5,
        "direct_documents": "00000041543",
        "direct_dates": "2025-07-09 18:44:29",
        "direct_amounts": "5495",
        "direct_sum": 5495,
        "accum_charge": 10990,
        "accum_paid": 5495,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "paid=5495/left=5495 выглядит корректно; автоправка всех 203 строк по одному числовому признаку опасна.",
        "question": "Подтвердите, что фактический остаток 5495; отдельно — почему в названии 9 месяцев, а _Fld1481=12.",
    },
    {
        "problem": 3,
        "contract_id": "00000146001",
        "client_id": "000025024",
        "reason": "Один _Document152 на 9900 связан с двумя продажами; регистр распределяет по 4950.",
        "subscription_ref": "918073D1D91D97EB45ECF55880CB0A11",
        "product_ref": "ABFFA4BF01266AD411EC1ADA0CBC08F3",
        "sale_number": "00000005198",
        "sale_line": 9900,
        "rg_price": 9900,
        "rg_paid_candidate": 4950,
        "rg_fld5963": 3,
        "direct_documents": "00000005344 (общий для двух продаж)",
        "direct_dates": "2026-01-27 12:46:48",
        "direct_amounts": "9900",
        "direct_sum": 9900,
        "accum_charge": 9900,
        "accum_paid": 4950,
        "refund_count": 0,
        "owner_change": "",
        "sibling_contract": "00000146003 / продажа 00000005200 / RecordKind=0: 4950",
        "registration_branch": "Фитнес Империя (Гоголевский)",
        "conclusion": "Нельзя присвоить всю _Fld1080=9900 каждой из двух продаж; нужна распределённая сумма _AccumRg3305=4950.",
        "question": "Подтвердите распределение 4950+4950 и укажите, можно ли брать распределённую сумму из _AccumRg3305.",
    },
]


def decimal_value(value: Any) -> Decimal:
    text = str(value or "").strip().replace(" ", "").replace(",", ".")
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def is_not_finished(row: dict[str, str]) -> bool:
    return (row.get("end_date") or "") >= CUTOFF_DATE


def matches_problem(problem: int, row: dict[str, str]) -> bool:
    if problem == 1:
        return (
            row.get("_product_class") == "full_subscription"
            and is_not_finished(row)
            and decimal_value(row.get("price")) > 0
            and (row.get("type_of_payment") or "") == "наличные"
            and not (row.get("_payment_match_source") or "").strip()
        )
    if problem == 2:
        return (
            row.get("_product_class") == "full_subscription"
            and (row.get("activation_date") or "") <= CUTOFF_DATE
            and is_not_finished(row)
            and decimal_value(row.get("price")) == 0
            and (row.get("_payment_match_source") or "").startswith("direct_doc152")
            and decimal_value(row.get("_document131_posted_unmarked_refund_count")) == 0
            and bool((row.get("type_of_payment") or "").strip())
        )
    if problem == 3:
        return (
            is_not_finished(row)
            and decimal_value(row.get("payment_left")) > 0
            and "рассроч" not in (row.get("contract_name") or "").lower()
        )
    raise ValueError(f"Unknown problem: {problem}")


def load_and_validate_source(status_path: Path, staging_path: Path) -> tuple[list[dict[str, str]], list[dict[str, Any]]]:
    status = json.loads(status_path.read_text(encoding="utf-8"))
    contract = status.get("cutoff_contract") or {}
    expected_contract = {
        "backup_finish_at": CUTOFF_AT,
        "cutoff_at": CUTOFF_AT,
        "cutoff_date": CUTOFF_DATE,
        "date_stamp": DATE_STAMP,
    }
    for key, expected in expected_contract.items():
        actual = contract.get(key)
        if actual != expected:
            raise RuntimeError(f"cutoff contract mismatch for {key}: {actual!r} != {expected!r}")
    if "validate" not in (status.get("completed_steps") or []):
        raise RuntimeError("full-cutoff run is not validated")

    with staging_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        missing_headers = [header for header in CLIENT_HEADERS if header not in (reader.fieldnames or [])]
        if missing_headers:
            raise RuntimeError(f"staging is missing Fitbase columns: {missing_headers}")
        rows = list(reader)

    problem_counts = {problem: sum(matches_problem(problem, row) for row in rows) for problem in PROBLEM_NAMES}
    if problem_counts != {1: 10, 2: 41, 3: 203}:
        raise RuntimeError(f"unexpected current problem counts: {problem_counts}")

    problem_contracts = {
        problem: {row.get("contract_id") for row in rows if matches_problem(problem, row)}
        for problem in PROBLEM_NAMES
    }
    union = set().union(*problem_contracts.values())
    if len(union) != 254:
        raise RuntimeError(f"problem contract union must contain 254 rows, got {len(union)}")

    by_contract: dict[str, dict[str, str]] = {}
    duplicates: set[str] = set()
    selected_ids = {case["contract_id"] for case in CASE_DEFINITIONS}
    for row in rows:
        contract_id = row.get("contract_id") or ""
        if contract_id not in selected_ids:
            continue
        if contract_id in by_contract:
            duplicates.add(contract_id)
        by_contract[contract_id] = row
    if duplicates:
        raise RuntimeError(f"selected contract IDs are duplicated in staging: {sorted(duplicates)}")

    selected_rows: list[dict[str, Any]] = []
    selected_problem_counts: Counter[int] = Counter()
    for case in CASE_DEFINITIONS:
        contract_id = case["contract_id"]
        row = by_contract.get(contract_id)
        if row is None:
            raise RuntimeError(f"selected contract is missing from staging: {contract_id}")
        if row.get("client_id") != case["client_id"]:
            raise RuntimeError(
                f"client mismatch for {contract_id}: {row.get('client_id')!r} != {case['client_id']!r}"
            )
        problem = int(case["problem"])
        if not matches_problem(problem, row):
            raise RuntimeError(f"selected contract {contract_id} no longer matches problem {problem}")
        selected_rows.append({"case": case, "row": row})
        selected_problem_counts[problem] += 1
    if selected_problem_counts != Counter({1: 3, 2: 3, 3: 3}):
        raise RuntimeError(f"expected 3 selected cases per problem, got {selected_problem_counts}")
    return rows, selected_rows


def typed_fitbase_value(header: str, value: str) -> Any:
    text = (value or "").strip()
    if not text:
        return None
    if header in {"create_date", "payment_date", "activation_date", "end_date"}:
        return datetime.strptime(text, "%Y-%m-%d").date()
    if header in {"duration", "freeze", "guests", "visits_left", "price", "amount_of_payments", "payment_left"}:
        number = decimal_value(text)
        return int(number) if number == number.to_integral_value() else float(number)
    return text


THIN_GREY = Side(style="thin", color="D9E2F3")
CELL_BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)
NAVY = "17365D"
BLUE = "4472C4"
LIGHT_BLUE = "D9EAF7"
WHITE = "FFFFFF"
GREY = "F2F2F2"


def style_header_row(ws, row_number: int, fill_color: str, font_color: str = WHITE) -> None:
    for cell in ws[row_number]:
        cell.fill = PatternFill("solid", fgColor=fill_color)
        cell.font = Font(name="Arial", size=10, bold=True, color=font_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER


def style_data_area(ws, min_row: int, max_row: int, max_col: int) -> None:
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font = Font(name="Arial", size=10, color="1F1F1F")
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = CELL_BORDER


def configure_print(ws, print_area: str, repeat_rows: str) -> None:
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.print_area = print_area
    ws.print_title_rows = repeat_rows
    ws.sheet_view.showGridLines = False


def build_fitbase_sheet(wb: Workbook, selected_rows: list[dict[str, Any]]) -> None:
    ws = wb.active
    ws.title = "Fitbase_9_cases"
    technical_headers = ["problem", *CLIENT_HEADERS]
    russian_headers = ["Проблема", *CLIENT_RUS_HEADERS]
    ws.append(technical_headers)
    ws.append(russian_headers)

    for item in selected_rows:
        case = item["case"]
        source = item["row"]
        ws.append(
            [
                PROBLEM_NAMES[int(case["problem"])],
                *[typed_fitbase_value(header, source.get(header, "")) for header in CLIENT_HEADERS],
            ]
        )

    max_row = ws.max_row
    max_col = ws.max_column
    style_header_row(ws, 1, NAVY)
    style_header_row(ws, 2, BLUE)
    style_data_area(ws, 3, max_row, max_col)

    for idx, item in enumerate(selected_rows, start=3):
        problem = int(item["case"]["problem"])
        ws.cell(idx, 1).fill = PatternFill("solid", fgColor=PROBLEM_FILLS[problem])
        ws.cell(idx, 1).font = Font(name="Arial", size=10, bold=True, color="1F1F1F")
        ws.row_dimensions[idx].height = 46
    ws.row_dimensions[1].height = 27
    ws.row_dimensions[2].height = 46

    widths = [
        38, 12, 18, 18, 28, 32, 50, 18, 15, 20, 15, 15, 15, 15, 19, 20, 19, 15, 15, 17, 18, 32, 33
    ]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width

    date_headers = {"create_date", "payment_date", "activation_date", "end_date"}
    money_headers = {"price", "amount_of_payments", "payment_left"}
    for column_index, header in enumerate(technical_headers, start=1):
        if header in date_headers:
            for row_index in range(3, max_row + 1):
                ws.cell(row_index, column_index).number_format = "yyyy-mm-dd"
        elif header in money_headers:
            for row_index in range(3, max_row + 1):
                ws.cell(row_index, column_index).number_format = "#,##0"

    ws.freeze_panes = "B3"
    ws.auto_filter.ref = f"A2:{get_column_letter(max_col)}{max_row}"
    ws.sheet_view.zoomScale = 75
    configure_print(ws, f"A1:{get_column_letter(max_col)}{max_row}", "1:2")


DIAGNOSTIC_HEADERS = [
    "Проблема",
    "contract_id",
    "client_id",
    "Клиент",
    "Зачем взят этот кейс",
    "_Document163._IDRRef",
    "_Reference72 product_ref",
    "_Document154._Number",
    "_Document154_VT1137._Fld1160",
    "_InfoRg3060._Fld3070",
    "_InfoRg3060._Fld3072",
    "_InfoRg3060._Fld5963",
    "_Document152: документы",
    "_Document152: даты",
    "_Document152._Fld1080: суммы",
    "Сумма уникальных direct-платежей",
    "_AccumRg3305 RecordKind=1: начислено",
    "_AccumRg3305 RecordKind=0: оплачено/закрыто",
    "_Document131: возвраты",
    "_Document138: переоформление",
    "Связанный договор/продажа",
    "Филиал регистрации клиента",
    "Филиал продажи (сырой)",
    "Филиал в Fitbase",
    "Что показывает кейс",
    "Что проверить",
]


def build_diagnostic_sheet(wb: Workbook, selected_rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Диагностика_1С")
    ws.append(DIAGNOSTIC_HEADERS)
    for item in selected_rows:
        case = item["case"]
        source = item["row"]
        ws.append(
            [
                PROBLEM_NAMES[int(case["problem"])],
                case["contract_id"],
                case["client_id"],
                source.get("client_fio", ""),
                case["reason"],
                case["subscription_ref"],
                case["product_ref"],
                case["sale_number"],
                case["sale_line"],
                case["rg_price"],
                case["rg_paid_candidate"],
                case["rg_fld5963"],
                case["direct_documents"],
                case["direct_dates"],
                case["direct_amounts"],
                case["direct_sum"],
                case["accum_charge"],
                case["accum_paid"],
                case["refund_count"],
                case["owner_change"],
                case["sibling_contract"],
                case["registration_branch"],
                source.get("_sale_branch_raw", ""),
                source.get("филиал", ""),
                case["conclusion"],
                case["question"],
            ]
        )

    style_header_row(ws, 1, NAVY)
    style_data_area(ws, 2, ws.max_row, ws.max_column)
    for row_index, item in enumerate(selected_rows, start=2):
        problem = int(item["case"]["problem"])
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=PROBLEM_FILLS[problem])
        ws.cell(row_index, 1).font = Font(name="Arial", size=10, bold=True)
        ws.row_dimensions[row_index].height = 84
    ws.row_dimensions[1].height = 58

    widths = [
        38, 18, 18, 32, 50, 35, 35, 20, 20, 18, 18, 17, 35, 35, 28, 22, 22, 24, 18, 45, 45, 33, 33, 33, 60, 60
    ]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for column_index in range(9, 19):
        if column_index != 12:
            for row_index in range(2, ws.max_row + 1):
                ws.cell(row_index, column_index).number_format = "#,##0"

    ws.freeze_panes = "D2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    ws.sheet_view.zoomScale = 70
    configure_print(ws, f"A1:{get_column_letter(ws.max_column)}{ws.max_row}", "1:1")


def add_note_row(ws, row: int, section: str, key: str, value: str, source: str = "") -> int:
    values = [section, key, value, source]
    for column, cell_value in enumerate(values, start=1):
        cell = ws.cell(row, column, cell_value)
        cell.border = CELL_BORDER
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.font = Font(name="Arial", size=10, color="1F1F1F")
    return row + 1


def build_notes_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("Памятка")
    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "9 тестовых кейсов: разбор денег абонементов 1С → Fitbase"
    title.fill = PatternFill("solid", fgColor=NAVY)
    title.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 34

    ws.append(["Раздел", "Параметр", "Значение / правило", "Источник"])
    style_header_row(ws, 2, BLUE)

    rows = [
        ("Срез", "backup", "data/Fitnes-30-06-26.bak", "RESTORE HEADERONLY"),
        ("Срез", "backup_finish_at = cutoff_at", CUTOFF_AT, "end-to-end-xlsx/work/20260630_full_cutoff/status.json"),
        ("Срез", "cutoff_date / date_stamp", f"{CUTOFF_DATE} / {DATE_STAMP}", "status.json"),
        ("Объём", "Актуальные problem-группы", "10 + 41 + 203 = 254 договора", "membership_import_rows.csv"),
        ("Объём", "Почему не 223", "223 = 3 + 41 + 179 относились к архивной смешанной сборке с майскими membership facts. На едином срезе 30 июня стало 254.", "README.md"),
        ("Проблема 1", "Фильтр", "full_subscription; end_date>=cutoff; price>0; type_of_payment=наличные; matched payment пуст.", "01_problem_no_payment_cash.md"),
        ("Проблема 1", "Расхождение", "«наличные» ставит скрипт, а не 1С. _Fld3072 выглядит как полная оплата, но Document152 и RecordKind=0 нет.", "01_problem_no_payment_cash.md"),
        ("Проблема 2", "Фильтр", "full_subscription; active; price=0; direct Document152 есть; проведённого непомеченного возврата нет.", "02_problem_zero_price_direct.md"),
        ("Проблема 2", "Расхождение", "Строка продажи и сумма уникальных direct-платежей равны, а _InfoRg3060._Fld3070/_Fld3072=0.", "02_problem_zero_price_direct.md"),
        ("Проблема 3", "Фильтр", "end_date>=cutoff; payment_left>0; в названии нет «рассроч».", "03_problem_non_named_payment_left.md"),
        ("Проблема 3", "Расхождение", "У 197 из 203 сумма direct-платежей равна текущему payment_left, но есть корректные 50%-кейсы и общие платежи.", "03_problem_non_named_payment_left.md"),
        ("Поля", "price / paid / left", "Сейчас price=_Fld3070; paid=min(_Fld3072,price) при _Fld3072>0; left=max(price-paid,0). Direct и _AccumRg3305 не являются основным денежным источником.", "04_field_mapping_1c_to_fitbase.md"),
        ("Поля", "payment_date", "Берётся _Document163._Date_Time, а не дата _Document152.", "04_field_mapping_1c_to_fitbase.md"),
        ("Филиалы", "Филиал в Fitbase", "Филиал продажи _Document154._Fld1116RRef; fallback — клуб _Document163. Это не филиал регистрации клиента.", "04_field_mapping_1c_to_fitbase.md"),
        ("Результат", "1", "По каждому contract_id указать: правильная цена, сколько оплачено, сколько осталось, тип оплаты и активен ли абонемент.", "README.md"),
        ("Результат", "2", "Указать таблицу и поле, из которых нужно брать цену и распределённую оплату.", "README.md"),
        ("Результат", "3", "Если неверны данные 1С, их нужно исправить и сделать новый backup. Если данные верны, мы поменяем правило выгрузки.", "README.md"),
    ]
    row_number = 3
    for section, key, value, source in rows:
        row_number = add_note_row(ws, row_number, section, key, value, source)

    current_section = None
    section_colors = {
        "Срез": LIGHT_BLUE,
        "Объём": GREY,
        "Проблема 1": PROBLEM_FILLS[1],
        "Проблема 2": PROBLEM_FILLS[2],
        "Проблема 3": PROBLEM_FILLS[3],
        "Поля": LIGHT_BLUE,
        "Филиалы": GREY,
        "Результат": "E2F0D9",
    }
    for row_index in range(3, ws.max_row + 1):
        section = ws.cell(row_index, 1).value
        if section != current_section:
            current_section = section
        fill = section_colors.get(str(section), GREY)
        ws.cell(row_index, 1).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row_index, 1).font = Font(name="Arial", size=10, bold=True)
        ws.cell(row_index, 2).font = Font(name="Arial", size=10, bold=True)
        ws.row_dimensions[row_index].height = 48

    for index, width in enumerate([18, 30, 105, 55], start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:D{ws.max_row}"
    ws.sheet_view.zoomScale = 85
    configure_print(ws, f"A1:D{ws.max_row}", "1:2")


def validate_output(path: Path, selected_rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(path, data_only=False)
    if wb.sheetnames != ["Fitbase_9_cases", "Диагностика_1С", "Памятка"]:
        raise RuntimeError(f"unexpected output sheets: {wb.sheetnames}")

    main = wb["Fitbase_9_cases"]
    if (main.max_row, main.max_column) != (11, 23):
        raise RuntimeError(f"unexpected main dimensions: {main.max_row}x{main.max_column}")
    if [cell.value for cell in main[1]] != ["problem", *CLIENT_HEADERS]:
        raise RuntimeError("technical header row mismatch")
    if [cell.value for cell in main[2]] != ["Проблема", *CLIENT_RUS_HEADERS]:
        raise RuntimeError("Russian header row mismatch")

    actual_contracts = [main.cell(row, 3).value for row in range(3, 12)]
    expected_contracts = [item["case"]["contract_id"] for item in selected_rows]
    if actual_contracts != expected_contracts:
        raise RuntimeError(f"contract order mismatch: {actual_contracts} != {expected_contracts}")
    if len(set(actual_contracts)) != 9:
        raise RuntimeError("main sheet must contain nine unique contracts")

    diag = wb["Диагностика_1С"]
    if (diag.max_row, diag.max_column) != (10, len(DIAGNOSTIC_HEADERS)):
        raise RuntimeError(f"unexpected diagnostic dimensions: {diag.max_row}x{diag.max_column}")
    if [cell.value for cell in diag[1]] != DIAGNOSTIC_HEADERS:
        raise RuntimeError("diagnostic headers mismatch")

    for sheet in wb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    raise RuntimeError(f"unexpected formula in {sheet.title}!{cell.coordinate}")
    wb.close()


def build(output_path: Path, status_path: Path, staging_path: Path) -> None:
    _, selected_rows = load_and_validate_source(status_path, staging_path)
    wb = Workbook()
    wb.properties.title = "9 тестовых кейсов 1С → Fitbase на 2026-06-30"
    wb.properties.subject = "Разбор трёх групп денежных проблем абонементов"
    wb.properties.creator = "Codex"
    wb.properties.description = f"cutoff_at={CUTOFF_AT}; 3 кейса на каждую из 3 проблем"
    build_fitbase_sheet(wb, selected_rows)
    build_diagnostic_sheet(wb, selected_rows)
    build_notes_sheet(wb)
    wb.active = 0

    output_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = output_path.with_name(f".{output_path.name}.tmp.xlsx")
    wb.save(temporary_path)
    wb.close()
    temporary_path.replace(output_path)
    validate_output(output_path, selected_rows)
    print(f"PASS: {output_path}")
    print("main_sheet: 9 rows, 23 columns (problem + 22 Fitbase columns)")
    print("diagnostic_sheet: 9 rows")
    print(f"cutoff_at: {CUTOFF_AT}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--status", type=Path, default=STATUS_PATH)
    parser.add_argument("--staging", type=Path, default=STAGING_PATH)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build(args.output.resolve(), args.status.resolve(), args.staging.resolve())


if __name__ == "__main__":
    main()
