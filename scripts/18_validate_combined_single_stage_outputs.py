#!/usr/bin/env python3
"""Validate combined Part 2 single-stage Fitbase XLSX outputs."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAIN_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
]
MAIN_RUS_HEADERS = [
    "Внутренний номер клиента ",
    "Телефон *",
    "ФИО клиента *",
    "Почта",
    "Воронка *",
    "Этап воронки *",
    "Бюджет ",
    "Дата создания *",
    "Менеджер ",
]
CARD_HEADERS = ["телефон", "фио", "номер пластиковой карты"]
ALLOWED_FINAL_PAIRS = {
    ("новые заявки", "неразобранные"),
    ("Действующие абонементы", "Все действующие абонементы"),
    ("Реактивация(годовые абонементы)", "Все закрытые абонементы"),
}
OLD_STEPS = {
    "60-31 день до окончания",
    "30-8 дней до окончания",
    "7-0 день до окончания",
    "Действующие клиенты",
    "1-6 дней",
    "7-29 дней",
    "30-59 дней",
    "60-89 дней",
    "более 90 дней",
    "Неразобранные",
}
REQUIRED_REPORTS = [
    "funnel_distribution.csv",
    "stage_distribution_by_funnel.csv",
    "fitbase_funnel_distribution.csv",
    "single_stage_distribution.csv",
    "manager_distribution_by_club.csv",
    "missing_phone_report.csv",
    "missing_card_report.csv",
    "missing_club_report.csv",
    "multiple_subscriptions_report.csv",
    "subscription_selection_report.csv",
    "subscription_overrides_report.csv",
    "multiple_cards_report.csv",
    "card_selection_report.csv",
    "product_classification_preflight.csv",
    "product_classification_report.csv",
    "product_classification_review_report.csv",
    "product_reclassification_impact.md",
    "product_reclassification_applied.csv",
    "product_reclassification_funnel_impact.csv",
]


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def workbook_rows(path: Path, first_data_row: int, width: int) -> tuple[list[object], list[tuple[object, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, width + 1)]
    rows = [
        tuple(row)
        for row in ws.iter_rows(min_row=first_data_row, max_col=width, values_only=True)
        if any(value not in (None, "") for value in row)
    ]
    wb.close()
    return headers, rows


def workbook_row_values(path: Path, row_number: int, width: int) -> list[object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = [ws.cell(row_number, col).value for col in range(1, width + 1)]
    wb.close()
    return values


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(float(str(value)))
    except ValueError:
        return default


def validate(args: argparse.Namespace) -> int:
    stage_dir = as_abs(args.stage_dir)
    output_dir = as_abs(args.output_dir)
    reports_dir = as_abs(args.reports_dir)
    date_stamp = args.date_stamp or args.cutoff_date.replace("-", "")

    errors: list[str] = []
    warnings: list[str] = []
    stage_path = stage_dir / "final_funnel_clients.csv"
    rows = read_csv(stage_path) if stage_path.exists() else []
    if not rows:
        errors.append(f"missing or empty final stage CSV: {stage_path.relative_to(ROOT)}")

    main_path = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}__all_funnels.xlsx"
    cards_path = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}__all_funnels.xlsx"
    xlsx_files = sorted(output_dir.glob("*.xlsx"))
    expected_xlsx = {main_path, cards_path}
    if set(xlsx_files) != expected_xlsx:
        errors.append(
            "final output directory must contain exactly two XLSX files: "
            f"found {[path.name for path in xlsx_files]}"
        )

    main_rows: list[tuple[object, ...]] = []
    card_rows: list[tuple[object, ...]] = []
    if main_path.exists():
        headers, main_rows = workbook_rows(main_path, 3, len(MAIN_HEADERS))
        if headers != MAIN_HEADERS:
            errors.append(f"main technical headers mismatch: {main_path.relative_to(ROOT)}")
        if workbook_row_values(main_path, 2, len(MAIN_RUS_HEADERS)) != MAIN_RUS_HEADERS:
            errors.append(f"main Russian headers mismatch: {main_path.relative_to(ROOT)}")
    else:
        errors.append(f"missing main XLSX: {main_path.relative_to(ROOT)}")

    if cards_path.exists():
        headers, card_rows = workbook_rows(cards_path, 2, len(CARD_HEADERS))
        if headers != CARD_HEADERS:
            errors.append(f"cards headers mismatch: {cards_path.relative_to(ROOT)}")
    else:
        errors.append(f"missing cards XLSX: {cards_path.relative_to(ROOT)}")

    if len(main_rows) != len(rows):
        errors.append(f"main row count mismatch: xlsx={len(main_rows)}, final_funnel_clients.csv={len(rows)}")
    if len(card_rows) != len(rows):
        errors.append(f"cards row count mismatch: xlsx={len(card_rows)}, final_funnel_clients.csv={len(rows)}")

    final_pairs = Counter((str(row[4] or ""), str(row[5] or "")) for row in main_rows)
    invalid_pairs = sorted(set(final_pairs) - ALLOWED_FINAL_PAIRS)
    if invalid_pairs:
        errors.append(f"invalid final funnel/funnel_step pairs: {invalid_pairs}")

    final_funnels = {pair[0] for pair in final_pairs}
    final_steps = {pair[1] for pair in final_pairs}
    if final_funnels != {pair[0] for pair in ALLOWED_FINAL_PAIRS}:
        errors.append(f"final XLSX funnels are not exactly the expected three values: {sorted(final_funnels)}")
    if final_steps != {pair[1] for pair in ALLOWED_FINAL_PAIRS}:
        errors.append(f"final XLSX steps are not exactly the expected three values: {sorted(final_steps)}")
    old_steps_found = sorted(step for step in final_steps if step in OLD_STEPS)
    if old_steps_found:
        errors.append(f"old multi-stage names found in final XLSX: {old_steps_found}")

    stage_refs = [row.get("client_ref", "") for row in rows if row.get("client_ref")]
    duplicate_refs = [key for key, count in Counter(stage_refs).items() if count > 1]
    if duplicate_refs:
        errors.append(f"duplicate client_ref rows in stage: {len(duplicate_refs)}")

    xlsx_client_ids = [str(row[0] or "") for row in main_rows if row and row[0] not in (None, "")]
    duplicate_xlsx_ids = [key for key, count in Counter(xlsx_client_ids).items() if count > 1]
    if duplicate_xlsx_ids:
        errors.append(f"duplicate client_id rows in final XLSX: {len(duplicate_xlsx_ids)}")

    comma_cards = sum(1 for row in card_rows if len(row) >= 3 and isinstance(row[2], str) and "," in row[2])
    if comma_cards:
        errors.append(f"card XLSX contains comma-separated card values: {comma_cards}")

    for report in REQUIRED_REPORTS:
        if not (reports_dir / report).exists():
            errors.append(f"missing required report: {(reports_dir / report).relative_to(ROOT)}")

    missing_phone = read_csv(reports_dir / "missing_phone_report.csv") if (reports_dir / "missing_phone_report.csv").exists() else []
    missing_card = read_csv(reports_dir / "missing_card_report.csv") if (reports_dir / "missing_card_report.csv").exists() else []
    missing_club = read_csv(reports_dir / "missing_club_report.csv") if (reports_dir / "missing_club_report.csv").exists() else []
    multiple_subs = read_csv(reports_dir / "multiple_subscriptions_report.csv") if (reports_dir / "multiple_subscriptions_report.csv").exists() else []
    card_selection = read_csv(reports_dir / "card_selection_report.csv") if (reports_dir / "card_selection_report.csv").exists() else []
    single_stage = read_csv(reports_dir / "single_stage_distribution.csv") if (reports_dir / "single_stage_distribution.csv").exists() else []

    if len(missing_phone) != sum(1 for row in rows if not (row.get("phones") or "").strip()):
        errors.append("missing_phone_report.csv count mismatch")
    if len(missing_card) != sum(1 for row in rows if not (row.get("selected_card_number") or "").strip()):
        errors.append("missing_card_report.csv count mismatch")
    if len(missing_club) != sum(1 for row in rows if not (row.get("normalized_club") or "").strip()):
        errors.append("missing_club_report.csv count mismatch")
    expected_multiple_subs = sum(
        1
        for row in rows
        if (
            row.get("funnel") == "Действующие клиенты"
            and int_value(row.get("active_full_subscription_count")) > 1
        )
        or (row.get("funnel") == "Реактивация" and int_value(row.get("finished_full_subscription_count")) > 1)
    )
    if len(multiple_subs) != expected_multiple_subs:
        errors.append("multiple_subscriptions_report.csv count mismatch")
    if len(card_selection) != len(rows):
        errors.append("card_selection_report.csv should contain one row per final client")

    reported_pairs = {
        (row.get("funnel", ""), row.get("funnel_step", "")): int_value(row.get("clients"))
        for row in single_stage
    }
    if reported_pairs != dict(final_pairs):
        errors.append("single_stage_distribution.csv does not match final XLSX distribution")

    review_rows = read_csv(reports_dir / "product_classification_review_report.csv") if (reports_dir / "product_classification_review_report.csv").exists() else []
    if review_rows:
        warnings.append(f"product classification rows needing business review: {len(review_rows)}")
    if missing_phone:
        warnings.append(f"clients without phone exported and reported: {len(missing_phone)}")
    if missing_card:
        warnings.append(f"clients without selected card exported and reported: {len(missing_card)}")
    if multiple_subs:
        warnings.append(f"clients with multiple subscription candidates reported: {len(multiple_subs)}")

    verdict = "PASS" if not errors else "FAIL"
    lines = [
        "# Part 2 Combined Single-Stage Validation Report",
        "",
        f"Run date: `{datetime.now().isoformat(timespec='seconds')}`",
        f"cutoff_date: `{args.cutoff_date}`",
        f"date_stamp: `{date_stamp}`",
        f"stage_rows: `{len(rows)}`",
        f"main_xlsx_rows: `{len(main_rows)}`",
        f"cards_xlsx_rows: `{len(card_rows)}`",
        "",
        "## Verdict",
        "",
        f"`{verdict}`",
        "",
        "## Final Single-Stage Distribution",
        "",
    ]
    for (funnel, step), count in final_pairs.most_common():
        lines.append(f"- `{funnel}` / `{step}`: `{count}`")
    lines.extend(["", "## Data Quality Counts", ""])
    lines.extend(
        [
            f"- missing_phone: `{len(missing_phone)}`",
            f"- missing_card: `{len(missing_card)}`",
            f"- missing_club: `{len(missing_club)}`",
            f"- multiple_subscription_clients: `{len(multiple_subs)}`",
            f"- product_review_rows: `{len(review_rows)}`",
        ]
    )
    lines.extend(["", "## Errors", ""])
    lines.extend([f"- {error}" for error in errors] if errors else ["None."])
    if warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in warnings)
    lines.append("")

    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "validation_report.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"verdict={verdict}")
    print(f"errors={len(errors)}")
    print(f"warnings={len(warnings)}")
    print(f"report={(reports_dir / 'validation_report.md').relative_to(ROOT)}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default="2026-05-25")
    parser.add_argument("--date-stamp", default="")
    parser.add_argument("--stage-dir", default="output/part2_20260525_0800_final/staging")
    parser.add_argument("--output-dir", default="output/part2_20260525_0800_final_combined")
    parser.add_argument("--reports-dir", default="output/part2_20260525_0800_final/reports")
    parser.add_argument("--main-template", default="task-desc/Копия Импорт_заявки.xlsx")
    parser.add_argument("--cards-template", default="task-desc/Пластиковая карта.xlsx")
    return parser.parse_args()


def main() -> None:
    raise SystemExit(validate(parse_args()))


if __name__ == "__main__":
    main()
