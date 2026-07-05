#!/usr/bin/env python3
"""Compare the previous May XLSX delivery with the 2026-06-30 rebuild."""

from __future__ import annotations

import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "output" / "20260630_xlsx_comparison"
DOC_PATH = ROOT / "docs" / "new-backup-30-06" / "06_xlsx_comparison.md"

OLD_BACKUP_DATE = date(2026, 5, 23)
OLD_CUTOFF_DATE = date(2026, 5, 25)
NEW_CUTOFF_DATE = date(2026, 6, 30)


@dataclass(frozen=True)
class WorkbookSpec:
    key: str
    title: str
    old_path: Path
    new_path: Path
    first_data_row: int
    key_fields: tuple[str, ...]
    client_field: str | None = None


SPECS = [
    WorkbookSpec(
        key="import_zayavki",
        title="Импорт заявок",
        old_path=ROOT / "output/20251115_0800_fix_owner/fitbase_active_clients_import_zayavki_20260525_0800_all_funnels.xlsx",
        new_path=ROOT / "output/20260630_fix_owner/fitbase_active_clients_import_zayavki_20260630_all_funnels.xlsx",
        first_data_row=3,
        key_fields=("client_id",),
        client_field="client_id",
    ),
    WorkbookSpec(
        key="plastic_cards",
        title="Пластиковые карты",
        old_path=ROOT / "output/20251115_0800_fix_owner/fitbase_active_clients_plastic_cards_20260525_0800_all_funnels.xlsx",
        new_path=ROOT / "output/20260630_fix_owner/fitbase_active_clients_plastic_cards_20260630_all_funnels.xlsx",
        first_data_row=2,
        key_fields=("телефон", "фио"),
        client_field=None,
    ),
    WorkbookSpec(
        key="membership_clients",
        title="Абонементы клиентов",
        old_path=ROOT / "output/20251115_0800_fix_owner_new_import/fitbase_import_abonementy_clientov_20260525_0800.xlsx",
        new_path=ROOT / "output/20260630_fix_owner_new_import/fitbase_import_abonementy_clientov_20260630.xlsx",
        first_data_row=3,
        key_fields=("contract_id",),
        client_field="client_id",
    ),
    WorkbookSpec(
        key="membership_templates",
        title="Шаблоны абонементов",
        old_path=ROOT / "output/20251115_0800_fix_owner_new_import/fitbase_import_shablony_abonementov_20260525_0800.xlsx",
        new_path=ROOT / "output/20260630_fix_owner_new_import/fitbase_import_shablony_abonementov_20260630.xlsx",
        first_data_row=3,
        key_fields=("name",),
    ),
    WorkbookSpec(
        key="service_clients",
        title="Услуги клиентов",
        old_path=ROOT / "output/20251115_0800_fix_owner_new_import/fitbase_import_uslugi_clientov_20260525_0800.xlsx",
        new_path=ROOT / "output/20260630_fix_owner_new_import/fitbase_import_uslugi_clientov_20260630.xlsx",
        first_data_row=3,
        key_fields=("service_id",),
        client_field="client_id",
    ),
    WorkbookSpec(
        key="service_templates",
        title="Шаблоны услуг",
        old_path=ROOT / "output/20251115_0800_fix_owner_new_import/fitbase_import_shablony_uslug_20260525_0800.xlsx",
        new_path=ROOT / "output/20260630_fix_owner_new_import/fitbase_import_shablony_uslug_20260630.xlsx",
        first_data_row=3,
        key_fields=("name",),
    ),
]


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == datetime.min.time():
            return value.strftime("%Y-%m-%d")
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_phone(value: str) -> str:
    digits = re.sub(r"\D+", "", value or "")
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return digits


def normalize_key_value(field: str, value: str) -> str:
    if field.lower() in {"телефон", "phone"}:
        return normalize_phone(value)
    return value.strip()


def parse_date(value: str | None) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        return datetime.strptime(text[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def read_xlsx(spec: WorkbookSpec, path: Path) -> tuple[list[str], list[dict[str, str]], Counter[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell_text(ws.cell(1, col).value) for col in range(1, ws.max_column + 1)]
    rows: list[dict[str, str]] = []
    duplicate_keys: Counter[str] = Counter()
    seen: set[str] = set()

    for values in ws.iter_rows(min_row=spec.first_data_row, max_col=len(headers), values_only=True):
        if not any(value not in (None, "") for value in values):
            continue
        row = {headers[idx]: cell_text(value) for idx, value in enumerate(values)}
        key = make_key(spec, row)
        if key in seen:
            duplicate_keys[key] += 1
        seen.add(key)
        rows.append(row)
    wb.close()
    return headers, rows, duplicate_keys


def make_key(spec: WorkbookSpec, row: dict[str, str]) -> str:
    return "||".join(normalize_key_value(field, row.get(field, "")) for field in spec.key_fields)


def rows_by_key(spec: WorkbookSpec, rows: list[dict[str, str]]) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    for row in rows:
        key = make_key(spec, row)
        result.setdefault(key, row)
    return result


def read_csv_by_key(path: Path, key_field: str) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return {row.get(key_field, ""): row for row in csv.DictReader(handle)}


def read_owner_change_clients(path: Path) -> set[str]:
    clients: set[str] = set()
    if not path.exists():
        return clients
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            for field in ("old_client_id", "new_client_id"):
                value = (row.get(field) or "").strip()
                if value:
                    clients.add(value)
    return clients


def read_final_stage_by_plastic_key(path: Path) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    if not path.exists():
        return result
    with path.open(encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            key = normalize_phone(row.get("phones", "")) + "||" + (row.get("client_fio") or "").strip()
            if not key:
                continue
            current = result.get(key)
            current_score = (
                int((current or {}).get("funnel") == "Действующие клиенты"),
                int(bool((current or {}).get("selected_card_number"))),
            )
            row_score = (
                int(row.get("funnel") == "Действующие клиенты"),
                int(bool(row.get("selected_card_number"))),
            )
            if current is None or row_score > current_score:
                result[key] = row
    return result


def changed_columns(headers: list[str], old_row: dict[str, str], new_row: dict[str, str]) -> list[str]:
    return [header for header in headers if old_row.get(header, "") != new_row.get(header, "")]


def date_fields(row: dict[str, str]) -> list[str]:
    return [field for field in row if "date" in field.lower() or field.startswith("Дата ")]


def has_date_after(row: dict[str, str], threshold: date) -> bool:
    for field in date_fields(row):
        parsed = parse_date(row.get(field))
        if parsed and parsed > threshold:
            return True
    return False


def any_date_changed(old_row: dict[str, str], new_row: dict[str, str]) -> bool:
    for field in set(date_fields(old_row)) | set(date_fields(new_row)):
        if old_row.get(field, "") != new_row.get(field, ""):
            return True
    return False


def classify_change(
    spec: WorkbookSpec,
    key: str,
    old_row: dict[str, str],
    new_row: dict[str, str],
    diff_cols: list[str],
    old_stage_final: dict[str, dict[str, str]],
    new_stage_final: dict[str, dict[str, str]],
    old_membership_rows: dict[str, dict[str, str]],
    new_membership_rows: dict[str, dict[str, str]],
    old_plastic_stage: dict[str, dict[str, str]],
    new_plastic_stage: dict[str, dict[str, str]],
    owner_change_clients: set[str],
) -> tuple[list[str], bool]:
    reasons: list[str] = []
    suspicious = False

    if has_date_after(new_row, OLD_BACKUP_DATE):
        reasons.append("new_export_date_after_2026-05-23")
    if any_date_changed(old_row, new_row):
        reasons.append("export_date_field_changed")

    if spec.key == "import_zayavki":
        client_id = old_row.get("client_id") or new_row.get("client_id") or key
        old_stage = old_stage_final.get(client_id, {})
        new_stage = new_stage_final.get(client_id, {})
        if client_id in owner_change_clients:
            reasons.append("owner_change_client")
        if new_stage:
            if parse_date(new_stage.get("create_date")) and parse_date(new_stage.get("create_date")) > OLD_BACKUP_DATE:
                reasons.append("client_created_after_2026-05-23")
            if parse_date(new_stage.get("selected_subscription_sale_date")) and parse_date(new_stage.get("selected_subscription_sale_date")) > OLD_BACKUP_DATE:
                reasons.append("selected_subscription_sale_after_2026-05-23")
        if old_stage and new_stage:
            if old_stage.get("selected_subscription_ref") != new_stage.get("selected_subscription_ref"):
                reasons.append("selected_subscription_changed")
            old_end = parse_date(old_stage.get("selected_subscription_end_date"))
            if (
                old_stage.get("selected_subscription_ref") == new_stage.get("selected_subscription_ref")
                and old_end
                and OLD_CUTOFF_DATE < old_end <= NEW_CUTOFF_DATE
            ):
                reasons.append("same_subscription_crossed_cutoff_window")
            exported_source_fields = [
                "phones",
                "client_fio",
                "email",
                "funnel",
                "funnel_step",
                "budget",
                "create_date",
                "manager",
                "normalized_club",
                "selected_card_number",
                "selected_subscription_ref",
            ]
            if any(old_stage.get(field, "") != new_stage.get(field, "") for field in exported_source_fields):
                reasons.append("stage_export_source_changed")
        if not reasons:
            suspicious = True

    elif spec.key == "plastic_cards":
        old_stage = old_plastic_stage.get(key, {})
        new_stage = new_plastic_stage.get(key, {})
        if old_stage or new_stage:
            if old_stage.get("selected_card_number", "") != new_stage.get("selected_card_number", ""):
                reasons.append("selected_card_changed_in_stage")
            if old_stage.get("selected_subscription_ref", "") != new_stage.get("selected_subscription_ref", ""):
                reasons.append("selected_subscription_changed")
            client_id = old_stage.get("client_id") or new_stage.get("client_id")
            if client_id in owner_change_clients:
                reasons.append("owner_change_client")
        if not reasons:
            suspicious = True

    elif spec.key == "membership_clients":
        old_stage = old_membership_rows.get(key, {})
        new_stage = new_membership_rows.get(key, {})
        if old_stage.get("_owner_change_ref") or new_stage.get("_owner_change_ref"):
            reasons.append("owner_change_membership")
        if any(column in diff_cols for column in ("price", "amount_of_payments", "payment_left", "type_of_payment")):
            reasons.append("money_fields_changed")
        if any(column in diff_cols for column in ("phone", "client_fio", "card", "manager")):
            reasons.append("client_or_card_fields_changed")
        if old_stage and new_stage:
            export_headers = [
                "contract_id",
                "client_id",
                "phone",
                "client_fio",
                "contract_name",
                "card",
                "duration",
                "duration_type",
                "create_date",
                "payment_date",
                "activation_date",
                "end_date",
                "freeze",
                "guests",
                "visits_left",
                "price",
                "amount_of_payments",
                "payment_left",
                "type_of_payment",
                "manager",
            ]
            if all(old_stage.get(field, "") == new_stage.get(field, "") for field in export_headers):
                reasons.append("membership_source_export_fields_unchanged")
                suspicious = True
            else:
                reasons.append("membership_source_export_fields_changed")
        if not reasons:
            suspicious = True

    elif spec.key == "service_clients":
        if any(column in diff_cols for column in ("price", "amount_of_payment", "payment_left", "type_of_payment")):
            reasons.append("service_money_fields_changed")
        if any(column in diff_cols for column in ("phone", "client_fio", "manager")):
            reasons.append("service_client_fields_changed")
        if any(column in diff_cols for column in ("count", "visits_left")):
            reasons.append("service_balance_fields_changed")
        if not reasons:
            suspicious = True

    elif "templates" in spec.key:
        if any(column in diff_cols for column in ("price", "duration", "visits", "freeze")):
            reasons.append("template_business_values_changed")
        if not reasons:
            suspicious = True

    else:
        if not reasons:
            suspicious = True

    return sorted(set(reasons)), suspicious


def write_rows(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    old_stage_final = read_csv_by_key(ROOT / "output/20251115_0800_fix_owner/staging/final_funnel_clients.csv", "client_id")
    new_stage_final = read_csv_by_key(ROOT / "output/20260630_fix_owner/staging/final_funnel_clients.csv", "client_id")
    old_membership_rows = read_csv_by_key(ROOT / "output/20251115_0800_fix_owner_new_import/staging/membership_import_rows.csv", "contract_id")
    new_membership_rows = read_csv_by_key(ROOT / "output/20260630_fix_owner_new_import/staging/membership_import_rows.csv", "contract_id")
    old_plastic_stage = read_final_stage_by_plastic_key(ROOT / "output/20251115_0800_fix_owner/staging/final_funnel_clients.csv")
    new_plastic_stage = read_final_stage_by_plastic_key(ROOT / "output/20260630_fix_owner/staging/final_funnel_clients.csv")
    owner_change_clients = read_owner_change_clients(ROOT / "output/20260630_fix_owner/staging/stg_membership_owner_changes.csv")

    summaries: list[dict[str, Any]] = []
    detail_paths: dict[str, dict[str, str]] = {}

    for spec in SPECS:
        old_headers, old_rows, old_dupes = read_xlsx(spec, spec.old_path)
        new_headers, new_rows, new_dupes = read_xlsx(spec, spec.new_path)
        headers = [header for header in old_headers if header in set(new_headers)]
        old_map = rows_by_key(spec, old_rows)
        new_map = rows_by_key(spec, new_rows)

        old_keys = set(old_map)
        new_keys = set(new_map)
        common_keys = old_keys & new_keys
        added_keys = sorted(new_keys - old_keys)
        removed_keys = sorted(old_keys - new_keys)

        changed_details: list[dict[str, Any]] = []
        field_counts: Counter[str] = Counter()
        reason_counts: Counter[str] = Counter()
        suspicious_count = 0

        for key in sorted(common_keys):
            old_row = old_map[key]
            new_row = new_map[key]
            diffs = changed_columns(headers, old_row, new_row)
            if not diffs:
                continue
            for column in diffs:
                field_counts[column] += 1
            reasons, suspicious = classify_change(
                spec,
                key,
                old_row,
                new_row,
                diffs,
                old_stage_final,
                new_stage_final,
                old_membership_rows,
                new_membership_rows,
                old_plastic_stage,
                new_plastic_stage,
                owner_change_clients,
            )
            for reason in reasons:
                reason_counts[reason] += 1
            if suspicious:
                suspicious_count += 1
            changed_details.append(
                {
                    "key": key,
                    "client_id": old_row.get("client_id") or new_row.get("client_id") or "",
                    "changed_columns": ";".join(diffs),
                    "reason_flags": ";".join(reasons),
                    "suspicious": "1" if suspicious else "0",
                    "old_values": json.dumps({field: old_row.get(field, "") for field in diffs}, ensure_ascii=False),
                    "new_values": json.dumps({field: new_row.get(field, "") for field in diffs}, ensure_ascii=False),
                }
            )

        added_rows = [
            {
                "key": key,
                "client_id": new_map[key].get("client_id", ""),
                "date_after_2026_05_23": "1" if has_date_after(new_map[key], OLD_BACKUP_DATE) else "0",
                "row_values": json.dumps(new_map[key], ensure_ascii=False),
            }
            for key in added_keys
        ]
        removed_rows = [
            {
                "key": key,
                "client_id": old_map[key].get("client_id", ""),
                "row_values": json.dumps(old_map[key], ensure_ascii=False),
            }
            for key in removed_keys
        ]

        write_rows(
            REPORT_DIR / f"{spec.key}__changed_common.csv",
            changed_details,
            ["key", "client_id", "changed_columns", "reason_flags", "suspicious", "old_values", "new_values"],
        )
        write_rows(
            REPORT_DIR / f"{spec.key}__added.csv",
            added_rows,
            ["key", "client_id", "date_after_2026_05_23", "row_values"],
        )
        write_rows(
            REPORT_DIR / f"{spec.key}__removed.csv",
            removed_rows,
            ["key", "client_id", "row_values"],
        )

        old_clients = {row.get(spec.client_field, "") for row in old_rows if spec.client_field and row.get(spec.client_field, "")}
        new_clients = {row.get(spec.client_field, "") for row in new_rows if spec.client_field and row.get(spec.client_field, "")}

        summary = {
            "key": spec.key,
            "title": spec.title,
            "old_rows": len(old_rows),
            "new_rows": len(new_rows),
            "row_delta": len(new_rows) - len(old_rows),
            "old_clients": len(old_clients) if spec.client_field else "",
            "new_clients": len(new_clients) if spec.client_field else "",
            "client_delta": (len(new_clients) - len(old_clients)) if spec.client_field else "",
            "common_keys": len(common_keys),
            "added_keys": len(added_keys),
            "removed_keys": len(removed_keys),
            "changed_common_keys": len(changed_details),
            "unchanged_common_keys": len(common_keys) - len(changed_details),
            "suspicious_changed_keys": suspicious_count,
            "old_duplicate_keys": sum(old_dupes.values()),
            "new_duplicate_keys": sum(new_dupes.values()),
            "top_changed_fields": ", ".join(f"{field}:{count}" for field, count in field_counts.most_common(8)),
            "top_reason_flags": ", ".join(f"{reason}:{count}" for reason, count in reason_counts.most_common(8)),
        }
        summaries.append(summary)
        detail_paths[spec.key] = {
            "changed": str((REPORT_DIR / f"{spec.key}__changed_common.csv").relative_to(ROOT)),
            "added": str((REPORT_DIR / f"{spec.key}__added.csv").relative_to(ROOT)),
            "removed": str((REPORT_DIR / f"{spec.key}__removed.csv").relative_to(ROOT)),
        }

    write_rows(
        REPORT_DIR / "summary.csv",
        summaries,
        [
            "key",
            "title",
            "old_rows",
            "new_rows",
            "row_delta",
            "old_clients",
            "new_clients",
            "client_delta",
            "common_keys",
            "added_keys",
            "removed_keys",
            "changed_common_keys",
            "unchanged_common_keys",
            "suspicious_changed_keys",
            "old_duplicate_keys",
            "new_duplicate_keys",
            "top_changed_fields",
            "top_reason_flags",
        ],
    )

    lines = [
        "# XLSX comparison: 2026-05-23/25 vs 2026-06-30",
        "",
        "Compared only the six requested XLSX pairs.",
        "",
        "Old package:",
        "",
        "- `output/20251115_0800_fix_owner/`",
        "- `output/20251115_0800_fix_owner_new_import/`",
        "",
        "New package:",
        "",
        "- `output/20260630_fix_owner/`",
        "- `output/20260630_fix_owner_new_import/`",
        "",
        "Dates used for guard checks:",
        "",
        "```text",
        "old backup date: 2026-05-23",
        "old export cutoff: 2026-05-25",
        "new export cutoff: 2026-06-30",
        "```",
        "",
        "The phrase `с 23 мая по 3 июня` was treated as the backup-to-backup",
        "window `2026-05-23` to `2026-06-30`, because the new file is",
        "`Fitnes-30-06-26.bak`.",
        "",
        "## Summary",
        "",
        "| File | Old rows | New rows | Delta | Old clients | New clients | Client delta | Common | Added | Removed | Changed common | Unchanged common | Suspicious changed |",
        "|---|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|---:|",
    ]
    for row in summaries:
        lines.append(
            "| "
            + " | ".join(
                [
                    str(row["title"]),
                    str(row["old_rows"]),
                    str(row["new_rows"]),
                    str(row["row_delta"]),
                    str(row["old_clients"]),
                    str(row["new_clients"]),
                    str(row["client_delta"]),
                    str(row["common_keys"]),
                    str(row["added_keys"]),
                    str(row["removed_keys"]),
                    str(row["changed_common_keys"]),
                    str(row["unchanged_common_keys"]),
                    str(row["suspicious_changed_keys"]),
                ]
            )
            + " |"
        )

    lines.extend(
        [
            "",
            "## Drift Guard",
            "",
            "For `import_заявки` and `абонементы клиентов`, the comparison also",
            "checked the staging/source rows used to build the XLSX. Rows whose",
            "export source fields stayed identical did not change in the XLSX.",
            "",
        ]
    )
    for row in summaries:
        verdict = "PASS" if int(row["suspicious_changed_keys"]) == 0 else "REVIEW"
        lines.append(
            f"- `{row['title']}`: {verdict}; suspicious changed keys = `{row['suspicious_changed_keys']}`."
        )

    lines.extend(["", "## Changed Field Highlights", ""])
    for row in summaries:
        lines.append(f"### {row['title']}")
        lines.append("")
        lines.append(f"- top changed fields: `{row['top_changed_fields'] or 'none'}`")
        lines.append(f"- reason flags: `{row['top_reason_flags'] or 'none'}`")
        lines.append(
            f"- details: `{detail_paths[row['key']]['changed']}`, `{detail_paths[row['key']]['added']}`, `{detail_paths[row['key']]['removed']}`"
        )
        lines.append("")

    DOC_PATH.parent.mkdir(parents=True, exist_ok=True)
    DOC_PATH.write_text("\n".join(lines), encoding="utf-8")

    print(f"summary={REPORT_DIR / 'summary.csv'}")
    print(f"report={DOC_PATH}")
    for row in summaries:
        print(
            f"{row['key']}: old={row['old_rows']} new={row['new_rows']} "
            f"delta={row['row_delta']} changed_common={row['changed_common_keys']} "
            f"suspicious={row['suspicious_changed_keys']}"
        )


if __name__ == "__main__":
    main()
