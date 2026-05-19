#!/usr/bin/env python3
"""Build Part 2 Fitbase XLSX files and reports for three funnels."""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Iterable

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAIN_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
]
MAIN_RUS_HEADERS = [
    "Внутренний номер клиента ",
    "Телефон *",
    "ФИО клиента *",
    "Почта",
    "Воронка *",
    "Этап воронки *",
    "Бюджет ",
    "Дата создания *",
    "Менеджер ",
]
CARD_HEADERS = ["телефон", "фио", "номер пластиковой карты"]

FUNNEL_SLUGS = {
    "Действующие клиенты": "deystvuyushchie_klienty",
    "Новые заявки": "novye_zayavki",
    "Реактивация": "reaktivatsiya",
}

FINAL_FUNNEL_FIELDS = [
    "client_ref",
    "client_id",
    "client_fio",
    "phones",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "create_date_source",
    "manager",
    "normalized_club",
    "club_source",
    "selected_subscription_ref",
    "selected_subscription_name",
    "selected_subscription_start_date",
    "selected_subscription_end_date",
    "selected_subscription_sale_date",
    "days_to_end",
    "days_since_end",
    "selected_card_number",
    "selected_card_ref",
    "active_full_subscription_count",
    "finished_full_subscription_count",
    "full_subscription_count",
    "trial_or_guest_sale_count",
    "selection_reason",
    "validation_status",
    "cutoff_date",
]


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(float(str(value)))
    except ValueError:
        return default


def parse_date(value: str | None):
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def copy_row_styles(ws, row_number: int, width: int) -> list[object]:
    return [copy.copy(ws.cell(row_number, col)._style) for col in range(1, width + 1)]


def clear_data_rows(ws, first_data_row: int) -> None:
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)


def trim_to_columns(ws, columns: int) -> None:
    if ws.max_column > columns:
        ws.delete_cols(columns + 1, ws.max_column - columns)


def assert_headers(ws, expected: list[str], row: int, label: str) -> None:
    actual = [ws.cell(row, col).value for col in range(1, len(expected) + 1)]
    if actual != expected:
        raise ValueError(f"{label} headers mismatch at row {row}: expected {expected}, got {actual}")
    extra = [ws.cell(row, col).value for col in range(len(expected) + 1, ws.max_column + 1)]
    if any(value not in (None, "") for value in extra):
        raise ValueError(f"{label} has non-empty extra columns after expected headers: {extra}")


def write_main_xlsx(template_path: Path, output_path: Path, rows: list[dict[str, str]]) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    assert_headers(ws, MAIN_HEADERS, 1, "main template")
    actual_ru = [ws.cell(2, col).value for col in range(1, len(MAIN_RUS_HEADERS) + 1)]
    if actual_ru != MAIN_RUS_HEADERS:
        raise ValueError(f"main template Russian headers mismatch: expected {MAIN_RUS_HEADERS}, got {actual_ru}")
    styles = copy_row_styles(ws, 3, len(MAIN_HEADERS))
    trim_to_columns(ws, len(MAIN_HEADERS))
    clear_data_rows(ws, 3)

    for index, row in enumerate(rows, start=3):
        values = [
            row.get("client_id", ""),
            row.get("phones", ""),
            row.get("client_fio", ""),
            row.get("email", ""),
            row.get("funnel", ""),
            row.get("funnel_step", ""),
            0,
            parse_date(row.get("create_date")),
            row.get("manager", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col, value)
            cell._style = copy.copy(styles[col - 1])
            if col in {1, 2, 3, 4, 5, 6, 9}:
                cell.number_format = "@"
            elif col == 7:
                cell.number_format = "0"
            elif col == 8:
                cell.number_format = "yyyy-mm-dd"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_cards_xlsx(template_path: Path, output_path: Path, rows: list[dict[str, str]]) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    assert_headers(ws, CARD_HEADERS, 1, "plastic cards template")
    styles = copy_row_styles(ws, 1, len(CARD_HEADERS))
    trim_to_columns(ws, len(CARD_HEADERS))
    clear_data_rows(ws, 2)

    for index, row in enumerate(rows, start=2):
        values = [row.get("phones", ""), row.get("client_fio", ""), row.get("selected_card_number", "")]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col, value)
            cell._style = copy.copy(styles[col - 1])
            cell.number_format = "@"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def load_managers(path: Path) -> dict[str, list[str]]:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    clubs = data.get("clubs", {})
    if not isinstance(clubs, dict) or not clubs:
        raise ValueError(f"No clubs found in {path}")
    return {str(club): [str(manager) for manager in managers] for club, managers in clubs.items()}


def stable_manager(client_id: str, managers: list[str]) -> str:
    digest = hashlib.sha256((client_id or "").encode("utf-8")).hexdigest()
    return managers[int(digest, 16) % len(managers)]


def assign_managers(rows: list[dict[str, str]], managers_by_club: dict[str, list[str]]) -> None:
    for row in rows:
        club = row.get("normalized_club", "")
        managers = managers_by_club.get(club)
        row["manager"] = stable_manager(row.get("client_id", ""), managers) if managers else ""


def sort_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    return sorted(rows, key=lambda r: (r.get("funnel", ""), r.get("funnel_step", ""), r.get("client_id", ""), r.get("client_ref", "")))


def rows_by_funnel(rows: list[dict[str, str]]) -> dict[str, list[dict[str, str]]]:
    grouped: dict[str, list[dict[str, str]]] = {funnel: [] for funnel in FUNNEL_SLUGS}
    for row in rows:
        grouped.setdefault(row.get("funnel", ""), []).append(row)
    return {funnel: sort_rows(items) for funnel, items in grouped.items()}


def build_subscription_selection_report(
    final_rows: list[dict[str, str]],
    candidates: list[dict[str, str]],
    subscriptions: list[dict[str, str]],
) -> list[dict[str, str]]:
    final_by_ref = {row["client_ref"]: row for row in final_rows}
    sub_by_key = {(row["client_ref"], row["subscription_ref"]): row for row in subscriptions}
    report_rows: list[dict[str, str]] = []
    for candidate in candidates:
        final = final_by_ref.get(candidate.get("client_ref", ""), {})
        sub = sub_by_key.get((candidate.get("client_ref", ""), candidate.get("subscription_ref", "")), {})
        selected = "1" if candidate.get("selection_status") == "selected" else "0"
        report_rows.append(
            {
                "client_ref": candidate.get("client_ref", ""),
                "client_id": final.get("client_id", candidate.get("client_id", "")),
                "client_fio": final.get("client_fio", ""),
                "candidate_for_funnel": candidate.get("candidate_for_funnel", ""),
                "subscription_ref": candidate.get("subscription_ref", ""),
                "subscription_name": sub.get("subscription_name", ""),
                "sale_date": sub.get("sale_date", ""),
                "start_date": sub.get("start_date", ""),
                "end_date": sub.get("end_date", ""),
                "status": sub.get("status", ""),
                "is_full_subscription": sub.get("is_full_subscription", ""),
                "is_active_on_cutoff": sub.get("is_active_on_cutoff", ""),
                "days_to_end": sub.get("days_to_end", ""),
                "days_since_end": sub.get("days_since_end", ""),
                "rank_number": candidate.get("rank_number", ""),
                "selected": selected,
                "selection_reason": candidate.get("selection_reason", ""),
                "manual_override_applied": candidate.get("manual_override_applied", ""),
            }
        )
    return report_rows


def build_multiple_cards_report(
    final_rows: list[dict[str, str]],
    cards: list[dict[str, str]],
    selected_cards: list[dict[str, str]],
) -> list[dict[str, str]]:
    final_by_ref = {row["client_ref"]: row for row in final_rows}
    selected_by_ref = {row["client_ref"]: row for row in selected_cards}
    multi_refs = {row["client_ref"] for row in selected_cards if int_value(row.get("active_card_count")) > 1}
    rows: list[dict[str, str]] = []
    for card in cards:
        client_ref = card.get("client_ref", "")
        if client_ref not in multi_refs:
            continue
        if card.get("is_unmarked") != "1" or not card.get("plastic_card_number"):
            continue
        final = final_by_ref.get(client_ref, {})
        selected = selected_by_ref.get(client_ref, {})
        rows.append(
            {
                "client_ref": client_ref,
                "client_id": final.get("client_id", ""),
                "client_fio": final.get("client_fio", ""),
                "funnel": final.get("funnel", ""),
                "card_ref": card.get("card_ref", ""),
                "plastic_card_number": card.get("plastic_card_number", ""),
                "issue_date": card.get("issue_date", ""),
                "is_selected": "1" if card.get("card_ref") == selected.get("selected_card_ref") else "0",
                "selection_reason": selected.get("card_selection_reason", ""),
            }
        )
    return rows


def build_active_diff(new_rows: list[dict[str, str]], previous_path: Path) -> list[dict[str, str]]:
    if not previous_path.exists():
        return []
    old_rows = read_csv(previous_path)
    old_by_ref = {row["client_ref"]: row for row in old_rows if row.get("client_ref")}
    new_active = {row["client_ref"]: row for row in new_rows if row.get("funnel") == "Действующие клиенты"}
    all_refs = sorted(set(old_by_ref) | set(new_active))
    diff_rows = []
    for client_ref in all_refs:
        old = old_by_ref.get(client_ref)
        new = new_active.get(client_ref)
        if old and new:
            changed = []
            if old.get("funnel_step") != new.get("funnel_step"):
                changed.append("stage_changed")
            if old.get("manager") != new.get("manager"):
                changed.append("manager_changed")
            if old.get("plastic_card_number") != new.get("selected_card_number"):
                changed.append("card_changed")
            if not changed:
                continue
            reason = ";".join(changed)
        elif old and not new:
            reason = "old_active_not_new_active"
        else:
            reason = "new_active_not_old_active"
        diff_rows.append(
            {
                "client_ref": client_ref,
                "client_id": (new or old or {}).get("client_id", ""),
                "client_fio": (new or old or {}).get("client_fio", ""),
                "old_present": "1" if old else "0",
                "new_present": "1" if new else "0",
                "old_funnel_step": (old or {}).get("funnel_step", ""),
                "new_funnel_step": (new or {}).get("funnel_step", ""),
                "old_manager": (old or {}).get("manager", ""),
                "new_manager": (new or {}).get("manager", ""),
                "old_card": (old or {}).get("plastic_card_number", ""),
                "new_card": (new or {}).get("selected_card_number", ""),
                "diff_reason": reason,
            }
        )
    return diff_rows


def write_reports(
    *,
    rows: list[dict[str, str]],
    stage_dir: Path,
    output_dir: Path,
    reports_dir: Path,
    csv_dir: Path,
) -> None:
    selected_cards = read_csv(stage_dir / "selected_cards.csv")
    cards = read_csv(stage_dir / "stg_plastic_cards.csv")
    candidates = read_csv(stage_dir / "subscription_candidates_ranked.csv")
    subscriptions = read_csv(stage_dir / "stg_subscriptions_all.csv")

    funnel_counts = Counter(row.get("funnel", "") for row in rows)
    write_csv(
        reports_dir / "funnel_distribution.csv",
        [{"funnel": funnel, "clients": count} for funnel, count in funnel_counts.most_common()],
        ["funnel", "clients"],
    )

    stage_counts = Counter((row.get("funnel", ""), row.get("funnel_step", "")) for row in rows)
    write_csv(
        reports_dir / "stage_distribution_by_funnel.csv",
        [
            {"funnel": funnel, "funnel_step": step, "clients": count}
            for (funnel, step), count in stage_counts.most_common()
        ],
        ["funnel", "funnel_step", "clients"],
    )

    manager_counts = Counter((row.get("normalized_club", ""), row.get("manager", "")) for row in rows)
    write_csv(
        reports_dir / "manager_distribution_by_club.csv",
        [
            {"normalized_club": club, "manager": manager, "clients": count}
            for (club, manager), count in manager_counts.most_common()
        ],
        ["normalized_club", "manager", "clients"],
    )

    missing_phone = [
        {
            **row,
            "reason": "phone is empty in 1C contact source",
        }
        for row in rows
        if not (row.get("phones") or "").strip()
    ]
    write_csv(
        reports_dir / "missing_phone_report.csv",
        missing_phone,
        ["client_ref", "client_id", "client_fio", "funnel", "funnel_step", "normalized_club", "manager", "reason"],
    )

    missing_card = [
        {**row, "reason": "no unmarked non-empty plastic card selected"}
        for row in rows
        if not (row.get("selected_card_number") or "").strip()
    ]
    write_csv(
        reports_dir / "missing_card_report.csv",
        missing_card,
        ["client_ref", "client_id", "client_fio", "funnel", "funnel_step", "reason"],
    )

    missing_club = [
        {
            **row,
            "last_sale_ref": "",
            "club_source_attempted": row.get("club_source", ""),
            "reason": "normalized club is empty or not mapped to configured managers",
        }
        for row in rows
        if not (row.get("normalized_club") or "").strip()
    ]
    write_csv(
        reports_dir / "missing_club_report.csv",
        missing_club,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "funnel",
            "funnel_step",
            "selected_subscription_ref",
            "last_sale_ref",
            "club_source_attempted",
            "reason",
        ],
    )

    multiple_subs = [
        {
            "client_ref": row.get("client_ref", ""),
            "client_id": row.get("client_id", ""),
            "client_fio": row.get("client_fio", ""),
            "funnel": row.get("funnel", ""),
            "candidate_count": row.get("active_full_subscription_count")
            if row.get("funnel") == "Действующие клиенты"
            else row.get("finished_full_subscription_count"),
            "selected_subscription_ref": row.get("selected_subscription_ref", ""),
            "selection_reason": row.get("selection_reason", ""),
        }
        for row in rows
        if (
            row.get("funnel") == "Действующие клиенты"
            and int_value(row.get("active_full_subscription_count")) > 1
        )
        or (row.get("funnel") == "Реактивация" and int_value(row.get("finished_full_subscription_count")) > 1)
    ]
    write_csv(
        reports_dir / "multiple_subscriptions_report.csv",
        multiple_subs,
        ["client_ref", "client_id", "client_fio", "funnel", "candidate_count", "selected_subscription_ref", "selection_reason"],
    )

    selection_rows = build_subscription_selection_report(rows, candidates, subscriptions)
    write_csv(
        reports_dir / "subscription_selection_report.csv",
        selection_rows,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "candidate_for_funnel",
            "subscription_ref",
            "subscription_name",
            "sale_date",
            "start_date",
            "end_date",
            "status",
            "is_full_subscription",
            "is_active_on_cutoff",
            "days_to_end",
            "days_since_end",
            "rank_number",
            "selected",
            "selection_reason",
            "manual_override_applied",
        ],
    )

    write_csv(
        reports_dir / "subscription_overrides_report.csv",
        [],
        [
            "client_ref",
            "client_id",
            "subscription_ref",
            "override_type",
            "applies_to_funnel",
            "applied",
            "result",
            "reason",
            "note",
        ],
    )

    multiple_card_rows = build_multiple_cards_report(rows, cards, selected_cards)
    write_csv(
        reports_dir / "multiple_cards_report.csv",
        multiple_card_rows,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "funnel",
            "card_ref",
            "plastic_card_number",
            "issue_date",
            "is_selected",
            "selection_reason",
        ],
    )

    selected_by_ref = {row["client_ref"]: row for row in selected_cards}
    card_selection_rows = []
    for row in rows:
        selected = selected_by_ref.get(row["client_ref"], {})
        card_selection_rows.append(
            {
                "client_ref": row.get("client_ref", ""),
                "client_id": row.get("client_id", ""),
                "client_fio": row.get("client_fio", ""),
                "funnel": row.get("funnel", ""),
                "selected_card_ref": selected.get("selected_card_ref", ""),
                "selected_card_number": selected.get("selected_card_number", ""),
                "selected_issue_date": selected.get("selected_issue_date", ""),
                "active_card_count": selected.get("active_card_count", "0"),
                "selection_reason": selected.get("card_selection_reason", ""),
                "has_future_issue_date_candidate": selected.get("has_future_issue_date_candidate", "0"),
                "has_issue_date_tie": selected.get("has_issue_date_tie", "0"),
            }
        )
    write_csv(
        reports_dir / "card_selection_report.csv",
        card_selection_rows,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "funnel",
            "selected_card_ref",
            "selected_card_number",
            "selected_issue_date",
            "active_card_count",
            "selection_reason",
            "has_future_issue_date_candidate",
            "has_issue_date_tie",
        ],
    )

    write_csv(
        reports_dir / "reactivation_boundary_anomalies.csv",
        [row for row in rows if "reactivation_boundary_anomaly" in row.get("validation_status", "")],
        ["client_ref", "client_id", "client_fio", "funnel", "funnel_step", "days_since_end", "selected_subscription_ref"],
    )

    write_csv(
        reports_dir / "new_application_create_date_review_report.csv",
        [row for row in rows if row.get("create_date_source") in {"first_non_full_sale_requires_review", "client_created_at_no_sales"}],
        ["client_ref", "client_id", "client_fio", "create_date", "create_date_source", "funnel", "normalized_club"],
    )

    write_csv(
        reports_dir / "missing_required_fields.csv",
        [
            {
                **row,
                "missing_fields": ",".join(
                    field
                    for field, value in [
                        ("client_fio", row.get("client_fio")),
                        ("create_date", row.get("create_date")),
                    ]
                    if not (value or "").strip()
                ),
            }
            for row in rows
            if not (row.get("client_fio") or "").strip() or not (row.get("create_date") or "").strip()
        ],
        ["client_ref", "client_id", "client_fio", "funnel", "funnel_step", "missing_fields", "create_date"],
    )

    active_diff = build_active_diff(rows, ROOT / "output" / "final_active_clients_20260429.csv")
    write_csv(
        reports_dir / "active_diff_vs_previous_export.csv",
        active_diff,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "old_present",
            "new_present",
            "old_funnel_step",
            "new_funnel_step",
            "old_manager",
            "new_manager",
            "old_card",
            "new_card",
            "diff_reason",
        ],
    )

    write_csv(csv_dir / "final_funnel_clients.csv", rows, FINAL_FUNNEL_FIELDS)
    for funnel, grouped_rows in rows_by_funnel(rows).items():
        slug = FUNNEL_SLUGS[funnel]
        write_csv(csv_dir / f"final_funnel_clients__{slug}.csv", grouped_rows, FINAL_FUNNEL_FIELDS)


def write_xlsx_files(
    rows: list[dict[str, str]],
    output_dir: Path,
    main_template: Path,
    cards_template: Path,
    date_stamp: str,
) -> None:
    for funnel, grouped_rows in rows_by_funnel(rows).items():
        slug = FUNNEL_SLUGS[funnel]
        main_xlsx = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}__{slug}.xlsx"
        cards_xlsx = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}__{slug}.xlsx"
        write_main_xlsx(main_template, main_xlsx, grouped_rows)
        write_cards_xlsx(cards_template, cards_xlsx, grouped_rows)
        print(f"{slug}_rows={len(grouped_rows)}")
        print(f"{slug}_main={main_xlsx.relative_to(ROOT)}")
        print(f"{slug}_cards={cards_xlsx.relative_to(ROOT)}")


def write_build_doc(
    output_dir: Path,
    reports_dir: Path,
    rows: list[dict[str, str]],
    cutoff_date: str,
    doc_suffix: str = "",
) -> None:
    funnel_counts = Counter(row.get("funnel", "") for row in rows)
    stage_counts = Counter((row.get("funnel", ""), row.get("funnel_step", "")) for row in rows)
    lines = [
        "# Part 2 XLSX build",
        "",
        f"cutoff_date: `{cutoff_date}`",
        f"built_at: `{datetime.now().isoformat(timespec='seconds')}`",
        "",
        "## Funnel distribution",
        "",
    ]
    for funnel, count in funnel_counts.most_common():
        lines.append(f"- `{funnel}`: `{count}`")
    lines.extend(["", "## Stage distribution", ""])
    for (funnel, step), count in stage_counts.most_common():
        lines.append(f"- `{funnel}` / `{step}`: `{count}`")
    lines.extend(["", "## Outputs", ""])
    for path in sorted(output_dir.glob("*.xlsx")):
        lines.append(f"- `{path.relative_to(ROOT)}`")
    lines.extend(["", f"Reports: `{reports_dir.relative_to(ROOT)}`", ""])
    safe_suffix = f"_{doc_suffix.strip('_')}" if doc_suffix else ""
    (ROOT / "docs" / f"part2_04_build_xlsx_{cutoff_date.replace('-', '')}{safe_suffix}.md").write_text(
        "\n".join(lines),
        encoding="utf-8",
    )


def build_outputs(args: argparse.Namespace) -> None:
    cutoff_date = args.cutoff_date
    date_stamp = args.date_stamp or cutoff_date.replace("-", "")
    stage_dir = as_abs(args.stage_dir)
    output_dir = as_abs(args.output_dir)
    reports_dir = as_abs(args.reports_dir) if args.reports_dir else output_dir / "reports"
    csv_dir = as_abs(args.csv_dir) if args.csv_dir else output_dir / "csv"
    main_template = as_abs(args.main_template)
    cards_template = as_abs(args.cards_template)
    managers_config = as_abs(args.managers_config)

    output_dir.mkdir(parents=True, exist_ok=True)
    reports_dir.mkdir(parents=True, exist_ok=True)
    csv_dir.mkdir(parents=True, exist_ok=True)

    rows = read_csv(stage_dir / "final_funnel_clients.csv")
    managers_by_club = load_managers(managers_config)
    assign_managers(rows, managers_by_club)
    rows = sort_rows(rows)

    write_csv(stage_dir / "final_funnel_clients.csv", rows, FINAL_FUNNEL_FIELDS)
    write_reports(rows=rows, stage_dir=stage_dir, output_dir=output_dir, reports_dir=reports_dir, csv_dir=csv_dir)
    write_xlsx_files(rows, output_dir, main_template, cards_template, date_stamp)
    write_build_doc(output_dir, reports_dir, rows, cutoff_date, args.doc_suffix)

    print(f"final_rows={len(rows)}")
    print(f"reports_dir={reports_dir.relative_to(ROOT)}")
    print(f"csv_dir={csv_dir.relative_to(ROOT)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default="2026-04-29")
    parser.add_argument("--date-stamp", default="")
    parser.add_argument("--stage-dir", default=str(ROOT / "output" / "part2_20260429" / "staging"))
    parser.add_argument("--output-dir", default=str(ROOT / "output" / "part2_20260429"))
    parser.add_argument("--reports-dir", default="")
    parser.add_argument("--csv-dir", default="")
    parser.add_argument("--main-template", default=str(ROOT / "task-desc" / "Копия Импорт_заявки.xlsx"))
    parser.add_argument("--cards-template", default=str(ROOT / "task-desc" / "Пластиковая карта.xlsx"))
    parser.add_argument("--managers-config", default=str(ROOT / "config" / "managers_by_club.yml"))
    parser.add_argument("--doc-suffix", default="")
    return parser.parse_args()


def main() -> None:
    build_outputs(parse_args())


if __name__ == "__main__":
    main()
