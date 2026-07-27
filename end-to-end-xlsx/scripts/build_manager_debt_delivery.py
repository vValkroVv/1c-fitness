#!/usr/bin/env python3
"""Resolve financial problem groups 1-3 from a manager-approved debt report.

The script is intentionally a post-processing stage.  It takes the already
verified ``manager_fixes_v2`` delivery plus the full pre-delivery membership
workbook, applies the manager's three financial values to problem1-3 rows, and
keeps problem4 separate.

Only these membership columns may change:

* ``price``;
* ``amount_of_payments``;
* ``payment_left``.

All other workbooks are copied byte-for-byte from the verified base delivery.
The output directory is immutable: an existing path is never overwritten.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import shutil
import tempfile
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
FINANCIAL_FIELDS = ("price", "amount_of_payments", "payment_left")
MANAGER_HEADERS = (
    "Структурная единица",
    "Продано",
    "Оплачено",
    "Задолженность",
)


@dataclass(frozen=True)
class MoneyTriple:
    """The three independent financial values transferred to FitBase."""

    price: Decimal
    paid: Decimal
    debt: Decimal

    def as_excel_values(self) -> tuple[int | float, int | float, int | float]:
        return tuple(excel_number(value) for value in (self.price, self.paid, self.debt))  # type: ignore[return-value]


@dataclass
class DebtSale:
    """One level-2 sale row and its level-3 phone row from the 1C report."""

    row_number: int
    document_number: str
    sale_at: datetime
    money: MoneyTriple
    phones: set[str]


@dataclass
class DebtClient:
    """One unique level-1 client row from the 1C hierarchy."""

    row_number: int
    fio: str
    normalized_fio: str
    money: MoneyTriple
    sales: list[DebtSale]

    @property
    def phones(self) -> set[str]:
        return set().union(*(sale.phones for sale in self.sales))


@dataclass(frozen=True)
class ProblemCase:
    """A row from one of the original problem1-3 workbooks."""

    group: int
    source_file: str
    values: tuple[Any, ...]
    contract_id: str
    client_id: str
    client_fio: str
    phone: str
    create_date: date | None
    payment_date: date | None
    old_money: MoneyTriple


@dataclass(frozen=True)
class Resolution:
    """The deterministic decision for one problem contract."""

    case: ProblemCase
    new_money: MoneyTriple
    source: str
    match_method: str
    phone_check: str
    manager_client_row: int | None
    manager_sale_row: int | None
    manager_document_number: str


def as_abs(path: str | Path) -> Path:
    """Resolve config paths relative to the reproducible package root."""

    value = Path(path)
    return value if value.is_absolute() else (ROOT / value).resolve()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_fio(value: Any) -> str:
    """Normalize exact names without fuzzy matching or transliteration."""

    text = unicodedata.normalize("NFKC", str(value or "")).casefold()
    text = text.replace("ё", "е")
    return " ".join(re.sub(r"[^0-9a-zа-я]+", " ", text).split())


def normalize_phones(value: Any) -> set[str]:
    """Return canonical Russian phone digits from a possibly multi-phone cell."""

    result: set[str] = set()
    for part in re.split(r"[,;/]", str(value or "")):
        digits = re.sub(r"\D", "", part)
        if len(digits) == 10:
            digits = f"7{digits}"
        elif len(digits) == 11 and digits.startswith("8"):
            digits = f"7{digits[1:]}"
        if len(digits) == 11:
            result.add(digits)
    return result


def decimal_value(value: Any, *, field: str, row_number: int | None = None) -> Decimal:
    """Read an authoritative numeric cell without silently replacing bad data."""

    if value is None or str(value).strip() == "":
        location = f" at row {row_number}" if row_number is not None else ""
        raise ValueError(f"Blank {field}{location}")
    text = str(value).strip().replace(" ", "").replace(",", ".")
    try:
        parsed = Decimal(text)
    except InvalidOperation as exc:
        location = f" at row {row_number}" if row_number is not None else ""
        raise ValueError(f"Invalid {field}{location}: {value!r}") from exc
    if not parsed.is_finite():
        raise ValueError(f"Non-finite {field}: {value!r}")
    return parsed


def money_from_values(
    price: Any,
    paid: Any,
    debt: Any,
    *,
    row_number: int | None = None,
) -> MoneyTriple:
    return MoneyTriple(
        price=decimal_value(price, field="price/Продано", row_number=row_number),
        paid=decimal_value(paid, field="amount_of_payments/Оплачено", row_number=row_number),
        debt=decimal_value(debt, field="payment_left/Задолженность", row_number=row_number),
    )


def excel_number(value: Decimal) -> int | float:
    """Keep whole roubles as integers while preserving possible fractional values."""

    integral = value.to_integral_value()
    return int(integral) if value == integral else float(value)


def parse_date_cell(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError(f"Unsupported date value: {value!r}")


def parse_manager_debt_report(path: Path) -> list[DebtClient]:
    """Parse the outline levels of the hierarchical 1C debt report."""

    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    actual_headers = tuple(worksheet.cell(1, column).value for column in range(1, 5))
    if actual_headers != MANAGER_HEADERS:
        workbook.close()
        raise ValueError(
            f"Unexpected manager report headers: {actual_headers!r}; "
            f"expected={MANAGER_HEADERS!r}"
        )
    hierarchy_labels = tuple(worksheet.cell(row, 1).value for row in range(2, 5))
    if hierarchy_labels != ("Клиент", "Документ продажи", "Клиент.Телефон"):
        workbook.close()
        raise ValueError(f"Unexpected hierarchy labels: {hierarchy_labels!r}")

    clients: list[DebtClient] = []
    current_client: DebtClient | None = None
    current_sale: DebtSale | None = None
    sale_pattern = re.compile(
        r"^Продажа\s+(?P<number>\d+)\s+от\s+"
        r"(?P<timestamp>\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})\s*$"
    )

    for row_number in range(5, worksheet.max_row + 1):
        value = str(worksheet.cell(row_number, 1).value or "").strip()
        level = int(worksheet.row_dimensions[row_number].outlineLevel or 0)
        row_money = (
            worksheet.cell(row_number, 2).value,
            worksheet.cell(row_number, 3).value,
            worksheet.cell(row_number, 4).value,
        )

        if level == 0:
            current_client = None
            current_sale = None
            continue
        if level == 1:
            if not value:
                workbook.close()
                raise ValueError(f"Blank client name at manager row {row_number}")
            current_client = DebtClient(
                row_number=row_number,
                fio=value,
                normalized_fio=normalize_fio(value),
                money=money_from_values(*row_money, row_number=row_number),
                sales=[],
            )
            clients.append(current_client)
            current_sale = None
            continue
        if level == 2:
            if current_client is None:
                workbook.close()
                raise ValueError(f"Sale row {row_number} has no level-1 client")
            match = sale_pattern.match(value)
            if not match:
                workbook.close()
                raise ValueError(f"Invalid sale label at row {row_number}: {value!r}")
            current_sale = DebtSale(
                row_number=row_number,
                document_number=match.group("number"),
                sale_at=datetime.strptime(
                    match.group("timestamp"), "%d.%m.%Y %H:%M"
                ),
                money=money_from_values(*row_money, row_number=row_number),
                phones=set(),
            )
            current_client.sales.append(current_sale)
            continue
        if level == 3:
            if current_sale is None:
                workbook.close()
                raise ValueError(f"Phone row {row_number} has no level-2 sale")
            current_sale.phones.update(normalize_phones(value))
            continue
        workbook.close()
        raise ValueError(f"Unsupported outline level {level} at row {row_number}")

    workbook.close()
    if not clients:
        raise ValueError(f"No level-1 clients found in {path}")
    missing_sales = [client.fio for client in clients if not client.sales]
    if missing_sales:
        raise ValueError(f"Manager clients without sale rows: {missing_sales[:10]}")

    by_name: dict[str, list[DebtClient]] = defaultdict(list)
    for client in clients:
        by_name[client.normalized_fio].append(client)
    duplicate_names = {
        name: rows for name, rows in by_name.items() if not name or len(rows) != 1
    }
    if duplicate_names:
        details = {
            name: [client.row_number for client in rows]
            for name, rows in duplicate_names.items()
        }
        raise ValueError(f"Manager report has blank/duplicate normalized FIO: {details}")
    return clients


def one_match(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one {pattern!r} in {directory}, found {len(matches)}"
        )
    return matches[0]


def read_problem_cases(
    base_delivery: Path,
) -> tuple[list[str], list[ProblemCase], dict[int, Path]]:
    """Read problem1-3 in their normal 22-column membership shape."""

    common_headers: list[str] | None = None
    cases: list[ProblemCase] = []
    source_paths: dict[int, Path] = {}
    seen_contracts: set[str] = set()

    for group in (1, 2, 3):
        path = one_match(base_delivery, f"problem_{group}_*.xlsx")
        source_paths[group] = path
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        headers = [
            str(value or "")
            for value in next(
                worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
            )
        ]
        if common_headers is None:
            common_headers = headers
        elif headers != common_headers:
            workbook.close()
            raise ValueError(f"Problem workbook headers differ: {path}")
        indexes = {header: index for index, header in enumerate(headers)}
        required = {
            "contract_id",
            "client_id",
            "client_fio",
            "phone",
            "create_date",
            "payment_date",
            *FINANCIAL_FIELDS,
        }
        missing = sorted(required - set(indexes))
        if missing:
            workbook.close()
            raise ValueError(f"{path} is missing required columns: {missing}")

        for values in worksheet.iter_rows(min_row=2, values_only=True):
            contract_id = str(values[indexes["contract_id"]] or "").strip()
            if not contract_id:
                workbook.close()
                raise ValueError(f"Blank contract_id in {path}")
            if contract_id in seen_contracts:
                workbook.close()
                raise ValueError(
                    f"contract_id belongs to multiple problem groups: {contract_id}"
                )
            seen_contracts.add(contract_id)
            cases.append(
                ProblemCase(
                    group=group,
                    source_file=path.name,
                    values=tuple(values[: len(headers)]),
                    contract_id=contract_id,
                    client_id=str(values[indexes["client_id"]] or "").strip(),
                    client_fio=str(values[indexes["client_fio"]] or "").strip(),
                    phone=str(values[indexes["phone"]] or "").strip(),
                    create_date=parse_date_cell(values[indexes["create_date"]]),
                    payment_date=parse_date_cell(values[indexes["payment_date"]]),
                    old_money=money_from_values(
                        values[indexes["price"]],
                        values[indexes["amount_of_payments"]],
                        values[indexes["payment_left"]],
                    ),
                )
            )
        workbook.close()

    assert common_headers is not None
    return common_headers, cases, source_paths


def read_problem_contracts(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    if "contract_id" not in headers:
        workbook.close()
        raise ValueError(f"No contract_id in {path}")
    index = headers.index("contract_id")
    values = {
        str(row[index] or "").strip()
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if str(row[index] or "").strip()
    }
    workbook.close()
    if not values:
        raise ValueError(f"No problem contracts in {path}")
    return values


def _candidate_sales(client: DebtClient, case: ProblemCase) -> list[DebtSale]:
    anchor_dates = {
        value for value in (case.create_date, case.payment_date) if value is not None
    }
    dated = [
        sale for sale in client.sales if sale.sale_at.date() in anchor_dates
    ]
    if dated:
        return dated
    if len(client.sales) == 1:
        return list(client.sales)
    return list(client.sales)


def assign_sales(
    client: DebtClient,
    cases: list[ProblemCase],
) -> list[tuple[ProblemCase, DebtSale, str]]:
    """Assign distinct sale rows while tolerating equivalent assignments.

    A multi-sale client must never receive the aggregate level-1 values on each
    contract.  We therefore solve a small one-to-one assignment against level-2
    sale rows.  Multiple assignments are accepted only when every possible
    assignment gives each contract the same financial triple.
    """

    ordered_cases = sorted(cases, key=lambda item: item.contract_id)
    sale_indexes = {id(sale): index for index, sale in enumerate(client.sales)}
    candidates = [
        [sale_indexes[id(sale)] for sale in _candidate_sales(client, case)]
        for case in ordered_cases
    ]
    assignments: list[tuple[int, ...]] = []

    def visit(position: int, used: set[int], selected: list[int]) -> None:
        if len(assignments) > 1000:
            raise ValueError(
                f"Too many sale assignments for manager client {client.fio!r}"
            )
        if position == len(ordered_cases):
            assignments.append(tuple(selected))
            return
        for sale_index in candidates[position]:
            if sale_index in used:
                continue
            used.add(sale_index)
            selected.append(sale_index)
            visit(position + 1, used, selected)
            selected.pop()
            used.remove(sale_index)

    visit(0, set(), [])
    if not assignments:
        raise ValueError(
            f"Cannot assign manager sale rows one-to-one for {client.fio!r}; "
            f"problem_contracts={[case.contract_id for case in ordered_cases]}; "
            f"sale_rows={[sale.row_number for sale in client.sales]}"
        )

    for position, case in enumerate(ordered_cases):
        possible_money = {
            client.sales[assignment[position]].money for assignment in assignments
        }
        if len(possible_money) != 1:
            raise ValueError(
                f"Ambiguous manager sale values for {client.fio!r}, "
                f"contract={case.contract_id}, "
                f"sale_rows={[client.sales[item[position]].row_number for item in assignments]}"
            )

    chosen = assignments[0]
    method = (
        "manager_sale_unique_assignment"
        if len(assignments) == 1
        else "manager_sale_equivalent_assignment"
    )
    return [
        (case, client.sales[chosen[position]], method)
        for position, case in enumerate(ordered_cases)
    ]


def resolve_cases(
    cases: list[ProblemCase],
    manager_clients: list[DebtClient],
    fallback: MoneyTriple,
) -> list[Resolution]:
    """Resolve every problem1-3 row exactly once."""

    manager_by_fio = {
        client.normalized_fio: client for client in manager_clients
    }
    cases_by_fio: dict[str, list[ProblemCase]] = defaultdict(list)
    for case in cases:
        cases_by_fio[normalize_fio(case.client_fio)].append(case)

    resolutions: list[Resolution] = []
    for normalized_fio in sorted(cases_by_fio):
        current_cases = cases_by_fio[normalized_fio]
        client = manager_by_fio.get(normalized_fio)
        if client is None:
            resolutions.extend(
                Resolution(
                    case=case,
                    new_money=fallback,
                    source="fallback_not_in_manager_debt_report",
                    match_method="normalized_fio_absent",
                    phone_check="not_applicable",
                    manager_client_row=None,
                    manager_sale_row=None,
                    manager_document_number="",
                )
                for case in current_cases
            )
            continue

        phone_checks: dict[str, str] = {}
        manager_phones = client.phones
        for case in current_cases:
            problem_phones = normalize_phones(case.phone)
            if problem_phones and manager_phones:
                if not (problem_phones & manager_phones):
                    raise ValueError(
                        f"FIO matched but phone conflicts for {case.client_fio!r}, "
                        f"contract={case.contract_id}; "
                        f"problem={sorted(problem_phones)}, manager={sorted(manager_phones)}"
                    )
                phone_checks[case.contract_id] = "fio_and_phone_match"
            elif not problem_phones and not manager_phones:
                phone_checks[case.contract_id] = "fio_match_both_phones_missing"
            elif not problem_phones:
                phone_checks[case.contract_id] = "fio_match_problem_phone_missing"
            else:
                phone_checks[case.contract_id] = "fio_match_manager_phone_missing"

        for case, sale, method in assign_sales(client, current_cases):
            resolutions.append(
                Resolution(
                    case=case,
                    new_money=sale.money,
                    source="manager_debt_report",
                    match_method=method,
                    phone_check=phone_checks[case.contract_id],
                    manager_client_row=client.row_number,
                    manager_sale_row=sale.row_number,
                    manager_document_number=sale.document_number,
                )
            )

    resolutions.sort(key=lambda item: (item.case.group, item.case.contract_id))
    if len(resolutions) != len(cases):
        raise RuntimeError(
            f"Resolution count mismatch: {len(resolutions)} != {len(cases)}"
        )
    if len({item.case.contract_id for item in resolutions}) != len(resolutions):
        raise RuntimeError("Duplicate contract_id values in resolutions")
    return resolutions


def canonical_cell(value: Any) -> tuple[str, str]:
    """Normalize values only for equality checks between equivalent workbooks."""

    if isinstance(value, datetime):
        return ("date", value.date().isoformat())
    if isinstance(value, date):
        return ("date", value.isoformat())
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return ("number", str(Decimal(str(value)).normalize()))
    if value is None:
        return ("blank", "")
    return ("text", str(value))


def rows_equal(left: Iterable[Any], right: Iterable[Any]) -> bool:
    return [canonical_cell(value) for value in left] == [
        canonical_cell(value) for value in right
    ]


def build_membership_workbook(
    source: Path,
    destination: Path,
    template: Path,
    problem_headers: list[str],
    resolutions: list[Resolution],
    excluded_contracts: set[str],
) -> dict[str, int]:
    """Stream the full membership workbook and apply only approved overrides."""

    resolution_by_contract = {
        item.case.contract_id: item for item in resolutions
    }
    source_workbook = load_workbook(source, read_only=True, data_only=False)
    source_sheet = source_workbook.active
    headers = [
        str(value or "")
        for value in next(
            source_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    if headers != problem_headers:
        source_workbook.close()
        raise ValueError("Full membership headers differ from problem workbook headers")
    indexes = {header: index for index, header in enumerate(headers)}
    for field in ("contract_id", *FINANCIAL_FIELDS):
        if field not in indexes:
            source_workbook.close()
            raise ValueError(f"Full membership is missing {field!r}")

    target_workbook = load_workbook(template)
    target_sheet = target_workbook.active
    width = len(headers)
    if target_sheet.max_column > width:
        target_sheet.delete_cols(width + 1, target_sheet.max_column - width)
    template_formats = [
        target_sheet.cell(3, column).number_format
        for column in range(1, width + 1)
    ]
    if target_sheet.max_row >= 3:
        target_sheet.delete_rows(3, target_sheet.max_row - 2)

    russian_headers = next(
        source_sheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    for column, value in enumerate(headers, start=1):
        target_sheet.cell(1, column).value = value
        target_sheet.cell(2, column).value = russian_headers[column - 1]

    format_columns: dict[int, str] = {}
    for column, header in enumerate(headers, start=1):
        if header in {"create_date", "payment_date", "activation_date", "end_date"}:
            format_columns[column] = "yyyy-mm-dd"
        elif header in FINANCIAL_FIELDS:
            format_columns[column] = "#,##0.00"
        elif template_formats[column - 1] not in ("", "General"):
            format_columns[column] = template_formats[column - 1]

    found_resolutions: set[str] = set()
    found_excluded: set[str] = set()
    seen_contracts: set[str] = set()
    source_rows = 0
    output_rows = 0
    changed_rows = 0
    changed_cells = 0

    for source_values in source_sheet.iter_rows(min_row=3, values_only=True):
        source_rows += 1
        values = list(source_values[:width])
        contract_id = str(values[indexes["contract_id"]] or "").strip()
        if contract_id:
            if contract_id in seen_contracts:
                source_workbook.close()
                target_workbook.close()
                raise ValueError(f"Duplicate contract_id in full membership: {contract_id}")
            seen_contracts.add(contract_id)
        if contract_id in excluded_contracts:
            found_excluded.add(contract_id)
            continue

        resolution = resolution_by_contract.get(contract_id)
        if resolution is not None:
            if not rows_equal(values, resolution.case.values):
                source_workbook.close()
                target_workbook.close()
                raise ValueError(
                    f"Problem row differs from full membership row: {contract_id}"
                )
            found_resolutions.add(contract_id)
            new_values = resolution.new_money.as_excel_values()
            row_changed = False
            for field, new_value in zip(FINANCIAL_FIELDS, new_values, strict=True):
                index = indexes[field]
                old_value = decimal_value(values[index], field=field)
                if old_value != Decimal(str(new_value)):
                    row_changed = True
                    changed_cells += 1
                values[index] = new_value
            if row_changed:
                changed_rows += 1

        target_sheet.append(values)
        output_rows += 1
        # ``Worksheet.max_row`` scans the full sparse-cell dictionary. Calling
        # it for every appended row turns a 121k-row export into quadratic
        # work, so derive the physical row directly from our linear counter.
        target_row = output_rows + 2
        for column, number_format in format_columns.items():
            target_sheet.cell(target_row, column).number_format = number_format

    source_workbook.close()
    missing_resolutions = sorted(set(resolution_by_contract) - found_resolutions)
    missing_excluded = sorted(excluded_contracts - found_excluded)
    if missing_resolutions or missing_excluded:
        target_workbook.close()
        raise ValueError(
            f"Contracts missing from full membership: "
            f"resolved={missing_resolutions[:10]}, excluded={missing_excluded[:10]}"
        )

    target_sheet.freeze_panes = "A3"
    for column in range(1, width + 1):
        letter = get_column_letter(column)
        current_width = target_sheet.column_dimensions[letter].width or 12
        target_sheet.column_dimensions[letter].width = min(
            max(current_width, 12), 34
        )
    destination.parent.mkdir(parents=True, exist_ok=True)
    target_workbook.save(destination)
    target_workbook.close()
    return {
        "source_rows": source_rows,
        "output_rows": output_rows,
        "resolved_rows": len(found_resolutions),
        "excluded_rows": len(found_excluded),
        "changed_rows": changed_rows,
        "changed_cells": changed_cells,
    }


def write_audit_csv(path: Path, resolutions: list[Resolution]) -> None:
    fieldnames = [
        "problem_group",
        "contract_id",
        "client_id",
        "client_fio",
        "phone",
        "resolution_source",
        "match_method",
        "phone_check",
        "manager_client_row",
        "manager_sale_row",
        "manager_document_number",
        "old_price",
        "old_amount_of_payments",
        "old_payment_left",
        "new_price",
        "new_amount_of_payments",
        "new_payment_left",
        "changed_fields",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fieldnames, lineterminator="\n"
        )
        writer.writeheader()
        for item in resolutions:
            old = item.case.old_money
            new = item.new_money
            changed = [
                field
                for field, old_value, new_value in zip(
                    FINANCIAL_FIELDS,
                    (old.price, old.paid, old.debt),
                    (new.price, new.paid, new.debt),
                    strict=True,
                )
                if old_value != new_value
            ]
            writer.writerow(
                {
                    "problem_group": item.case.group,
                    "contract_id": item.case.contract_id,
                    "client_id": item.case.client_id,
                    "client_fio": item.case.client_fio,
                    "phone": item.case.phone,
                    "resolution_source": item.source,
                    "match_method": item.match_method,
                    "phone_check": item.phone_check,
                    "manager_client_row": item.manager_client_row or "",
                    "manager_sale_row": item.manager_sale_row or "",
                    "manager_document_number": item.manager_document_number,
                    "old_price": old.price,
                    "old_amount_of_payments": old.paid,
                    "old_payment_left": old.debt,
                    "new_price": new.price,
                    "new_amount_of_payments": new.paid,
                    "new_payment_left": new.debt,
                    "changed_fields": ",".join(changed),
                }
            )


def resolution_summary(
    *,
    config: dict[str, Any],
    manager_debt_path: Path,
    manager_clients: list[DebtClient],
    resolutions: list[Resolution],
    membership_stats: dict[str, int],
    copied_xlsx: list[Path],
    source_problem_paths: dict[int, Path],
    problem4_path: Path,
    staging_dir: Path,
) -> dict[str, Any]:
    by_group: dict[str, dict[str, int]] = {}
    for group in (1, 2, 3):
        group_rows = [item for item in resolutions if item.case.group == group]
        by_group[str(group)] = {
            "total": len(group_rows),
            "manager_debt_report": sum(
                item.source == "manager_debt_report" for item in group_rows
            ),
            "fallback_not_in_manager_debt_report": sum(
                item.source == "fallback_not_in_manager_debt_report"
                for item in group_rows
            ),
        }

    manager_rows = [
        item for item in resolutions if item.source == "manager_debt_report"
    ]
    fallback_rows = [
        item
        for item in resolutions
        if item.source == "fallback_not_in_manager_debt_report"
    ]
    arithmetic_anomalies = [
        item
        for item in manager_rows
        if item.new_money.price - item.new_money.paid != item.new_money.debt
    ]

    def sums(rows: list[Resolution]) -> dict[str, str]:
        return {
            "price": str(sum((item.new_money.price for item in rows), Decimal("0"))),
            "amount_of_payments": str(
                sum((item.new_money.paid for item in rows), Decimal("0"))
            ),
            "payment_left": str(
                sum((item.new_money.debt for item in rows), Decimal("0"))
            ),
        }

    xlsx_hashes = {
        path.name: file_sha256(staging_dir / path.name) for path in copied_xlsx
    }
    membership_name = config["output"]["membership_file_name"]
    xlsx_hashes[membership_name] = file_sha256(staging_dir / membership_name)
    xlsx_hashes[problem4_path.name] = file_sha256(staging_dir / problem4_path.name)

    return {
        "cutoff_at": config["run"]["cutoff_at"],
        "cutoff_date": config["run"]["cutoff_date"],
        "date_stamp": config["run"]["date_stamp"],
        "manager_debt_report": {
            "path": str(manager_debt_path),
            "sha256": file_sha256(manager_debt_path),
            "clients": len(manager_clients),
            "sales": sum(len(client.sales) for client in manager_clients),
        },
        "resolved_problem_contracts": len(resolutions),
        "resolution_sources": {
            "manager_debt_report": len(manager_rows),
            "fallback_not_in_manager_debt_report": len(fallback_rows),
        },
        "by_problem_group": by_group,
        "manager_arithmetic_anomaly_contracts": len(arithmetic_anomalies),
        "target_sums": {
            "manager_debt_report": sums(manager_rows),
            "fallback": sums(fallback_rows),
            "combined": sums(resolutions),
        },
        "membership": membership_stats,
        "problem4": {
            "file": problem4_path.name,
            "contracts": sorted(read_problem_contracts(problem4_path)),
            "copied_unchanged": True,
        },
        "source_problem_sha256": {
            str(group): file_sha256(path)
            for group, path in sorted(source_problem_paths.items())
        },
        "output_xlsx_sha256": xlsx_hashes,
    }


def write_summary_markdown(path: Path, summary: dict[str, Any]) -> None:
    sources = summary["resolution_sources"]
    membership = summary["membership"]
    lines = [
        "# Применение достоверной задолженности менеджера",
        "",
        f"- единый cutoff: `{summary['cutoff_at']}`;",
        f"- договоров problem1–3 обработано: `{summary['resolved_problem_contracts']}`;",
        f"- найдено в менеджерском XLSX: `{sources['manager_debt_report']}`;",
        f"- отсутствовало и получило `12000 / 12000 / 0`: "
        f"`{sources['fallback_not_in_manager_debt_report']}`;",
        f"- строк в итоговом clean membership: `{membership['output_rows']}`;",
        f"- договоров problem4 исключено из clean: `{membership['excluded_rows']}`;",
        "",
        "## По группам",
        "",
        "| группа | всего | из XLSX менеджера | fallback |",
        "| --- | ---: | ---: | ---: |",
    ]
    for group, values in summary["by_problem_group"].items():
        lines.append(
            f"| problem{group} | {values['total']} | "
            f"{values['manager_debt_report']} | "
            f"{values['fallback_not_in_manager_debt_report']} |"
        )
    lines.extend(
        [
            "",
            "## Контроль",
            "",
            "- менялись только `price`, `amount_of_payments`, `payment_left`;",
            f"- реально изменено строк: `{membership['changed_rows']}`;",
            f"- реально изменено финансовых ячеек: `{membership['changed_cells']}`;",
            "- `Задолженность` переносилась как независимое достоверное поле, "
            "без пересчёта через `Продано - Оплачено`;",
            f"- таких арифметически нестандартных выбранных строк: "
            f"`{summary['manager_arithmetic_anomaly_contracts']}`;",
            "- problem1–3 возвращены в clean membership; problem4 скопирован "
            "без изменений и остаётся единственной problem-группой.",
            "",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def load_config(path: Path) -> dict[str, Any]:
    config = yaml.safe_load(path.read_text(encoding="utf-8"))
    required_sections = {"run", "inputs", "rule", "output"}
    missing = sorted(required_sections - set(config or {}))
    if missing:
        raise ValueError(f"Config is missing sections: {missing}")
    date_stamp = str(config["run"]["date_stamp"])
    cutoff_date = datetime.strptime(
        str(config["run"]["cutoff_date"]), "%Y-%m-%d"
    ).date()
    cutoff_at = datetime.strptime(
        str(config["run"]["cutoff_at"]), "%Y-%m-%d %H:%M:%S"
    )
    backup_finish_at = datetime.strptime(
        str(config["run"]["backup_finish_at"]), "%Y-%m-%d %H:%M:%S"
    )
    if date_stamp != cutoff_date.strftime("%Y%m%d"):
        raise ValueError("date_stamp must equal YYYYMMDD(cutoff_date)")
    if cutoff_at.date() != cutoff_date:
        raise ValueError("cutoff_at and cutoff_date must be the same slice")
    if cutoff_at != backup_finish_at:
        raise ValueError(
            "cutoff_at must exactly equal RESTORE HEADERONLY.BackupFinishDate"
        )
    return config


def build(config_path: Path) -> Path:
    config = load_config(config_path)
    inputs = config["inputs"]
    output = config["output"]
    base_delivery = as_abs(inputs["base_delivery"])
    full_membership = as_abs(inputs["full_membership"])
    manager_debt_path = as_abs(inputs["manager_debt_xlsx"])
    membership_template = as_abs(inputs["membership_template"])
    source_manifest_path = as_abs(inputs["source_manifest"])
    output_dir = as_abs(output["directory"])

    for path in (
        base_delivery,
        full_membership,
        manager_debt_path,
        membership_template,
        source_manifest_path,
    ):
        if not path.exists():
            raise FileNotFoundError(path)
    if output_dir.exists():
        raise FileExistsError(f"Output directory already exists: {output_dir}")

    source_manifest = yaml.safe_load(
        source_manifest_path.read_text(encoding="utf-8")
    )
    manifest_backup_finish = str(
        source_manifest["backup"]["backup_finish_at"]
    )
    if manifest_backup_finish != str(config["run"]["backup_finish_at"]):
        raise ValueError(
            "Config backup_finish_at differs from the pinned source manifest"
        )

    pinned_files = {
        full_membership: str(inputs["full_membership_sha256"]),
        membership_template: str(inputs["membership_template_sha256"]),
    }
    for path, expected_hash in pinned_files.items():
        actual_hash = file_sha256(path)
        if actual_hash != expected_hash.strip().lower():
            raise ValueError(
                f"Input SHA-256 mismatch for {path}: "
                f"actual={actual_hash}, expected={expected_hash}"
            )
    for name, expected_hash in inputs["base_delivery_sha256"].items():
        path = base_delivery / name
        if not path.is_file():
            raise FileNotFoundError(path)
        actual_hash = file_sha256(path)
        if actual_hash != str(expected_hash).strip().lower():
            raise ValueError(
                f"Base delivery SHA-256 mismatch for {name}: "
                f"actual={actual_hash}, expected={expected_hash}"
            )

    actual_manager_hash = file_sha256(manager_debt_path)
    expected_manager_hash = str(inputs["manager_debt_sha256"]).strip().lower()
    if actual_manager_hash != expected_manager_hash:
        raise ValueError(
            f"Manager debt XLSX SHA-256 mismatch: "
            f"actual={actual_manager_hash}, expected={expected_manager_hash}"
        )
    cutoff = datetime.strptime(str(config["run"]["cutoff_date"]), "%Y-%m-%d").date()
    expected_name_date = cutoff.strftime("%d.%m.%Y")
    if expected_name_date not in manager_debt_path.name:
        raise ValueError(
            f"Manager report filename must contain cutoff {expected_name_date}: "
            f"{manager_debt_path.name}"
        )

    manager_clients = parse_manager_debt_report(manager_debt_path)
    problem_headers, problem_cases, source_problem_paths = read_problem_cases(
        base_delivery
    )
    fallback_config = config["rule"]["not_in_manager_debt_report"]
    fallback = money_from_values(
        fallback_config["price"],
        fallback_config["amount_of_payments"],
        fallback_config["payment_left"],
    )
    resolutions = resolve_cases(problem_cases, manager_clients, fallback)

    problem4_path = one_match(base_delivery, "problem_4_*.xlsx")
    problem4_contracts = read_problem_contracts(problem4_path)
    resolved_contracts = {item.case.contract_id for item in resolutions}
    if resolved_contracts & problem4_contracts:
        raise ValueError("A contract occurs in both problem1-3 and problem4")

    date_stamp = str(config["run"]["date_stamp"])
    membership_name = str(output["membership_file_name"])
    expected_membership_name = (
        f"fitbase_import_abonementy_clientov_{date_stamp}.xlsx"
    )
    if membership_name != expected_membership_name:
        raise ValueError(
            f"Unexpected membership output name: {membership_name!r}; "
            f"expected={expected_membership_name!r}"
        )
    static_names = [
        f"fitbase_active_clients_import_zayavki_{date_stamp}_all_funnels.xlsx",
        f"fitbase_active_clients_plastic_cards_{date_stamp}_all_funnels.xlsx",
        f"fitbase_import_shablony_abonementov_{date_stamp}.xlsx",
        f"fitbase_import_shablony_uslug_{date_stamp}.xlsx",
        f"fitbase_import_uslugi_clientov_{date_stamp}.xlsx",
    ]
    static_sources = [base_delivery / name for name in static_names]
    for path in static_sources:
        if not path.is_file():
            raise FileNotFoundError(path)

    output_dir.parent.mkdir(parents=True, exist_ok=True)
    staging_dir = Path(
        tempfile.mkdtemp(prefix=f".{output_dir.name}.", dir=output_dir.parent)
    )
    try:
        copied_xlsx: list[Path] = []
        for source in static_sources:
            shutil.copy2(source, staging_dir / source.name)
            copied_xlsx.append(source)
        shutil.copy2(problem4_path, staging_dir / problem4_path.name)

        membership_stats = build_membership_workbook(
            source=full_membership,
            destination=staging_dir / membership_name,
            template=membership_template,
            problem_headers=problem_headers,
            resolutions=resolutions,
            excluded_contracts=problem4_contracts,
        )
        reports_dir = staging_dir / "reports"
        write_audit_csv(reports_dir / "financial_resolution_audit.csv", resolutions)
        summary = resolution_summary(
            config=config,
            manager_debt_path=manager_debt_path,
            manager_clients=manager_clients,
            resolutions=resolutions,
            membership_stats=membership_stats,
            copied_xlsx=copied_xlsx,
            source_problem_paths=source_problem_paths,
            problem4_path=problem4_path,
            staging_dir=staging_dir,
        )
        (reports_dir / "manager_debt_resolution.json").write_text(
            json.dumps(summary, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
        write_summary_markdown(
            reports_dir / "manager_debt_resolution.md", summary
        )
        staging_dir.rename(output_dir)
    except BaseException:
        # The directory is a validated mkdtemp child created by this process.
        # Cleaning it also makes Ctrl-C safe without exposing a partial delivery.
        shutil.rmtree(staging_dir, ignore_errors=True)
        raise

    print(f"output_dir={output_dir}")
    print(f"resolved_contracts={len(resolutions)}")
    print(
        "manager_matches="
        f"{sum(item.source == 'manager_debt_report' for item in resolutions)}"
    )
    print(
        "fallback_rows="
        f"{sum(item.source != 'manager_debt_report' for item in resolutions)}"
    )
    print(f"membership_rows={membership_stats['output_rows']}")
    print(f"problem4_contracts={len(problem4_contracts)}")
    return output_dir


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True)
    return parser.parse_args()


if __name__ == "__main__":
    arguments = parse_args()
    build(as_abs(arguments.config))
