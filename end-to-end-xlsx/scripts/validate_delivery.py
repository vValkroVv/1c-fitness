#!/usr/bin/env python3
"""Validate a complete XLSX delivery against structural invariants."""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def as_abs(path: str | Path) -> Path:
    value = Path(path)
    return value if value.is_absolute() else ROOT / value


def workbook_summary(path: Path, header_rows: int) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    worksheet = workbook.active
    row_count = 0
    first_rows: list[list[Any]] = []
    formula_cells: list[str] = []
    blank_data_rows: list[int] = []
    for row in worksheet.iter_rows():
        row_count += 1
        values = [cell.value for cell in row]
        if len(first_rows) < header_rows:
            first_rows.append(values)
        elif not any(value not in (None, "") for value in values):
            blank_data_rows.append(row_count)
        for cell in row:
            if cell.data_type == "f":
                formula_cells.append(cell.coordinate)
    summary = {
        "total_rows": row_count,
        "data_rows": max(row_count - header_rows, 0),
        "columns": worksheet.max_column,
        "headers": first_rows,
        "sheet_names": workbook.sheetnames,
        "formula_cells": formula_cells,
        "blank_data_rows": blank_data_rows,
    }
    workbook.close()
    return summary


def problem_format_summary(path: Path, header_rows: int) -> dict[str, Any]:
    """Read layout metadata from a small problem workbook.

    ``ReadOnlyWorksheet`` deliberately omits freeze panes and some formatting
    metadata, so this helper uses normal mode only for the tiny problem files.
    The large clean membership workbook remains streamed by ``workbook_summary``.
    """

    workbook = load_workbook(path, read_only=False, data_only=False)
    worksheet = workbook.active
    summary = {
        "freeze_panes": str(worksheet.freeze_panes or ""),
        "auto_filter": str(worksheet.auto_filter.ref or ""),
        "dimension": worksheet.calculate_dimension(),
        "header_style_ids": [cell.style_id for cell in worksheet[1]],
        "first_data_style_ids": (
            [cell.style_id for cell in worksheet[header_rows + 1]]
            if worksheet.max_row > header_rows
            else []
        ),
    }
    workbook.close()
    return summary


def read_contract_ids(path: Path, data_start_row: int) -> tuple[set[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if "contract_id" not in headers:
        workbook.close()
        raise RuntimeError(f"No contract_id header in {path}")
    index = headers.index("contract_id")
    values: list[str] = []
    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        value = str(row[index] or "").strip()
        if value:
            values.append(value)
    workbook.close()
    return set(values), len(values)


def branch_values(path: Path, data_start_row: int) -> Counter[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if "филиал" not in headers:
        workbook.close()
        raise RuntimeError(f"No филиал header in {path}")
    index = headers.index("филиал")
    counts: Counter[str] = Counter()
    for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
        counts[str(row[index] or "").strip()] += 1
    workbook.close()
    return counts


def validate(args: argparse.Namespace) -> int:
    output_dir = as_abs(args.output_dir)
    expected_path = as_abs(args.expected)
    report_path = as_abs(args.report)
    json_path = as_abs(args.json_report)
    expected = yaml.safe_load(expected_path.read_text(encoding="utf-8"))
    expected_files: dict[str, dict[str, Any]] = expected["files"]
    enforce_counts = args.enforce_reference_counts

    errors: list[str] = []
    warnings: list[str] = []
    actual_names = {path.name for path in output_dir.glob("*.xlsx")}
    wanted_names = set(expected_files)
    if actual_names != wanted_names:
        errors.append(
            f"XLSX file set mismatch; missing={sorted(wanted_names - actual_names)}, "
            f"unexpected={sorted(actual_names - wanted_names)}"
        )

    summaries: dict[str, dict[str, Any]] = {}
    for name, spec in expected_files.items():
        path = output_dir / name
        if not path.is_file():
            continue
        summary = workbook_summary(path, int(spec["header_rows"]))
        summaries[name] = summary
        if summary["columns"] != int(spec["columns"]):
            errors.append(
                f"{name}: columns={summary['columns']}, expected={spec['columns']}"
            )
        if len(summary["sheet_names"]) != 1:
            errors.append(
                f"{name}: sheets={summary['sheet_names']!r}, expected exactly one sheet"
            )
        if summary["formula_cells"]:
            errors.append(
                f"{name}: formula cells are not allowed: "
                f"{summary['formula_cells'][:10]}"
            )
        if summary["blank_data_rows"]:
            errors.append(
                f"{name}: fully blank data rows: {summary['blank_data_rows'][:10]}"
            )
        if enforce_counts and summary["data_rows"] != int(spec["data_rows"]):
            errors.append(
                f"{name}: data_rows={summary['data_rows']}, expected={spec['data_rows']}"
            )
        expected_contracts = {
            str(value).strip() for value in spec.get("contract_ids", [])
        }
        if expected_contracts:
            actual_contracts, contract_rows = read_contract_ids(
                path, int(spec["header_rows"]) + 1
            )
            if actual_contracts != expected_contracts:
                errors.append(
                    f"{name}: contract_id set={sorted(actual_contracts)}, "
                    f"expected={sorted(expected_contracts)}"
                )
            if contract_rows != len(expected_contracts):
                errors.append(
                    f"{name}: contract_id rows={contract_rows}, "
                    f"expected exactly={len(expected_contracts)}"
                )

    problem_names = sorted(name for name in wanted_names if name.startswith("problem_"))
    available_problem_names = [
        name for name in problem_names if (output_dir / name).is_file()
    ]
    if available_problem_names:
        baseline_name = available_problem_names[0]
        baseline = summaries[baseline_name]
        problem_formats = {
            name: problem_format_summary(
                output_dir / name, int(expected_files[name]["header_rows"])
            )
            for name in available_problem_names
        }
        baseline_format = problem_formats[baseline_name]
        baseline_headers = baseline["headers"]
        baseline_header_styles = baseline_format["header_style_ids"]
        baseline_data_styles = baseline_format["first_data_style_ids"]
        baseline_sheets = baseline["sheet_names"]
        for name in available_problem_names:
            summary = summaries[name]
            format_summary = problem_formats[name]
            if summary["headers"] != baseline_headers:
                errors.append(
                    f"{name}: problem workbook headers differ from {baseline_name}"
                )
            if format_summary["header_style_ids"] != baseline_header_styles:
                errors.append(
                    f"{name}: problem header format differs from {baseline_name}"
                )
            if format_summary["first_data_style_ids"] != baseline_data_styles:
                errors.append(
                    f"{name}: problem data-row format differs from {baseline_name}"
                )
            if summary["sheet_names"] != baseline_sheets:
                errors.append(
                    f"{name}: problem sheet names differ from {baseline_name}"
                )
            if format_summary["freeze_panes"] != "A2":
                errors.append(
                    f"{name}: freeze_panes={format_summary['freeze_panes']!r}, expected='A2'"
                )
            if format_summary["auto_filter"] != format_summary["dimension"]:
                errors.append(
                    f"{name}: auto_filter={format_summary['auto_filter']!r}, "
                    f"expected full range {format_summary['dimension']!r}"
                )

    problem_sets: list[set[str]] = []
    for name in problem_names:
        path = output_dir / name
        if not path.is_file():
            continue
        values, total = read_contract_ids(path, 2)
        if len(values) != total:
            errors.append(
                f"{name}: duplicate contract_id values: rows={total}, unique={len(values)}"
            )
        problem_sets.append(values)
    problem_union = set().union(*problem_sets) if problem_sets else set()
    if sum(len(values) for values in problem_sets) != len(problem_union):
        errors.append("Some contract_id values occur in more than one problem workbook")
    expected_problem_total = int(expected["problem_contracts"]["unique_total"])
    if enforce_counts and len(problem_union) != expected_problem_total:
        errors.append(
            f"unique problem contract_id={len(problem_union)}, expected={expected_problem_total}"
        )

    membership_name = next(
        (
            name
            for name in wanted_names
            if name.startswith("fitbase_import_abonementy_clientov_")
        ),
        "",
    )
    if membership_name and (output_dir / membership_name).is_file():
        clean_contracts, clean_total = read_contract_ids(
            output_dir / membership_name, 3
        )
        overlap = sorted(clean_contracts & problem_union)
        if overlap:
            errors.append(
                f"Problem contract_id values remain in clean membership XLSX: {overlap[:10]}"
            )
        # Empty contract_id is allowed for refuser placeholder rows and is not
        # returned by read_contract_ids. Non-empty IDs must still be unique.
        if len(clean_contracts) != clean_total:
            errors.append(
                f"Clean membership has duplicate non-empty contract_id values: "
                f"rows={clean_total}, unique={len(clean_contracts)}"
            )

    allowed_branches = set(expected["allowed_branches"])
    branch_files = [
        (name, 3 if not name.startswith("problem_") else 2)
        for name in wanted_names
        if name.startswith("problem_")
        or name.startswith("fitbase_active_clients_import_zayavki_")
        or name.startswith("fitbase_import_abonementy_clientov_")
        or name.startswith("fitbase_import_uslugi_clientov_")
    ]
    branch_summaries: dict[str, dict[str, int]] = {}
    for name, start_row in branch_files:
        path = output_dir / name
        if not path.is_file():
            continue
        counts = branch_values(path, start_row)
        branch_summaries[name] = dict(counts)
        if counts.get("", 0):
            errors.append(f"{name}: blank филиал values={counts['']}")
        unexpected = sorted(set(counts) - allowed_branches - {""})
        if unexpected:
            errors.append(f"{name}: unexpected филиал values={unexpected}")

    verdict = "PASS" if not errors else "FAIL"
    report_lines = [
        f"# {len(expected_files)}-XLSX delivery validation",
        "",
        f"- verdict: **{verdict}**",
        f"- directory: `{output_dir}`",
        f"- XLSX files: `{len(actual_names)}`",
        f"- unique problem contract_id: `{len(problem_union)}`",
        "",
        "## Files",
        "",
        "| file | data rows | columns |",
        "| --- | ---: | ---: |",
    ]
    for name in sorted(summaries):
        summary = summaries[name]
        report_lines.append(
            f"| `{name}` | {summary['data_rows']} | {summary['columns']} |"
        )
    report_lines.extend(["", "## Errors", ""])
    report_lines.extend([f"- {item}" for item in errors] or ["- none"])
    report_lines.extend(["", "## Warnings", ""])
    report_lines.extend([f"- {item}" for item in warnings] or ["- none"])
    report_lines.append("")

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(
            {
                "verdict": verdict,
                "errors": errors,
                "warnings": warnings,
                "files": summaries,
                "problem_contract_ids": len(problem_union),
                "branches": branch_summaries,
            },
            ensure_ascii=False,
            indent=2,
            default=str,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"verdict={verdict}")
    print(f"report={report_path}")
    for error in errors:
        print(f"ERROR: {error}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--expected", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--json-report", required=True)
    parser.add_argument("--enforce-reference-counts", action="store_true")
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(validate(parse_args()))
