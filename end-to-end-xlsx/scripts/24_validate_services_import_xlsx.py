#!/usr/bin/env python3
"""Validate generated Fitbase services import XLSX files."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260630"
ACTIVE_MEMBERSHIP_FUNNEL = "Действующие абонементы"
REGISTER_END_DATE_SOURCE = "dbo._InfoRg3060._Fld3064"
SALE_DATE_FALLBACK_SOURCE = "sale_date_conservative_fallback"
CLIENT_HEADERS = [
    "service_id",
    "client_id",
    "phone",
    "client_fio",
    "service_name",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "count",
    "visits_left",
    "price",
    "amount_of_payment",
    "payment_left",
    "type_of_payment",
    "manager",
    "филиал",
]
ALLOWED_BRANCHES = {
    "Фитнес Империя (Гоголевский)",
    "Фитнес Империя (Промышленная)",
    "Фитнес Империя (Ровио)",
    "Фитнес Империя (Столица)",
}
TEMPLATE_HEADERS = [
    "name",
    "price",
    "duration",
    "visits",
    "do_enter",
    "first_visit_activation",
    "archive",
    "category",
    "legal_entity",
]


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def iter_data_rows(path: Path, width: int) -> tuple[list[Any], list[tuple[Any, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    headers = list(next(rows_iter)[:width])
    next(rows_iter, None)
    rows = [tuple(row[:width]) for row in rows_iter if any(value not in (None, "") for value in row[:width])]
    wb.close()
    return headers, rows


def read_service_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = [str(row[0] or "").strip() for row in ws.iter_rows(values_only=True) if str(row[0] or "").strip()]
    wb.close()
    return names


def read_source_clients(path: Path) -> dict[str, str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    client_col = headers.index("client_id")
    funnel_col = headers.index("funnel")
    clients = {
        str(row[client_col]).strip(): str(row[funnel_col] or "").strip()
        for row in ws.iter_rows(min_row=3, values_only=True)
        if row[client_col] not in (None, "")
    }
    wb.close()
    return clients


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt).date()
        except ValueError:
            continue
    return None


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def blank(value: Any) -> bool:
    return value is None or value == ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-output-dir", default="work/20260630/owner")
    parser.add_argument("--output-dir", default="work/20260630/imports")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    parser.add_argument(
        "--cutoff-date",
        help="Single backup cutoff date; defaults to the first YYYYMMDD in --date-stamp.",
    )
    parser.add_argument("--services-list", default="templates/services_required.xlsx")
    parser.add_argument("--expected-real-end-rows", type=int)
    parser.add_argument("--expected-fallback-rows", type=int)
    parser.add_argument("--expected-live-rows", type=int)
    parser.add_argument("--expected-case-service-id")
    parser.add_argument("--expected-case-end-date")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_output_dir = as_abs(args.source_output_dir)
    output_dir = as_abs(args.output_dir)
    source_clients_xlsx = source_output_dir / f"fitbase_active_clients_import_zayavki_{args.date_stamp}_all_funnels.xlsx"
    client_xlsx = output_dir / f"fitbase_import_uslugi_clientov_{args.date_stamp}.xlsx"
    template_xlsx = output_dir / f"fitbase_import_shablony_uslug_{args.date_stamp}.xlsx"
    coverage_path = output_dir / "reports" / "services_coverage_report.csv"

    service_names = read_service_names(as_abs(args.services_list))
    source_clients = read_source_clients(source_clients_xlsx)
    source_client_ids = set(source_clients)
    client_headers, client_rows = iter_data_rows(client_xlsx, len(CLIENT_HEADERS))
    template_headers, template_rows = iter_data_rows(template_xlsx, len(TEMPLATE_HEADERS))
    cutoff_date = parse_date(args.cutoff_date) if args.cutoff_date else None
    if cutoff_date is None:
        try:
            cutoff_date = datetime.strptime(args.date_stamp[:8], "%Y%m%d").date()
        except ValueError as exc:
            raise ValueError(
                "--cutoff-date is required when --date-stamp does not start with YYYYMMDD"
            ) from exc

    errors: list[str] = []
    warnings: list[str] = []
    if client_headers != CLIENT_HEADERS:
        errors.append(f"client headers mismatch: {client_headers}")
    if template_headers != TEMPLATE_HEADERS:
        errors.append(f"template headers mismatch: {template_headers}")

    allowed_services = set(service_names)
    template_names = [str(row[0]).strip() for row in template_rows]
    if template_names != service_names:
        errors.append("template service list does not exactly match required 51 services in order")

    client_service_names = {str(row[4]).strip() for row in client_rows}
    outside_services = client_service_names - allowed_services
    if outside_services:
        errors.append(f"client rows outside required service list: {len(outside_services)}")

    client_ids = {str(row[1]).strip() for row in client_rows}
    outside_clients = client_ids - source_client_ids
    if outside_clients:
        warnings.append(f"client rows outside final import_zayavki: {len(outside_clients)}")

    service_ids = [str(row[0]).strip() for row in client_rows]
    duplicated_service_ids = [item for item, count in Counter(service_ids).items() if item and count > 1]
    if duplicated_service_ids:
        errors.append(f"duplicate service_id values: {len(duplicated_service_ids)}")

    required_indexes = {
        "service_id": 0,
        "client_id": 1,
        "client_fio": 3,
        "service_name": 4,
        "create_date": 5,
        "payment_date": 6,
        "end_date": 8,
        "price": 11,
        "amount_of_payment": 12,
        "payment_left": 13,
        "manager": 15,
        "филиал": 16,
    }
    blanks = Counter()
    for row in client_rows:
        for field, idx in required_indexes.items():
            if blank(row[idx]):
                blanks[field] += 1
    if blanks:
        errors.append(f"required blanks: {dict(blanks)}")

    invalid_date_order = 0
    unparsable_dates = 0
    for row in client_rows:
        activation_date = parse_date(row[7])
        end_date = parse_date(row[8])
        if end_date is None or (not blank(row[7]) and activation_date is None):
            unparsable_dates += 1
        elif activation_date is not None and activation_date > end_date:
            invalid_date_order += 1
    if unparsable_dates:
        errors.append(f"rows with non-date activation/end values: {unparsable_dates}")
    if invalid_date_order:
        errors.append(f"rows with activation_date after end_date: {invalid_date_order}")

    payment_type_counts = Counter(str(row[14] or "blank") for row in client_rows)
    branch_counts = Counter(str(row[16] or "blank") for row in client_rows)
    invalid_branches = sorted(branch for branch in branch_counts if branch not in ALLOWED_BRANCHES)
    if invalid_branches:
        errors.append(f"invalid branch values: {invalid_branches}")
    blank_payment_positive = 0
    for row in client_rows:
        if row[14] in (None, "") and (row[11] or 0) > 0:
            blank_payment_positive += 1
    if blank_payment_positive:
        warnings.append(f"positive-price rows with blank payment type: {blank_payment_positive}")

    coverage_missing_selected = []
    if coverage_path.exists():
        with coverage_path.open("r", encoding="utf-8-sig", newline="") as handle:
            for row in csv.DictReader(handle):
                if int(row.get("selected_rows") or 0) == 0:
                    coverage_missing_selected.append(row.get("service_name", ""))
    if coverage_missing_selected:
        warnings.append(f"services without client rows, template only: {len(coverage_missing_selected)}")

    row_by_service_id = {
        str(row[0]).strip(): row
        for row in client_rows
        if str(row[0] or "").strip()
    }
    reports_dir = output_dir / "reports"
    full_date_audit = read_csv_rows(reports_dir / "services_end_dates_audit.csv")
    active_audit = read_csv_rows(reports_dir / "services_active_rows_audit.csv")
    live_audit = read_csv_rows(
        reports_dir / "services_live_active_membership_audit.csv"
    )
    fallback_audit = read_csv_rows(reports_dir / "services_end_date_fallbacks.csv")
    for name in [
        "services_end_dates_audit.csv",
        "services_active_rows_audit.csv",
        "services_live_active_membership_audit.csv",
        "services_end_date_fallbacks.csv",
    ]:
        if not (reports_dir / name).is_file():
            errors.append(f"missing required date audit: {name}")

    audit_ids = [
        str(row.get("service_id") or "").strip()
        for row in full_date_audit
    ]
    if len(audit_ids) != len(set(audit_ids)):
        errors.append("duplicate service_id values in full end-date audit")
    if set(audit_ids) != set(row_by_service_id):
        errors.append(
            "full end-date audit does not exactly cover every XLSX service_id"
        )
    if len(full_date_audit) != len(client_rows):
        errors.append(
            "full end-date audit row count does not match service XLSX: "
            f"{len(full_date_audit)} != {len(client_rows)}"
        )
    full_audit_by_service_id = {
        str(row.get("service_id") or "").strip(): row
        for row in full_date_audit
    }

    audit_date_mismatches = 0
    for audit_row in full_date_audit:
        service_id = str(audit_row.get("service_id") or "").strip()
        xlsx_row = row_by_service_id.get(service_id)
        if xlsx_row is None:
            audit_date_mismatches += 1
            continue
        if (
            parse_date(xlsx_row[7]) != parse_date(audit_row.get("activation_date"))
            or parse_date(xlsx_row[8]) != parse_date(audit_row.get("end_date"))
        ):
            audit_date_mismatches += 1
    if audit_date_mismatches:
        errors.append(f"XLSX/date-audit mismatches: {audit_date_mismatches}")

    active_ids = {
        str(row.get("service_id") or "").strip()
        for row in active_audit
    }
    expected_active_ids = {
        str(row.get("service_id") or "").strip()
        for row in full_date_audit
        if row.get("row_kind") == "active"
    }
    if active_ids != expected_active_ids or len(active_audit) != len(active_ids):
        errors.append(
            "active service audit does not exactly match full-audit active rows"
        )

    expected_live_ids = {
        str(row.get("service_id") or "").strip()
        for row in full_date_audit
        if row.get("is_active_by_date") == "1"
    }
    live_ids = {
        str(row.get("service_id") or "").strip()
        for row in live_audit
    }
    if live_ids != expected_live_ids or len(live_audit) != len(live_ids):
        errors.append(
            "live active audit does not exactly match active rows with is_active_by_date=1"
        )

    invalid_live_rows = 0
    for row in live_audit:
        end_date = parse_date(row.get("end_date"))
        register_start_date = parse_date(row.get("register_start_date"))
        client_id = str(row.get("client_id") or "").strip()
        if (
            row.get("end_date_source") != REGISTER_END_DATE_SOURCE
            or row.get("source_funnel") != ACTIVE_MEMBERSHIP_FUNNEL
            or source_clients.get(client_id) != ACTIVE_MEMBERSHIP_FUNNEL
            or row.get("is_active_by_date") != "1"
            or register_start_date is None
            or end_date is None
            or register_start_date > cutoff_date
            or end_date < cutoff_date
        ):
            invalid_live_rows += 1
    if invalid_live_rows:
        errors.append(
            "live service rows without real dates/current coverage/active membership: "
            f"{invalid_live_rows}"
        )

    expected_fallback_ids = {
        str(row.get("service_id") or "").strip()
        for row in full_date_audit
        if row.get("end_date_source") == SALE_DATE_FALLBACK_SOURCE
    }
    fallback_ids = {
        str(row.get("service_id") or "").strip()
        for row in fallback_audit
    }
    if (
        fallback_ids != expected_fallback_ids
        or len(fallback_audit) != len(fallback_ids)
    ):
        errors.append(
            "end-date fallback audit does not exactly match full-audit fallbacks"
        )

    invalid_fallback_rows = 0
    for row in fallback_audit:
        if (
            row.get("end_date_source") != SALE_DATE_FALLBACK_SOURCE
            or parse_date(row.get("end_date")) != parse_date(row.get("create_date"))
            or row.get("is_active_by_date") == "1"
        ):
            invalid_fallback_rows += 1
    if invalid_fallback_rows:
        errors.append(f"invalid conservative end-date fallbacks: {invalid_fallback_rows}")

    real_end_date_rows = sum(
        row.get("end_date_source") == REGISTER_END_DATE_SOURCE
        for row in full_date_audit
    )
    unknown_end_sources = sorted(
        {
            str(row.get("end_date_source") or "")
            for row in full_date_audit
        }
        - {REGISTER_END_DATE_SOURCE, SALE_DATE_FALLBACK_SOURCE}
    )
    if unknown_end_sources:
        errors.append(f"unknown end-date sources: {unknown_end_sources}")
    if real_end_date_rows + len(fallback_audit) != len(client_rows):
        errors.append(
            "real and fallback end-date source counts do not cover the XLSX"
        )
    if (
        args.expected_real_end_rows is not None
        and real_end_date_rows != args.expected_real_end_rows
    ):
        errors.append(
            "real InfoRg end-date regression: "
            f"{real_end_date_rows} != {args.expected_real_end_rows}"
        )
    if (
        args.expected_fallback_rows is not None
        and len(fallback_audit) != args.expected_fallback_rows
    ):
        errors.append(
            "end-date fallback regression: "
            f"{len(fallback_audit)} != {args.expected_fallback_rows}"
        )
    if (
        args.expected_live_rows is not None
        and len(live_audit) != args.expected_live_rows
    ):
        errors.append(
            f"live service regression: {len(live_audit)} != {args.expected_live_rows}"
        )
    if bool(args.expected_case_service_id) != bool(args.expected_case_end_date):
        errors.append(
            "expected case requires both --expected-case-service-id and "
            "--expected-case-end-date"
        )
    elif args.expected_case_service_id:
        expected_case_date = parse_date(args.expected_case_end_date)
        case_id = str(args.expected_case_service_id).strip()
        case_row = row_by_service_id.get(case_id)
        case_audit = full_audit_by_service_id.get(case_id)
        if (
            expected_case_date is None
            or case_row is None
            or parse_date(case_row[8]) != expected_case_date
            or case_audit is None
            or case_audit.get("end_date_source") != REGISTER_END_DATE_SOURCE
        ):
            errors.append(
                "named service end-date regression: "
                f"service_id={case_id!r}, expected={args.expected_case_end_date!r}"
            )

    report = [
        "# Services import XLSX validation",
        "",
        f"- client rows: {len(client_rows)}",
        f"- template rows: {len(template_rows)}",
        f"- required services: {len(service_names)}",
        f"- services represented in client rows: {len(client_service_names)}",
        f"- source final clients: {len(source_client_ids)}",
        f"- row clients: {len(client_ids)}",
        f"- duplicate service_id values: {len(duplicated_service_ids)}",
        f"- rows with a real InfoRg3060 end date: {real_end_date_rows}",
        f"- rows with conservative sale-date fallback: {len(fallback_audit)}",
        f"- selected balance/date-active rows audited: {len(active_audit)}",
        f"- live services on cutoff: {len(live_audit)}",
        f"- live services belonging to active-membership clients: {sum(row.get('source_funnel') == ACTIVE_MEMBERSHIP_FUNNEL for row in live_audit)}",
        f"- date audit mismatches: {audit_date_mismatches}",
        f"- full end-date audit coverage: {len(full_date_audit)}/{len(client_rows)}",
        f"- named regression case: {args.expected_case_service_id or 'not configured'}",
        f"- payment types: {dict(payment_type_counts)}",
        f"- branches: {dict(branch_counts)}",
        f"- template-only services: {len(coverage_missing_selected)}",
        f"- status: {'PASS' if not errors else 'FAIL'}",
        "",
        "## Errors",
        "",
    ]
    report.extend(f"- {item}" for item in errors)
    if not errors:
        report.append("- none")
    report.extend(["", "## Warnings", ""])
    report.extend(f"- {item}" for item in warnings)
    if not warnings:
        report.append("- none")

    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "services_validation_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    print("\n".join(report))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
