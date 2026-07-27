#!/usr/bin/env python3
"""Build full active-problem workbooks in the normal membership import shape."""

from __future__ import annotations

import csv
import os
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = os.environ.get("ACTIVE_PROBLEM_DATE_STAMP", "").strip()
CUTOFF_DATE = os.environ.get("ACTIVE_PROBLEM_CUTOFF_DATE", "").strip()

if not DATE_STAMP or not CUTOFF_DATE:
    raise RuntimeError(
        "ACTIVE_PROBLEM_DATE_STAMP and ACTIVE_PROBLEM_CUTOFF_DATE are required"
    )
if DATE_STAMP != CUTOFF_DATE.replace("-", ""):
    raise RuntimeError("ACTIVE_PROBLEM_DATE_STAMP must match ACTIVE_PROBLEM_CUTOFF_DATE")

OUTPUT_DIR = ROOT / os.environ.get("ACTIVE_PROBLEM_OUTPUT_DIR", "output/20251115_0800_fix_owner_new_import")
MAIN_XLSX_PATH = OUTPUT_DIR / f"fitbase_import_abonementy_clientov_{DATE_STAMP}.xlsx"
STAGING_CSV_PATH = OUTPUT_DIR / "staging/membership_import_rows.csv"


@dataclass(frozen=True)
class CaseWorkbookSpec:
    key: str
    output_prefix: str
    description: str
    predicate: Callable[[dict[str, str]], bool]
    sort_key: Callable[[dict[str, str]], tuple[Any, ...]]


def decimal_value(value: str | None) -> Decimal:
    text = (value or "").replace(" ", "").replace(",", ".").strip()
    if not text:
        return Decimal("0")
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal("0")


def is_not_finished(row: dict[str, str]) -> bool:
    return (row.get("end_date") or "") >= CUTOFF_DATE


def is_no_payment_cash_active(row: dict[str, str]) -> bool:
    return (
        row.get("_product_class") == "full_subscription"
        and is_not_finished(row)
        and decimal_value(row.get("price")) > 0
        and (row.get("type_of_payment") or "") == "наличные"
        and not (row.get("_payment_match_source") or "").strip()
    )


def is_zero_price_direct_active_full(row: dict[str, str]) -> bool:
    return (
        row.get("_product_class") == "full_subscription"
        and (row.get("activation_date") or "") <= CUTOFF_DATE
        and is_not_finished(row)
        and decimal_value(row.get("price")) == 0
        and (row.get("_payment_match_source") or "").startswith("direct_doc152")
        and decimal_value(row.get("_document131_posted_unmarked_refund_count")) == 0
        and bool((row.get("type_of_payment") or "").strip())
    )


def is_non_named_payment_left_active(row: dict[str, str]) -> bool:
    return (
        is_not_finished(row)
        and decimal_value(row.get("payment_left")) > 0
        and "рассроч" not in (row.get("contract_name") or "").lower()
    )


def sort_by_client_and_date(row: dict[str, str]) -> tuple[Any, ...]:
    return (
        (row.get("client_fio") or "").lower(),
        row.get("client_id") or "",
        row.get("payment_date") or "",
        row.get("contract_id") or "",
    )


def sort_by_payment_left_desc(row: dict[str, str]) -> tuple[Any, ...]:
    return (
        -float(decimal_value(row.get("payment_left"))),
        (row.get("client_fio") or "").lower(),
        row.get("payment_date") or "",
        row.get("contract_id") or "",
    )


SPECS = [
    CaseWorkbookSpec(
        key="active_no_payment_cash",
        output_prefix="active_problem_1_no_payment_cash",
        description="Активные/будущие full-членства: price>0, type_of_payment=наличные, платеж не найден.",
        predicate=is_no_payment_cash_active,
        sort_key=sort_by_client_and_date,
    ),
    CaseWorkbookSpec(
        key="active_zero_price_direct_full",
        output_prefix="active_problem_2_zero_price_direct_full",
        description="Активные full-членства: price=0, direct payment есть, возврата Document131 нет.",
        predicate=is_zero_price_direct_active_full,
        sort_key=sort_by_client_and_date,
    ),
    CaseWorkbookSpec(
        key="active_non_named_payment_left",
        output_prefix="active_problem_3_non_named_payment_left",
        description="Активные строки с payment_left>0 без слова рассрочка в названии.",
        predicate=is_non_named_payment_left_active,
        sort_key=sort_by_payment_left_desc,
    ),
]


def read_staging_rows() -> list[dict[str, str]]:
    with STAGING_CSV_PATH.open(newline="", encoding="utf-8-sig") as handle:
        return list(csv.DictReader(handle))


def read_main_rows() -> tuple[list[str], dict[str, list[Any]]]:
    wb = load_workbook(MAIN_XLSX_PATH, read_only=True, data_only=True)
    ws = wb.active
    headers = [str(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    contract_idx = headers.index("contract_id")

    rows: dict[str, list[Any]] = {}
    for values in ws.iter_rows(min_row=3, values_only=True):
        contract_id = str(values[contract_idx] or "")
        rows[contract_id] = list(values)

    wb.close()
    return headers, rows


def autosize(ws) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_workbook(headers: list[str], rows: list[list[Any]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт_абонементы"
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_ws(ws)
    return wb


def main() -> None:
    staging_rows = read_staging_rows()
    headers, main_rows_by_contract = read_main_rows()

    for spec in SPECS:
        selected = sorted((row for row in staging_rows if spec.predicate(row)), key=spec.sort_key)
        missing = [row["contract_id"] for row in selected if row["contract_id"] not in main_rows_by_contract]
        if missing:
            raise RuntimeError(f"{spec.key}: rows not found in main workbook: {missing[:10]}")

        workbook_rows = [main_rows_by_contract[row["contract_id"]] for row in selected]
        wb = build_workbook(headers, workbook_rows)
        output_path = OUTPUT_DIR / f"{spec.output_prefix}_{len(selected)}_cases_{DATE_STAMP}.xlsx"
        wb.save(output_path)

        unique_clients = len({row["client_id"] for row in selected})
        print(
            f"{spec.key}: {len(selected)} rows, {unique_clients} clients -> "
            f"{output_path.relative_to(ROOT)}"
        )


if __name__ == "__main__":
    main()
