#!/usr/bin/env python3
"""Extract customer checked-file decisions for Part 2 reports."""

from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def cell_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value).strip()


def read_rows(path: Path, min_cols: int) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        values = [cell_value(value) for value in row[:min_cols]]
        if any(values[:3]):
            rows.append(values)
    wb.close()
    return rows


def read_card_rows(path: Path) -> list[list[str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [cell_value(value) for value in row[:4]]
        if any(values[:3]):
            rows.append(values)
    wb.close()
    return rows


def write_csv(path: Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fieldnames})


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def normalize_fio(value: str) -> str:
    return " ".join((value or "").lower().split())


def normalize_phone_token(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) == 11 and digits.startswith("8"):
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return digits


def normalize_phones(value: str) -> str:
    phones = []
    for part in (value or "").split(","):
        phone = normalize_phone_token(part)
        if phone and phone not in phones:
            phones.append(phone)
    return ",".join(sorted(phones))


def normalize_date_token(value: str) -> str:
    value = (value or "").strip()
    if not value:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(value, fmt).date().isoformat()
        except ValueError:
            continue
    return ""


def extract_date_options(value: str) -> list[str]:
    raw = (value or "").strip()
    if not raw:
        return []
    tokens = re.findall(r"\d{4}-\d{2}-\d{2}|\d{1,2}\.\d{1,2}\.\d{4}", raw)
    dates: list[str] = []
    for token in tokens:
        normalized = normalize_date_token(token)
        if normalized and normalized not in dates:
            dates.append(normalized)
    return dates


def build_card_rule_map(detail_path: Path) -> dict[tuple[str, str], dict[str, str]]:
    rows = read_csv(detail_path)
    grouped: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        key = (normalize_fio(row.get("client_fio", "")), normalize_phones(row.get("phones", "")))
        if key[0] or key[1]:
            grouped[key].append(row)
    result: dict[tuple[str, str], dict[str, str]] = {}
    for key, items in grouped.items():
        selected = sorted(
            items,
            key=lambda row: (row.get("issue_date", ""), row.get("card_ref", "")),
            reverse=True,
        )[0]
        result[key] = selected
    return result


def classify_subscription_comment(comment: str) -> str:
    lowered = comment.lower()
    flags = []
    if "ошиб" in lowered or "невер" in lowered:
        flags.append("system_or_operator_error")
    if "владел" in lowered:
        flags.append("owner_change")
    if "модифик" in lowered:
        flags.append("modifier")
    if "удален" in lowered or "удаление" in lowered:
        flags.append("marked_for_deletion")
    if "одно действующее" in lowered:
        flags.append("single_active_confirmed")
    return ";".join(flags) or "comment_only"


def build_subscription_date_match_rows(
    checked_rows: list[dict[str, str]],
    final_rows_path: Path,
    selection_report_path: Path,
) -> list[dict[str, str]]:
    if not final_rows_path.exists() or not selection_report_path.exists():
        return []

    final_by_id = {row["client_id"]: row for row in read_csv(final_rows_path) if row.get("client_id")}
    active_candidates = [
        row
        for row in read_csv(selection_report_path)
        if row.get("candidate_for_funnel") == "active"
    ]
    selected_active_by_id = {
        row["client_id"]: row
        for row in active_candidates
        if row.get("selected") == "1"
    }
    candidates_by_id: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in active_candidates:
        candidates_by_id[row.get("client_id", "")].append(row)

    match_rows: list[dict[str, str]] = []
    for checked in checked_rows:
        checked_sale = extract_date_options(checked.get("checked_sale_date", ""))
        checked_start = extract_date_options(checked.get("checked_start_date", ""))
        checked_end = extract_date_options(checked.get("checked_end_date", ""))
        if not (checked_sale or checked_start or checked_end):
            continue

        client_id = checked.get("client_id", "")
        selected = selected_active_by_id.get(client_id, {})
        final = final_by_id.get(client_id, {})
        sale_match = bool(selected.get("sale_date") and selected.get("sale_date") in checked_sale) if checked_sale else True
        start_match = bool(selected.get("start_date") and selected.get("start_date") in checked_start) if checked_start else True
        end_match = bool(selected.get("end_date") and selected.get("end_date") in checked_end) if checked_end else True
        all_match = sale_match and start_match and end_match

        matching_candidate: dict[str, str] = {}
        for candidate in candidates_by_id.get(client_id, []):
            candidate_matches = (
                (not checked_sale or candidate.get("sale_date") in checked_sale)
                and (not checked_start or candidate.get("start_date") in checked_start)
                and (not checked_end or candidate.get("end_date") in checked_end)
            )
            if candidate_matches:
                matching_candidate = candidate
                break

        match_rows.append(
            {
                "client_id": client_id,
                "client_fio": checked.get("client_fio", ""),
                "checked_sale_date_raw": checked.get("checked_sale_date", ""),
                "checked_start_date_raw": checked.get("checked_start_date", ""),
                "checked_end_date_raw": checked.get("checked_end_date", ""),
                "checked_sale_date_options": "|".join(checked_sale),
                "checked_start_date_options": "|".join(checked_start),
                "checked_end_date_options": "|".join(checked_end),
                "selected_subscription_ref": selected.get("subscription_ref", ""),
                "selected_subscription_name": selected.get("subscription_name", ""),
                "selected_sale_date": selected.get("sale_date", ""),
                "selected_start_date": selected.get("start_date", ""),
                "selected_end_date": selected.get("end_date", ""),
                "selected_rank_number": selected.get("rank_number", ""),
                "active_candidate_count": final.get("active_full_subscription_count", ""),
                "sale_date_match": "1" if sale_match else "0",
                "start_date_match": "1" if start_match else "0",
                "end_date_match": "1" if end_match else "0",
                "all_checked_dates_match_selected": "1" if all_match else "0",
                "any_active_candidate_matches_checked_dates": "1" if matching_candidate else "0",
                "matching_candidate_ref": matching_candidate.get("subscription_ref", ""),
                "matching_candidate_rank": matching_candidate.get("rank_number", ""),
                "comment": checked.get("comment", ""),
                "decision_flags": checked.get("decision_flags", ""),
            }
        )
    return match_rows


def review(args: argparse.Namespace) -> None:
    checked_dir = as_abs(args.checked_dir)
    reports_dir = as_abs(args.reports_dir)
    reports_dir.mkdir(parents=True, exist_ok=True)

    sub_path = checked_dir / "fitbase_active_clients_import_zayavki_20260429__05_tolko_neskolko_abonementov(проверено).xlsx"
    phone_path = checked_dir / "fitbase_active_clients_import_zayavki_20260429__07_tolko_net_telefona(проверено).xlsx"
    card_path = checked_dir / "fitbase_active_clients_plastic_cards_20260429__02_tolko_neskolko_kart (проверено).xlsx"

    subscription_rows = []
    for values in read_rows(sub_path, 13):
        subscription_rows.append(
            {
                "client_id": values[0],
                "phone": values[1],
                "client_fio": values[2],
                "checked_sale_date": values[9] if len(values) > 9 else "",
                "checked_start_date": values[10] if len(values) > 10 else "",
                "checked_end_date": values[11] if len(values) > 11 else "",
                "comment": values[12] if len(values) > 12 else "",
                "decision_flags": classify_subscription_comment(values[12] if len(values) > 12 else ""),
            }
        )
    write_csv(
        reports_dir / "checked_subscription_decisions.csv",
        subscription_rows,
        [
            "client_id",
            "phone",
            "client_fio",
            "checked_sale_date",
            "checked_start_date",
            "checked_end_date",
            "comment",
            "decision_flags",
        ],
    )

    subscription_date_match_rows = build_subscription_date_match_rows(
        subscription_rows,
        as_abs(args.final_funnel_clients),
        as_abs(args.subscription_selection_report),
    )
    write_csv(
        reports_dir / "checked_subscription_date_match.csv",
        subscription_date_match_rows,
        [
            "client_id",
            "client_fio",
            "checked_sale_date_raw",
            "checked_start_date_raw",
            "checked_end_date_raw",
            "checked_sale_date_options",
            "checked_start_date_options",
            "checked_end_date_options",
            "selected_subscription_ref",
            "selected_subscription_name",
            "selected_sale_date",
            "selected_start_date",
            "selected_end_date",
            "selected_rank_number",
            "active_candidate_count",
            "sale_date_match",
            "start_date_match",
            "end_date_match",
            "all_checked_dates_match_selected",
            "any_active_candidate_matches_checked_dates",
            "matching_candidate_ref",
            "matching_candidate_rank",
            "comment",
            "decision_flags",
        ],
    )

    missing_phone_rows = []
    for values in read_rows(phone_path, 10):
        missing_phone_rows.append(
            {
                "client_id": values[0],
                "phone": values[1],
                "client_fio": values[2],
                "checked_phone_comment": values[9] if len(values) > 9 else "",
                "confirmed_missing_phone": "1" if (values[9] if len(values) > 9 else "").lower() == "нет" else "0",
            }
        )
    write_csv(
        reports_dir / "checked_missing_phone_confirmations.csv",
        missing_phone_rows,
        ["client_id", "phone", "client_fio", "checked_phone_comment", "confirmed_missing_phone"],
    )

    rule_map = build_card_rule_map(as_abs(args.previous_multiple_cards_detail))
    card_decision_rows = []
    for phone, fio, card_list, checked_selected in read_card_rows(card_path):
        key = (normalize_fio(fio), normalize_phones(phone))
        rule = rule_map.get(key, {})
        rule_selected = rule.get("plastic_card_number", "")
        if checked_selected:
            match_status = "matches_rule" if checked_selected == rule_selected else "differs_from_rule"
        else:
            match_status = "no_checked_selection"
        card_decision_rows.append(
            {
                "phone": phone,
                "client_fio": fio,
                "all_cards_from_checked_file": card_list,
                "checked_selected_card": checked_selected,
                "rule_selected_card": rule_selected,
                "rule_selected_card_ref": rule.get("card_ref", ""),
                "rule_selected_issue_date": rule.get("issue_date", ""),
                "selection_rule": "issue_date DESC, card_ref DESC",
                "match_status": match_status,
            }
        )
    write_csv(
        reports_dir / "checked_card_decisions.csv",
        card_decision_rows,
        [
            "phone",
            "client_fio",
            "all_cards_from_checked_file",
            "checked_selected_card",
            "rule_selected_card",
            "rule_selected_card_ref",
            "rule_selected_issue_date",
            "selection_rule",
            "match_status",
        ],
    )

    summary_lines = [
        "# Checked Files Review Summary",
        "",
        f"reviewed_at: `{datetime.now().isoformat(timespec='seconds')}`",
        "",
        "## Files",
        "",
        f"- `{sub_path.relative_to(ROOT)}`",
        f"- `{phone_path.relative_to(ROOT)}`",
        f"- `{card_path.relative_to(ROOT)}`",
        "",
        "## Counts",
        "",
        f"- subscription checked rows: `{len(subscription_rows)}`",
        f"- subscription checked rows with explicit dates: `{len(subscription_date_match_rows)}`",
        f"- checked subscription date rows matching selected subscription: `{sum(1 for row in subscription_date_match_rows if row['all_checked_dates_match_selected'] == '1')}`",
        f"- checked subscription date rows differing from selected subscription: `{sum(1 for row in subscription_date_match_rows if row['all_checked_dates_match_selected'] != '1')}`",
        f"- missing-phone checked rows: `{len(missing_phone_rows)}`",
        f"- card checked rows: `{len(card_decision_rows)}`",
        f"- card rows with checked selected card: `{sum(1 for row in card_decision_rows if row['checked_selected_card'])}`",
        f"- checked selected cards matching rule: `{sum(1 for row in card_decision_rows if row['match_status'] == 'matches_rule')}`",
        f"- checked selected cards differing from rule: `{sum(1 for row in card_decision_rows if row['match_status'] == 'differs_from_rule')}`",
        "",
        "## Decision",
        "",
        "Checked-file notes were extracted into CSV reports. Missing-phone confirmations are treated as authoritative report-only rows; the pipeline does not try to auto-fill those phones from backup data.",
        "",
    ]
    (reports_dir / "checked_review_summary.md").write_text("\n".join(summary_lines), encoding="utf-8")
    print(f"subscription_rows={len(subscription_rows)}")
    print(f"subscription_date_match_rows={len(subscription_date_match_rows)}")
    print(f"missing_phone_rows={len(missing_phone_rows)}")
    print(f"card_rows={len(card_decision_rows)}")
    print(f"summary={(reports_dir / 'checked_review_summary.md').relative_to(ROOT)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--checked-dir", default=str(ROOT / "output" / "splits" / "checked"))
    parser.add_argument("--reports-dir", default=str(ROOT / "output" / "part2_20260429" / "reports"))
    parser.add_argument("--previous-multiple-cards-detail", default=str(ROOT / "output" / "multiple_cards_detail.csv"))
    parser.add_argument("--final-funnel-clients", default=str(ROOT / "output" / "part2_20260429" / "staging" / "final_funnel_clients.csv"))
    parser.add_argument("--subscription-selection-report", default=str(ROOT / "output" / "part2_20260429" / "reports" / "subscription_selection_report.csv"))
    return parser.parse_args()


def main() -> None:
    review(parse_args())


if __name__ == "__main__":
    main()
