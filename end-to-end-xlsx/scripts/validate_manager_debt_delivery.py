#!/usr/bin/env python3
"""Independently validate the manager-debt post-processed delivery.

This validator does not import the builder.  It compares the final membership
row-by-row with the full source membership, permits only the three audited
financial overrides, confirms problem4 is the sole exclusion, and verifies
that all other XLSX files are byte-identical to the base delivery.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
FINANCIAL_FIELDS = ("price", "amount_of_payments", "payment_left")


def as_abs(path: str | Path) -> Path:
    value = Path(path)
    return value if value.is_absolute() else (ROOT / value).resolve()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def decimal_value(value: Any) -> Decimal:
    try:
        return Decimal(str(value).strip().replace(" ", "").replace(",", "."))
    except (InvalidOperation, AttributeError) as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc


def canonical_cell(value: Any) -> tuple[str, str]:
    if isinstance(value, datetime):
        return ("date", value.date().isoformat())
    if isinstance(value, date):
        return ("date", value.isoformat())
    if isinstance(value, (int, float, Decimal)) and not isinstance(value, bool):
        return ("number", str(Decimal(str(value)).normalize()))
    if value is None:
        return ("blank", "")
    return ("text", str(value))


def equal_values(left: Any, right: Any) -> bool:
    return canonical_cell(left) == canonical_cell(right)


def read_contract_ids(path: Path, data_start_row: int) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    if "contract_id" not in headers:
        workbook.close()
        raise ValueError(f"No contract_id in {path}")
    index = headers.index("contract_id")
    values = {
        str(row[index] or "").strip()
        for row in worksheet.iter_rows(min_row=data_start_row, values_only=True)
        if str(row[index] or "").strip()
    }
    workbook.close()
    return values


def one_match(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one {pattern!r} in {directory}, found {len(matches)}"
        )
    return matches[0]


def read_audit(path: Path) -> dict[str, dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    contracts = [row["contract_id"].strip() for row in rows]
    if not all(contracts) or len(set(contracts)) != len(contracts):
        raise ValueError("Audit has blank or duplicate contract_id values")
    return {row["contract_id"].strip(): row for row in rows}


def read_problem123_contracts(base_delivery: Path) -> dict[int, set[str]]:
    return {
        group: read_contract_ids(
            one_match(base_delivery, f"problem_{group}_*.xlsx"), 2
        )
        for group in (1, 2, 3)
    }


def validate_audit(
    *,
    audit: dict[str, dict[str, str]],
    problem_sets: dict[int, set[str]],
    expected: dict[str, Any],
    errors: list[str],
) -> dict[str, Any]:
    expected_resolution = expected["manager_debt_resolution"]
    problem_union = set().union(*problem_sets.values())
    if set(audit) != problem_union:
        errors.append(
            "audit contract set differs from base problem1-3: "
            f"missing={sorted(problem_union - set(audit))[:10]}, "
            f"unexpected={sorted(set(audit) - problem_union)[:10]}"
        )

    source_counts = Counter(row["resolution_source"] for row in audit.values())
    group_counts: dict[str, dict[str, int]] = {}
    for group in (1, 2, 3):
        rows = [
            row for row in audit.values() if int(row["problem_group"]) == group
        ]
        group_counts[str(group)] = {
            "total": len(rows),
            "manager_debt_report": sum(
                row["resolution_source"] == "manager_debt_report" for row in rows
            ),
            "fallback_not_in_manager_debt_report": sum(
                row["resolution_source"]
                == "fallback_not_in_manager_debt_report"
                for row in rows
            ),
        }
        if {row["contract_id"] for row in rows} != problem_sets[group]:
            errors.append(f"audit problem{group} contract set mismatch")

    changed_rows = 0
    changed_cells = 0
    sums = {field: Decimal("0") for field in FINANCIAL_FIELDS}
    for row in audit.values():
        row_changed = False
        for field, audit_name in zip(
            FINANCIAL_FIELDS,
            ("new_price", "new_amount_of_payments", "new_payment_left"),
            strict=True,
        ):
            sums[field] += decimal_value(row[audit_name])
        for old_name, new_name in (
            ("old_price", "new_price"),
            ("old_amount_of_payments", "new_amount_of_payments"),
            ("old_payment_left", "new_payment_left"),
        ):
            if decimal_value(row[old_name]) != decimal_value(row[new_name]):
                changed_cells += 1
                row_changed = True
        changed_rows += row_changed

    actual = {
        "resolved_contracts": len(audit),
        "manager_debt_report": source_counts["manager_debt_report"],
        "fallback_not_in_manager_debt_report": source_counts[
            "fallback_not_in_manager_debt_report"
        ],
        "changed_rows": changed_rows,
        "changed_cells": changed_cells,
        "by_problem_group": group_counts,
        "target_sums": {field: str(value) for field, value in sums.items()},
    }
    for key in (
        "resolved_contracts",
        "manager_debt_report",
        "fallback_not_in_manager_debt_report",
        "changed_rows",
        "changed_cells",
        "by_problem_group",
        "target_sums",
    ):
        if actual[key] != expected_resolution[key]:
            errors.append(
                f"audit {key}={actual[key]!r}, expected={expected_resolution[key]!r}"
            )
    return actual


def compare_membership(
    *,
    source_path: Path,
    output_path: Path,
    audit: dict[str, dict[str, str]],
    excluded_contracts: set[str],
    errors: list[str],
) -> dict[str, int]:
    source_workbook = load_workbook(source_path, read_only=True, data_only=False)
    output_workbook = load_workbook(output_path, read_only=True, data_only=False)
    source_sheet = source_workbook.active
    output_sheet = output_workbook.active
    source_headers = [
        str(value or "")
        for value in next(
            source_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    output_headers = [
        str(value or "")
        for value in next(
            output_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    if source_headers != output_headers:
        errors.append("output membership headers differ from full source")
    indexes = {header: index for index, header in enumerate(source_headers)}
    required = {"contract_id", *FINANCIAL_FIELDS}
    if not required <= set(indexes):
        source_workbook.close()
        output_workbook.close()
        raise ValueError(f"Membership is missing fields: {sorted(required - set(indexes))}")

    output_iterator = iter(output_sheet.iter_rows(min_row=3, values_only=True))
    seen_audit: set[str] = set()
    seen_excluded: set[str] = set()
    compared_rows = 0
    nonfinancial_differences = 0
    financial_differences_outside_audit = 0

    for source_row in source_sheet.iter_rows(min_row=3, values_only=True):
        contract_id = str(source_row[indexes["contract_id"]] or "").strip()
        if contract_id in excluded_contracts:
            seen_excluded.add(contract_id)
            continue
        try:
            output_row = next(output_iterator)
        except StopIteration:
            errors.append("output membership ended before full source")
            break
        compared_rows += 1
        output_contract = str(output_row[indexes["contract_id"]] or "").strip()
        if output_contract != contract_id:
            errors.append(
                f"membership order/contract mismatch at compared row {compared_rows}: "
                f"source={contract_id!r}, output={output_contract!r}"
            )
            break

        audit_row = audit.get(contract_id)
        if audit_row is not None:
            seen_audit.add(contract_id)
            for index, header in enumerate(source_headers):
                if header in FINANCIAL_FIELDS:
                    continue
                if not equal_values(source_row[index], output_row[index]):
                    nonfinancial_differences += 1
                    if nonfinancial_differences <= 10:
                        errors.append(
                            f"{contract_id}: non-financial field changed: {header}"
                        )
            expected_new = {
                "price": decimal_value(audit_row["new_price"]),
                "amount_of_payments": decimal_value(
                    audit_row["new_amount_of_payments"]
                ),
                "payment_left": decimal_value(audit_row["new_payment_left"]),
            }
            expected_old = {
                "price": decimal_value(audit_row["old_price"]),
                "amount_of_payments": decimal_value(
                    audit_row["old_amount_of_payments"]
                ),
                "payment_left": decimal_value(audit_row["old_payment_left"]),
            }
            for field in FINANCIAL_FIELDS:
                index = indexes[field]
                if decimal_value(source_row[index]) != expected_old[field]:
                    errors.append(f"{contract_id}: audit old {field} mismatch")
                if decimal_value(output_row[index]) != expected_new[field]:
                    errors.append(f"{contract_id}: output {field} differs from audit")
        else:
            for index, header in enumerate(source_headers):
                if equal_values(source_row[index], output_row[index]):
                    continue
                if header in FINANCIAL_FIELDS:
                    financial_differences_outside_audit += 1
                else:
                    nonfinancial_differences += 1
                if (
                    nonfinancial_differences
                    + financial_differences_outside_audit
                    <= 10
                ):
                    errors.append(
                        f"{contract_id or '<placeholder>'}: unaudited field changed: {header}"
                    )

    try:
        extra_row = next(output_iterator)
        errors.append(f"output membership has extra row beginning {extra_row[:3]!r}")
    except StopIteration:
        pass
    source_workbook.close()
    output_workbook.close()

    missing_audit = sorted(set(audit) - seen_audit)
    missing_excluded = sorted(excluded_contracts - seen_excluded)
    if missing_audit:
        errors.append(f"audited contracts missing from membership: {missing_audit[:10]}")
    if missing_excluded:
        errors.append(
            f"problem4 contracts missing from full source: {missing_excluded[:10]}"
        )
    return {
        "compared_rows": compared_rows,
        "audited_rows": len(seen_audit),
        "excluded_rows": len(seen_excluded),
        "nonfinancial_differences": nonfinancial_differences,
        "financial_differences_outside_audit": financial_differences_outside_audit,
    }


def validate(args: argparse.Namespace) -> int:
    config = yaml.safe_load(as_abs(args.config).read_text(encoding="utf-8"))
    expected = yaml.safe_load(as_abs(args.expected).read_text(encoding="utf-8"))
    base_delivery = as_abs(config["inputs"]["base_delivery"])
    full_membership = as_abs(config["inputs"]["full_membership"])
    output_dir = as_abs(config["output"]["directory"])
    membership_name = str(config["output"]["membership_file_name"])
    audit_path = output_dir / "reports" / "financial_resolution_audit.csv"
    report_path = as_abs(args.report)
    json_path = as_abs(args.json_report)

    errors: list[str] = []
    audit = read_audit(audit_path)
    problem_sets = read_problem123_contracts(base_delivery)
    audit_summary = validate_audit(
        audit=audit,
        problem_sets=problem_sets,
        expected=expected,
        errors=errors,
    )

    problem4_source = one_match(base_delivery, "problem_4_*.xlsx")
    problem4_output = output_dir / problem4_source.name
    problem4_contracts = read_contract_ids(problem4_source, 2)
    if sha256(problem4_source) != sha256(problem4_output):
        errors.append("problem4 workbook is not byte-identical to base delivery")

    static_names = [
        name
        for name in expected["files"]
        if name != membership_name and not name.startswith("problem_")
    ]
    static_hashes: dict[str, str] = {}
    for name in static_names:
        source = base_delivery / name
        destination = output_dir / name
        source_hash = sha256(source)
        output_hash = sha256(destination)
        static_hashes[name] = output_hash
        if source_hash != output_hash:
            errors.append(f"unchanged workbook differs from base: {name}")

    membership_summary = compare_membership(
        source_path=full_membership,
        output_path=output_dir / membership_name,
        audit=audit,
        excluded_contracts=problem4_contracts,
        errors=errors,
    )
    expected_rows = int(expected["files"][membership_name]["data_rows"])
    if membership_summary["compared_rows"] != expected_rows:
        errors.append(
            f"membership compared_rows={membership_summary['compared_rows']}, "
            f"expected={expected_rows}"
        )

    actual_xlsx_names = {path.name for path in output_dir.glob("*.xlsx")}
    expected_xlsx_names = set(expected["files"])
    if actual_xlsx_names != expected_xlsx_names:
        errors.append(
            f"XLSX set mismatch: missing={sorted(expected_xlsx_names - actual_xlsx_names)}, "
            f"unexpected={sorted(actual_xlsx_names - expected_xlsx_names)}"
        )

    verdict = "PASS" if not errors else "FAIL"
    result = {
        "verdict": verdict,
        "errors": errors,
        "audit": audit_summary,
        "membership": membership_summary,
        "problem4_contracts": sorted(problem4_contracts),
        "unchanged_workbook_sha256": static_hashes,
    }
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    report_lines = [
        "# Independent manager-debt delivery validation",
        "",
        f"- verdict: **{verdict}**",
        f"- audited problem1–3 contracts: `{audit_summary['resolved_contracts']}`",
        f"- manager debt rows: `{audit_summary['manager_debt_report']}`",
        f"- fallback rows: "
        f"`{audit_summary['fallback_not_in_manager_debt_report']}`",
        f"- clean membership rows compared: `{membership_summary['compared_rows']}`",
        f"- problem4 contracts excluded: `{membership_summary['excluded_rows']}`",
        f"- non-financial differences: "
        f"`{membership_summary['nonfinancial_differences']}`",
        f"- unaudited financial differences: "
        f"`{membership_summary['financial_differences_outside_audit']}`",
        "",
        "## Errors",
        "",
        *([f"- {error}" for error in errors] or ["- none"]),
        "",
    ]
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    print(f"verdict={verdict}")
    print(f"report={report_path}")
    for error in errors:
        print(f"ERROR: {error}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True)
    parser.add_argument("--expected", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--json-report", required=True)
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(validate(parse_args()))
