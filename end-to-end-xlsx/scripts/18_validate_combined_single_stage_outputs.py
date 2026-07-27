#!/usr/bin/env python3
"""Validate combined Part 2 single-stage Fitbase XLSX outputs."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAIN_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
    "филиал",
]
MAIN_RUS_HEADERS = [
    "Внутренний номер клиента ",
    "Телефон *",
    "ФИО клиента *",
    "Почта",
    "Воронка *",
    "Этап воронки *",
    "Бюджет ",
    "Дата создания *",
    "Менеджер ",
    "филиал",
]
CARD_HEADERS = ["телефон", "фио", "номер пластиковой карты"]
ALLOWED_FINAL_PAIRS = {
    ("новые заявки", "неразобранные"),
    ("Действующие абонементы", "Все действующие абонементы"),
    ("Реактивация(годовые абонементы)", "Все закрытые абонементы"),
}
FITBASE_LABELS = {
    "Новые заявки": ("новые заявки", "неразобранные"),
    "Действующие клиенты": ("Действующие абонементы", "Все действующие абонементы"),
    "Реактивация": ("Реактивация(годовые абонементы)", "Все закрытые абонементы"),
}
ALLOWED_BRANCHES = {
    "Фитнес Империя (Гоголевский)",
    "Фитнес Империя (Промышленная)",
    "Фитнес Империя (Ровио)",
    "Фитнес Империя (Столица)",
}
DEFAULT_BRANCHES_CONFIG = ROOT / "config" / "branches_by_club.yml"
OLD_STEPS = {
    "60-31 день до окончания",
    "30-8 дней до окончания",
    "7-0 день до окончания",
    "Действующие клиенты",
    "1-6 дней",
    "7-29 дней",
    "30-59 дней",
    "60-89 дней",
    "более 90 дней",
    "Неразобранные",
}
REQUIRED_REPORTS = [
    "funnel_distribution.csv",
    "stage_distribution_by_funnel.csv",
    "fitbase_funnel_distribution.csv",
    "single_stage_distribution.csv",
    "manager_distribution_by_club.csv",
    "missing_phone_report.csv",
    "missing_card_report.csv",
    "missing_club_report.csv",
    "multiple_subscriptions_report.csv",
    "subscription_selection_report.csv",
    "subscription_overrides_report.csv",
    "multiple_cards_report.csv",
    "card_selection_report.csv",
    "product_classification_preflight.csv",
    "product_classification_report.csv",
    "product_classification_review_report.csv",
    "product_reclassification_impact.md",
    "product_reclassification_applied.csv",
    "product_reclassification_funnel_impact.csv",
    "export_filter_summary.csv",
    "export_filter_funnel_distribution.csv",
    "export_filter_rules.md",
    "phone_deduplication_removed_clients.csv",
    "phone_deduplication_summary.csv",
    "branch_distribution.csv",
    "branch_distribution_by_club.csv",
]
PHONE_SPLIT_RE = re.compile(r"[,;]\s*")


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_branches(path: Path) -> dict[str, str]:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    branches = data.get("branches", {})
    if not isinstance(branches, dict) or not branches:
        raise ValueError(f"No branches found in {path}")
    return {str(club): str(branch) for club, branch in branches.items()}


def workbook_rows(path: Path, first_data_row: int, width: int) -> tuple[list[object], list[tuple[object, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, width + 1)]
    rows = [
        tuple(row)
        for row in ws.iter_rows(min_row=first_data_row, max_col=width, values_only=True)
        if any(value not in (None, "") for value in row)
    ]
    wb.close()
    return headers, rows


def workbook_row_values(path: Path, row_number: int, width: int) -> list[object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    values = [ws.cell(row_number, col).value for col in range(1, width + 1)]
    wb.close()
    return values


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(float(str(value)))
    except ValueError:
        return default


def has_phone(row: dict[str, str]) -> bool:
    return bool((row.get("phones") or "").strip())


def parse_date(value: str | None) -> date:
    if not value:
        return date.min
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except ValueError:
        return date.min


def int_client_id(value: str | None) -> int:
    digits = re.sub(r"\D+", "", value or "")
    return int(digits) if digits else 0


def normalize_phone_token(value: str) -> str:
    digits = re.sub(r"\D+", "", value or "")
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return ""


def normalize_phones(value: str) -> list[str]:
    phones: list[str] = []
    for part in PHONE_SPLIT_RE.split(value or ""):
        phone = normalize_phone_token(part)
        if phone and phone not in phones:
            phones.append(phone)
    return phones


def phone_dedupe_score(row: dict[str, str]) -> tuple[object, ...]:
    subscription_sale_date = parse_date(row.get("selected_subscription_sale_date"))
    subscription_end_date = parse_date(row.get("selected_subscription_end_date"))
    subscription_start_date = parse_date(row.get("selected_subscription_start_date"))
    create_date = parse_date(row.get("create_date"))
    has_subscription = int(
        bool(row.get("selected_subscription_ref"))
        or subscription_sale_date != date.min
        or subscription_end_date != date.min
    )
    funnel_priority = {
        "Действующие клиенты": 2,
        "Реактивация": 1,
        "Новые заявки": 0,
    }.get(row.get("funnel", ""), 0)
    has_card = int(bool((row.get("selected_card_number") or "").strip()))
    return (
        has_subscription,
        subscription_sale_date,
        subscription_end_date,
        subscription_start_date,
        funnel_priority,
        has_card,
        create_date,
        int_client_id(row.get("client_id")),
    )


def phone_component_key(row: dict[str, str]) -> tuple[object, ...]:
    phones = normalize_phones(row.get("phones", ""))
    return (min(phones) if phones else "", row.get("client_id", ""), row.get("client_ref", ""))


def dedupe_by_phone_keep_latest_subscription(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    phone_to_indexes: dict[str, list[int]] = defaultdict(list)
    for index, row in enumerate(rows):
        for phone in normalize_phones(row.get("phones", "")):
            phone_to_indexes[phone].append(index)

    parent = list(range(len(rows)))

    def find(index: int) -> int:
        while parent[index] != index:
            parent[index] = parent[parent[index]]
            index = parent[index]
        return index

    def union(left: int, right: int) -> None:
        left_root = find(left)
        right_root = find(right)
        if left_root != right_root:
            parent[right_root] = left_root

    for indexes in phone_to_indexes.values():
        if len(indexes) < 2:
            continue
        first = indexes[0]
        for index in indexes[1:]:
            union(first, index)

    components: dict[int, list[int]] = defaultdict(list)
    for index in range(len(rows)):
        components[find(index)].append(index)

    kept_indexes: set[int] = set()
    for component_indexes in components.values():
        kept_indexes.add(max(component_indexes, key=lambda index: phone_dedupe_score(rows[index])))
    return sorted((dict(row) for index, row in enumerate(rows) if index in kept_indexes), key=phone_component_key)


def duplicate_normalized_phone_count_from_stage_rows(rows: list[dict[str, str]]) -> int:
    phone_rows: Counter[str] = Counter()
    for row in rows:
        for phone in normalize_phones(row.get("phones", "")):
            phone_rows[phone] += 1
    return sum(1 for count in phone_rows.values() if count > 1)


def duplicate_normalized_phone_count_from_xlsx_rows(rows: list[tuple[object, ...]], phone_col: int) -> int:
    phone_rows: Counter[str] = Counter()
    for row in rows:
        for phone in normalize_phones(str(row[phone_col] or "")):
            phone_rows[phone] += 1
    return sum(1 for count in phone_rows.values() if count > 1)


def filter_expected_main_rows(
    rows: list[dict[str, str]],
    require_phone_for_new_applications: bool,
) -> list[dict[str, str]]:
    if not require_phone_for_new_applications:
        return rows
    return [
        row
        for row in rows
        if row.get("funnel") != "Новые заявки" or has_phone(row)
    ]


def is_new_application_refuser(row: dict[str, str]) -> bool:
    return row.get("funnel") == "Новые заявки" and row.get("funnel_step") == "Неразобранные"


def split_new_application_refusers(
    rows: list[dict[str, str]],
    transfer_new_applications_to_memberships: bool,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    if not transfer_new_applications_to_memberships:
        return rows, []
    return (
        [row for row in rows if not is_new_application_refuser(row)],
        [row for row in rows if is_new_application_refuser(row)],
    )


def filter_expected_card_rows(rows: list[dict[str, str]], funnel_filter: str) -> list[dict[str, str]]:
    if not funnel_filter:
        return rows
    return [row for row in rows if row.get("funnel") == funnel_filter]


def expected_fitbase_pair(row: dict[str, str]) -> tuple[str, str]:
    try:
        return FITBASE_LABELS[row.get("funnel", "")]
    except KeyError as exc:
        raise ValueError(f"Unknown internal funnel: {row.get('funnel', '')!r}") from exc


def expected_branch(row: dict[str, str], branches_by_club: dict[str, str]) -> str:
    return branches_by_club.get(row.get("normalized_club", ""), "")


def validate(args: argparse.Namespace) -> int:
    stage_dir = as_abs(args.stage_dir)
    output_dir = as_abs(args.output_dir)
    reports_dir = as_abs(args.reports_dir)
    branches_by_club = load_branches(as_abs(args.branches_config))
    date_stamp = args.date_stamp or args.cutoff_date.replace("-", "")

    errors: list[str] = []
    warnings: list[str] = []
    stage_path = stage_dir / "final_funnel_clients.csv"
    rows = read_csv(stage_path) if stage_path.exists() else []
    if not rows:
        errors.append(f"missing or empty final stage CSV: {stage_path.relative_to(ROOT)}")
    expected_main_rows = filter_expected_main_rows(rows, args.main_require_phone_for_new_applications)
    if args.dedupe_by_phone_keep_latest_subscription:
        expected_main_rows = dedupe_by_phone_keep_latest_subscription(expected_main_rows)
    expected_main_rows, expected_refuser_rows = split_new_application_refusers(
        expected_main_rows,
        args.main_transfer_new_applications_to_memberships,
    )
    card_source_rows = (
        expected_main_rows
        if args.dedupe_by_phone_keep_latest_subscription or args.main_transfer_new_applications_to_memberships
        else rows
    )
    expected_card_rows = filter_expected_card_rows(card_source_rows, args.cards_funnel_filter)

    main_path = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}__all_funnels.xlsx"
    cards_path = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}__all_funnels.xlsx"
    xlsx_files = sorted(output_dir.glob("*.xlsx"))
    expected_xlsx = {main_path, cards_path}
    if set(xlsx_files) != expected_xlsx:
        errors.append(
            "final output directory must contain exactly two XLSX files: "
            f"found {[path.name for path in xlsx_files]}"
        )

    main_rows: list[tuple[object, ...]] = []
    card_rows: list[tuple[object, ...]] = []
    if main_path.exists():
        headers, main_rows = workbook_rows(main_path, 3, len(MAIN_HEADERS))
        if headers != MAIN_HEADERS:
            errors.append(f"main technical headers mismatch: {main_path.relative_to(ROOT)}")
        if workbook_row_values(main_path, 2, len(MAIN_RUS_HEADERS)) != MAIN_RUS_HEADERS:
            errors.append(f"main Russian headers mismatch: {main_path.relative_to(ROOT)}")
    else:
        errors.append(f"missing main XLSX: {main_path.relative_to(ROOT)}")

    if cards_path.exists():
        headers, card_rows = workbook_rows(cards_path, 2, len(CARD_HEADERS))
        if headers != CARD_HEADERS:
            errors.append(f"cards headers mismatch: {cards_path.relative_to(ROOT)}")
    else:
        errors.append(f"missing cards XLSX: {cards_path.relative_to(ROOT)}")

    if len(main_rows) != len(expected_main_rows):
        errors.append(f"main row count mismatch: xlsx={len(main_rows)}, expected_after_filters={len(expected_main_rows)}")
    if len(card_rows) != len(expected_card_rows):
        errors.append(f"cards row count mismatch: xlsx={len(card_rows)}, expected_after_filters={len(expected_card_rows)}")

    final_pairs = Counter((str(row[4] or ""), str(row[5] or "")) for row in main_rows)
    expected_final_pairs = Counter(expected_fitbase_pair(row) for row in expected_main_rows)
    if final_pairs != expected_final_pairs:
        errors.append("main XLSX funnel/stage distribution does not match expected filtered stage rows")

    invalid_pairs = sorted(set(final_pairs) - ALLOWED_FINAL_PAIRS)
    if invalid_pairs:
        errors.append(f"invalid final funnel/funnel_step pairs: {invalid_pairs}")

    allowed_final_pairs = set(ALLOWED_FINAL_PAIRS)
    if args.main_transfer_new_applications_to_memberships:
        allowed_final_pairs.discard(FITBASE_LABELS["Новые заявки"])
    final_funnels = {pair[0] for pair in final_pairs}
    final_steps = {pair[1] for pair in final_pairs}
    if final_funnels != {pair[0] for pair in allowed_final_pairs}:
        errors.append(f"final XLSX funnels are not exactly the expected values: {sorted(final_funnels)}")
    if final_steps != {pair[1] for pair in allowed_final_pairs}:
        errors.append(f"final XLSX steps are not exactly the expected values: {sorted(final_steps)}")
    old_steps_found = sorted(step for step in final_steps if step in OLD_STEPS)
    if old_steps_found:
        errors.append(f"old multi-stage names found in final XLSX: {old_steps_found}")

    stage_refs = [row.get("client_ref", "") for row in rows if row.get("client_ref")]
    duplicate_refs = [key for key, count in Counter(stage_refs).items() if count > 1]
    if duplicate_refs:
        errors.append(f"duplicate client_ref rows in stage: {len(duplicate_refs)}")

    xlsx_client_ids = [str(row[0] or "") for row in main_rows if row and row[0] not in (None, "")]
    duplicate_xlsx_ids = [key for key, count in Counter(xlsx_client_ids).items() if count > 1]
    if duplicate_xlsx_ids:
        errors.append(f"duplicate client_id rows in final XLSX: {len(duplicate_xlsx_ids)}")
    expected_main_ids = Counter(row.get("client_id", "") for row in expected_main_rows if row.get("client_id"))
    if Counter(xlsx_client_ids) != expected_main_ids:
        errors.append("main XLSX client_id set does not match expected filtered stage rows")
    actual_branch_counts = Counter(str(row[9] or "") for row in main_rows if len(row) >= 10)
    expected_branch_counts = Counter(expected_branch(row, branches_by_club) for row in expected_main_rows)
    if actual_branch_counts != expected_branch_counts:
        errors.append("main XLSX branch distribution does not match expected normalized_club mapping")
    invalid_branches = sorted(branch for branch in actual_branch_counts if branch not in ALLOWED_BRANCHES)
    if invalid_branches:
        errors.append(f"invalid branch values in main XLSX: {invalid_branches}")
    blank_branch_rows = sum(1 for row in main_rows if len(row) >= 10 and not str(row[9] or "").strip())
    if blank_branch_rows:
        errors.append(f"blank branch values in main XLSX: {blank_branch_rows}")

    comma_cards = sum(1 for row in card_rows if len(row) >= 3 and isinstance(row[2], str) and "," in row[2])
    if comma_cards:
        errors.append(f"card XLSX contains comma-separated card values: {comma_cards}")
    expected_card_values = Counter(
        (
            row.get("phones", ""),
            row.get("client_fio", ""),
            row.get("selected_card_number", ""),
        )
        for row in expected_card_rows
    )
    actual_card_values = Counter(
        (
            str(row[0] or ""),
            str(row[1] or ""),
            str(row[2] or ""),
        )
        for row in card_rows
    )
    if actual_card_values != expected_card_values:
        errors.append("card XLSX rows do not match expected filtered stage rows")
    card_numbers = [str(row[2] or "").strip() for row in card_rows if len(row) >= 3 and str(row[2] or "").strip()]
    duplicate_card_numbers = [key for key, count in Counter(card_numbers).items() if count > 1]
    if duplicate_card_numbers:
        errors.append(f"duplicate non-empty plastic card numbers in card XLSX: {len(duplicate_card_numbers)}")
    new_application_rows_without_phone = [
        row for row in main_rows if str(row[4] or "") == "новые заявки" and not str(row[1] or "").strip()
    ]
    if args.main_require_phone_for_new_applications and new_application_rows_without_phone:
        errors.append(f"new application rows without phone still exported: {len(new_application_rows_without_phone)}")
    if args.main_transfer_new_applications_to_memberships:
        remaining_new_application_rows = [
            row
            for row in main_rows
            if (str(row[4] or ""), str(row[5] or "")) == FITBASE_LABELS["Новые заявки"]
        ]
        if remaining_new_application_rows:
            errors.append(
                "new application/refuser rows still exported to main XLSX: "
                f"{len(remaining_new_application_rows)}"
            )
    if args.dedupe_by_phone_keep_latest_subscription:
        main_duplicate_phones = duplicate_normalized_phone_count_from_xlsx_rows(main_rows, 1)
        card_duplicate_phones = duplicate_normalized_phone_count_from_xlsx_rows(card_rows, 0)
        expected_duplicate_phones = duplicate_normalized_phone_count_from_stage_rows(expected_main_rows)
        if main_duplicate_phones:
            errors.append(f"main XLSX still contains duplicate normalized phone groups: {main_duplicate_phones}")
        if card_duplicate_phones:
            errors.append(f"card XLSX still contains duplicate normalized phone groups: {card_duplicate_phones}")
        if expected_duplicate_phones:
            errors.append(f"expected deduplicated stage rows still contain duplicate phone groups: {expected_duplicate_phones}")

    for report in REQUIRED_REPORTS:
        if not (reports_dir / report).exists():
            errors.append(f"missing required report: {(reports_dir / report).relative_to(ROOT)}")

    missing_phone = read_csv(reports_dir / "missing_phone_report.csv") if (reports_dir / "missing_phone_report.csv").exists() else []
    missing_card = read_csv(reports_dir / "missing_card_report.csv") if (reports_dir / "missing_card_report.csv").exists() else []
    missing_club = read_csv(reports_dir / "missing_club_report.csv") if (reports_dir / "missing_club_report.csv").exists() else []
    multiple_subs = read_csv(reports_dir / "multiple_subscriptions_report.csv") if (reports_dir / "multiple_subscriptions_report.csv").exists() else []
    card_selection = read_csv(reports_dir / "card_selection_report.csv") if (reports_dir / "card_selection_report.csv").exists() else []
    single_stage = read_csv(reports_dir / "single_stage_distribution.csv") if (reports_dir / "single_stage_distribution.csv").exists() else []
    branch_distribution = read_csv(reports_dir / "branch_distribution.csv") if (reports_dir / "branch_distribution.csv").exists() else []
    phone_dedupe_removed = (
        read_csv(reports_dir / "phone_deduplication_removed_clients.csv")
        if (reports_dir / "phone_deduplication_removed_clients.csv").exists()
        else []
    )
    refuser_csv_path = output_dir / "csv" / "new_application_refusers.csv"
    refuser_rows = read_csv(refuser_csv_path) if refuser_csv_path.exists() else []
    if args.main_transfer_new_applications_to_memberships:
        if not refuser_csv_path.exists():
            errors.append(f"missing new application refusers CSV: {refuser_csv_path.relative_to(ROOT)}")
        expected_refuser_ids = Counter(row.get("client_id", "") for row in expected_refuser_rows)
        actual_refuser_ids = Counter(row.get("client_id", "") for row in refuser_rows)
        if actual_refuser_ids != expected_refuser_ids:
            errors.append(
                "new_application_refusers.csv client_id set does not match final transferred new applications"
            )

    if len(missing_phone) != sum(1 for row in rows if not (row.get("phones") or "").strip()):
        errors.append("missing_phone_report.csv count mismatch")
    if len(missing_card) != sum(1 for row in rows if not (row.get("selected_card_number") or "").strip()):
        errors.append("missing_card_report.csv count mismatch")
    if len(missing_club) != sum(1 for row in rows if not (row.get("normalized_club") or "").strip()):
        errors.append("missing_club_report.csv count mismatch")
    expected_multiple_subs = sum(
        1
        for row in rows
        if (
            row.get("funnel") == "Действующие клиенты"
            and int_value(row.get("active_full_subscription_count")) > 1
        )
        or (row.get("funnel") == "Реактивация" and int_value(row.get("finished_full_subscription_count")) > 1)
    )
    if len(multiple_subs) != expected_multiple_subs:
        errors.append("multiple_subscriptions_report.csv count mismatch")
    if len(card_selection) != len(rows):
        errors.append("card_selection_report.csv should contain one row per final client")

    reported_pairs = {
        (row.get("funnel", ""), row.get("funnel_step", "")): int_value(row.get("clients"))
        for row in single_stage
    }
    if reported_pairs != dict(final_pairs):
        errors.append("single_stage_distribution.csv does not match final XLSX distribution")
    reported_branches = {row.get("branch", ""): int_value(row.get("clients")) for row in branch_distribution}
    if reported_branches and reported_branches != dict(actual_branch_counts):
        errors.append("branch_distribution.csv does not match final XLSX branch distribution")

    review_rows = read_csv(reports_dir / "product_classification_review_report.csv") if (reports_dir / "product_classification_review_report.csv").exists() else []
    if review_rows:
        warnings.append(f"product classification rows needing business review: {len(review_rows)}")
    excluded_new_without_phone = sum(
        1
        for row in rows
        if row.get("funnel") == "Новые заявки" and not has_phone(row)
    )
    exported_main_missing_phone = sum(1 for row in main_rows if len(row) >= 2 and not str(row[1] or "").strip())
    if args.main_require_phone_for_new_applications and excluded_new_without_phone:
        warnings.append(f"new application rows without phone excluded from main XLSX: {excluded_new_without_phone}")
    if args.main_transfer_new_applications_to_memberships and expected_refuser_rows:
        warnings.append(f"new application/refuser rows moved to membership import: {len(expected_refuser_rows)}")
    if args.dedupe_by_phone_keep_latest_subscription and phone_dedupe_removed:
        warnings.append(f"same-phone duplicate clients excluded from main XLSX: {len(phone_dedupe_removed)}")
    if exported_main_missing_phone:
        warnings.append(f"clients without phone still present outside new applications in main XLSX: {exported_main_missing_phone}")
    if missing_card:
        warnings.append(f"clients without selected card in full stage and reported: {len(missing_card)}")
    if multiple_subs:
        warnings.append(f"clients with multiple subscription candidates reported: {len(multiple_subs)}")

    verdict = "PASS" if not errors else "FAIL"
    lines = [
        "# Part 2 Combined Single-Stage Validation Report",
        "",
        f"Run date: `{datetime.now().isoformat(timespec='seconds')}`",
        f"cutoff_date: `{args.cutoff_date}`",
        f"date_stamp: `{date_stamp}`",
        f"stage_rows: `{len(rows)}`",
        f"main_expected_rows_after_filters: `{len(expected_main_rows)}`",
        f"cards_expected_rows_after_filters: `{len(expected_card_rows)}`",
        f"main_xlsx_rows: `{len(main_rows)}`",
        f"cards_xlsx_rows: `{len(card_rows)}`",
        f"same_phone_deduplication_removed: `{len(phone_dedupe_removed)}`",
        f"new_application_refusers_to_membership: `{len(expected_refuser_rows)}`",
        "",
        "## Verdict",
        "",
        f"`{verdict}`",
        "",
        "## Final Single-Stage Distribution",
        "",
    ]
    for (funnel, step), count in final_pairs.most_common():
        lines.append(f"- `{funnel}` / `{step}`: `{count}`")
    lines.extend(["", "## Branch Distribution", ""])
    for branch, count in actual_branch_counts.most_common():
        lines.append(f"- `{branch}`: `{count}`")
    lines.extend(["", "## Data Quality Counts", ""])
    lines.extend(
        [
            f"- missing_phone: `{len(missing_phone)}`",
            f"- exported_main_missing_phone: `{exported_main_missing_phone}`",
            f"- excluded_new_applications_without_phone: `{excluded_new_without_phone}`",
            f"- new_application_refusers_to_membership: `{len(expected_refuser_rows)}`",
            f"- same_phone_deduplication_removed: `{len(phone_dedupe_removed)}`",
            f"- missing_card: `{len(missing_card)}`",
            f"- missing_club: `{len(missing_club)}`",
            f"- multiple_subscription_clients: `{len(multiple_subs)}`",
            f"- product_review_rows: `{len(review_rows)}`",
        ]
    )
    lines.extend(["", "## Errors", ""])
    lines.extend([f"- {error}" for error in errors] if errors else ["None."])
    if warnings:
        lines.extend(["", "## Warnings", ""])
        lines.extend(f"- {warning}" for warning in warnings)
    lines.append("")

    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "validation_report.md").write_text("\n".join(lines), encoding="utf-8")
    print(f"verdict={verdict}")
    print(f"errors={len(errors)}")
    print(f"warnings={len(warnings)}")
    print(f"report={(reports_dir / 'validation_report.md').relative_to(ROOT)}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default="2026-06-30")
    parser.add_argument("--date-stamp", default="20260630")
    parser.add_argument("--stage-dir", default="work/20260630/owner/staging")
    parser.add_argument("--output-dir", default="work/20260630/owner")
    parser.add_argument("--reports-dir", default="work/20260630/owner/reports")
    parser.add_argument("--main-template", default="templates/import_zayavki.xlsx")
    parser.add_argument("--cards-template", default="templates/plastic_cards.xlsx")
    parser.add_argument("--branches-config", default=str(DEFAULT_BRANCHES_CONFIG))
    parser.add_argument(
        "--main-require-phone-for-new-applications",
        action="store_true",
        help="Validate that internal `Новые заявки` rows without phone are excluded from the main XLSX.",
    )
    parser.add_argument(
        "--main-transfer-new-applications-to-memberships",
        action="store_true",
        help="Validate that final `Новые заявки / Неразобранные` rows are moved to membership tag import.",
    )
    parser.add_argument(
        "--cards-funnel-filter",
        default="",
        help="Validate that the plastic-card XLSX contains only this internal funnel.",
    )
    parser.add_argument(
        "--dedupe-by-phone-keep-latest-subscription",
        action="store_true",
        help="Validate that final XLSX outputs keep one client per normalized phone component.",
    )
    return parser.parse_args()


def main() -> None:
    raise SystemExit(validate(parse_args()))


if __name__ == "__main__":
    main()
