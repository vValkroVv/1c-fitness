#!/usr/bin/env python3
"""Build a targeted delivery where only service-client end dates may change."""

from __future__ import annotations

import argparse
import csv
import hashlib
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
SERVICE_PREFIX = "fitbase_import_uslugi_clientov_"
ALLOWED_CHANGED_FIELDS = {"end_date"}
PRESERVED_REPORT_FILES = (
    Path("reports/financial_source_coverage.csv"),
    Path("reports/manager_debt_comparison.json"),
    Path("reports/manager_debt_comparison.md"),
)


def as_abs(path: str | Path) -> Path:
    value = Path(path)
    return value.resolve() if value.is_absolute() else (ROOT / value).resolve()


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def blank(value: Any) -> bool:
    return value in (None, "")


def comparable(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date()
    return value


def read_rows(path: Path) -> tuple[list[str], list[str], list[tuple[Any, ...]], str]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    if len(workbook.sheetnames) != 1:
        workbook.close()
        raise ValueError(f"Expected exactly one sheet in {path}")
    worksheet = workbook.active
    iterator = worksheet.iter_rows(values_only=True)
    headers = [str(value or "") for value in next(iterator)]
    russian_headers = [str(value or "") for value in next(iterator)]
    rows = [
        tuple(comparable(value) for value in values[: len(headers)])
        for values in iterator
        if any(not blank(value) for value in values[: len(headers)])
    ]
    title = worksheet.title
    workbook.close()
    return headers, russian_headers, rows, title


def style_without_number_format(cell: Any) -> tuple[Any, ...]:
    style = cell._style
    if style is None:
        return (0, 0, 0, 0, 0, 0)
    return (
        style.fontId,
        style.fillId,
        style.borderId,
        style.alignmentId,
        style.protectionId,
        style.xfId,
    )


def worksheet_layout(worksheet: Any) -> dict[str, Any]:
    return {
        "title": worksheet.title,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "freeze_panes": str(worksheet.freeze_panes or ""),
        "auto_filter": str(worksheet.auto_filter.ref or ""),
        "merged_cells": sorted(str(value) for value in worksheet.merged_cells.ranges),
        "show_grid_lines": worksheet.sheet_view.showGridLines,
        "column_dimensions": {
            key: (
                value.width,
                value.hidden,
                value.bestFit,
                value.outlineLevel,
                value.collapsed,
            )
            for key, value in worksheet.column_dimensions.items()
        },
        "row_dimensions": {
            key: (
                value.height,
                value.hidden,
                value.outlineLevel,
                value.collapsed,
            )
            for key, value in worksheet.row_dimensions.items()
        },
    }


def validate_layout_and_styles(
    old_path: Path,
    corrected_path: Path,
    end_column: int,
) -> int:
    old_workbook = load_workbook(old_path, read_only=False, data_only=False)
    new_workbook = load_workbook(corrected_path, read_only=False, data_only=False)
    if old_workbook.sheetnames != new_workbook.sheetnames:
        old_workbook.close()
        new_workbook.close()
        raise ValueError("Corrected service workbook changed sheet topology")
    old_sheet = old_workbook.active
    new_sheet = new_workbook.active
    if worksheet_layout(old_sheet) != worksheet_layout(new_sheet):
        old_workbook.close()
        new_workbook.close()
        raise ValueError("Corrected service workbook changed worksheet layout")

    end_style_changes = 0
    for row_number in range(1, old_sheet.max_row + 1):
        for column_number in range(1, old_sheet.max_column + 1):
            old_cell = old_sheet.cell(row_number, column_number)
            new_cell = new_sheet.cell(row_number, column_number)
            if column_number != end_column or row_number <= 2:
                if old_cell._style != new_cell._style:
                    old_workbook.close()
                    new_workbook.close()
                    raise ValueError(
                        "Corrected service workbook changed a non-authorized style: "
                        f"cell={old_cell.coordinate}"
                    )
                continue
            if style_without_number_format(old_cell) != style_without_number_format(
                new_cell
            ):
                old_workbook.close()
                new_workbook.close()
                raise ValueError(
                    "Corrected service workbook changed end_date styling beyond "
                    f"the number format: cell={old_cell.coordinate}"
                )
            if old_cell.number_format != new_cell.number_format:
                end_style_changes += 1
            if not blank(new_cell.value) and new_cell.number_format != "yyyy-mm-dd":
                old_workbook.close()
                new_workbook.close()
                raise ValueError(
                    "Corrected end_date cell does not use yyyy-mm-dd format: "
                    f"cell={new_cell.coordinate}, format={new_cell.number_format!r}"
                )

    old_workbook.close()
    new_workbook.close()
    return end_style_changes


def validate_targeted_change(
    old_path: Path,
    corrected_path: Path,
) -> dict[str, Any]:
    old_headers, old_russian, old_rows, old_sheet = read_rows(old_path)
    new_headers, new_russian, new_rows, new_sheet = read_rows(corrected_path)
    if old_headers != new_headers:
        raise ValueError("Corrected service workbook changed technical headers")
    if old_russian != new_russian:
        raise ValueError("Corrected service workbook changed Russian headers")
    if old_sheet != new_sheet:
        raise ValueError("Corrected service workbook changed sheet name")
    if len(old_rows) != len(new_rows):
        raise ValueError(
            "Corrected service workbook changed row count: "
            f"{len(old_rows)} -> {len(new_rows)}"
        )

    indexes = {name: index for index, name in enumerate(old_headers)}
    required = {"service_id", "client_id", "end_date"}
    missing = sorted(required - set(indexes))
    if missing:
        raise ValueError(f"Missing service workbook columns: {missing}")
    allowed_indexes = {indexes[name] for name in ALLOWED_CHANGED_FIELDS}
    service_index = indexes["service_id"]
    end_index = indexes["end_date"]
    end_style_changes = validate_layout_and_styles(
        old_path,
        corrected_path,
        end_index + 1,
    )

    changed_cells = 0
    changed_rows = 0
    blank_end_rows: list[str] = []
    for row_number, (old_row, new_row) in enumerate(
        zip(old_rows, new_rows, strict=True),
        start=3,
    ):
        old_id = str(old_row[service_index] or "").strip()
        new_id = str(new_row[service_index] or "").strip()
        if old_id != new_id:
            raise ValueError(
                f"Service row identity/order changed at Excel row {row_number}: "
                f"{old_id!r} -> {new_id!r}"
            )
        row_changed = False
        for index, (old_value, new_value) in enumerate(
            zip(old_row, new_row, strict=True)
        ):
            if old_value == new_value:
                continue
            if index not in allowed_indexes:
                raise ValueError(
                    "Corrected service workbook changed a non-authorized cell: "
                    f"row={row_number}, field={old_headers[index]!r}, "
                    f"old={old_value!r}, new={new_value!r}"
                )
            changed_cells += 1
            row_changed = True
        if row_changed:
            changed_rows += 1
        if blank(new_row[end_index]):
            blank_end_rows.append(new_id or f"row-{row_number}")
        elif not isinstance(new_row[end_index], date):
            raise ValueError(
                f"end_date is not an Excel date at row {row_number}: "
                f"{new_row[end_index]!r}"
            )

    if blank_end_rows:
        raise ValueError(
            "Corrected service workbook still has blank end dates: "
            f"{blank_end_rows[:10]}"
        )
    if changed_cells == 0:
        raise ValueError("Corrected service workbook has no end-date changes")
    return {
        "rows": len(new_rows),
        "changed_rows": changed_rows,
        "changed_cells": changed_cells,
        "blank_end_dates": len(blank_end_rows),
        "end_date_number_format_changes": end_style_changes,
    }


def write_manifest(path: Path, rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "file_name",
                "source_sha256",
                "delivery_sha256",
                "status",
            ],
            lineterminator="\n",
        )
        writer.writeheader()
        writer.writerows(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-delivery", required=True)
    parser.add_argument("--corrected-service", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--date-stamp", default="20260630")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_dir = as_abs(args.source_delivery)
    corrected_service = as_abs(args.corrected_service)
    output_dir = as_abs(args.output_dir)
    service_name = f"{SERVICE_PREFIX}{args.date_stamp}.xlsx"
    old_service = source_dir / service_name

    if not source_dir.is_dir():
        raise FileNotFoundError(source_dir)
    if not old_service.is_file():
        raise FileNotFoundError(old_service)
    if not corrected_service.is_file():
        raise FileNotFoundError(corrected_service)
    if corrected_service.name != service_name:
        raise ValueError(
            f"Corrected service file must be named {service_name!r}, "
            f"got {corrected_service.name!r}"
        )
    source_xlsx = sorted(source_dir.rglob("*.xlsx"))
    if not source_xlsx:
        raise ValueError(f"No XLSX files found in {source_dir}")
    if output_dir.exists() and any(output_dir.iterdir()):
        raise FileExistsError(
            f"Targeted delivery already exists and is not empty: {output_dir}"
        )

    service_validation = validate_targeted_change(old_service, corrected_service)
    output_dir.mkdir(parents=True, exist_ok=True)
    reports_dir = output_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)

    manifest_rows: list[dict[str, str]] = []
    for source_path in source_xlsx:
        relative_path = source_path.relative_to(source_dir)
        replacement = (
            corrected_service
            if relative_path == Path(service_name)
            else source_path
        )
        destination = output_dir / relative_path
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(replacement, destination)
        source_hash = sha256(source_path)
        delivery_hash = sha256(destination)
        is_service = relative_path == Path(service_name)
        if is_service and source_hash == delivery_hash:
            raise ValueError("Corrected service workbook is byte-identical to the old one")
        if not is_service and source_hash != delivery_hash:
            raise ValueError(
                f"Untargeted workbook changed while copying: {source_path.name}"
            )
        manifest_rows.append(
            {
                "file_name": relative_path.as_posix(),
                "source_sha256": source_hash,
                "delivery_sha256": delivery_hash,
                "status": "changed_end_date_only" if is_service else "byte_identical",
            }
        )

    preserved_reports = 0
    for relative_path in PRESERVED_REPORT_FILES:
        source_path = source_dir / relative_path
        if not source_path.is_file():
            continue
        destination = output_dir / relative_path
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source_path, destination)
        if sha256(source_path) != sha256(destination):
            raise ValueError(f"Preserved report changed while copying: {relative_path}")
        preserved_reports += 1

    manifest_path = reports_dir / "xlsx_hash_manifest.csv"
    write_manifest(manifest_path, manifest_rows)
    root_xlsx_count = sum(path.parent == source_dir for path in source_xlsx)
    report = [
        "# Targeted service end-date delivery",
        "",
        f"- source delivery: `{source_dir}`",
        f"- corrected delivery: `{output_dir}`",
        f"- root delivery XLSX files: {root_xlsx_count}",
        f"- XLSX files compared recursively: {len(manifest_rows)}",
        f"- service data rows: {service_validation['rows']}",
        f"- rows with changed end_date: {service_validation['changed_rows']}",
        f"- changed cells (end_date only): {service_validation['changed_cells']}",
        f"- blank end_date cells after correction: {service_validation['blank_end_dates']}",
        f"- end_date number-format changes: {service_validation['end_date_number_format_changes']}",
        f"- unchanged XLSX verified byte-identical: {len(manifest_rows) - 1}",
        f"- preserved non-XLSX baseline reports: {preserved_reports}",
        "- status: PASS",
        "",
        "Only `end_date` is authorized to differ inside the service-client workbook.",
        "Every other cell and every other XLSX is checked before the delivery is accepted.",
    ]
    (reports_dir / "targeted_change_validation.md").write_text(
        "\n".join(report) + "\n",
        encoding="utf-8",
    )
    print("\n".join(report))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
