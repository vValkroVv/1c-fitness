#!/usr/bin/env python3
"""Build a separate membership workbook for the 12000/12000/0 fallback rows.

The authoritative source of membership values is the already validated clean
membership workbook. The audit CSV is used only to select the exact contracts
that were resolved because the person was absent from the manager debt report.
"""

from __future__ import annotations

import argparse
import csv
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


FALLBACK_SOURCE = "fallback_not_in_manager_debt_report"
EXPECTED_MONEY = {
    "price": Decimal("12000"),
    "amount_of_payments": Decimal("12000"),
    "payment_left": Decimal("0"),
}
EXPECTED_ROWS = 62


def decimal_value(value: object, *, field: str, contract_id: str) -> Decimal:
    """Convert an Excel numeric value without losing exact decimal meaning."""

    try:
        return Decimal(str(value if value is not None else 0))
    except InvalidOperation as exc:
        raise ValueError(
            f"Invalid {field} for contract {contract_id}: {value!r}"
        ) from exc


def read_fallback_contracts(audit_path: Path) -> set[str]:
    """Read and validate the exact fallback contract set from the audit."""

    with audit_path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = [
            row
            for row in csv.DictReader(handle)
            if row["resolution_source"] == FALLBACK_SOURCE
        ]

    if len(rows) != EXPECTED_ROWS:
        raise ValueError(
            f"Expected {EXPECTED_ROWS} fallback audit rows, found {len(rows)}"
        )

    contract_ids: set[str] = set()
    for row in rows:
        contract_id = str(row["contract_id"] or "").strip()
        if not contract_id:
            raise ValueError("Fallback audit row has an empty contract_id")
        if contract_id in contract_ids:
            raise ValueError(f"Duplicate fallback contract_id: {contract_id}")
        contract_ids.add(contract_id)

        for audit_field, expected in (
            ("new_price", EXPECTED_MONEY["price"]),
            ("new_amount_of_payments", EXPECTED_MONEY["amount_of_payments"]),
            ("new_payment_left", EXPECTED_MONEY["payment_left"]),
        ):
            actual = decimal_value(
                row[audit_field],
                field=audit_field,
                contract_id=contract_id,
            )
            if actual != expected:
                raise ValueError(
                    f"Unexpected {audit_field} for {contract_id}: "
                    f"{actual} != {expected}"
                )

    return contract_ids


def build_subset(
    membership_path: Path,
    audit_path: Path,
    template_path: Path,
    output_path: Path,
) -> None:
    """Create a template-compatible XLSX containing only fallback contracts."""

    wanted_contracts = read_fallback_contracts(audit_path)

    source_workbook = load_workbook(
        membership_path,
        read_only=True,
        data_only=False,
    )
    source_sheet = source_workbook.active
    headers = [
        str(value or "")
        for value in next(
            source_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    indexes = {header: index for index, header in enumerate(headers)}
    required_fields = {"contract_id", *EXPECTED_MONEY}
    missing_fields = sorted(required_fields - set(indexes))
    if missing_fields:
        source_workbook.close()
        raise ValueError(f"Membership is missing fields: {missing_fields}")

    target_workbook = load_workbook(template_path)
    target_sheet = target_workbook.active
    width = len(headers)
    if target_sheet.max_column > width:
        target_sheet.delete_cols(width + 1, target_sheet.max_column - width)

    template_formats = [
        target_sheet.cell(3, column).number_format
        for column in range(1, width + 1)
    ]
    if target_sheet.max_row >= 3:
        target_sheet.delete_rows(3, target_sheet.max_row - 2)

    russian_headers = next(
        source_sheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    for column, header in enumerate(headers, start=1):
        target_sheet.cell(1, column).value = header
        target_sheet.cell(2, column).value = russian_headers[column - 1]

    format_columns: dict[int, str] = {}
    for column, header in enumerate(headers, start=1):
        if header in {"create_date", "payment_date", "activation_date", "end_date"}:
            format_columns[column] = "yyyy-mm-dd"
        elif header in EXPECTED_MONEY:
            format_columns[column] = "#,##0.00"
        elif template_formats[column - 1] not in ("", "General"):
            format_columns[column] = template_formats[column - 1]

    found_contracts: set[str] = set()
    for source_values in source_sheet.iter_rows(min_row=3, values_only=True):
        values = list(source_values[:width])
        contract_id = str(values[indexes["contract_id"]] or "").strip()
        if contract_id not in wanted_contracts:
            continue
        if contract_id in found_contracts:
            source_workbook.close()
            target_workbook.close()
            raise ValueError(f"Duplicate contract in membership: {contract_id}")

        for field, expected in EXPECTED_MONEY.items():
            actual = decimal_value(
                values[indexes[field]],
                field=field,
                contract_id=contract_id,
            )
            if actual != expected:
                source_workbook.close()
                target_workbook.close()
                raise ValueError(
                    f"Membership {field} mismatch for {contract_id}: "
                    f"{actual} != {expected}"
                )

        target_sheet.append(values)
        found_contracts.add(contract_id)
        target_row = len(found_contracts) + 2
        for column, number_format in format_columns.items():
            target_sheet.cell(target_row, column).number_format = number_format

    source_workbook.close()
    missing_contracts = sorted(wanted_contracts - found_contracts)
    if missing_contracts:
        target_workbook.close()
        raise ValueError(
            f"Fallback contracts missing from clean membership: "
            f"{missing_contracts[:10]}"
        )

    target_sheet.freeze_panes = "A3"
    for column in range(1, width + 1):
        letter = get_column_letter(column)
        current_width = target_sheet.column_dimensions[letter].width or 12
        target_sheet.column_dimensions[letter].width = min(
            max(current_width, 12),
            34,
        )

    if output_path.exists():
        target_workbook.close()
        raise FileExistsError(f"Output already exists: {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    target_workbook.save(output_path)
    target_workbook.close()

    print(f"output={output_path.resolve()}")
    print(f"rows={len(found_contracts)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--membership", required=True, type=Path)
    parser.add_argument("--audit", required=True, type=Path)
    parser.add_argument("--template", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    build_subset(
        membership_path=args.membership.resolve(),
        audit_path=args.audit.resolve(),
        template_path=args.template.resolve(),
        output_path=args.output.resolve(),
    )


if __name__ == "__main__":
    main()
