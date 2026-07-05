#!/usr/bin/env python3
"""Build Fitbase membership import workbooks from fixed owner-change outputs."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]

DATE_STAMP = "20260525_0800"
CLIENT_HEADERS = [
    "tag",
    "contract_id",
    "client_id",
    "phone",
    "client_fio",
    "contract_name",
    "card",
    "duration",
    "duration_type",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "freeze",
    "guests",
    "visits_left",
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "manager",
]
CLIENT_RUS_HEADERS = [
    "Тег",
    "Внутренний номер абонемента",
    "Внутренний номер клиента",
    "Телефон клиента",
    "ФИО клиента  *",
    "Название абонемента *",
    "Номер карты",
    "Продолжительность",
    "Тип продолжительности",
    "Дата  добавления *",
    "Дата оплаты *",
    "Дата активации ",
    "Дата окончания",
    "Осталось дней для заморозки",
    "Осталось гостевых визитов",
    "Осталось посещений *",
    "Стоимость *",
    "Оплачено *",
    "Осталось оплатить *",
    "Тип оплаты",
    "Менеджер ",
]
TEMPLATE_HEADERS = [
    "branches_access",
    "name",
    "price",
    "duration",
    "duration_type",
    "visits",
    "guests",
    "freeze",
    "first_visit_activation",
    "archive",
    "category",
    "legal_entity",
]
TEMPLATE_RUS_HEADERS = [
    "Доступ в филиалах",
    "Название абонемента *",
    "Стоимость *",
    "Продолжительность  *",
    "Тип продолжительности *",
    "Посещений ",
    "Гостевые визиты",
    "Дни для заморозки",
    "Активация с первого посещения?",
    "В архиве?",
    "Категория в структуре",
    "Юр.лицо",
]
FACT_FIELDS = [
    "client_ref",
    "client_id",
    "original_client_ref",
    "original_client_id",
    "original_client_fio",
    "effective_client_ref",
    "effective_client_id",
    "effective_client_fio",
    "owner_change_ref",
    "owner_change_number",
    "owner_change_datetime",
    "owner_change_old_client_ref",
    "owner_change_new_client_ref",
    "owner_change_modifier_name",
    "owner_change_count_for_membership",
    "subscription_ref",
    "document_number",
    "holder_client_ref",
    "payer_client_ref",
    "client_role_source",
    "product_ref",
    "product_code",
    "subscription_name",
    "product_class",
    "is_full_subscription",
    "is_trial_or_guest",
    "is_subrent",
    "is_limited_subrent",
    "sale_date",
    "sale_datetime",
    "start_date",
    "end_date",
    "duration_days",
    "status",
    "booking_status_ref",
    "booking_status_name",
    "doc_posted",
    "doc_marked",
    "register_duration_days",
    "is_active_on_cutoff",
    "is_finished_before_cutoff",
    "days_to_end",
    "days_since_end",
    "raw_club",
    "normalized_club",
    "club_source",
    "raw_source",
    "doc_duration_value",
    "rg_duration_days",
    "rg_freeze_days",
    "rg_guests",
    "rg_price",
    "rg_paid_candidate",
    "rg_payment_count_candidate",
    "rg_visits_candidate_8007",
    "rg_visits_candidate_8008",
    "rg_visits_candidate_8009",
    "subrent_visit_limit",
    "subrent_active_by_dates_on_cutoff",
    "subrent_finished_by_dates_before_cutoff",
    "subrent_rg3336_receipt_qty",
    "subrent_rg3336_expense_qty",
    "subrent_rg3336_signed_balance",
    "subrent_rg3336_visit_doc_expense_qty",
    "subrent_rg3336_receipt_rows",
    "subrent_rg3336_expense_rows",
    "subrent_rg3336_case_group",
    "matched_payment_ref",
    "matched_payment_datetime",
    "matched_payment_amount",
    "matched_payment_method",
    "matched_payment_operation",
    "matched_payment_match_source",
    "membership_sale_line_amount",
    "membership_sale_line_count",
    "document131_refund_count",
    "document131_posted_unmarked_refund_count",
    "cutoff_at",
]


@dataclass(frozen=True)
class SourceClient:
    client_id: str
    phone: str
    client_fio: str
    create_date: date | None
    manager: str
    branch: str
    tag: str = ""


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt).date()
        except ValueError:
            continue
    return None


def excel_date(value: Any) -> date | None:
    parsed = parse_date(value)
    if parsed == date(2001, 1, 1):
        return None
    return parsed


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def excel_number(value: Decimal | int | float | None) -> int | float | None:
    if value is None:
        return None
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    if value == value.to_integral_value():
        return int(value)
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def int_or_blank(value: Decimal) -> int | None:
    if value <= 0:
        return None
    return int(value.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


FORCED_FREE_TRIAL_NAMES = {
    normalize_key(name)
    for name in [
        "Абонемент НЕДЕЛЯ ДРУГ",
        "Абонемент 10 ДНЕЙ",
        "Абонемент НЕДЕЛЯ ХОЛОДНЫЕ",
        "Абонемент Неделя марафон",
        "Абонемент НЕДЕЛЯ КАРЕЛЬСКИЙ",
        "Абонемент УЛЬТРА 1 МЕСЯЦ БЕСПЛАТНО",
    ]
}
ZERO_PRICE_CONFIRMED_FREE_TRIAL_NAMES = {
    normalize_key(name)
    for name in [
        "Абонемент НЕДЕЛЯ САЙТ",
        "Абонемент НЕДЕЛЯ ФИТНЕСА БЕСПЛАТНО",
    ]
}
ZERO_PRICE_DIRECT_SALE_LINE_FREE_TRIAL_NAMES = {
    normalize_key(name)
    for name in [
        "Абонемент НЕДЕЛЯ САЙТ",
    ]
}


def business_zero_override_reason(fact: dict[str, str]) -> str:
    """Return the explicit business rule forcing price/payment to blank."""

    name = normalize_key(fact.get("subscription_name", ""))
    sale_date = (fact.get("sale_date") or "").strip()
    product_class = (fact.get("product_class") or "").strip()
    price = decimal_value(fact.get("rg_price"))
    matched_payment_ref = (fact.get("matched_payment_ref") or "").strip()
    matched_payment_source = (fact.get("matched_payment_match_source") or "").strip()
    matched_payment_method = (fact.get("matched_payment_method") or "").strip()
    membership_sale_line_amount = decimal_value(fact.get("membership_sale_line_amount"))
    membership_sale_line_count = decimal_value(fact.get("membership_sale_line_count"))
    document131_posted_unmarked_refund_count = decimal_value(
        fact.get("document131_posted_unmarked_refund_count")
    )

    if sale_date.startswith("2018") and product_class == "full_subscription":
        return "business_legacy_2018_full_subscription_zero_price_blank_payment"
    if name in FORCED_FREE_TRIAL_NAMES:
        return "business_free_trial_zero_price_blank_payment"
    if (
        price <= 0
        and name in ZERO_PRICE_CONFIRMED_FREE_TRIAL_NAMES
        and not matched_payment_source.startswith("direct")
    ):
        return "business_confirmed_free_trial_zero_price_blank_payment"
    if (
        price <= 0
        and name in ZERO_PRICE_DIRECT_SALE_LINE_FREE_TRIAL_NAMES
        and matched_payment_ref
        and matched_payment_source.startswith("direct")
        and membership_sale_line_count > 0
        and membership_sale_line_amount <= 0
    ):
        return "business_direct_free_site_week_sale_line_zero_blank_payment"
    if (
        price <= 0
        and matched_payment_ref
        and matched_payment_source.startswith("direct")
        and document131_posted_unmarked_refund_count > 0
        and fact.get("is_active_on_cutoff") != "1"
    ):
        return "business_historical_document131_refund_zero_direct_blank_payment"
    if price <= 0 and product_class == "full_subscription" and not matched_payment_ref:
        return "business_full_zero_no_payment_initial_balance_corporate_or_modifier"
    if price <= 0 and not matched_payment_ref:
        return "business_zero_no_payment_blank_payment_type"
    if price <= 0 and matched_payment_source == "client_date_14_days":
        return "business_zero_fallback_payment_type_blank"
    if price <= 0 and matched_payment_ref and not matched_payment_method:
        return "business_zero_raw_blank_payment_type_blank"
    return ""


def read_source_clients(path: Path, default_tag: str = "") -> dict[str, SourceClient]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = list(header_row)
    indexes = {str(name): idx + 1 for idx, name in enumerate(headers) if name}
    required = ["client_id", "phone", "client_fio", "create_date", "manager", "филиал"]
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"Missing required columns in {path}: {missing}")

    clients: dict[str, SourceClient] = {}
    zero_based = {name: indexes[name] - 1 for name in required}
    for values in ws.iter_rows(min_row=3, values_only=True):
        client_id = str(values[zero_based["client_id"]] or "").strip()
        if not client_id:
            continue
        clients[client_id] = SourceClient(
            client_id=client_id,
            phone=str(values[zero_based["phone"]] or "").strip(),
            client_fio=str(values[zero_based["client_fio"]] or "").strip(),
            create_date=parse_date(values[zero_based["create_date"]]),
            manager=str(values[zero_based["manager"]] or "").strip(),
            branch=str(values[zero_based["филиал"]] or "").strip(),
            tag=default_tag,
        )
    wb.close()
    return clients


def read_refuser_clients(path: Path) -> dict[str, SourceClient]:
    if not path.exists():
        return {}
    clients: dict[str, SourceClient] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"client_id", "phone", "client_fio", "create_date", "manager", "branch"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise ValueError(f"Missing required columns in {path}: {missing}")
        for row in reader:
            client_id = (row.get("client_id") or "").strip()
            if not client_id:
                continue
            clients[client_id] = SourceClient(
                client_id=client_id,
                phone=(row.get("phone") or "").strip(),
                client_fio=(row.get("client_fio") or "").strip(),
                create_date=parse_date(row.get("create_date")),
                manager=(row.get("manager") or "").strip(),
                branch=(row.get("branch") or "").strip(),
                tag="отказники",
            )
    return clients


def read_cards(stage_dir: Path) -> dict[str, str]:
    path = stage_dir / "final_funnel_clients.csv"
    if not path.exists():
        return {}
    cards: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            client_id = (row.get("client_id") or "").strip()
            card = (row.get("selected_card_number") or "").strip()
            if client_id and card:
                cards[client_id] = card
    return cards


def read_facts(path: Path) -> list[dict[str, str]]:
    def is_hex_ref(value: str) -> bool:
        return bool(re.fullmatch(r"[0-9A-F]{32}", value or ""))

    def repair_row(raw: list[str]) -> list[str]:
        if len(raw) == len(FACT_FIELDS):
            return raw
        effective_ref_idx = next((idx for idx in range(5, len(raw)) if is_hex_ref(raw[idx])), -1)
        if effective_ref_idx < 0:
            raise ValueError(f"Cannot repair TSV row with {len(raw)} columns from {path}")
        effective_fio_start = effective_ref_idx + 2
        tail_after_effective_fio = len(FACT_FIELDS) - (FACT_FIELDS.index("effective_client_fio") + 1)
        effective_fio_end = len(raw) - tail_after_effective_fio
        fixed = (
            raw[:4]
            + [" ".join(part.strip() for part in raw[4:effective_ref_idx] if part.strip())]
            + [raw[effective_ref_idx], raw[effective_ref_idx + 1]]
            + [" ".join(part.strip() for part in raw[effective_fio_start:effective_fio_end] if part.strip())]
            + raw[effective_fio_end:]
        )
        if len(fixed) != len(FACT_FIELDS):
            raise ValueError(f"Cannot repair TSV row with {len(raw)} columns from {path}; repaired to {len(fixed)}")
        return fixed

    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-16", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for raw in reader:
            if not raw:
                continue
            repaired = repair_row(raw)
            rows.append({key: value.replace("\t", " ").replace("\r", " ").replace("\n", " ") for key, value in zip(FACT_FIELDS, repaired, strict=True)})
    return rows


def map_payment_type(method: str) -> str:
    text = normalize_key(method)
    if not text:
        return ""
    if "для ошибок" in text or "бар ип иконников андрей анатольевич" in text:
        return "безналичные"
    if "сбп" in text or "сбпр" in text:
        return "сбп"
    if "налич" in text and "безнал" not in text:
        return "наличные"
    if (
        "эквайр" in text
        or "банк" in text
        or "безнал" in text
        or "терминал" in text
        or "карта" in text
        or "р/с" in text
    ):
        return "безналичные"
    return ""


def payment_type_from_direct_blank_method(fact: dict[str, str], price: Decimal, business_override: str) -> str:
    """Map confirmed old direct payments with an empty method to cash."""

    if business_override or price <= 0:
        return ""
    match_source = (fact.get("matched_payment_match_source") or "").strip()
    raw_method = (fact.get("matched_payment_method") or "").strip()
    operation = normalize_key(fact.get("matched_payment_operation", ""))
    amount = decimal_value(fact.get("matched_payment_amount"))
    if (
        match_source.startswith("direct")
        and not raw_method
        and amount > 0
        and operation == "оплата от клиента"
    ):
        return "наличные"
    return ""


def payment_type_from_positive_no_payment(price: Decimal, business_override: str, matched_payment_ref: str) -> str:
    """Apply the business default for paid rows without a matched payment."""

    if business_override or price <= 0 or matched_payment_ref:
        return ""
    return "наличные"


def parse_template_visits(name: str) -> int | None:
    match = re.search(r"(\d+)\s*посещ", name.lower())
    if not match:
        return None
    return int(match.group(1))


def compute_visits_left(fact: dict[str, str], contract_name: str) -> tuple[int | float | None, str, str]:
    """Return visits_left and the source used for limited subrent rows."""

    if fact.get("is_limited_subrent") != "1":
        return None, "not_limited_subrent", ""

    if fact.get("subrent_finished_by_dates_before_cutoff") == "1":
        return 0, "business_expired_limited_subrent_zero_visits_left", ""

    visit_limit = decimal_value(fact.get("subrent_visit_limit"))
    if visit_limit <= 0:
        parsed_limit = parse_template_visits(contract_name)
        visit_limit = Decimal(parsed_limit or 0)

    balance = decimal_value(fact.get("subrent_rg3336_signed_balance"))
    case_group = (fact.get("subrent_rg3336_case_group") or "").strip()

    if fact.get("subrent_active_by_dates_on_cutoff") == "1":
        if case_group == "clean_register_balance" and Decimal("0") <= balance <= visit_limit:
            return excel_number(balance), "rg3336_correct_dimension_balance", ""
        if balance >= 0:
            return excel_number(balance), "rg3336_correct_dimension_balance_needs_review", (
                f"Active limited subrent has non-clean register balance: "
                f"case_group={case_group}; limit={visit_limit}; balance={balance}."
            )
        return 0, "rg3336_correct_dimension_negative_active_clamped_to_zero", (
            f"Active limited subrent has negative register balance: "
            f"case_group={case_group}; limit={visit_limit}; balance={balance}."
        )

    return None, "limited_subrent_no_cutoff_rule", (
        "Limited subrent is neither active nor expired by date flags; visits_left left blank."
    )


def compute_duration_months(fact: dict[str, str]) -> tuple[int, str]:
    name = normalize_key(fact.get("subscription_name", ""))
    duration_days = decimal_value(fact.get("duration_days"))
    doc_duration = decimal_value(fact.get("doc_duration_value"))

    if duration_days > 0 and duration_days < 28:
        return 0, "duration_days<28"
    if "недел" in name or re.search(r"\b10\s*дн", name):
        return 0, "short_membership_name"
    if doc_duration > 0 and doc_duration <= 60:
        return int(doc_duration.quantize(Decimal("1"), rounding=ROUND_HALF_UP)), "document163_fld1481"
    if duration_days > 0:
        months = int((duration_days / Decimal("30.4375")).quantize(Decimal("1"), rounding=ROUND_HALF_UP))
        return max(months, 0), "duration_days_rounded_to_months"
    return 0, "missing_duration"


def compute_money(fact: dict[str, str]) -> tuple[Decimal, Decimal, Decimal, str]:
    price = decimal_value(fact.get("rg_price"))
    paid_candidate = decimal_value(fact.get("rg_paid_candidate"))
    name = normalize_key(fact.get("subscription_name", ""))

    if price <= 0:
        return Decimal("0"), Decimal("0"), Decimal("0"), "zero_price"
    if paid_candidate > 0:
        paid = min(paid_candidate, price)
        return price, paid, max(price - paid, Decimal("0")), "rg_fld3072_paid_candidate"
    if "рассроч" in name:
        matched_payment_amount = decimal_value(fact.get("matched_payment_amount"))
        if matched_payment_amount > 0:
            paid = min(matched_payment_amount, price)
            return price, paid, max(price - paid, Decimal("0")), "matched_payment_amount_for_installment"
        return price, Decimal("0"), price, "installment_without_paid_candidate"
    return price, price, Decimal("0"), "assume_paid_full_when_no_installment_marker"


def is_full_subscription_fact(fact: dict[str, str]) -> bool:
    return fact.get("is_full_subscription") == "1"


def is_active_on_cutoff_fact(fact: dict[str, str]) -> bool:
    return fact.get("is_active_on_cutoff") == "1"


def normalized_status(fact: dict[str, str]) -> str:
    return (fact.get("status") or "").strip()


def find_contact_next_exclusions(facts: list[dict[str, str]]) -> tuple[set[str], list[dict[str, str]]]:
    """Find active/later full rows in status Contact that should not be imported.

    Business rule from the Popova case: if a client already has a full
    membership active on the cutoff, and a later full membership for the same
    client is in status "Контакт с клиентом", the later row is a hanging sale
    candidate and is excluded from membership import.
    """

    cutoff_date = parse_date(next((fact.get("cutoff_at") for fact in facts if fact.get("cutoff_at")), ""))
    if cutoff_date is None:
        return set(), []

    full_facts_by_client: dict[str, list[dict[str, str]]] = defaultdict(list)
    for fact in facts:
        if is_full_subscription_fact(fact):
            full_facts_by_client[(fact.get("client_id") or "").strip()].append(fact)

    excluded_refs: set[str] = set()
    excluded_rows: dict[str, dict[str, str]] = {}
    for client_id, client_facts in full_facts_by_client.items():
        if not client_id:
            continue
        current_active = [fact for fact in client_facts if is_active_on_cutoff_fact(fact)]
        contact_candidates = [
            fact
            for fact in client_facts
            if normalized_status(fact) == "Контакт с клиентом" and parse_date(fact.get("end_date"))
        ]
        for candidate in contact_candidates:
            candidate_ref = (candidate.get("subscription_ref") or "").strip()
            candidate_doc = (candidate.get("document_number") or "").strip()
            candidate_start = parse_date(candidate.get("start_date"))
            candidate_end = parse_date(candidate.get("end_date"))
            candidate_sale = (candidate.get("sale_datetime") or "").strip()
            if not candidate_ref or not candidate_start or not candidate_end:
                continue
            best_current: dict[str, str] | None = None
            for current in current_active:
                current_ref = (current.get("subscription_ref") or "").strip()
                if current_ref == candidate_ref:
                    continue
                current_start = parse_date(current.get("start_date"))
                current_end = parse_date(current.get("end_date"))
                current_sale = (current.get("sale_datetime") or "").strip()
                if not current_start or not current_end:
                    continue
                if candidate_sale < current_sale:
                    continue
                if candidate_start < current_start:
                    continue
                if candidate_end < cutoff_date:
                    continue
                if best_current is None or str(current.get("end_date", "")) > str(best_current.get("end_date", "")):
                    best_current = current
            if best_current is None:
                continue
            excluded_refs.add(candidate_ref)
            excluded_rows[candidate_ref] = {
                "rule": "exclude_active_later_contact_full",
                "contract_id": candidate_doc,
                "client_id": client_id,
                "client_fio": candidate.get("effective_client_fio", ""),
                "contract_name": candidate.get("subscription_name", ""),
                "sale_datetime": candidate.get("sale_datetime", ""),
                "start_date": candidate.get("start_date", ""),
                "end_date": candidate.get("end_date", ""),
                "status": candidate.get("status", ""),
                "price_candidate": candidate.get("rg_price", ""),
                "paid_candidate": candidate.get("rg_paid_candidate", ""),
                "matched_payment_ref": candidate.get("matched_payment_ref", ""),
                "matched_payment_amount": candidate.get("matched_payment_amount", ""),
                "matched_payment_method": candidate.get("matched_payment_method", ""),
                "current_contract_id": best_current.get("document_number", ""),
                "current_contract_name": best_current.get("subscription_name", ""),
                "current_start_date": best_current.get("start_date", ""),
                "current_end_date": best_current.get("end_date", ""),
                "current_status": best_current.get("status", ""),
                "current_matched_payment_ref": best_current.get("matched_payment_ref", ""),
                "current_matched_payment_amount": best_current.get("matched_payment_amount", ""),
                "current_matched_payment_method": best_current.get("matched_payment_method", ""),
            }
    return excluded_refs, sorted(excluded_rows.values(), key=lambda row: (row["client_id"], row["contract_id"]))


def build_rows(
    source_clients: dict[str, SourceClient],
    cards: dict[str, str],
    facts: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]], list[dict[str, str]], dict[str, Counter]]:
    client_rows: list[dict[str, Any]] = []
    uncertainties: list[dict[str, str]] = []
    counters: dict[str, Counter] = defaultdict(Counter)
    template_by_name: dict[str, dict[str, Any]] = {}
    template_variants: dict[str, set[tuple[Any, ...]]] = defaultdict(set)
    excluded_refs, excluded_rows = find_contact_next_exclusions(facts)
    counters["exclusions"]["exclude_active_later_contact_full"] = len(excluded_rows)

    for fact in facts:
        if (fact.get("subscription_ref") or "").strip() in excluded_refs:
            continue
        client_id = (fact.get("client_id") or "").strip()
        source = source_clients.get(client_id)
        if not source:
            counters["facts"]["not_in_final_client_xlsx"] += 1
            continue

        contract_name = (fact.get("subscription_name") or "").strip()
        contract_id = (fact.get("document_number") or "").strip() or (fact.get("subscription_ref") or "").strip()
        duration, duration_source = compute_duration_months(fact)
        price, paid, payment_left, money_source = compute_money(fact)
        business_override = business_zero_override_reason(fact)
        if business_override:
            price = Decimal("0")
            paid = Decimal("0")
            payment_left = Decimal("0")
            money_source = business_override
        is_subrent = fact.get("is_subrent") == "1"
        is_limited_subrent = fact.get("is_limited_subrent") == "1"
        visits = parse_template_visits(contract_name) if is_limited_subrent else None
        payment_type = map_payment_type(fact.get("matched_payment_method", ""))
        if not payment_type:
            payment_type = payment_type_from_direct_blank_method(fact, price, business_override)
        if not payment_type:
            payment_type = payment_type_from_positive_no_payment(
                price,
                business_override,
                (fact.get("matched_payment_ref") or "").strip(),
            )
        if business_override:
            payment_type = ""
        freeze_days = int_or_blank(decimal_value(fact.get("rg_freeze_days")))
        activation_date = excel_date(fact.get("start_date"))
        end_date = excel_date(fact.get("end_date"))
        visits_left, visits_left_source, visits_left_issue = compute_visits_left(fact, contract_name)

        row = {
            "tag": source.tag,
            "contract_id": contract_id,
            "client_id": client_id,
            "phone": source.phone,
            "client_fio": source.client_fio,
            "contract_name": contract_name,
            "card": cards.get(client_id, ""),
            "duration": duration,
            "duration_type": "месяц",
            "create_date": source.create_date,
            "payment_date": excel_date(fact.get("sale_date")),
            "activation_date": activation_date,
            "end_date": end_date,
            "freeze": freeze_days,
            "guests": None,
            "visits_left": visits_left,
            "price": excel_number(price),
            "amount_of_payments": excel_number(paid),
            "payment_left": excel_number(payment_left),
            "type_of_payment": payment_type,
            "manager": source.manager,
            "_subscription_ref": fact.get("subscription_ref", ""),
            "_product_ref": fact.get("product_ref", ""),
            "_product_class": fact.get("product_class", ""),
            "_is_subrent": "1" if is_subrent else "0",
            "_is_limited_subrent": "1" if is_limited_subrent else "0",
            "_duration_source": duration_source,
            "_money_source": money_source,
            "_payment_method_raw": fact.get("matched_payment_method", ""),
            "_payment_match_source": fact.get("matched_payment_match_source", ""),
            "_business_override": business_override,
            "_membership_sale_line_amount": fact.get("membership_sale_line_amount", ""),
            "_membership_sale_line_count": fact.get("membership_sale_line_count", ""),
            "_document131_refund_count": fact.get("document131_refund_count", ""),
            "_document131_posted_unmarked_refund_count": fact.get(
                "document131_posted_unmarked_refund_count", ""
            ),
            "_owner_change_ref": fact.get("owner_change_ref", ""),
            "_visits_left_source": visits_left_source,
            "_subrent_rg3336_case_group": fact.get("subrent_rg3336_case_group", ""),
            "_refuser_placeholder": "0",
        }
        client_rows.append(row)

        counters["rows_by_product_class"][fact.get("product_class", "")] += 1
        if source.tag == "отказники":
            counters["refusers"]["real_membership_rows"] += 1
        counters["duration_source"][duration_source] += 1
        counters["money_source"][money_source] += 1
        counters["payment_type"][payment_type or "blank"] += 1
        counters["visits_left_source"][visits_left_source] += 1
        if business_override:
            counters["business_override"][business_override] += 1
        if is_subrent:
            counters["special"]["subrent_rows"] += 1
        if is_limited_subrent:
            counters["special"]["limited_subrent_rows"] += 1
            counters["subrent_rg3336_case_group"][fact.get("subrent_rg3336_case_group", "") or "blank"] += 1
            if visits_left_issue:
                uncertainties.append(
                    {
                        "issue_type": "limited_subrent_visits_left_rule_review",
                        "contract_id": contract_id,
                        "client_id": client_id,
                        "client_fio": source.client_fio,
                        "contract_name": contract_name,
                        "details": visits_left_issue,
                    }
                )
        if "installment_without_paid_candidate" == money_source:
            uncertainties.append(
                {
                    "issue_type": "installment_without_paid_candidate",
                    "contract_id": contract_id,
                    "client_id": client_id,
                    "client_fio": source.client_fio,
                    "contract_name": contract_name,
                    "details": "Name contains installment marker but InfoRg3060._Fld3072 is zero; payment_left set to full price.",
                }
            )
        if not payment_type and price > 0:
            uncertainties.append(
                {
                    "issue_type": "payment_type_not_matched",
                    "contract_id": contract_id,
                    "client_id": client_id,
                    "client_fio": source.client_fio,
                    "contract_name": contract_name,
                    "details": "No nearby payment document with a mappable method was found.",
                }
            )
        if not activation_date or not end_date:
            uncertainties.append(
                {
                    "issue_type": "blank_activation_or_end_date",
                    "contract_id": contract_id,
                    "client_id": client_id,
                    "client_fio": source.client_fio,
                    "contract_name": contract_name,
                    "details": f"activation_date={fact.get('start_date', '')}; end_date={fact.get('end_date', '')}",
                }
            )

        branches_access = "Все" if "мультикарта" in normalize_key(contract_name) else "Продажа"
        template_key = normalize_key(contract_name)
        template_variant = (
            excel_number(price),
            duration,
            visits,
            freeze_days,
            branches_access,
        )
        template_variants[template_key].add(template_variant)
        template_row = {
            "branches_access": branches_access,
            "name": contract_name,
            "price": excel_number(price),
            "duration": duration,
            "duration_type": "месяц",
            "visits": visits,
            "guests": None,
            "freeze": freeze_days,
            "first_visit_activation": None,
            "archive": None,
            "category": None,
            "legal_entity": None,
            "_last_sale_datetime": fact.get("sale_datetime", ""),
        }
        current = template_by_name.get(template_key)
        if current is None or str(template_row["_last_sale_datetime"]) >= str(current.get("_last_sale_datetime", "")):
            template_by_name[template_key] = template_row

    row_client_ids = {str(row.get("client_id", "")) for row in client_rows}
    for source in source_clients.values():
        if source.tag != "отказники" or source.client_id in row_client_ids:
            continue
        client_rows.append(
            {
                "tag": "отказники",
                "contract_id": "",
                "client_id": source.client_id,
                "phone": source.phone,
                "client_fio": source.client_fio,
                "contract_name": "",
                "card": cards.get(source.client_id, ""),
                "duration": None,
                "duration_type": None,
                "create_date": source.create_date,
                "payment_date": None,
                "activation_date": None,
                "end_date": None,
                "freeze": None,
                "guests": None,
                "visits_left": None,
                "price": None,
                "amount_of_payments": None,
                "payment_left": None,
                "type_of_payment": "",
                "manager": source.manager,
                "_subscription_ref": "",
                "_product_ref": "",
                "_product_class": "refuser_without_membership",
                "_is_subrent": "0",
                "_is_limited_subrent": "0",
                "_duration_source": "refuser_without_membership",
                "_money_source": "refuser_without_membership",
                "_payment_method_raw": "",
                "_payment_match_source": "",
                "_business_override": "",
                "_membership_sale_line_amount": "",
                "_membership_sale_line_count": "",
                "_document131_refund_count": "",
                "_document131_posted_unmarked_refund_count": "",
                "_owner_change_ref": "",
                "_visits_left_source": "refuser_without_membership",
                "_subrent_rg3336_case_group": "",
                "_refuser_placeholder": "1",
            }
        )
        counters["rows_by_product_class"]["refuser_without_membership"] += 1
        counters["duration_source"]["refuser_without_membership"] += 1
        counters["money_source"]["refuser_without_membership"] += 1
        counters["payment_type"]["blank"] += 1
        counters["visits_left_source"]["refuser_without_membership"] += 1
        counters["refusers"]["placeholder_rows"] += 1

    for template_key, variants in template_variants.items():
        if len(variants) > 1:
            template = template_by_name[template_key]
            uncertainties.append(
                {
                    "issue_type": "template_variants_collapsed_to_latest_sale",
                    "contract_id": "",
                    "client_id": "",
                    "client_fio": "",
                    "contract_name": str(template.get("name", "")),
                    "details": f"{len(variants)} price/duration/freeze/access variants found for one template name; latest sale variant used.",
                }
            )

    canonical_template_names = {key: str(row.get("name", "")) for key, row in template_by_name.items()}
    for row in client_rows:
        canonical_name = canonical_template_names.get(normalize_key(str(row.get("contract_name", ""))))
        if canonical_name:
            row["contract_name"] = canonical_name

    template_rows = sorted(
        ({k: v for k, v in row.items() if not k.startswith("_")} for row in template_by_name.values()),
        key=lambda item: normalize_key(str(item.get("name", ""))),
    )
    client_rows.sort(key=lambda item: (str(item["client_id"]), str(item["payment_date"] or ""), str(item["contract_id"])))
    return client_rows, template_rows, uncertainties, excluded_rows, counters


def write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def snapshot_row_style(ws, row_number: int, width: int) -> list[dict[str, Any]]:
    styles = []
    for col in range(1, width + 1):
        cell = ws.cell(row_number, col)
        styles.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )
    return styles


def apply_style(cell, style: dict[str, Any]) -> None:
    cell.font = copy(style["font"])
    cell.fill = copy(style["fill"])
    cell.border = copy(style["border"])
    cell.alignment = copy(style["alignment"])
    cell.number_format = style["number_format"]
    cell.protection = copy(style["protection"])


def clear_data_rows(ws, first_data_row: int = 3) -> None:
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)


def trim_columns(ws, width: int) -> None:
    if ws.max_column > width:
        ws.delete_cols(width + 1, ws.max_column - width)


def write_workbook(
    template_path: Path,
    output_path: Path,
    headers: list[str],
    rows: list[dict[str, Any]],
    russian_headers: list[str] | None = None,
) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    width = len(headers)
    trim_columns(ws, max(width, 1))

    data_number_formats = [ws.cell(3, col).number_format for col in range(1, width + 1)]
    clear_data_rows(ws, 3)

    for col_idx, header in enumerate(headers, start=1):
        ws.cell(1, col_idx).value = header
        if russian_headers:
            ws.cell(2, col_idx).value = russian_headers[col_idx - 1]

    for row in rows:
        ws.append([row.get(header) for header in headers])

    date_headers = {"create_date", "payment_date", "activation_date", "end_date"}
    money_headers = {"price", "amount_of_payments", "payment_left"}
    format_cols: dict[int, str] = {}
    for col_idx, header in enumerate(headers, start=1):
        if header in date_headers:
            format_cols[col_idx] = "yyyy-mm-dd"
        elif header in money_headers:
            format_cols[col_idx] = "#,##0.00"
        elif data_number_formats[col_idx - 1] and data_number_formats[col_idx - 1] != "General":
            format_cols[col_idx] = data_number_formats[col_idx - 1]
    for row_idx in range(3, 3 + len(rows)):
        for col_idx, number_format in format_cols.items():
            ws.cell(row_idx, col_idx).number_format = number_format

    ws.freeze_panes = "A3"
    for col_idx in range(1, width + 1):
        letter = get_column_letter(col_idx)
        current_width = ws.column_dimensions[letter].width or 12
        ws.column_dimensions[letter].width = min(max(current_width, 12), 34)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    wb.close()


def build_validation(
    source_clients: dict[str, SourceClient],
    client_rows: list[dict[str, Any]],
    template_rows: list[dict[str, Any]],
    uncertainties: list[dict[str, str]],
    counters: dict[str, Counter],
) -> str:
    contract_ids = [str(row["contract_id"]) for row in client_rows if str(row.get("contract_id") or "").strip()]
    duplicate_contract_ids = [item for item, count in Counter(contract_ids).items() if count > 1]
    row_client_ids = {str(row["client_id"]) for row in client_rows}
    template_names = {str(row["name"]) for row in template_rows}
    missing_template_names = sorted(
        {str(row["contract_name"]) for row in client_rows if str(row.get("contract_name") or "").strip()}
        - template_names
    )
    required_blank_counts = Counter()
    for row in client_rows:
        required_fields = ["tag", "client_id", "client_fio", "create_date", "manager"]
        if row.get("_refuser_placeholder") != "1":
            required_fields = [
                "contract_id",
                "client_id",
                "client_fio",
                "contract_name",
                "create_date",
                "payment_date",
                "price",
                "manager",
            ]
        for field in required_fields:
            if row.get(field) in (None, ""):
                required_blank_counts[field] += 1
    refuser_client_ids = {client.client_id for client in source_clients.values() if client.tag == "отказники"}
    tagged_refuser_client_ids = {str(row["client_id"]) for row in client_rows if row.get("tag") == "отказники"}
    refuser_placeholder_rows = sum(1 for row in client_rows if row.get("_refuser_placeholder") == "1")
    refuser_real_rows = sum(
        1
        for row in client_rows
        if row.get("tag") == "отказники" and row.get("_refuser_placeholder") != "1"
    )

    lines = [
        "# Membership import validation",
        "",
        f"- source final clients: {len(source_clients)}",
        f"- clients with at least one membership row: {len(row_client_ids)}",
        f"- source clients without membership rows: {len(source_clients) - len(row_client_ids)}",
        f"- client membership rows: {len(client_rows)}",
        f"- membership template rows: {len(template_rows)}",
        f"- duplicate contract_id count: {len(duplicate_contract_ids)}",
        f"- contract names missing in template file: {len(missing_template_names)}",
        f"- uncertainty rows: {len(uncertainties)}",
        f"- refuser source clients: {len(refuser_client_ids)}",
        f"- refuser clients present in membership rows: {len(tagged_refuser_client_ids)}",
        f"- refuser real membership rows: {refuser_real_rows}",
        f"- refuser placeholder rows: {refuser_placeholder_rows}",
        "",
        "## Row Classes",
        "",
    ]
    for key, value in counters["rows_by_product_class"].most_common():
        lines.append(f"- {key or '<blank>'}: {value}")
    lines.extend(["", "## Money Sources", ""])
    for key, value in counters["money_source"].most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Business Overrides", ""])
    if counters["business_override"]:
        for key, value in counters["business_override"].most_common():
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- none")
    lines.extend(["", "## Exclusions", ""])
    if counters["exclusions"]:
        for key, value in counters["exclusions"].most_common():
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- none")
    lines.extend(["", "## Payment Types", ""])
    for key, value in counters["payment_type"].most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Visits Left Sources", ""])
    for key, value in counters["visits_left_source"].most_common():
        lines.append(f"- {key}: {value}")
    lines.extend(["", "## Limited Subrent Register Balance Groups", ""])
    if counters["subrent_rg3336_case_group"]:
        for key, value in counters["subrent_rg3336_case_group"].most_common():
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- none")
    lines.extend(["", "## Refusers", ""])
    if counters["refusers"]:
        for key, value in counters["refusers"].most_common():
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- none")
    lines.extend(["", "## Required Blank Counts", ""])
    if required_blank_counts:
        for key, value in required_blank_counts.items():
            lines.append(f"- {key}: {value}")
    else:
        lines.append("- none")
    lines.extend(["", "## Hard Checks", ""])
    lines.append(f"- all row clients are from source final XLSX: {'yes' if row_client_ids <= set(source_clients) else 'no'}")
    lines.append(f"- contract_id unique: {'yes' if not duplicate_contract_ids else 'no'}")
    lines.append(f"- every contract_name exists in templates: {'yes' if not missing_template_names else 'no'}")
    lines.append(f"- every refuser client has a tagged row: {'yes' if refuser_client_ids <= tagged_refuser_client_ids else 'no'}")
    return "\n".join(lines) + "\n"


def build_rassrochka_report(client_rows: list[dict[str, Any]], uncertainties: list[dict[str, str]]) -> str:
    installment_rows = [row for row in client_rows if "рассроч" in normalize_key(str(row.get("contract_name", "")))]
    unresolved = [row for row in uncertainties if row["issue_type"] == "installment_without_paid_candidate"]
    payment_left_positive = [row for row in installment_rows if decimal_value(row.get("payment_left")) > 0]
    lines = [
        "# Rassrochka validation",
        "",
        f"- installment rows by name marker: {len(installment_rows)}",
        f"- installment rows with positive payment_left: {len(payment_left_positive)}",
        f"- installment rows with missing paid candidate: {len(unresolved)}",
        "",
        "Rule used: `amount_of_payments = InfoRg3060._Fld3072` when it is positive; otherwise, for rows with `рассрочка` in the name, fallback to the nearest matched `Document152` payment amount. `payment_left = price - amount_of_payments`. Rows are flagged only when both sources are empty.",
        "",
        "## First Flagged Rows",
        "",
    ]
    for item in unresolved[:30]:
        lines.append(f"- {item['client_id']} {item['client_fio']} | {item['contract_id']} | {item['contract_name']}")
    if not unresolved:
        lines.append("- none")
    return "\n".join(lines) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-output-dir", default="output/20251115_0800_fix_owner")
    parser.add_argument("--output-dir", default="output/20251115_0800_fix_owner_new_import")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    parser.add_argument("--facts-tsv", default=None)
    parser.add_argument("--client-template", default="new-changes/Импорт_абонементы_клиентов.xlsx")
    parser.add_argument("--membership-template", default="new-changes/Импорт_шаблоны_абонементов.xlsx")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_output_dir = as_abs(args.source_output_dir)
    output_dir = as_abs(args.output_dir)
    reports_dir = output_dir / "reports"
    staging_dir = output_dir / "staging"
    source_clients_xlsx = source_output_dir / f"fitbase_active_clients_import_zayavki_{args.date_stamp}_all_funnels.xlsx"
    refuser_clients_csv = source_output_dir / "csv" / "new_application_refusers.csv"
    facts_tsv = as_abs(args.facts_tsv) if args.facts_tsv else staging_dir / "membership_import_facts.tsv"

    source_clients = read_source_clients(source_clients_xlsx)
    source_clients.update(read_refuser_clients(refuser_clients_csv))
    cards = read_cards(source_output_dir / "staging")
    facts = read_facts(facts_tsv)
    client_rows, template_rows, uncertainties, excluded_rows, counters = build_rows(source_clients, cards, facts)

    client_xlsx = output_dir / f"fitbase_import_abonementy_clientov_{args.date_stamp}.xlsx"
    template_xlsx = output_dir / f"fitbase_import_shablony_abonementov_{args.date_stamp}.xlsx"
    write_workbook(as_abs(args.client_template), client_xlsx, CLIENT_HEADERS, client_rows, CLIENT_RUS_HEADERS)
    write_workbook(as_abs(args.membership_template), template_xlsx, TEMPLATE_HEADERS, template_rows, TEMPLATE_RUS_HEADERS)

    write_csv(staging_dir / "membership_import_rows.csv", client_rows, CLIENT_HEADERS + [
        "_subscription_ref",
        "_product_ref",
        "_product_class",
        "_is_subrent",
        "_is_limited_subrent",
        "_duration_source",
        "_money_source",
        "_payment_method_raw",
        "_payment_match_source",
        "_business_override",
        "_membership_sale_line_amount",
        "_membership_sale_line_count",
        "_document131_refund_count",
        "_document131_posted_unmarked_refund_count",
        "_owner_change_ref",
        "_visits_left_source",
        "_subrent_rg3336_case_group",
        "_refuser_placeholder",
    ])
    write_csv(staging_dir / "membership_template_rows.csv", template_rows, TEMPLATE_HEADERS)
    write_csv(staging_dir / "membership_import_excluded_rows.csv", excluded_rows, [
        "rule",
        "contract_id",
        "client_id",
        "client_fio",
        "contract_name",
        "sale_datetime",
        "start_date",
        "end_date",
        "status",
        "price_candidate",
        "paid_candidate",
        "matched_payment_ref",
        "matched_payment_amount",
        "matched_payment_method",
        "current_contract_id",
        "current_contract_name",
        "current_start_date",
        "current_end_date",
        "current_status",
        "current_matched_payment_ref",
        "current_matched_payment_amount",
        "current_matched_payment_method",
    ])
    write_csv(
        reports_dir / "membership_import_uncertainties.csv",
        uncertainties,
        ["issue_type", "contract_id", "client_id", "client_fio", "contract_name", "details"],
    )
    zero_price_counts: Counter[tuple[str, str]] = Counter()
    for row in client_rows:
        if decimal_value(row.get("price")) == 0:
            zero_price_counts[(str(row.get("_product_class", "")), str(row.get("contract_name", "")))] += 1
    zero_price_rows = [
        {"product_class": product_class, "contract_name": contract_name, "rows_count": rows_count}
        for (product_class, contract_name), rows_count in zero_price_counts.most_common()
    ]
    write_csv(reports_dir / "zero_price_report.csv", zero_price_rows, ["product_class", "contract_name", "rows_count"])

    validation_report = build_validation(source_clients, client_rows, template_rows, uncertainties, counters)
    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "validation_report.md").write_text(validation_report, encoding="utf-8")
    (reports_dir / "rassrochka_validation.md").write_text(
        build_rassrochka_report(client_rows, uncertainties),
        encoding="utf-8",
    )

    print(validation_report)
    print(f"client_xlsx={client_xlsx}")
    print(f"template_xlsx={template_xlsx}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
