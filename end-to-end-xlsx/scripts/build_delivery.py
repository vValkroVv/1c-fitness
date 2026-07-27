#!/usr/bin/env python3
"""Assemble a versioned clean delivery and its separate problem XLSX files."""

from __future__ import annotations

import argparse
import shutil
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]


def as_abs(path: str | Path) -> Path:
    value = Path(path)
    return value if value.is_absolute() else ROOT / value


def one_match(directory: Path, pattern: str) -> Path:
    matches = sorted(directory.glob(pattern))
    if len(matches) != 1:
        raise RuntimeError(
            f"Expected exactly one {pattern!r} in {directory}, found {len(matches)}"
        )
    return matches[0]


def read_problem_contracts(path: Path) -> set[str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    if "contract_id" not in headers:
        workbook.close()
        raise RuntimeError(f"No contract_id column in {path}")
    contract_index = headers.index("contract_id")
    contracts = {
        str(row[contract_index] or "").strip()
        for row in worksheet.iter_rows(min_row=2, values_only=True)
        if str(row[contract_index] or "").strip()
    }
    workbook.close()
    return contracts


def canonical_contract_id(value: str) -> str:
    """Return the zero-padded contract identifier used in membership XLSX files."""

    text = str(value or "").strip()
    if not text or not text.isdigit():
        raise ValueError(f"Contract ID must contain digits only, got {value!r}")
    return text.zfill(11)


def validate_run_id(value: str) -> str:
    """Accept a single safe path component for an immutable delivery run."""

    run_id = str(value or "").strip()
    if not run_id or run_id in {".", ".."} or Path(run_id).name != run_id:
        raise ValueError(
            "--delivery-run-id must be one non-empty directory name without path separators"
        )
    return run_id


def autosize_problem_sheet(worksheet: Any) -> None:
    """Match the sizing rules used by the existing problem1-3 builder."""

    for column in worksheet.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        worksheet.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def write_single_contract_problem_workbook(
    destination: Path,
    headers: list[str],
    values: list[Any],
) -> None:
    """Write one membership row in exactly the same shape as problem1-3."""

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Импорт_абонементы"
    worksheet.append(headers)
    worksheet.append(values)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in worksheet[2]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    autosize_problem_sheet(worksheet)
    destination.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(destination)
    workbook.close()


def filter_membership_workbook(
    source: Path,
    destination: Path,
    problem_ids: set[str],
    template_path: Path,
    capture_ids: set[str] | None = None,
) -> tuple[int, list[str], dict[str, list[Any]]]:
    """Rewrite the large workbook once while excluding problem contracts.

    Calling Worksheet.delete_rows for many scattered rows repeatedly shifts
    millions of cells.  A streaming source plus the original template keeps
    the same workbook structure and makes runtime linear in the row count.
    """

    source_workbook = load_workbook(source, read_only=True, data_only=False)
    source_sheet = source_workbook.active
    headers = [
        str(value or "")
        for value in next(
            source_sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    if "contract_id" not in headers:
        source_workbook.close()
        raise RuntimeError(f"No contract_id column in {source}")
    contract_index = headers.index("contract_id")

    target_workbook = load_workbook(template_path)
    target_sheet = target_workbook.active
    width = len(headers)
    if target_sheet.max_column > width:
        target_sheet.delete_cols(width + 1, target_sheet.max_column - width)
    number_formats = [
        target_sheet.cell(3, column).number_format for column in range(1, width + 1)
    ]
    if target_sheet.max_row >= 3:
        target_sheet.delete_rows(3, target_sheet.max_row - 2)

    source_russian_headers = next(
        source_sheet.iter_rows(min_row=2, max_row=2, values_only=True)
    )
    for column, value in enumerate(headers, start=1):
        target_sheet.cell(1, column).value = value
        target_sheet.cell(2, column).value = source_russian_headers[column - 1]

    date_headers = {"create_date", "payment_date", "activation_date", "end_date"}
    money_headers = {"price", "amount_of_payments", "payment_left"}
    format_columns: dict[int, str] = {}
    for column, header in enumerate(headers, start=1):
        if header in date_headers:
            format_columns[column] = "yyyy-mm-dd"
        elif header in money_headers:
            format_columns[column] = "#,##0.00"
        elif number_formats[column - 1] and number_formats[column - 1] != "General":
            format_columns[column] = number_formats[column - 1]

    capture_ids = capture_ids or set()
    captured_rows: dict[str, list[Any]] = {}
    found: set[str] = set()
    removed = 0
    target_row = 2
    for values in source_sheet.iter_rows(min_row=3, values_only=True):
        contract_id = str(values[contract_index] or "").strip()
        if contract_id in problem_ids:
            found.add(contract_id)
            removed += 1
            if contract_id in capture_ids:
                if contract_id in captured_rows:
                    source_workbook.close()
                    target_workbook.close()
                    raise RuntimeError(
                        f"Captured contract_id occurs more than once: {contract_id}"
                    )
                captured_rows[contract_id] = list(values[:width])
            continue
        target_sheet.append(list(values[:width]))
        target_row += 1
        for column, number_format in format_columns.items():
            target_sheet.cell(target_row, column).number_format = number_format

    missing = sorted(problem_ids - found)
    if missing:
        source_workbook.close()
        target_workbook.close()
        raise RuntimeError(
            f"Problem contract_id values not found in full membership workbook: {missing[:10]}"
        )
    if removed != len(problem_ids):
        source_workbook.close()
        target_workbook.close()
        raise RuntimeError(
            "A problem contract_id occurs more than once in the full membership workbook: "
            f"rows={removed}, unique={len(problem_ids)}"
        )

    target_sheet.freeze_panes = "A3"
    for column in range(1, width + 1):
        letter = get_column_letter(column)
        current_width = target_sheet.column_dimensions[letter].width or 12
        target_sheet.column_dimensions[letter].width = min(max(current_width, 12), 34)
    destination.parent.mkdir(parents=True, exist_ok=True)
    target_workbook.save(destination)
    source_workbook.close()
    target_workbook.close()
    missing_captures = sorted(capture_ids - set(captured_rows))
    if missing_captures:
        raise RuntimeError(
            f"Requested problem rows were not captured: {missing_captures[:10]}"
        )
    return removed, headers, captured_rows


def build(args: argparse.Namespace) -> None:
    owner_dir = as_abs(args.owner_dir)
    imports_dir = as_abs(args.imports_dir)
    output_base = as_abs(args.output_dir)
    if args.delivery_run_id:
        run_id = validate_run_id(args.delivery_run_id)
        output_dir = (output_base / run_id).resolve()
        if output_dir.exists() and any(output_dir.iterdir()):
            raise FileExistsError(
                f"Versioned delivery already exists and will not be overwritten: {output_dir}"
            )
    else:
        run_id = ""
        output_dir = output_base
    report_path = as_abs(args.report)
    membership_template = as_abs(args.membership_template)
    date_stamp = args.date_stamp

    output_dir.mkdir(parents=True, exist_ok=True)
    if not run_id:
        # Backward-compatible behaviour for the original pipeline. Versioned
        # runs are immutable and fail above instead of deleting prior output.
        for stale in output_dir.glob("*.xlsx"):
            stale.unlink()

    main_name = f"fitbase_active_clients_import_zayavki_{date_stamp}_all_funnels.xlsx"
    cards_name = f"fitbase_active_clients_plastic_cards_{date_stamp}_all_funnels.xlsx"
    membership_name = f"fitbase_import_abonementy_clientov_{date_stamp}.xlsx"
    membership_templates_name = f"fitbase_import_shablony_abonementov_{date_stamp}.xlsx"
    services_name = f"fitbase_import_uslugi_clientov_{date_stamp}.xlsx"
    service_templates_name = f"fitbase_import_shablony_uslug_{date_stamp}.xlsx"

    static_sources = {
        main_name: owner_dir / main_name,
        cards_name: owner_dir / cards_name,
        membership_templates_name: imports_dir / membership_templates_name,
        services_name: imports_dir / services_name,
        service_templates_name: imports_dir / service_templates_name,
    }
    for name, source in static_sources.items():
        if not source.is_file():
            raise FileNotFoundError(source)
        shutil.copy2(source, output_dir / name)

    base_problem_sources = (
        []
        if args.financial_problem_groups_resolved
        else [
            one_match(
                imports_dir,
                f"active_problem_1_no_payment_cash_*_cases_{date_stamp}.xlsx",
            ),
            one_match(
                imports_dir,
                f"active_problem_2_zero_price_direct_full_*_cases_{date_stamp}.xlsx",
            ),
            one_match(
                imports_dir,
                f"active_problem_3_non_named_payment_left_*_cases_{date_stamp}.xlsx",
            ),
        ]
    )
    problem_sets = [read_problem_contracts(path) for path in base_problem_sources]
    problem_labels = [path.name for path in base_problem_sources]

    problem4_contract_id = ""
    problem4_path: Path | None = None
    problem4_source: Path | None = None
    if args.problem4_contract_id:
        problem4_contract_id = canonical_contract_id(args.problem4_contract_id)
        short_contract_id = problem4_contract_id.lstrip("0") or "0"
        problem4_path = output_dir / (
            "problem_4_subrent_visits_left_contract_"
            f"{short_contract_id}_1_case_{date_stamp}.xlsx"
        )
        problem_sets.append({problem4_contract_id})
        problem_labels.append(problem4_path.name)
        if args.problem4_source:
            problem4_source = as_abs(args.problem4_source)
            if not problem4_source.is_file():
                raise FileNotFoundError(problem4_source)
            source_contracts = read_problem_contracts(problem4_source)
            if source_contracts != {problem4_contract_id}:
                raise RuntimeError(
                    "Configured problem4 source contract set differs from "
                    f"{problem4_contract_id}: {sorted(source_contracts)}"
                )

    problem_ids = set().union(*problem_sets) if problem_sets else set()
    summed = sum(len(values) for values in problem_sets)
    if len(problem_ids) != summed:
        raise RuntimeError(
            f"A contract_id belongs to more than one problem group: summed={summed}, union={len(problem_ids)}"
        )

    for source in base_problem_sources:
        output_name = source.name.removeprefix("active_")
        shutil.copy2(source, output_dir / output_name)

    full_membership = imports_dir / membership_name
    if not full_membership.is_file():
        raise FileNotFoundError(full_membership)
    capture_ids = (
        {problem4_contract_id}
        if problem4_contract_id and problem4_source is None
        else set()
    )
    removed, membership_headers, captured_rows = filter_membership_workbook(
        full_membership,
        output_dir / membership_name,
        problem_ids,
        membership_template,
        capture_ids,
    )
    if problem4_path is not None and problem4_source is not None:
        shutil.copy2(problem4_source, problem4_path)
    elif problem4_path is not None:
        write_single_contract_problem_workbook(
            problem4_path,
            membership_headers,
            captured_rows[problem4_contract_id],
        )

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        "\n".join(
            [
                "# Delivery build",
                "",
                f"- date stamp: `{date_stamp}`",
                f"- output: `{output_dir}`",
                f"- delivery run id: `{run_id or 'legacy/unversioned'}`",
                "- clean XLSX: `6`",
                f"- problem XLSX: `{len(problem_sets)}`",
                f"- unique problem contract_id: `{len(problem_ids)}`",
                f"- rows removed from clean membership XLSX: `{removed}`",
                "- financial problem groups 1-3 resolved in clean membership: "
                f"`{'yes' if args.financial_problem_groups_resolved else 'no'}`",
                "",
                "## Problem groups",
                "",
                *[
                    f"- `{name}`: `{len(values)}`"
                    for name, values in zip(problem_labels, problem_sets, strict=True)
                ],
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(f"delivery_dir={output_dir}")
    print(f"problem_contract_ids={len(problem_ids)}")
    print(f"membership_rows_removed={removed}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--owner-dir", required=True)
    parser.add_argument("--imports-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument(
        "--delivery-run-id",
        help=(
            "Optional child directory under --output-dir. A non-empty versioned "
            "directory is never overwritten."
        ),
    )
    parser.add_argument("--date-stamp", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--membership-template", required=True)
    parser.add_argument(
        "--problem4-contract-id",
        help=(
            "Optional contract ID for a one-row problem4 workbook; short numeric "
            "IDs are zero-padded to the membership XLSX convention."
        ),
    )
    parser.add_argument(
        "--problem4-source",
        help=(
            "Optional previously approved one-row problem4 XLSX to copy "
            "byte-for-byte instead of regenerating it."
        ),
    )
    parser.add_argument(
        "--financial-problem-groups-resolved",
        action="store_true",
        help=(
            "Keep former financial problem groups 1-3 in the clean membership "
            "workbook. Use only when their money fields were rebuilt from an "
            "authoritative cutoff-aware source."
        ),
    )
    return parser.parse_args()


if __name__ == "__main__":
    build(parse_args())
