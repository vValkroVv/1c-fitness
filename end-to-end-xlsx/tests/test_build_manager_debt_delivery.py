from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MODULE_PATH = ROOT / "scripts" / "build_manager_debt_delivery.py"
SPEC = importlib.util.spec_from_file_location("manager_debt_builder", MODULE_PATH)
assert SPEC is not None and SPEC.loader is not None
builder = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = builder
SPEC.loader.exec_module(builder)


def money(price: int, paid: int, debt: int):
    return builder.MoneyTriple(
        Decimal(price), Decimal(paid), Decimal(debt)
    )


def problem(
    *,
    contract_id: str,
    fio: str,
    phone: str,
    create_date: date,
):
    return builder.ProblemCase(
        group=3,
        source_file="problem_3.xlsx",
        values=(),
        contract_id=contract_id,
        client_id="CLIENT",
        client_fio=fio,
        phone=phone,
        create_date=create_date,
        payment_date=create_date,
        old_money=money(100, 50, 50),
    )


class NormalizationTests(unittest.TestCase):
    def test_fio_normalization_is_exact_but_format_tolerant(self) -> None:
        self.assertEqual(
            builder.normalize_fio("  ФЁДОРОВА—АННА  "),
            "федорова анна",
        )

    def test_phone_normalization_supports_multiple_russian_numbers(self) -> None:
        self.assertEqual(
            builder.normalize_phones(
                "+7 (911) 111-22-33, 8 (921) 444-55-66"
            ),
            {"79111112233", "79214445566"},
        )

    def test_config_rejects_cutoff_different_from_backup_finish(self) -> None:
        config = {
            "run": {
                "date_stamp": "20260630",
                "cutoff_date": "2026-06-30",
                "cutoff_at": "2026-06-30 00:00:00",
                "backup_finish_at": "2026-06-30 23:27:03",
            },
            "inputs": {},
            "rule": {},
            "output": {},
        }
        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "config.yml"
            path.write_text(
                yaml.safe_dump(config, allow_unicode=True), encoding="utf-8"
            )
            with self.assertRaisesRegex(
                ValueError, "RESTORE HEADERONLY.BackupFinishDate"
            ):
                builder.load_config(path)


class ManagerReportTests(unittest.TestCase):
    def test_outline_report_parser_uses_sale_rows_not_client_total(self) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(list(builder.MANAGER_HEADERS))
        sheet.append(["Клиент", None, None, None])
        sheet.append(["Документ продажи", None, None, None])
        sheet.append(["Клиент.Телефон", None, None, None])
        sheet.append(["Филиал", 19800, 9900, 9900])
        sheet.append(["Тестов Иван Иванович", 19800, 9900, 9900])
        sheet.row_dimensions[6].outlineLevel = 1
        sheet.append(["Продажа 1 от 01.01.2026 10:00", 9900, 4950, 4950])
        sheet.row_dimensions[7].outlineLevel = 2
        sheet.append(["+7 (911) 111-22-33", 9900, 4950, 4950])
        sheet.row_dimensions[8].outlineLevel = 3
        sheet.append(["Продажа 2 от 01.01.2026 10:01", 9900, 4950, 4950])
        sheet.row_dimensions[9].outlineLevel = 2
        sheet.append(["+7 (911) 111-22-33", 9900, 4950, 4950])
        sheet.row_dimensions[10].outlineLevel = 3
        sheet.append(["Итого", 19800, 9900, 9900])

        with tempfile.TemporaryDirectory() as directory:
            path = Path(directory) / "report.xlsx"
            workbook.save(path)
            clients = builder.parse_manager_debt_report(path)

        self.assertEqual(len(clients), 1)
        self.assertEqual(len(clients[0].sales), 2)
        self.assertEqual(clients[0].money, money(19800, 9900, 9900))
        self.assertEqual(clients[0].sales[0].money, money(9900, 4950, 4950))


class ResolutionTests(unittest.TestCase):
    def test_multi_contract_equivalent_sales_are_assigned_once_each(self) -> None:
        client = builder.DebtClient(
            row_number=6,
            fio="Тестов Иван Иванович",
            normalized_fio="тестов иван иванович",
            money=money(19800, 9900, 9900),
            sales=[
                builder.DebtSale(
                    row_number=7,
                    document_number="1",
                    sale_at=datetime(2026, 1, 1, 10, 0),
                    money=money(9900, 4950, 4950),
                    phones={"79111112233"},
                ),
                builder.DebtSale(
                    row_number=9,
                    document_number="2",
                    sale_at=datetime(2026, 1, 1, 10, 1),
                    money=money(9900, 4950, 4950),
                    phones={"79111112233"},
                ),
            ],
        )
        cases = [
            problem(
                contract_id="A",
                fio=client.fio,
                phone="79111112233",
                create_date=date(2026, 1, 1),
            ),
            problem(
                contract_id="B",
                fio=client.fio,
                phone="79111112233",
                create_date=date(2026, 1, 1),
            ),
        ]
        resolved = builder.resolve_cases(cases, [client], money(12000, 12000, 0))
        self.assertEqual(
            {item.manager_sale_row for item in resolved},
            {7, 9},
        )
        self.assertTrue(
            all(item.new_money == money(9900, 4950, 4950) for item in resolved)
        )

    def test_absent_person_gets_exact_fallback(self) -> None:
        case = problem(
            contract_id="A",
            fio="Нет В Отчете",
            phone="79990000000",
            create_date=date(2026, 1, 1),
        )
        fallback = money(12000, 12000, 0)
        resolved = builder.resolve_cases([case], [], fallback)
        self.assertEqual(resolved[0].new_money, fallback)
        self.assertEqual(
            resolved[0].source, "fallback_not_in_manager_debt_report"
        )

    def test_phone_conflict_fails_instead_of_silent_match(self) -> None:
        client = builder.DebtClient(
            row_number=6,
            fio="Тестов Иван Иванович",
            normalized_fio="тестов иван иванович",
            money=money(100, 50, 50),
            sales=[
                builder.DebtSale(
                    row_number=7,
                    document_number="1",
                    sale_at=datetime(2026, 1, 1, 10, 0),
                    money=money(100, 50, 50),
                    phones={"79111112233"},
                )
            ],
        )
        case = problem(
            contract_id="A",
            fio=client.fio,
            phone="79214445566",
            create_date=date(2026, 1, 1),
        )
        with self.assertRaisesRegex(ValueError, "phone conflicts"):
            builder.resolve_cases([case], [client], money(12000, 12000, 0))


class MembershipRewriteTests(unittest.TestCase):
    def test_only_financial_fields_change_and_problem4_is_excluded(self) -> None:
        headers = [
            "contract_id",
            "client_id",
            "client_fio",
            "price",
            "amount_of_payments",
            "payment_left",
            "manager",
        ]
        russian_headers = [
            "Договор",
            "Клиент",
            "ФИО",
            "Цена",
            "Оплачено",
            "Долг",
            "Менеджер",
        ]
        source_rows = [
            ["A", "C1", "Первый Клиент", 100, 50, 50, "Менеджер 1"],
            ["B", "C2", "Второй Клиент", 200, 200, 0, "Менеджер 2"],
            ["P4", "C3", "Problem4 Клиент", 300, 300, 0, "Менеджер 3"],
        ]
        current_case = builder.ProblemCase(
            group=3,
            source_file="problem_3.xlsx",
            values=tuple(source_rows[0]),
            contract_id="A",
            client_id="C1",
            client_fio="Первый Клиент",
            phone="",
            create_date=date(2026, 1, 1),
            payment_date=date(2026, 1, 1),
            old_money=money(100, 50, 50),
        )
        resolution = builder.Resolution(
            case=current_case,
            new_money=money(120, 120, 0),
            source="manager_debt_report",
            match_method="test",
            phone_check="test",
            manager_client_row=1,
            manager_sale_row=2,
            manager_document_number="1",
        )

        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            source_path = root / "source.xlsx"
            template_path = root / "template.xlsx"
            output_path = root / "output.xlsx"

            source_workbook = Workbook()
            source_sheet = source_workbook.active
            source_sheet.append(headers)
            source_sheet.append(russian_headers)
            for row in source_rows:
                source_sheet.append(row)
            source_workbook.save(source_path)

            template_workbook = Workbook()
            template_sheet = template_workbook.active
            template_sheet.append(headers)
            template_sheet.append(russian_headers)
            template_sheet.append([""] * len(headers))
            template_workbook.save(template_path)

            stats = builder.build_membership_workbook(
                source=source_path,
                destination=output_path,
                template=template_path,
                problem_headers=headers,
                resolutions=[resolution],
                excluded_contracts={"P4"},
            )
            output_workbook = load_workbook(
                output_path, read_only=True, data_only=True
            )
            values = list(
                output_workbook.active.iter_rows(min_row=3, values_only=True)
            )
            output_workbook.close()

        self.assertEqual(
            values,
            [
                ("A", "C1", "Первый Клиент", 120, 120, 0, "Менеджер 1"),
                ("B", "C2", "Второй Клиент", 200, 200, 0, "Менеджер 2"),
            ],
        )
        self.assertEqual(stats["resolved_rows"], 1)
        self.assertEqual(stats["excluded_rows"], 1)
        self.assertEqual(stats["changed_rows"], 1)
        self.assertEqual(stats["changed_cells"], 3)


if __name__ == "__main__":
    unittest.main()
