#!/usr/bin/env python3
"""Validate Part 2 three-funnel Fitbase outputs."""

from __future__ import annotations

import argparse
import csv
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

import yaml
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAIN_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
]
MAIN_RUS_HEADERS = [
    "Внутренний номер клиента ",
    "Телефон *",
    "ФИО клиента *",
    "Почта",
    "Воронка *",
    "Этап воронки *",
    "Бюджет ",
    "Дата создания *",
    "Менеджер ",
]
CARD_HEADERS = ["телефон", "фио", "номер пластиковой карты"]
FUNNEL_SLUGS = {
    "Действующие клиенты": "deystvuyushchie_klienty",
    "Новые заявки": "novye_zayavki",
    "Реактивация": "reaktivatsiya",
}
ALLOWED_STEPS = {
    "Действующие клиенты": {
        "60-31 день до окончания",
        "30-8 дней до окончания",
        "7-0 день до окончания",
        "Действующие клиенты",
    },
    "Новые заявки": {"Неразобранные"},
    "Реактивация": {"1-6 дней", "7-29 дней", "30-59 дней", "60-89 дней", "более 90 дней"},
}


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in ("", None):
            return default
        return int(float(str(value)))
    except ValueError:
        return default


def workbook_headers(path: Path, row: int, width: int) -> list[object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(row, col).value for col in range(1, width + 1)]
    wb.close()
    return headers


def workbook_row_count(path: Path, first_data_row: int, width: int) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=first_data_row, max_col=width, values_only=True):
        if any(value not in (None, "") for value in row):
            count += 1
    wb.close()
    return count


def card_xlsx_comma_count(path: Path) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=2, max_col=len(CARD_HEADERS), values_only=True):
        card = row[2] if len(row) >= 3 else None
        if isinstance(card, str) and "," in card:
            count += 1
    wb.close()
    return count


def load_managers(path: Path) -> dict[str, set[str]]:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    return {str(club): {str(manager) for manager in managers} for club, managers in data.get("clubs", {}).items()}


def validate(args: argparse.Namespace) -> int:
    cutoff_date = args.cutoff_date
    date_stamp = cutoff_date.replace("-", "")
    stage_dir = as_abs(args.stage_dir)
    output_dir = as_abs(args.output_dir)
    reports_dir = as_abs(args.reports_dir) if args.reports_dir else output_dir / "reports"
    managers_by_club = load_managers(as_abs(args.managers_config))

    errors: list[str] = []
    warnings: list[str] = []
    required_reports = [
        "funnel_distribution.csv",
        "stage_distribution_by_funnel.csv",
        "manager_distribution_by_club.csv",
        "missing_phone_report.csv",
        "missing_card_report.csv",
        "missing_club_report.csv",
        "multiple_subscriptions_report.csv",
        "subscription_selection_report.csv",
        "subscription_overrides_report.csv",
        "multiple_cards_report.csv",
        "card_selection_report.csv",
        "product_classification_preflight.csv",
        "product_classification_report.csv",
        "product_classification_review_report.csv",
        "club_reference_candidates.csv",
        "club_link_candidates.csv",
        "active_diff_vs_previous_export.csv",
    ]

    final_path = stage_dir / "final_funnel_clients.csv"
    if not final_path.exists():
        errors.append(f"missing final stage CSV: {final_path.relative_to(ROOT)}")
        rows: list[dict[str, str]] = []
    else:
        rows = read_csv(final_path)
    if not rows:
        errors.append("final_funnel_clients.csv is empty")

    for report in required_reports:
        if not (reports_dir / report).exists():
            errors.append(f"missing required report: {(reports_dir / report).relative_to(ROOT)}")

    by_funnel: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in rows:
        by_funnel[row.get("funnel", "")].append(row)

    for funnel, slug in FUNNEL_SLUGS.items():
        main_xlsx = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}__{slug}.xlsx"
        cards_xlsx = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}__{slug}.xlsx"
        if not main_xlsx.exists():
            errors.append(f"missing main XLSX: {main_xlsx.relative_to(ROOT)}")
            continue
        if not cards_xlsx.exists():
            errors.append(f"missing cards XLSX: {cards_xlsx.relative_to(ROOT)}")
            continue
        expected_rows = len(by_funnel.get(funnel, []))
        if workbook_headers(main_xlsx, 1, len(MAIN_HEADERS)) != MAIN_HEADERS:
            errors.append(f"main technical headers mismatch: {main_xlsx.name}")
        if workbook_headers(main_xlsx, 2, len(MAIN_RUS_HEADERS)) != MAIN_RUS_HEADERS:
            errors.append(f"main Russian headers mismatch: {main_xlsx.name}")
        if workbook_headers(cards_xlsx, 1, len(CARD_HEADERS)) != CARD_HEADERS:
            errors.append(f"card headers mismatch: {cards_xlsx.name}")
        main_rows = workbook_row_count(main_xlsx, 3, len(MAIN_HEADERS))
        card_rows = workbook_row_count(cards_xlsx, 2, len(CARD_HEADERS))
        if main_rows != expected_rows:
            errors.append(f"main row count mismatch for {funnel}: xlsx={main_rows}, csv={expected_rows}")
        if card_rows != expected_rows:
            errors.append(f"card row count mismatch for {funnel}: xlsx={card_rows}, csv={expected_rows}")
        comma_cards = card_xlsx_comma_count(cards_xlsx)
        if comma_cards:
            errors.append(f"card XLSX contains comma-separated card numbers for {funnel}: {comma_cards}")

    client_refs = [row.get("client_ref", "") for row in rows]
    duplicate_refs = [key for key, count in Counter(client_refs).items() if key and count > 1]
    if duplicate_refs:
        errors.append(f"duplicate client_ref rows: {len(duplicate_refs)}")

    client_ids = [row.get("client_id", "") for row in rows]
    duplicate_ids = [key for key, count in Counter(client_ids).items() if key and count > 1]
    if duplicate_ids:
        errors.append(f"duplicate client_id rows: {len(duplicate_ids)}")

    for row in rows:
        funnel = row.get("funnel", "")
        step = row.get("funnel_step", "")
        if funnel not in FUNNEL_SLUGS:
            errors.append(f"unknown funnel: {funnel}")
            break
        if step not in ALLOWED_STEPS[funnel]:
            errors.append(f"invalid step for funnel {funnel}: {step}")
            break
        if step == "Бронь":
            errors.append("booking stage remains in final output")
            break
        if row.get("manager") in {"A1", "A2", "A3"}:
            errors.append(f"temporary manager remains: client_id={row.get('client_id')}")
            break
        manager = row.get("manager", "")
        club = row.get("normalized_club", "")
        if manager and manager not in managers_by_club.get(club, set()):
            errors.append(f"manager is not configured for club: client_id={row.get('client_id')}, club={club}, manager={manager}")
            break
        if row.get("funnel") == "Действующие клиенты":
            days = int_value(row.get("days_to_end"), -999999)
            if step == "60-31 день до окончания" and not (31 <= days <= 60):
                errors.append(f"active 60-31 boundary violation: client_id={row.get('client_id')}, days={days}")
                break
            if step == "30-8 дней до окончания" and not (8 <= days <= 30):
                errors.append(f"active 30-8 boundary violation: client_id={row.get('client_id')}, days={days}")
                break
            if step == "7-0 день до окончания" and not (0 <= days <= 7):
                errors.append(f"active 7-0 boundary violation: client_id={row.get('client_id')}, days={days}")
                break
        if row.get("funnel") == "Реактивация":
            days = int_value(row.get("days_since_end"), -999999)
            if days <= 0:
                errors.append(f"reactivation non-positive days_since_end: client_id={row.get('client_id')}, days={days}")
                break

    missing_phone = read_csv(reports_dir / "missing_phone_report.csv") if (reports_dir / "missing_phone_report.csv").exists() else []
    missing_card = read_csv(reports_dir / "missing_card_report.csv") if (reports_dir / "missing_card_report.csv").exists() else []
    missing_club = read_csv(reports_dir / "missing_club_report.csv") if (reports_dir / "missing_club_report.csv").exists() else []
    multiple_subs = read_csv(reports_dir / "multiple_subscriptions_report.csv") if (reports_dir / "multiple_subscriptions_report.csv").exists() else []
    card_selection = read_csv(reports_dir / "card_selection_report.csv") if (reports_dir / "card_selection_report.csv").exists() else []

    if len(missing_phone) != sum(1 for row in rows if not (row.get("phones") or "").strip()):
        errors.append("missing_phone_report.csv count mismatch")
    if len(missing_card) != sum(1 for row in rows if not (row.get("selected_card_number") or "").strip()):
        errors.append("missing_card_report.csv count mismatch")
    if len(missing_club) != sum(1 for row in rows if not (row.get("normalized_club") or "").strip()):
        errors.append("missing_club_report.csv count mismatch")
    expected_multiple_subs = sum(
        1
        for row in rows
        if (
            row.get("funnel") == "Действующие клиенты"
            and int_value(row.get("active_full_subscription_count")) > 1
        )
        or (row.get("funnel") == "Реактивация" and int_value(row.get("finished_full_subscription_count")) > 1)
    )
    if len(multiple_subs) != expected_multiple_subs:
        errors.append("multiple_subscriptions_report.csv count mismatch")
    if len(card_selection) != len(rows):
        errors.append("card_selection_report.csv should contain one row per final client")

    review_rows = read_csv(reports_dir / "product_classification_review_report.csv") if (reports_dir / "product_classification_review_report.csv").exists() else []
    if review_rows:
        warnings.append(f"product classification rows needing business review: {len(review_rows)}")
    if missing_phone:
        warnings.append(f"clients without phone exported and reported: {len(missing_phone)}")
    if missing_card:
        warnings.append(f"clients without selected card exported and reported: {len(missing_card)}")
    if missing_club:
        warnings.append(f"clients without discovered club/manager exported and reported: {len(missing_club)}")
    if multiple_subs:
        warnings.append(f"clients with multiple selected-funnel subscription candidates reported: {len(multiple_subs)}")

    verdict = "PASS" if not errors else "FAIL"
    report_lines = [
        "# Part 2 Validation Report",
        "",
        f"Run date: `{datetime.now().isoformat(timespec='seconds')}`",
        f"cutoff_date: `{cutoff_date}`",
        f"final_rows: `{len(rows)}`",
        "",
        "## Verdict",
        "",
        f"`{verdict}`",
        "",
        "## Funnel Distribution",
        "",
    ]
    for funnel, count in Counter(row.get("funnel", "") for row in rows).most_common():
        report_lines.append(f"- `{funnel}`: `{count}`")
    report_lines.extend(["", "## Stage Distribution", ""])
    for (funnel, step), count in Counter((row.get("funnel", ""), row.get("funnel_step", "")) for row in rows).most_common():
        report_lines.append(f"- `{funnel}` / `{step}`: `{count}`")
    report_lines.extend(["", "## Data Quality Counts", ""])
    report_lines.extend(
        [
            f"- missing_phone: `{len(missing_phone)}`",
            f"- missing_card: `{len(missing_card)}`",
            f"- missing_club: `{len(missing_club)}`",
            f"- multiple_subscription_clients: `{len(multiple_subs)}`",
            f"- product_review_rows: `{len(review_rows)}`",
        ]
    )
    if errors:
        report_lines.extend(["", "## Errors", ""])
        report_lines.extend(f"- {error}" for error in errors)
    else:
        report_lines.extend(["", "## Errors", "", "None."])
    if warnings:
        report_lines.extend(["", "## Warnings", ""])
        report_lines.extend(f"- {warning}" for warning in warnings)
    report_lines.append("")

    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "validation_report.md").write_text("\n".join(report_lines), encoding="utf-8")
    print(f"verdict={verdict}")
    print(f"errors={len(errors)}")
    print(f"warnings={len(warnings)}")
    print(f"report={(reports_dir / 'validation_report.md').relative_to(ROOT)}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default="2026-04-29")
    parser.add_argument("--stage-dir", default=str(ROOT / "output" / "part2_20260429" / "staging"))
    parser.add_argument("--output-dir", default=str(ROOT / "output" / "part2_20260429"))
    parser.add_argument("--reports-dir", default="")
    parser.add_argument("--managers-config", default=str(ROOT / "config" / "managers_by_club.yml"))
    return parser.parse_args()


def main() -> None:
    raise SystemExit(validate(parse_args()))


if __name__ == "__main__":
    main()
