#!/usr/bin/env python3
"""Build Fitbase services import workbooks from service staging facts."""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260630"

CLIENT_HEADERS = [
    "service_id",
    "client_id",
    "phone",
    "client_fio",
    "service_name",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "count",
    "visits_left",
    "price",
    "amount_of_payment",
    "payment_left",
    "type_of_payment",
    "manager",
    "филиал",
]
CLIENT_RUS_HEADERS = [
    "Внутренний номер услуги",
    "Внутренний номер клиента",
    "Телефон клиента",
    "ФИО клиента *",
    "Название услуги *",
    "Дата добавления *",
    "Дата оплаты *",
    "Дата активации ",
    "Дата окончания",
    "Количество услуг",
    "Осталось посещений ",
    "Цена *",
    "Оплачено *",
    "Осталось оплатить *",
    "Тип оплаты",
    "Менеджер ",
    "Филиал продажи",
]
TEMPLATE_HEADERS = [
    "name",
    "price",
    "duration",
    "visits",
    "do_enter",
    "first_visit_activation",
    "archive",
    "category",
    "legal_entity",
]
TEMPLATE_RUS_HEADERS = [
    "Название услуги *",
    "Цена *",
    "Продолжительность ( в днях ) *",
    "Количество тренировок *",
    "Осуществлять вход в клуб?",
    "Активация с первого посещения?",
    "В архиве?",
    "Категория в структуре",
    "Юр.лицо",
]
FACT_FIELDS = [
    "service_order",
    "service_name",
    "product_ref",
    "product_code",
    "product_name",
    "sale_doc_ref",
    "sale_number",
    "sale_line_no",
    "sale_line_id",
    "sale_datetime",
    "sale_date",
    "sale_client_ref",
    "sale_client_id",
    "sale_client_fio",
    "sale_client_phone",
    "sale_branch_raw",
    "sale_branch",
    "sale_branch_source",
    "linked_service_doc_ref",
    "linked_object_rtref",
    "service_doc_number",
    "service_doc_datetime",
    "service_doc_holder_ref",
    "service_doc_holder_id",
    "service_doc_holder_fio",
    "service_start_date",
    "service_end_date",
    "service_doc_duration_value",
    "service_doc_posted",
    "service_doc_marked",
    "line_quantity",
    "line_total_amount",
    "unit_price",
    "vat_amount",
    "line_comment",
    "rg_duration_days",
    "rg_price",
    "rg_paid_candidate",
    "rg_payment_count_candidate",
    "rg_visits_candidate_8007",
    "rg_visits_candidate_8008",
    "rg_visits_candidate_8009",
    "rg3336_receipt_qty",
    "rg3336_expense_qty",
    "rg3336_signed_balance",
    "rg3336_movement_rows",
    "rg3336_receipt_rows",
    "rg3336_expense_rows",
    "has_linked_service_doc",
    "is_active_by_balance",
    "is_active_by_date",
    "is_active_on_cutoff",
    "payment_ref",
    "payment_datetime",
    "payment_amount",
    "payment_method",
    "payment_operation",
    "payment_match_source",
    "cutoff_at",
    "raw_source",
]


@dataclass(frozen=True)
class SourceClient:
    client_id: str
    phone: str
    client_fio: str
    create_date: date | None
    manager: str
    branch: str


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def parse_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S", "%d.%m.%Y"):
        try:
            return datetime.strptime(text[:19] if "%H" in fmt else text[:10], fmt).date()
        except ValueError:
            continue
    return None


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def excel_number(value: Decimal | int | float | None) -> int | float | None:
    if value is None:
        return None
    if not isinstance(value, Decimal):
        value = Decimal(str(value))
    if value == value.to_integral_value():
        return int(value)
    return float(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").strip().lower())


def read_source_clients(path: Path) -> dict[str, SourceClient]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    indexes = {str(name): idx for idx, name in enumerate(headers) if name}
    required = ["client_id", "phone", "client_fio", "create_date", "manager", "филиал"]
    missing = [name for name in required if name not in indexes]
    if missing:
        raise ValueError(f"Missing required columns in {path}: {missing}")

    clients: dict[str, SourceClient] = {}
    for values in ws.iter_rows(min_row=3, values_only=True):
        client_id = str(values[indexes["client_id"]] or "").strip()
        if not client_id:
            continue
        clients[client_id] = SourceClient(
            client_id=client_id,
            phone=str(values[indexes["phone"]] or "").strip(),
            client_fio=str(values[indexes["client_fio"]] or "").strip(),
            create_date=parse_date(values[indexes["create_date"]]),
            manager=str(values[indexes["manager"]] or "").strip(),
            branch=str(values[indexes["филиал"]] or "").strip(),
        )
    wb.close()
    return clients


def read_service_names(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    names = [str(row[0] or "").strip() for row in ws.iter_rows(values_only=True) if str(row[0] or "").strip()]
    wb.close()
    if len(names) != 51:
        raise ValueError(f"Expected 51 service names in {path}, got {len(names)}")
    if len(set(names)) != len(names):
        raise ValueError(f"Duplicate service names in {path}")
    return names


def read_facts(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-16", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for raw in reader:
            if not raw:
                continue
            if len(raw) != len(FACT_FIELDS):
                raise ValueError(f"Unexpected TSV column count {len(raw)} in {path}; expected {len(FACT_FIELDS)}")
            rows.append(
                {
                    key: value.replace("\t", " ").replace("\r", " ").replace("\n", " ")
                    for key, value in zip(FACT_FIELDS, raw, strict=True)
                }
            )
    return rows


def map_payment_type(method: str) -> str:
    text = normalize_key(method)
    if not text:
        return ""
    if "для ошибок" in text or "бар ип иконников андрей анатольевич" in text:
        return "безналичные"
    if "сбп" in text or "сбпр" in text:
        return "сбп"
    if "налич" in text and "безнал" not in text:
        return "наличные"
    if (
        "эквайр" in text
        or "банк" in text
        or "безнал" in text
        or "терминал" in text
        or "карта" in text
        or "р/с" in text
    ):
        return "безналичные"
    return ""


def fact_client_id(fact: dict[str, str]) -> str:
    return (fact.get("service_doc_holder_id") or fact.get("sale_client_id") or "").strip()


def fact_client_fio(fact: dict[str, str]) -> str:
    return (fact.get("service_doc_holder_fio") or fact.get("sale_client_fio") or "").strip()


def source_client_for_fact(fact: dict[str, str], source_clients: dict[str, SourceClient]) -> SourceClient | None:
    client_id = fact_client_id(fact)
    source = source_clients.get(client_id)
    if source:
        return source
    if not client_id:
        return None
    return SourceClient(
        client_id=client_id,
        phone=(fact.get("sale_client_phone") or "").strip(),
        client_fio=fact_client_fio(fact),
        create_date=None,
        manager="УТОЧНИТЬ: вне import_заявки",
        branch="",
    )


def service_id(fact: dict[str, str]) -> str:
    return (fact.get("service_doc_number") or "").strip() or (fact.get("sale_line_id") or "").strip()


def is_admin_money_service(name: str) -> bool:
    text = normalize_key(name)
    return any(
        marker in text
        for marker in [
            "доплата",
            "заморозка",
            "перевод из клуба",
            "переоформление",
            "восстановление пластиковой",
            "подарочный сертификат",
            "утеря",
            "аренда рекламного",
        ]
    )


def parse_visits_from_name(name: str) -> int | None:
    text = normalize_key(name)
    match = re.search(r"(\d+)\s*(?:пос|посещ)", text)
    if match:
        return int(match.group(1))
    match = re.search(r"пакет\s+(\d+)", text)
    if match:
        return int(match.group(1))
    if "солярий 5" in text:
        return 5
    if "солярий 1" in text:
        return 1
    if any(marker in text for marker in ["разов", "пробная", "персональная тренировка", "йога в гамаках", "стрип", "сайкл"]):
        return 1
    if "субаренда 1" in text:
        return 1
    return None


def template_duration(name: str, facts: list[dict[str, str]]) -> int | None:
    if is_admin_money_service(name):
        return None
    for fact in sorted(facts, key=lambda item: item.get("sale_datetime", ""), reverse=True):
        duration = decimal_value(fact.get("rg_duration_days")) or decimal_value(fact.get("service_doc_duration_value"))
        if duration > 0:
            return int(duration.quantize(Decimal("1"), rounding=ROUND_HALF_UP))
    if parse_visits_from_name(name):
        return 30 if "пакет" in normalize_key(name) else 1
    return None


def do_enter(name: str) -> int | None:
    if is_admin_money_service(name):
        return 0
    return 1


def first_visit_activation(name: str) -> int | None:
    if is_admin_money_service(name):
        return 0
    return 1


def active_visits_left(fact: dict[str, str]) -> int | float | None:
    balance = decimal_value(fact.get("rg3336_signed_balance"))
    if balance > 0:
        return excel_number(balance)
    if fact.get("is_active_by_date") == "1":
        parsed = parse_visits_from_name(fact.get("service_name", ""))
        return parsed
    return None


def build_rows(
    source_clients: dict[str, SourceClient],
    service_names: list[str],
    facts: list[dict[str, str]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, str]], list[dict[str, Any]], dict[str, Counter]]:
    facts_by_service: dict[str, list[dict[str, str]]] = defaultdict(list)
    facts_for_final_by_service: dict[str, list[dict[str, str]]] = defaultdict(list)
    active_for_final_by_service: dict[str, list[dict[str, str]]] = defaultdict(list)
    counters: dict[str, Counter] = defaultdict(Counter)
    uncertainties: list[dict[str, str]] = []

    for fact in facts:
        name = fact.get("service_name", "")
        facts_by_service[name].append(fact)
        client_id = fact_client_id(fact)
        if client_id in source_clients:
            facts_for_final_by_service[name].append(fact)
            if fact.get("is_active_on_cutoff") == "1":
                active_for_final_by_service[name].append(fact)
        else:
            counters["facts"]["not_in_final_client_xlsx"] += 1

    selected_facts: list[tuple[dict[str, str], str]] = []
    coverage_rows: list[dict[str, Any]] = []
    for order, name in enumerate(service_names, start=1):
        service_facts = sorted(facts_by_service.get(name, []), key=lambda item: item.get("sale_datetime", ""), reverse=True)
        final_facts = sorted(facts_for_final_by_service.get(name, []), key=lambda item: item.get("sale_datetime", ""), reverse=True)
        active_facts = sorted(active_for_final_by_service.get(name, []), key=lambda item: item.get("sale_datetime", ""), reverse=True)
        row_kind = "active"
        selected_for_service = active_facts
        if not selected_for_service:
            row_kind = "historical_fallback"
            selected_for_service = final_facts[:5]
        if not selected_for_service and service_facts:
            row_kind = "historical_fallback_outside_import_zayavki"
            selected_for_service = service_facts[:5]
            uncertainties.append(
                {
                    "issue_type": "service_fallback_taken_outside_import_zayavki",
                    "service_id": "",
                    "client_id": "",
                    "client_fio": "",
                    "service_name": name,
                    "details": (
                        "No active or historical rows were found for final import_zayavki clients; "
                        f"selected historical rows outside import_zayavki: {len(selected_for_service)}."
                    ),
                }
            )
        elif not active_facts:
            uncertainties.append(
                {
                    "issue_type": "service_has_no_active_rows_for_final_clients",
                    "service_id": "",
                    "client_id": "",
                    "client_fio": "",
                    "service_name": name,
                    "details": f"Active rows for final clients: 0; historical fallback rows selected: {len(selected_for_service)}.",
                }
            )
        if not selected_for_service:
            uncertainties.append(
                {
                    "issue_type": "service_has_no_rows_for_final_clients",
                    "service_id": "",
                    "client_id": "",
                    "client_fio": "",
                    "service_name": name,
                    "details": "No active or historical sale rows were found for clients from final import_zayavki.",
                }
            )
        for fact in selected_for_service:
            selected_facts.append((fact, row_kind))
        coverage_rows.append(
            {
                "service_order": order,
                "service_name": name,
                "total_sale_rows": len(service_facts),
                "final_client_sale_rows": len(final_facts),
                "active_final_client_rows": len(active_facts),
                "selected_rows": len(selected_for_service),
                "selected_kind": row_kind if selected_for_service else "template_only_no_final_client_rows",
                "selected_outside_import_zayavki": 1 if row_kind == "historical_fallback_outside_import_zayavki" else 0,
                "latest_sale_datetime": service_facts[0].get("sale_datetime", "") if service_facts else "",
            }
        )

    client_rows: list[dict[str, Any]] = []
    seen_ids: Counter[str] = Counter()
    for fact, row_kind in selected_facts:
        client_id = fact_client_id(fact)
        source = source_client_for_fact(fact, source_clients)
        if not source:
            continue

        sid = service_id(fact)
        seen_ids[sid] += 1
        price = decimal_value(fact.get("line_total_amount")) or decimal_value(fact.get("rg_price"))
        payment_type = map_payment_type(fact.get("payment_method", ""))
        if not payment_type and price > 0:
            payment_type = "наличные"
            counters["payment_type"]["default_cash_no_payment_method"] += 1
        elif payment_type:
            counters["payment_type"][payment_type] += 1
        else:
            counters["payment_type"]["blank"] += 1

        sale_date = parse_date(fact.get("sale_date"))
        payment_date = parse_date(fact.get("payment_datetime")) or sale_date
        activation_date = parse_date(fact.get("service_start_date"))
        end_date = parse_date(fact.get("service_end_date"))
        visits_left = active_visits_left(fact)
        if row_kind == "historical_fallback":
            activation_date = activation_date or sale_date
            end_date = end_date or sale_date
            visits_left = 0

        row = {
            "service_id": sid,
            "client_id": source.client_id,
            "phone": source.phone,
            "client_fio": source.client_fio,
            "service_name": fact.get("service_name", ""),
            "create_date": sale_date,
            "payment_date": payment_date,
            "activation_date": activation_date,
            "end_date": end_date,
            "count": excel_number(decimal_value(fact.get("line_quantity"))) or 1,
            "visits_left": visits_left,
            "price": excel_number(price),
            "amount_of_payment": excel_number(price),
            "payment_left": 0,
            "type_of_payment": payment_type,
            "manager": source.manager,
            "филиал": (fact.get("sale_branch") or "").strip() or source.branch,
            "_row_kind": row_kind,
            "_outside_import_zayavki": "0" if client_id in source_clients else "1",
            "_sale_doc_ref": fact.get("sale_doc_ref", ""),
            "_sale_branch_raw": fact.get("sale_branch_raw", ""),
            "_sale_branch_source": fact.get("sale_branch_source", ""),
            "_linked_service_doc_ref": fact.get("linked_service_doc_ref", ""),
            "_payment_method_raw": fact.get("payment_method", ""),
            "_is_active_by_balance": fact.get("is_active_by_balance", ""),
            "_is_active_by_date": fact.get("is_active_by_date", ""),
            "_rg3336_signed_balance": fact.get("rg3336_signed_balance", ""),
            "_sale_datetime": fact.get("sale_datetime", ""),
        }
        client_rows.append(row)
        counters["rows_by_kind"][row_kind] += 1
        if row.get("_outside_import_zayavki") == "1":
            counters["special"]["outside_import_zayavki_rows"] += 1
        counters["rows_by_service"][row["service_name"]] += 1

    for duplicate_id, count in seen_ids.items():
        if duplicate_id and count > 1:
            # Some service document numbers can theoretically repeat across branches; make IDs stable and unique.
            matching = [row for row in client_rows if row["service_id"] == duplicate_id]
            for idx, row in enumerate(matching, start=1):
                row["service_id"] = f"{duplicate_id}-{idx}"
            uncertainties.append(
                {
                    "issue_type": "duplicate_service_id_disambiguated",
                    "service_id": duplicate_id,
                    "client_id": "",
                    "client_fio": "",
                    "service_name": "",
                    "details": f"Service id appeared {count} times; suffixed with row number.",
                }
            )

    template_rows: list[dict[str, Any]] = []
    for name in service_names:
        service_facts = sorted(facts_by_service.get(name, []), key=lambda item: item.get("sale_datetime", ""), reverse=True)
        latest = service_facts[0] if service_facts else {}
        price = decimal_value(latest.get("unit_price")) if latest else Decimal("0")
        if not service_facts:
            uncertainties.append(
                {
                    "issue_type": "template_service_has_no_sales",
                    "service_id": "",
                    "client_id": "",
                    "client_fio": "",
                    "service_name": name,
                    "details": "Service exists in required list but no sale rows were found in Document154_VT1137 before cutoff.",
                }
            )
        template_rows.append(
            {
                "name": name,
                "price": excel_number(price),
                "duration": template_duration(name, service_facts),
                "visits": None if is_admin_money_service(name) else parse_visits_from_name(name),
                "do_enter": do_enter(name),
                "first_visit_activation": first_visit_activation(name),
                "archive": None,
                "category": None,
                "legal_entity": None,
            }
        )

    client_rows.sort(key=lambda row: (row["_row_kind"] != "active", row["service_name"], row.get("client_id", ""), row.get("_sale_datetime", "")))
    return client_rows, template_rows, uncertainties, coverage_rows, counters


def write_workbook(path: Path, headers: list[str], rus_headers: list[str], rows: list[dict[str, Any]]) -> None:
    wb = load_workbook(path) if path.exists() else None
    if wb is None:
        raise ValueError(f"Template workbook does not exist: {path}")
    ws = wb.active
    ws.delete_rows(1, ws.max_row)
    ws.append(headers)
    ws.append(rus_headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    for col_idx in range(1, len(headers) + 1):
        source_cell = ws.cell(1, col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row_idx, col_idx)
            if row_idx <= 2:
                continue
            cell.number_format = "yyyy-mm-dd" if isinstance(cell.value, date) else "General"
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(len(str(headers[col_idx - 1])) + 2, 12), 28)
        if source_cell.has_style:
            for row_idx in (1, 2):
                cell = ws.cell(row_idx, col_idx)
                cell.font = copy(source_cell.font)
                cell.fill = copy(source_cell.fill)
                cell.border = copy(source_cell.border)
                cell.alignment = copy(source_cell.alignment)
    ws.freeze_panes = "A3"
    ws.auto_filter.ref = ws.dimensions
    wb.save(path)
    wb.close()


def write_csv(path: Path, rows: list[dict[str, Any]], fields: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({field: row.get(field, "") for field in fields})


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-output-dir", default="work/20260630/owner")
    parser.add_argument("--output-dir", default="work/20260630/imports")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    parser.add_argument("--services-list", default="templates/services_required.xlsx")
    parser.add_argument("--client-template", default="templates/service_clients.xlsx")
    parser.add_argument("--template-template", default="templates/service_templates.xlsx")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_output_dir = as_abs(args.source_output_dir)
    output_dir = as_abs(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    reports_dir = output_dir / "reports"
    staging_dir = output_dir / "staging"
    reports_dir.mkdir(parents=True, exist_ok=True)

    source_clients_xlsx = source_output_dir / f"fitbase_active_clients_import_zayavki_{args.date_stamp}_all_funnels.xlsx"
    facts_tsv = staging_dir / "services_import_facts.tsv"
    source_clients = read_source_clients(source_clients_xlsx)
    service_names = read_service_names(as_abs(args.services_list))
    facts = read_facts(facts_tsv)

    client_rows, template_rows, uncertainties, coverage_rows, counters = build_rows(source_clients, service_names, facts)

    client_xlsx = output_dir / f"fitbase_import_uslugi_clientov_{args.date_stamp}.xlsx"
    template_xlsx = output_dir / f"fitbase_import_shablony_uslug_{args.date_stamp}.xlsx"
    client_template = as_abs(args.client_template)
    template_template = as_abs(args.template_template)
    client_xlsx.write_bytes(client_template.read_bytes())
    template_xlsx.write_bytes(template_template.read_bytes())
    write_workbook(client_xlsx, CLIENT_HEADERS, CLIENT_RUS_HEADERS, client_rows)
    write_workbook(template_xlsx, TEMPLATE_HEADERS, TEMPLATE_RUS_HEADERS, template_rows)

    write_csv(
        reports_dir / "services_coverage_report.csv",
        coverage_rows,
        [
            "service_order",
            "service_name",
            "total_sale_rows",
            "final_client_sale_rows",
            "active_final_client_rows",
            "selected_rows",
            "selected_kind",
            "selected_outside_import_zayavki",
            "latest_sale_datetime",
        ],
    )
    write_csv(
        reports_dir / "services_import_uncertainties.csv",
        uncertainties,
        ["issue_type", "service_id", "client_id", "client_fio", "service_name", "details"],
    )
    active_rows = [row for row in client_rows if row.get("_row_kind") == "active"]
    write_csv(
        reports_dir / "services_active_rows_audit.csv",
        [
            {
                "service_id": row["service_id"],
                "client_id": row["client_id"],
                "client_fio": row["client_fio"],
                "service_name": row["service_name"],
                "visits_left": row["visits_left"],
                "price": row["price"],
                "type_of_payment": row["type_of_payment"],
                "филиал": row["филиал"],
                "sale_doc_ref": row["_sale_doc_ref"],
                "sale_branch_raw": row["_sale_branch_raw"],
                "sale_branch_source": row["_sale_branch_source"],
                "linked_service_doc_ref": row["_linked_service_doc_ref"],
                "is_active_by_balance": row["_is_active_by_balance"],
                "is_active_by_date": row["_is_active_by_date"],
                "rg3336_signed_balance": row["_rg3336_signed_balance"],
            }
            for row in active_rows
        ],
        [
            "service_id",
            "client_id",
            "client_fio",
            "service_name",
            "visits_left",
            "price",
            "type_of_payment",
            "филиал",
            "sale_doc_ref",
            "sale_branch_raw",
            "sale_branch_source",
            "linked_service_doc_ref",
            "is_active_by_balance",
            "is_active_by_date",
            "rg3336_signed_balance",
        ],
    )
    branch_counts = Counter(str(row.get("филиал") or "blank") for row in client_rows)
    write_csv(
        reports_dir / "services_branch_distribution.csv",
        [{"branch": branch, "rows_count": rows_count} for branch, rows_count in branch_counts.most_common()],
        ["branch", "rows_count"],
    )

    report = [
        "# Services import build report",
        "",
        f"- source final clients: {len(source_clients)}",
        f"- raw service facts: {len(facts)}",
        f"- client rows selected: {len(client_rows)}",
        f"- active client rows selected: {counters['rows_by_kind'].get('active', 0)}",
        f"- historical fallback rows selected: {counters['rows_by_kind'].get('historical_fallback', 0)}",
        f"- outside import_zayavki fallback rows selected: {counters['rows_by_kind'].get('historical_fallback_outside_import_zayavki', 0)}",
        f"- template rows: {len(template_rows)}",
        f"- services with selected rows: {sum(1 for row in coverage_rows if row['selected_rows'])}",
        f"- services template-only/no final-client rows: {sum(1 for row in coverage_rows if not row['selected_rows'])}",
        f"- payment type counts: {dict(counters['payment_type'])}",
        f"- branch counts: {dict(branch_counts)}",
        "",
        "## Output",
        "",
        f"- `{client_xlsx.relative_to(ROOT)}`",
        f"- `{template_xlsx.relative_to(ROOT)}`",
        "",
        "## Reports",
        "",
        f"- `{(reports_dir / 'services_coverage_report.csv').relative_to(ROOT)}`",
        f"- `{(reports_dir / 'services_import_uncertainties.csv').relative_to(ROOT)}`",
        f"- `{(reports_dir / 'services_active_rows_audit.csv').relative_to(ROOT)}`",
        f"- `{(reports_dir / 'services_branch_distribution.csv').relative_to(ROOT)}`",
    ]
    (reports_dir / "services_build_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    print("\n".join(report))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
