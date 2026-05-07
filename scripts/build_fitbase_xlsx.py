#!/usr/bin/env python3
"""Build final Fitbase XLSX exports from staging CSV files."""

from __future__ import annotations

import argparse
import copy
import csv
import hashlib
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]

DEFAULT_CUTOFF_DATE = "2026-04-29"
DEFAULT_DATE_STAMP = DEFAULT_CUTOFF_DATE.replace("-", "")
DEFAULT_STAGING_DIR = ROOT / "output" / f"staging_{DEFAULT_CUTOFF_DATE}"
DEFAULT_OUTPUT_DIR = ROOT / "output"
DEFAULT_MAIN_TEMPLATE = ROOT / "task-desc" / "Копия Импорт_заявки.xlsx"
DEFAULT_CARDS_TEMPLATE = ROOT / "task-desc" / "Пластиковая карта.xlsx"
DEFAULT_MANAGERS_CONFIG = ROOT / "config" / "managers.yml"

MAIN_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
]

MAIN_RUS_HEADERS = [
    "Внутренний номер клиента ",
    "Телефон *",
    "ФИО клиента *",
    "Почта",
    "Воронка *",
    "Этап воронки *",
    "Бюджет ",
    "Дата создания *",
    "Менеджер ",
]

CARD_HEADERS = ["телефон", "фио", "номер пластиковой карты"]

ALLOWED_STEPS = {
    "Бронь",
    "60-31 день до окончания",
    "30-8 дней до окончания",
    "7-0 день до окончания",
    "Действующие клиенты",
}

FINAL_FIELDS = [
    "client_ref",
    "client_id",
    "client_fio",
    "phones",
    "email",
    "first_sale_date",
    "client_created_at",
    "create_date",
    "create_date_source",
    "active_subscription_ref",
    "active_subscription_name",
    "active_subscription_sale_date",
    "active_subscription_start_date",
    "active_subscription_end_date",
    "active_subscription_duration_days",
    "is_short_duration_active",
    "days_to_end",
    "has_active_booking",
    "plastic_card_number",
    "funnel",
    "funnel_step",
    "budget",
    "manager",
    "dedupe_status",
    "validation_status",
    "active_subscription_count",
    "active_card_count",
    "cutoff_date",
    "normalized_fio",
    "normalized_phones_set",
    "merged_client_refs",
]


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, rows: Iterable[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore", lineterminator="\n")
        writer.writeheader()
        for row in rows:
            writer.writerow({name: row.get(name, "") for name in fieldnames})


def load_managers(path: Path) -> list[str]:
    managers: list[str] = []
    in_managers = False
    for raw_line in path.read_text(encoding="utf-8").splitlines():
        line = raw_line.strip()
        if not line or line.startswith("#"):
            continue
        if line == "managers:":
            in_managers = True
            continue
        if in_managers and line.startswith("-"):
            value = line[1:].strip().strip("'\"")
            if value:
                managers.append(value)
    if not managers:
        raise ValueError(f"No managers found in {path}")
    return managers


def normalized_fio(value: str | None) -> str:
    return " ".join((value or "").strip().lower().split())


def split_raw_values(value: str | None) -> list[str]:
    result: list[str] = []
    for part in (value or "").split(","):
        item = part.strip()
        if item and item not in result:
            result.append(item)
    return result


def normalize_phone_token(value: str) -> str:
    digits = re.sub(r"\D", "", value or "")
    if len(digits) == 11 and digits.startswith("8"):
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return digits


def normalized_phones(value: str | None) -> list[str]:
    phones: list[str] = []
    for raw_phone in split_raw_values(value):
        phone = normalize_phone_token(raw_phone)
        if phone and phone not in phones:
            phones.append(phone)
    return sorted(phones)


def normalized_phones_set(value: str | None) -> str:
    return ",".join(normalized_phones(value))


def client_sort_key(row: dict[str, str]) -> tuple[int, str, str]:
    client_id = row.get("client_id", "")
    digits = re.sub(r"\D", "", client_id)
    numeric = int(digits) if digits else 10**18
    return numeric, client_id, row.get("client_ref", "")


def stable_manager(client_id: str, managers: list[str]) -> str:
    digest = hashlib.sha256((client_id or "").encode("utf-8")).hexdigest()
    return managers[int(digest, 16) % len(managers)]


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(str(value))
    except ValueError:
        return default


def parse_date(value: str | None) -> date | None:
    if not value:
        return None
    return datetime.strptime(value, "%Y-%m-%d").date()


def unique_join(values: Iterable[str | None]) -> str:
    result: list[str] = []
    for value in values:
        for item in split_raw_values(value):
            if item not in result:
                result.append(item)
    return ", ".join(result)


def card_sort_key(row: dict[str, str]) -> tuple[str, str]:
    return row.get("issue_date", ""), row.get("card_ref", "")


def build_active_card_numbers_by_client(cards: list[dict[str, str]]) -> dict[str, str]:
    grouped: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    for row in cards:
        if row.get("is_unmarked") != "1":
            continue
        if not (row.get("plastic_card_number") or "").strip():
            continue
        grouped[row.get("client_ref", "")].append(row)

    result: dict[str, str] = {}
    for client_ref, rows in grouped.items():
        ordered = sorted(rows, key=card_sort_key, reverse=True)
        result[client_ref] = unique_join(row.get("plastic_card_number") for row in ordered)
    return result


def apply_active_card_number_lists(
    rows: list[dict[str, str]], card_numbers_by_client: dict[str, str]
) -> None:
    for row in rows:
        row["plastic_card_number"] = card_numbers_by_client.get(row.get("client_ref", ""), "")


def annotate_rows(rows: list[dict[str, str]], managers: list[str]) -> None:
    for row in rows:
        row["normalized_fio"] = normalized_fio(row.get("client_fio"))
        row["normalized_phones_set"] = normalized_phones_set(row.get("phones"))
        row["funnel"] = "Действующие клиенты"
        row["budget"] = "0"
        row["manager"] = stable_manager(row.get("client_id", ""), managers)
        row.setdefault("merged_client_refs", "")


def merge_exact_duplicate_group(group: list[dict[str, str]]) -> dict[str, str]:
    ordered = sorted(group, key=client_sort_key)
    base = dict(ordered[0])
    base["phones"] = unique_join(row.get("phones") for row in ordered)
    base["email"] = unique_join(row.get("email") for row in ordered)
    base["merged_client_refs"] = ",".join(row.get("client_ref", "") for row in ordered[1:])
    base["dedupe_status"] = "deduped_exact_fio_phone"
    base["active_subscription_count"] = str(max(int_value(row.get("active_subscription_count")) for row in ordered))
    base["active_card_count"] = str(max(int_value(row.get("active_card_count")) for row in ordered))
    base["plastic_card_number"] = unique_join(row.get("plastic_card_number") for row in ordered)
    return base


def build_duplicate_reports(
    rows: list[dict[str, str]],
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    exact_groups: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    by_phone: defaultdict[str, list[dict[str, str]]] = defaultdict(list)
    by_fio: defaultdict[str, list[dict[str, str]]] = defaultdict(list)

    for row in rows:
        fio_key = row["normalized_fio"]
        phones_key = row["normalized_phones_set"]
        if fio_key and phones_key:
            exact_groups[(fio_key, phones_key)].append(row)
        for phone in normalized_phones(row.get("phones")):
            by_phone[phone].append(row)
        if fio_key:
            by_fio[fio_key].append(row)

    report_rows: list[dict[str, str]] = []
    exact_duplicate_rows: list[dict[str, str]] = []

    for (fio_key, phones_key), group in sorted(exact_groups.items()):
        if len(group) <= 1:
            continue
        ordered = sorted(group, key=client_sort_key)
        kept = ordered[0]
        exact_duplicate_rows.extend(ordered)
        for row in ordered:
            action = "kept" if row is kept else "merged_into_kept"
            report_rows.append(
                {
                    "report_type": "auto_merged_exact_fio_phone_set",
                    "group_key": f"{fio_key}|{phones_key}",
                    "normalized_fio": fio_key,
                    "normalized_phone": "",
                    "normalized_phones_set": phones_key,
                    "kept_client_id": kept.get("client_id", ""),
                    "kept_client_ref": kept.get("client_ref", ""),
                    "client_ids": row.get("client_id", ""),
                    "client_refs": row.get("client_ref", ""),
                    "client_fios": row.get("client_fio", ""),
                    "phones": row.get("phones", ""),
                    "action": action,
                    "note": "automatic merge by exact normalized_fio + normalized_phones_set",
                }
            )

    for phone, group in sorted(by_phone.items()):
        fios = sorted({row["normalized_fio"] for row in group if row["normalized_fio"]})
        if len(fios) <= 1:
            continue
        ordered = sorted(group, key=client_sort_key)
        report_rows.append(
            {
                "report_type": "same_phone_different_fio",
                "group_key": phone,
                "normalized_fio": "; ".join(fios),
                "normalized_phone": phone,
                "normalized_phones_set": "",
                "kept_client_id": "",
                "kept_client_ref": "",
                "client_ids": ", ".join(row.get("client_id", "") for row in ordered),
                "client_refs": ", ".join(row.get("client_ref", "") for row in ordered),
                "client_fios": " | ".join(row.get("client_fio", "") for row in ordered),
                "phones": " | ".join(row.get("phones", "") for row in ordered),
                "action": "report_only",
                "note": "same normalized phone, different FIO; not auto-merged",
            }
        )

    for fio_key, group in sorted(by_fio.items()):
        phone_sets = sorted({row["normalized_phones_set"] for row in group if row["normalized_phones_set"]})
        if len(phone_sets) <= 1:
            continue
        ordered = sorted(group, key=client_sort_key)
        report_rows.append(
            {
                "report_type": "same_fio_different_phones",
                "group_key": fio_key,
                "normalized_fio": fio_key,
                "normalized_phone": "",
                "normalized_phones_set": "; ".join(phone_sets),
                "kept_client_id": "",
                "kept_client_ref": "",
                "client_ids": ", ".join(row.get("client_id", "") for row in ordered),
                "client_refs": ", ".join(row.get("client_ref", "") for row in ordered),
                "client_fios": " | ".join(row.get("client_fio", "") for row in ordered),
                "phones": " | ".join(row.get("phones", "") for row in ordered),
                "action": "report_only",
                "note": "same normalized FIO, different phone set; not auto-merged",
            }
        )

    return report_rows, exact_duplicate_rows


def deduplicate_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    groups: defaultdict[tuple[str, str], list[dict[str, str]]] = defaultdict(list)
    no_key_rows: list[dict[str, str]] = []
    for row in rows:
        fio_key = row["normalized_fio"]
        phones_key = row["normalized_phones_set"]
        if fio_key and phones_key:
            groups[(fio_key, phones_key)].append(row)
        else:
            no_key_rows.append(row)

    final_rows: list[dict[str, str]] = []
    for group in groups.values():
        if len(group) == 1:
            row = dict(group[0])
            row["dedupe_status"] = "unique_by_phone_fio"
            final_rows.append(row)
        else:
            final_rows.append(merge_exact_duplicate_group(group))

    for row in no_key_rows:
        copied = dict(row)
        copied["dedupe_status"] = "missing_dedup_key"
        final_rows.append(copied)

    return sorted(final_rows, key=client_sort_key)


def rows_with_missing_required(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    result: list[dict[str, str]] = []
    for row in rows:
        missing = []
        if not (row.get("phones") or "").strip():
            missing.append("phone")
        if not (row.get("client_fio") or "").strip():
            missing.append("client_fio")
        if not (row.get("create_date") or "").strip():
            missing.append("create_date")
        if missing:
            result.append({**row, "missing_fields": ",".join(missing)})
    return result


def build_booking_without_active_report(
    final_rows: list[dict[str, str]], stg_bookings: list[dict[str, str]]
) -> list[dict[str, str]]:
    final_client_refs = {row.get("client_ref", "") for row in final_rows}
    return [
        row
        for row in stg_bookings
        if row.get("is_active_booking") == "1" and row.get("client_ref", "") not in final_client_refs
    ]


def copy_row_styles(ws, row_number: int, width: int) -> list[object]:
    return [copy.copy(ws.cell(row_number, col)._style) for col in range(1, width + 1)]


def clear_data_rows(ws, first_data_row: int) -> None:
    if ws.max_row >= first_data_row:
        ws.delete_rows(first_data_row, ws.max_row - first_data_row + 1)


def trim_to_columns(ws, columns: int) -> None:
    if ws.max_column > columns:
        ws.delete_cols(columns + 1, ws.max_column - columns)


def assert_headers(ws, expected: list[str], row: int, label: str) -> None:
    actual = [ws.cell(row, col).value for col in range(1, len(expected) + 1)]
    if actual != expected:
        raise ValueError(f"{label} headers mismatch at row {row}: expected {expected}, got {actual}")
    extra = [ws.cell(row, col).value for col in range(len(expected) + 1, ws.max_column + 1)]
    if any(value not in (None, "") for value in extra):
        raise ValueError(f"{label} has non-empty extra columns after expected headers: {extra}")


def write_main_xlsx(template_path: Path, output_path: Path, rows: list[dict[str, str]]) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    assert_headers(ws, MAIN_HEADERS, 1, "main template")
    actual_ru = [ws.cell(2, col).value for col in range(1, len(MAIN_RUS_HEADERS) + 1)]
    if actual_ru != MAIN_RUS_HEADERS:
        raise ValueError(f"main template Russian headers mismatch: expected {MAIN_RUS_HEADERS}, got {actual_ru}")
    styles = copy_row_styles(ws, 3, len(MAIN_HEADERS))
    trim_to_columns(ws, len(MAIN_HEADERS))
    clear_data_rows(ws, 3)

    for index, row in enumerate(rows, start=3):
        values = [
            row.get("client_id", ""),
            row.get("phones", ""),
            row.get("client_fio", ""),
            row.get("email", ""),
            "Действующие клиенты",
            row.get("funnel_step", ""),
            0,
            parse_date(row.get("create_date")),
            row.get("manager", ""),
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col, value)
            cell._style = copy.copy(styles[col - 1])
            if col in {1, 2, 3, 4, 5, 6, 9}:
                cell.number_format = "@"
            if col == 7:
                cell.number_format = "0"
            if col == 8:
                cell.number_format = "yyyy-mm-dd"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def write_cards_xlsx(template_path: Path, output_path: Path, rows: list[dict[str, str]]) -> None:
    wb = load_workbook(template_path)
    ws = wb.active
    assert_headers(ws, CARD_HEADERS, 1, "plastic cards template")
    styles = copy_row_styles(ws, 1, len(CARD_HEADERS))
    trim_to_columns(ws, len(CARD_HEADERS))
    clear_data_rows(ws, 2)

    for index, row in enumerate(rows, start=2):
        values = [row.get("phones", ""), row.get("client_fio", ""), row.get("plastic_card_number", "")]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(index, col, value)
            cell._style = copy.copy(styles[col - 1])
            cell.number_format = "@"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def workbook_data_row_count(path: Path, first_data_row: int, width: int) -> int:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=first_data_row, max_col=width, values_only=True):
        if any(value not in (None, "") for value in row):
            count += 1
    wb.close()
    return count


def count_schema_tables(schema_tables_csv: Path) -> int:
    if not schema_tables_csv.exists():
        return 0
    with schema_tables_csv.open("r", encoding="utf-8-sig", newline="") as handle:
        return max(sum(1 for _ in csv.DictReader(handle)), 0)


def make_validation_report(
    *,
    output_path: Path,
    cutoff_date: str,
    date_stamp: str,
    source_count: int,
    final_rows: list[dict[str, str]],
    duplicate_report_rows: list[dict[str, str]],
    exact_duplicate_rows: list[dict[str, str]],
    missing_required: list[dict[str, str]],
    missing_sales: list[dict[str, str]],
    missing_cards: list[dict[str, str]],
    multiple_active_subscriptions: list[dict[str, str]],
    multiple_cards: list[dict[str, str]],
    booking_without_active: list[dict[str, str]],
    stage_counts: Counter[str],
    manager_counts: Counter[str],
    main_xlsx: Path,
    cards_xlsx: Path,
) -> None:
    errors: list[str] = []
    warnings: list[str] = []

    if workbook_data_row_count(main_xlsx, 3, len(MAIN_HEADERS)) != len(final_rows):
        errors.append("main XLSX row count does not match deduped client count")
    if workbook_data_row_count(cards_xlsx, 2, len(CARD_HEADERS)) != len(final_rows):
        errors.append("plastic cards XLSX row count does not match deduped client count")

    client_ids = [row.get("client_id", "") for row in final_rows]
    duplicated_client_ids = [client_id for client_id, count in Counter(client_ids).items() if client_id and count > 1]
    if duplicated_client_ids:
        errors.append(f"duplicate client_id values remain: {len(duplicated_client_ids)}")

    exact_final_keys = Counter(
        (row.get("normalized_fio", ""), row.get("normalized_phones_set", ""))
        for row in final_rows
        if row.get("normalized_fio") and row.get("normalized_phones_set")
    )
    remaining_exact_dups = [key for key, count in exact_final_keys.items() if count > 1]
    if remaining_exact_dups:
        errors.append(f"exact FIO+phone duplicate keys remain: {len(remaining_exact_dups)}")

    invalid_steps = sorted({row.get("funnel_step", "") for row in final_rows if row.get("funnel_step", "") not in ALLOWED_STEPS})
    if invalid_steps:
        errors.append(f"invalid funnel_step values: {invalid_steps}")

    if any(row.get("funnel") != "Действующие клиенты" for row in final_rows):
        errors.append("some rows have unexpected funnel")
    if any(int_value(row.get("budget")) != 0 for row in final_rows):
        errors.append("some rows have non-zero budget")

    for row in final_rows:
        step = row.get("funnel_step", "")
        days = int_value(row.get("days_to_end"), -999999)
        has_booking = row.get("has_active_booking") == "1"
        if has_booking and step != "Бронь":
            errors.append(f"booking row not in booking step: client_id={row.get('client_id')}")
            break
        if step == "60-31 день до окончания" and not (31 <= days <= 60):
            errors.append(f"60-31 boundary violation: client_id={row.get('client_id')}, days={days}")
            break
        if step == "30-8 дней до окончания" and not (8 <= days <= 30):
            errors.append(f"30-8 boundary violation: client_id={row.get('client_id')}, days={days}")
            break
        if step == "7-0 день до окончания" and not (0 <= days <= 7):
            errors.append(f"7-0 boundary violation: client_id={row.get('client_id')}, days={days}")
            break

    card_numbers = [
        card_number
        for row in final_rows
        for card_number in split_raw_values(row.get("plastic_card_number"))
    ]
    duplicated_cards = [card for card, count in Counter(card_numbers).items() if count > 1]
    if duplicated_cards:
        errors.append(f"duplicate plastic card numbers remain: {len(duplicated_cards)}")

    if missing_required:
        warnings.append(f"missing required field rows exported and reported: {len(missing_required)}")
    if missing_sales:
        warnings.append(f"missing first sale/create_date rows exported and reported: {len(missing_sales)}")
    if missing_cards:
        warnings.append(f"clients without plastic card exported and reported: {len(missing_cards)}")
    if multiple_active_subscriptions:
        warnings.append(
            f"clients with multiple active subscriptions reported: {len(multiple_active_subscriptions)}"
        )
    if multiple_cards:
        warnings.append(f"clients with multiple plastic cards reported: {len(multiple_cards)}")
    if duplicate_report_rows:
        warnings.append(f"duplicate/potential duplicate signals reported: {len(duplicate_report_rows)}")

    verdict = "FAIL" if errors else "PASS"
    backup_path = ROOT / "data" / "Fitnes.bak"
    backup_size = backup_path.stat().st_size if backup_path.exists() else 0
    schema_table_count = count_schema_tables(ROOT / "output" / "schema_tables.csv")

    exact_auto_merge_groups = len(
        {
            (row.get("normalized_fio", ""), row.get("normalized_phones_set", ""))
            for row in exact_duplicate_rows
        }
    )
    same_phone_groups = sum(1 for row in duplicate_report_rows if row.get("report_type") == "same_phone_different_fio")
    same_fio_groups = sum(1 for row in duplicate_report_rows if row.get("report_type") == "same_fio_different_phones")

    lines = [
        "# Fitbase Final Export Validation",
        "",
        f"Run date: {datetime.now().isoformat(timespec='seconds')}",
        f"cutoff_date: `{cutoff_date}`",
        f"date_stamp: `{date_stamp}`",
        f"backup: `data/Fitnes.bak` ({backup_size:,} bytes)",
        "DatabaseName: `FitnessRestored`",
        f"Restored DB user tables: `{schema_table_count}`",
        "",
        "## Output Files",
        "",
        f"- Main XLSX: `{main_xlsx.relative_to(ROOT)}`",
        f"- Plastic cards XLSX: `{cards_xlsx.relative_to(ROOT)}`",
        "- Validation/report files are in `output/`.",
        "",
        "## Counts",
        "",
        f"- active-client candidates before deduplication: `{source_count}`",
        f"- clients after exact FIO+phone deduplication: `{len(final_rows)}`",
        f"- rows in main XLSX: `{workbook_data_row_count(main_xlsx, 3, len(MAIN_HEADERS))}`",
        f"- rows in plastic cards XLSX: `{workbook_data_row_count(cards_xlsx, 2, len(CARD_HEADERS))}`",
        f"- clients without phone: `{sum(1 for row in final_rows if not row.get('phones'))}`",
        f"- clients without FIO: `{sum(1 for row in final_rows if not row.get('client_fio'))}`",
        f"- clients without first sale: `{len(missing_sales)}`",
        f"- clients without create_date: `{sum(1 for row in final_rows if not row.get('create_date'))}`",
        f"- clients without plastic card: `{len(missing_cards)}`",
        f"- clients with multiple active subscriptions: `{len(multiple_active_subscriptions)}`",
        f"- clients with multiple plastic cards: `{len(multiple_cards)}`",
        f"- clients with booking but without active subscription: `{len(booking_without_active)}`",
        f"- exact duplicate groups auto-merged: `{exact_auto_merge_groups}`",
        f"- same-phone/different-FIO groups reported: `{same_phone_groups}`",
        f"- same-FIO/different-phone groups reported: `{same_fio_groups}`",
        "",
        "## Funnel Step Distribution",
        "",
    ]
    for step, count in stage_counts.most_common():
        lines.append(f"- `{step}`: `{count}`")
    lines.extend(["", "## Manager Distribution", ""])
    for manager, count in manager_counts.most_common():
        lines.append(f"- `{manager}`: `{count}`")
    lines.extend(["", "## Remaining Technical Questions", ""])
    lines.extend(
        [
            "- Fitbase date-format acceptance still needs a mini-test; current XLSX writes `create_date` as an Excel date with `yyyy-mm-dd` format.",
            "- Multiple active subscriptions are reported and not silently resolved by business logic.",
            "- Multiple plastic cards are reported; the export writes all active/unmarked card numbers comma-separated.",
        ]
    )
    lines.extend(["", "## Validation", ""])
    if errors:
        lines.append("Errors:")
        for error in errors:
            lines.append(f"- {error}")
    else:
        lines.append("Errors: none.")
    if warnings:
        lines.append("")
        lines.append("Data-quality warnings:")
        for warning in warnings:
            lines.append(f"- {warning}")
    lines.extend(["", f"Verdict: `{verdict}`", ""])

    output_path.write_text("\n".join(lines), encoding="utf-8")


def build_outputs(args: argparse.Namespace) -> None:
    staging_dir = Path(args.staging_dir)
    output_dir = Path(args.output_dir)
    cutoff_date = args.cutoff_date
    date_stamp = args.date_stamp or cutoff_date.replace("-", "")

    managers = load_managers(Path(args.managers_config))
    source_rows = read_csv(staging_dir / "mart_active_clients.csv")
    stg_bookings = read_csv(staging_dir / "stg_bookings.csv")
    stg_plastic_cards = read_csv(staging_dir / "stg_plastic_cards.csv")
    apply_active_card_number_lists(source_rows, build_active_card_numbers_by_client(stg_plastic_cards))
    annotate_rows(source_rows, managers)

    duplicate_report_rows, exact_duplicate_rows = build_duplicate_reports(source_rows)
    final_rows = deduplicate_rows(source_rows)

    stage_counts = Counter(row.get("funnel_step", "") for row in final_rows)
    manager_counts = Counter(row.get("manager", "") for row in final_rows)

    missing_required = rows_with_missing_required(final_rows)
    missing_sales = [row for row in final_rows if not row.get("first_sale_date") or not row.get("create_date")]
    missing_cards = [
        row for row in final_rows if not row.get("plastic_card_number") or int_value(row.get("active_card_count")) == 0
    ]
    multiple_active_subscriptions = [
        row for row in final_rows if int_value(row.get("active_subscription_count")) > 1
    ]
    multiple_cards = [row for row in final_rows if int_value(row.get("active_card_count")) > 1]
    booking_without_active = build_booking_without_active_report(final_rows, stg_bookings)

    main_xlsx = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}.xlsx"
    cards_xlsx = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}.xlsx"

    final_csv = output_dir / f"final_active_clients_{date_stamp}.csv"
    write_csv(final_csv, final_rows, FINAL_FIELDS)
    write_csv(
        output_dir / "stage_distribution.csv",
        [{"funnel_step": step, "clients": count} for step, count in stage_counts.most_common()],
        ["funnel_step", "clients"],
    )
    write_csv(
        output_dir / "manager_distribution.csv",
        [{"manager": manager, "clients": count} for manager, count in manager_counts.most_common()],
        ["manager", "clients"],
    )
    write_csv(
        output_dir / "duplicates_report.csv",
        duplicate_report_rows,
        [
            "report_type",
            "group_key",
            "normalized_fio",
            "normalized_phone",
            "normalized_phones_set",
            "kept_client_id",
            "kept_client_ref",
            "client_ids",
            "client_refs",
            "client_fios",
            "phones",
            "action",
            "note",
        ],
    )
    write_csv(
        output_dir / "missing_required_fields.csv",
        missing_required,
        ["client_ref", "client_id", "client_fio", "phones", "email", "missing_fields", "create_date", "funnel_step"],
    )
    write_csv(
        output_dir / "missing_sales_report.csv",
        missing_sales,
        ["client_ref", "client_id", "client_fio", "phones", "first_sale_date", "create_date", "create_date_source"],
    )
    write_csv(
        output_dir / "missing_cards_report.csv",
        missing_cards,
        ["client_ref", "client_id", "client_fio", "phones", "plastic_card_number", "active_card_count"],
    )
    write_csv(
        output_dir / "multiple_active_subscriptions_report.csv",
        multiple_active_subscriptions,
        [
            "client_ref",
            "client_id",
            "client_fio",
            "phones",
            "active_subscription_count",
            "active_subscription_ref",
            "active_subscription_name",
            "active_subscription_start_date",
            "active_subscription_end_date",
        ],
    )
    write_csv(
        output_dir / "multiple_cards_report.csv",
        multiple_cards,
        ["client_ref", "client_id", "client_fio", "phones", "plastic_card_number", "active_card_count"],
    )
    write_csv(
        output_dir / "booking_without_active_subscription_report.csv",
        booking_without_active,
        ["client_ref", "booking_ref", "booking_date", "booking_status", "is_active_booking", "raw_source"],
    )

    write_main_xlsx(Path(args.main_template), main_xlsx, final_rows)
    write_cards_xlsx(Path(args.cards_template), cards_xlsx, final_rows)

    make_validation_report(
        output_path=output_dir / "validation_report.md",
        cutoff_date=cutoff_date,
        date_stamp=date_stamp,
        source_count=len(source_rows),
        final_rows=final_rows,
        duplicate_report_rows=duplicate_report_rows,
        exact_duplicate_rows=exact_duplicate_rows,
        missing_required=missing_required,
        missing_sales=missing_sales,
        missing_cards=missing_cards,
        multiple_active_subscriptions=multiple_active_subscriptions,
        multiple_cards=multiple_cards,
        booking_without_active=booking_without_active,
        stage_counts=stage_counts,
        manager_counts=manager_counts,
        main_xlsx=main_xlsx,
        cards_xlsx=cards_xlsx,
    )

    print(f"source_rows={len(source_rows)}")
    print(f"final_rows={len(final_rows)}")
    print(f"main_xlsx={main_xlsx.relative_to(ROOT)}")
    print(f"cards_xlsx={cards_xlsx.relative_to(ROOT)}")
    print(f"duplicates_report_rows={len(duplicate_report_rows)}")
    print(f"missing_required_rows={len(missing_required)}")
    print(f"missing_sales_rows={len(missing_sales)}")
    print(f"missing_cards_rows={len(missing_cards)}")
    print(f"multiple_active_subscriptions_rows={len(multiple_active_subscriptions)}")
    print(f"multiple_cards_rows={len(multiple_cards)}")
    print(f"booking_without_active_subscription_rows={len(booking_without_active)}")
    print(f"validation_report={output_dir.joinpath('validation_report.md').relative_to(ROOT)}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default=DEFAULT_CUTOFF_DATE)
    parser.add_argument("--date-stamp", default=DEFAULT_DATE_STAMP)
    parser.add_argument("--staging-dir", default=str(DEFAULT_STAGING_DIR))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    parser.add_argument("--main-template", default=str(DEFAULT_MAIN_TEMPLATE))
    parser.add_argument("--cards-template", default=str(DEFAULT_CARDS_TEMPLATE))
    parser.add_argument("--managers-config", default=str(DEFAULT_MANAGERS_CONFIG))
    return parser.parse_args()


def main() -> None:
    build_outputs(parse_args())


if __name__ == "__main__":
    main()
