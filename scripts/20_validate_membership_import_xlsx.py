#!/usr/bin/env python3
"""Validate generated Fitbase membership import XLSX files."""

from __future__ import annotations

import argparse
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"
CLIENT_HEADERS = [
    "tag",
    "contract_id",
    "client_id",
    "phone",
    "client_fio",
    "contract_name",
    "card",
    "duration",
    "duration_type",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "freeze",
    "guests",
    "visits_left",
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "manager",
]
TEMPLATE_HEADERS = [
    "branches_access",
    "name",
    "price",
    "duration",
    "duration_type",
    "visits",
    "guests",
    "freeze",
    "first_visit_activation",
    "archive",
    "category",
    "legal_entity",
]


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def iter_data_rows(path: Path, width: int) -> tuple[list[Any], list[tuple[Any, ...]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    headers = list(next(rows_iter)[:width])
    next(rows_iter, None)  # Russian header row.
    rows = [tuple(row[:width]) for row in rows_iter if any(value not in (None, "") for value in row[:width])]
    wb.close()
    return headers, rows


def read_source_client_ids(path: Path) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
    client_col = headers.index("client_id")
    ids = {
        str(row[client_col]).strip()
        for row in ws.iter_rows(min_row=3, values_only=True)
        if row[client_col] not in (None, "")
    }
    wb.close()
    return ids


def read_refuser_client_ids(path: Path) -> set[str]:
    if not path.exists():
        return set()
    import csv

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return {
            str(row.get("client_id") or "").strip()
            for row in csv.DictReader(handle)
            if str(row.get("client_id") or "").strip()
        }


def blank(value: Any) -> bool:
    return value is None or value == ""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source-output-dir", default="output/20251115_0800_fix_owner")
    parser.add_argument("--output-dir", default="output/20251115_0800_fix_owner_new_import")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source_output_dir = as_abs(args.source_output_dir)
    output_dir = as_abs(args.output_dir)
    source_clients_xlsx = source_output_dir / f"fitbase_active_clients_import_zayavki_{args.date_stamp}_all_funnels.xlsx"
    client_xlsx = output_dir / f"fitbase_import_abonementy_clientov_{args.date_stamp}.xlsx"
    template_xlsx = output_dir / f"fitbase_import_shablony_abonementov_{args.date_stamp}.xlsx"
    refuser_clients_csv = source_output_dir / "csv" / "new_application_refusers.csv"

    client_headers, client_rows = iter_data_rows(client_xlsx, len(CLIENT_HEADERS))
    template_headers, template_rows = iter_data_rows(template_xlsx, len(TEMPLATE_HEADERS))
    refuser_client_ids = read_refuser_client_ids(refuser_clients_csv)
    source_client_ids = read_source_client_ids(source_clients_xlsx) | refuser_client_ids
    client_idx = {header: index for index, header in enumerate(CLIENT_HEADERS)}
    template_idx = {header: index for index, header in enumerate(TEMPLATE_HEADERS)}

    errors: list[str] = []
    warnings: list[str] = []
    if client_headers != CLIENT_HEADERS:
        errors.append(f"client headers mismatch: {client_headers}")
    if template_headers != TEMPLATE_HEADERS:
        errors.append(f"template headers mismatch: {template_headers}")

    contract_ids = [str(row[client_idx["contract_id"]]).strip() for row in client_rows if not blank(row[client_idx["contract_id"]])]
    duplicated_contracts = [item for item, count in Counter(contract_ids).items() if count > 1]
    if duplicated_contracts:
        errors.append(f"duplicate contract_id values: {len(duplicated_contracts)}")

    client_ids = {str(row[client_idx["client_id"]]).strip() for row in client_rows}
    unknown_clients = client_ids - source_client_ids
    if unknown_clients:
        errors.append(f"client rows outside source final XLSX: {len(unknown_clients)}")

    template_names = {str(row[template_idx["name"]]).strip() for row in template_rows}
    missing_templates = {
        str(row[client_idx["contract_name"]]).strip()
        for row in client_rows
        if not blank(row[client_idx["contract_name"]])
    } - template_names
    if missing_templates:
        errors.append(f"client contract_name missing in templates: {len(missing_templates)}")

    blanks = Counter()
    refuser_rows = [row for row in client_rows if str(row[client_idx["tag"]] or "").strip() == "отказники"]
    refuser_row_client_ids = {str(row[client_idx["client_id"]]).strip() for row in refuser_rows}
    placeholder_refuser_rows = [
        row
        for row in refuser_rows
        if blank(row[client_idx["contract_id"]]) and blank(row[client_idx["contract_name"]])
    ]
    for row in client_rows:
        is_refuser_placeholder = (
            str(row[client_idx["tag"]] or "").strip() == "отказники"
            and blank(row[client_idx["contract_id"]])
            and blank(row[client_idx["contract_name"]])
        )
        required_fields = ["tag", "client_id", "client_fio", "create_date", "manager"] if is_refuser_placeholder else [
            "contract_id",
            "client_id",
            "client_fio",
            "contract_name",
            "create_date",
            "payment_date",
            "price",
            "amount_of_payments",
            "payment_left",
            "manager",
        ]
        for field in required_fields:
            if blank(row[client_idx[field]]):
                blanks[field] += 1
    if blanks:
        errors.append(f"required blanks: {dict(blanks)}")
    if refuser_client_ids and not (refuser_client_ids <= refuser_row_client_ids):
        errors.append(f"refuser clients missing tagged membership rows: {len(refuser_client_ids - refuser_row_client_ids)}")
    untagged_blank_contract_rows = [
        row
        for row in client_rows
        if blank(row[client_idx["contract_id"]]) and str(row[client_idx["tag"]] or "").strip() != "отказники"
    ]
    if untagged_blank_contract_rows:
        errors.append(f"blank contract_id rows without tag=отказники: {len(untagged_blank_contract_rows)}")

    payment_type_counts = Counter(str(row[client_idx["type_of_payment"]] or "blank") for row in client_rows)
    if payment_type_counts.get("blank", 0):
        warnings.append(f"blank payment type rows: {payment_type_counts['blank']}")

    report = [
        "# Membership import XLSX recheck",
        "",
        f"- client rows: {len(client_rows)}",
        f"- template rows: {len(template_rows)}",
        f"- source final clients: {len(source_client_ids)}",
        f"- refuser source clients: {len(refuser_client_ids)}",
        f"- refuser tagged rows: {len(refuser_rows)}",
        f"- refuser placeholder rows: {len(placeholder_refuser_rows)}",
        f"- row clients: {len(client_ids)}",
        f"- duplicate contract_id values: {len(duplicated_contracts)}",
        f"- missing template names: {len(missing_templates)}",
        f"- payment types: {dict(payment_type_counts)}",
        f"- status: {'PASS' if not errors else 'FAIL'}",
        "",
        "## Errors",
        "",
    ]
    report.extend(f"- {item}" for item in errors)
    if not errors:
        report.append("- none")
    report.extend(["", "## Warnings", ""])
    report.extend(f"- {item}" for item in warnings)
    if not warnings:
        report.append("- none")

    reports_dir = output_dir / "reports"
    reports_dir.mkdir(parents=True, exist_ok=True)
    (reports_dir / "validation_recheck.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    print("\n".join(report))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())
