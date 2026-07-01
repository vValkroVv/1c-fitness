#!/usr/bin/env python3
"""Build a 7-row payment-left/retail sample in the membership import format."""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"

MAIN_XLSX_PATH = ROOT / (
    "output/20251115_0800_fix_owner_new_import/"
    f"fitbase_import_abonementy_clientov_{DATE_STAMP}.xlsx"
)
OUTPUT_PATH = ROOT / (
    "output/20251115_0800_fix_owner_new_import/"
    f"payment_left_retail_review_7_examples_{DATE_STAMP}.xlsx"
)

# The workbook is intentionally business-facing: one sheet, same 20 columns as
# the main membership import, no diagnostic columns and no extra sheets.
EXAMPLE_CONTRACT_IDS = [
    # Named installment, active, positive debt, cashless.
    "00000150281",
    # Named installment, active, positive debt, SBP.
    "00000150311",
    # Named installment, active, fully paid after payment fallback.
    "00000150494",
    # No "рассрочка" in the name, but active and payment_left > 0.
    "00000145694",
    # Retail client, historical zero-price refund, linked Document131.
    "00000055819",
    # Retail by FIO with another client_id, paid historical row without Document131 refund.
    "00000118791",
    # Retail client, zero-price non-refund fallback/free historical row.
    "00000117808",
]


def read_main_rows(path: Path) -> tuple[list[str], dict[str, list[Any]]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    contract_idx = headers.index("contract_id")

    rows: dict[str, list[Any]] = {}
    for values in ws.iter_rows(min_row=2, values_only=True):
        contract_id = str(values[contract_idx] or "")
        if contract_id in EXAMPLE_CONTRACT_IDS:
            rows[contract_id] = list(values)

    wb.close()
    return [str(header) for header in headers], rows


def autosize(ws) -> None:
    for column in ws.columns:
        letter = get_column_letter(column[0].column)
        max_len = 0
        for cell in column:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 45)


def style_ws(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    autosize(ws)


def build_workbook() -> Workbook:
    headers, rows_by_contract = read_main_rows(MAIN_XLSX_PATH)
    missing = [contract_id for contract_id in EXAMPLE_CONTRACT_IDS if contract_id not in rows_by_contract]
    if missing:
        raise RuntimeError(f"Example rows not found in main membership import: {missing}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт_абонементы"
    ws.append(headers)
    for contract_id in EXAMPLE_CONTRACT_IDS:
        ws.append(rows_by_contract[contract_id])

    style_ws(ws)
    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
