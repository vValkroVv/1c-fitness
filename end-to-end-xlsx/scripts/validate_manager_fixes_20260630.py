#!/usr/bin/env python3
"""Read-only business validation for the 2026-06-30 manager-fixes delivery.

The validator deliberately reads only the final XLSX files, exported staging
CSV/TSV files, and the immutable baseline XLSX files.  It does not query SQL
and does not import production generator code, so a generator regression
cannot silently change the expected business rules in this audit.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import zipfile
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass, field
from datetime import date, datetime, timedelta, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Iterable

from openpyxl import load_workbook


EXPECTED_CUTOFF = datetime(2026, 6, 30, 23, 27, 3)
EXPECTED_CUTOFF_TEXT = "2026-06-30 23:27:03"
EXPECTED_CUTOFF_DATE = date(2026, 6, 30)
DATE_STAMP = "20260630"
EXPECTED_MEMBERSHIP_FACTS = 101_436
EXPECTED_BASELINE_OLD_ONLY = 45
EXPECTED_BASELINE_NEW_ONLY = 41

APPLICATIONS_FILE = (
    "fitbase_active_clients_import_zayavki_20260630_all_funnels.xlsx"
)
CARDS_FILE = "fitbase_active_clients_plastic_cards_20260630_all_funnels.xlsx"
MEMBERSHIP_FILE = "fitbase_import_abonementy_clientov_20260630.xlsx"
MEMBERSHIP_TEMPLATES_FILE = "fitbase_import_shablony_abonementov_20260630.xlsx"
SERVICE_TEMPLATES_FILE = "fitbase_import_shablony_uslug_20260630.xlsx"
CLIENT_SERVICES_FILE = "fitbase_import_uslugi_clientov_20260630.xlsx"
PROBLEM_FILES = {
    1: "problem_1_no_payment_cash_10_cases_20260630.xlsx",
    2: "problem_2_zero_price_direct_full_41_cases_20260630.xlsx",
    3: "problem_3_non_named_payment_left_203_cases_20260630.xlsx",
    4: "problem_4_subrent_visits_left_contract_151350_1_case_20260630.xlsx",
}
PROBLEM4_CONTRACT_ID = "00000151350"

EXPECTED_PROBLEM_CREATE_DATE_CHANGES = {1: 9, 2: 27, 3: 147}
EXPECTED_PROBLEM_OWNER_VALUES: dict[int, dict[str, dict[str, str]]] = {
    2: {
        "00000150337": {
            "client_id": "000071125",
            "phone": "+7 (929) 1714793",
            "client_fio": "Шалыгина София Эдуардовна",
            "card": "115000290105",
        }
    },
    3: {
        "00000137855": {
            "client_id": "000068958",
            "phone": "+7 (911) 4320667",
            "client_fio": "Шмаков Данил Олегович",
            "card": "115000229969",
            "manager": "Яковлева Александра Владимировна",
        }
    },
}
FINANCIAL_MEMBERSHIP_FIELDS = {
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "payment_date",
}

OWNER_CASES = {
    "00000133547": {
        "client_id": "000004598",
        "client_fio": "Дворжицкая Анна Александровна",
        "old_client_id": "000034737",
        "old_client_fio": "Дворжицкий Владислав Станиславович",
    },
    "00000144947": {
        "client_id": "000074154",
        "client_fio": "Галаничева Карина Павловна",
        "old_client_id": "000059399",
        "old_client_fio": "Позолотин Никита Олегович",
    },
}

# These contracts prove that the previous owners still have their own active
# memberships and therefore must not disappear from applications/cards.
OLD_OWNER_CONTROL_CONTRACTS = {
    "00000144918": {
        "client_id": "000034737",
        "client_fio": "Дворжицкий Владислав Станиславович",
    },
    "00000140996": {
        "client_id": "000059399",
        "client_fio": "Позолотин Никита Олегович",
    },
}

APPLICATION_CARD_CONTROLS = {
    "000004598": {
        "client_fio": "Дворжицкая Анна Александровна",
        "card": "115000284449",
        "role": "новый владелец 133547",
    },
    "000074154": {
        "client_fio": "Галаничева Карина Павловна",
        "card": "115000273238",
        "role": "новый владелец 144947",
    },
    "000034737": {
        "client_fio": "Дворжицкий Владислав Станиславович",
        "card": "115000284463",
        "role": "прежний владелец 133547",
    },
    "000059399": {
        "client_fio": "Позолотин Никита Олегович",
        "card": "115000256767",
        "role": "прежний владелец 144947",
    },
}

CYCLE_BALANCE_CONTROLS = {
    "00000138687": 2,
    "00000141600": 3,
    "00000144782": 3,
    "00000144816": 0,
    "00000145361": 2,
    "00000147786": 6,
    "00000147787": 5,
    "00000151241": 7,
}

SHULEYKO_CONTRACT_ID = "00000140663"
SHULEYKO_NAME = "Шулейко Екатерина Витальевна"
SHULEYKO_SALE_DATE = date(2025, 9, 23)
SHULEYKO_DURATION = Decimal("12")
SHULEYKO_PRICE = Decimal("12990")

EXPECTED_SERVICE_REMOVED = {"00000032932-1", "00000033161-1"}
EXPECTED_SERVICE_ADDED = {
    "00000032113-1": {
        "client_id": "000044977",
        "client_fio": "Ларионов Виктор Юрьевич",
        "service_name": "Переоформление платное абонемента на фитнес",
        "create_date": date(2026, 6, 1),
        "price": Decimal("3500"),
        "manager": "Соколова Анастасия Александровна",
        "филиал": "Фитнес Империя (Ровио)",
    },
    "00000032795-1": {
        "client_id": "000061410",
        "client_fio": "Саурин Артем Сергеевич",
        "service_name": "Переоформление платное абонемента на фитнес",
        "create_date": date(2026, 6, 7),
        "price": Decimal("3500"),
        "manager": "Васильева Яна Денисовна",
        "филиал": "Фитнес Империя (Гоголевский)",
    },
}
EXPECTED_SERVICE_MANAGER_CHANGES = {
    "00000066249": (
        "Васильева Яна Денисовна",
        "Мартынова Дарья Дмитриевна",
    ),
    "00000068067": (
        "Васьковская Виктория Петровна",
        "Абраамян Татьяна Викторовна",
    ),
    "00000068248": (
        "Фёдорова Надежда Сергеевна",
        "Абраамян Татьяна Викторовна",
    ),
}

# This is a frozen copy of the SQL export contract, not an import from the
# production builder.  It lets this validator locate the business fields in a
# headerless UTF-16 TSV independently.
MEMBERSHIP_FACT_FIELDS = [
    "client_ref",
    "client_id",
    "original_client_ref",
    "original_client_id",
    "original_client_fio",
    "effective_client_ref",
    "effective_client_id",
    "effective_client_fio",
    "owner_change_ref",
    "owner_change_number",
    "owner_change_datetime",
    "owner_change_old_client_ref",
    "owner_change_new_client_ref",
    "owner_change_modifier_name",
    "owner_change_count_for_membership",
    "subscription_ref",
    "document_number",
    "holder_client_ref",
    "payer_client_ref",
    "client_role_source",
    "product_ref",
    "product_code",
    "subscription_name",
    "product_class",
    "is_full_subscription",
    "is_trial_or_guest",
    "is_subrent",
    "is_limited_subrent",
    "sale_date",
    "sale_datetime",
    "start_date",
    "end_date",
    "duration_days",
    "status",
    "booking_status_ref",
    "booking_status_name",
    "doc_posted",
    "doc_marked",
    "register_duration_days",
    "is_active_on_cutoff",
    "is_finished_before_cutoff",
    "days_to_end",
    "days_since_end",
    "raw_club",
    "normalized_club",
    "club_source",
    "sale_branch_raw",
    "sale_branch",
    "sale_branch_source",
    "raw_source",
    "doc_duration_value",
    "rg_duration_days",
    "rg_freeze_days",
    "rg_guests",
    "rg_price",
    "rg_paid_candidate",
    "rg_payment_count_candidate",
    "rg_visits_candidate_8007",
    "rg_visits_candidate_8008",
    "rg_visits_candidate_8009",
    "subrent_visit_limit",
    "subrent_active_by_dates_on_cutoff",
    "subrent_finished_by_dates_before_cutoff",
    "subrent_rg3336_receipt_qty",
    "subrent_rg3336_expense_qty",
    "subrent_rg3336_signed_balance",
    "subrent_rg3336_visit_doc_expense_qty",
    "subrent_rg3336_receipt_rows",
    "subrent_rg3336_expense_rows",
    "subrent_rg3336_case_group",
    "matched_payment_ref",
    "matched_payment_datetime",
    "matched_payment_amount",
    "matched_payment_method",
    "matched_payment_operation",
    "matched_payment_match_source",
    "membership_sale_line_amount",
    "membership_sale_line_count",
    "document131_refund_count",
    "document131_posted_unmarked_refund_count",
    "cutoff_at",
]


@dataclass(frozen=True)
class Fact:
    client_id: str
    original_client_id: str
    effective_client_id: str
    subscription_ref: str
    owner_change_ref: str
    sale_date: date | None
    is_active_on_cutoff: str


@dataclass
class FactsSummary:
    path: Path
    row_count: int = 0
    repaired_rows: int = 0
    by_contract: dict[str, Fact] = field(default_factory=dict)
    duplicate_contracts: dict[str, int] = field(default_factory=dict)
    conflicting_contracts: dict[str, list[tuple[str, str]]] = field(default_factory=dict)
    cutoff_values: Counter[str] = field(default_factory=Counter)


@dataclass
class MembershipRow:
    contract_id: str
    client_id: str
    client_fio: str
    contract_name: str
    create_date: Any
    duration: Any
    visits_left: Any
    price: Any
    source_file: str
    row_number: int


@dataclass
class MembershipDeliverySummary:
    file_sets: dict[str, set[str]] = field(default_factory=dict)
    file_counts: dict[str, int] = field(default_factory=dict)
    duplicate_contracts: dict[str, list[str]] = field(default_factory=dict)
    missing_files: list[str] = field(default_factory=list)
    union: set[str] = field(default_factory=set)
    partition_overlaps: dict[str, list[str]] = field(default_factory=dict)
    contract_clients: dict[str, str] = field(default_factory=dict)
    contract_client_mismatches: list[str] = field(default_factory=list)
    selected_rows: dict[str, list[MembershipRow]] = field(
        default_factory=lambda: defaultdict(list)
    )
    cycle_rows: list[MembershipRow] = field(default_factory=list)
    real_row_count: int = 0
    matched_create_dates: int = 0
    missing_staging_facts: list[str] = field(default_factory=list)
    invalid_create_dates: list[str] = field(default_factory=list)
    mismatched_create_dates: list[str] = field(default_factory=list)


@dataclass
class ContractFileSummary:
    contract_ids: set[str]
    row_count: int
    duplicates: set[str]
    client_ids: dict[str, str]


@dataclass
class CheckResult:
    check_id: str
    title: str
    status: str
    summary: str
    details: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    metrics: dict[str, Any] = field(default_factory=dict)


def clean_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized_text(value: Any) -> str:
    return re.sub(r"\s+", " ", clean_text(value)).casefold().replace("ё", "е")


def normalize_contract_id(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(11) if text.isdigit() else text


def normalize_client_id(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    if re.fullmatch(r"\d+\.0", text):
        text = text[:-2]
    return text.zfill(9) if text.isdigit() else text


def as_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    try:
        return Decimal(clean_text(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def as_date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    for candidate in (text[:10], text):
        try:
            return date.fromisoformat(candidate)
        except ValueError:
            pass
    for fmt in ("%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def as_cutoff_datetime(value: Any) -> datetime | None:
    """Parse SQL-style timestamps, including seven fractional digits."""

    text = clean_text(value)
    match = re.fullmatch(
        r"(\d{4}-\d{2}-\d{2})[ T](\d{2}:\d{2}:\d{2})(?:\.(\d+))?",
        text,
    )
    if not match:
        return None
    try:
        parsed = datetime.strptime(
            f"{match.group(1)} {match.group(2)}", "%Y-%m-%d %H:%M:%S"
        )
    except ValueError:
        return None
    fraction = match.group(3) or ""
    if fraction:
        microseconds = int((fraction + "000000")[:6])
        parsed += timedelta(microseconds=microseconds)
        # Datetime cannot retain sub-microsecond precision.  Reject a non-zero
        # remainder rather than silently considering it the expected cutoff.
        if any(character != "0" for character in fraction[6:]):
            return None
    return parsed


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def sample(values: Iterable[Any], limit: int = 20) -> list[Any]:
    result: list[Any] = []
    for value in values:
        result.append(value)
        if len(result) >= limit:
            break
    return result


def format_sample(values: Iterable[Any], limit: int = 20) -> str:
    materialized = list(values)
    head = materialized[:limit]
    suffix = f"; ещё {len(materialized) - limit}" if len(materialized) > limit else ""
    return ", ".join(str(value) for value in head) + suffix


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def locate_work_file(work_dir: Path, candidates: list[str]) -> Path:
    for relative in candidates:
        candidate = work_dir / relative
        if candidate.is_file():
            return candidate
    # Return the canonical location so the error identifies the expected path.
    return work_dir / candidates[0]


def read_csv_record(path: Path) -> dict[str, str]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        rows = list(reader)
    if len(rows) != 1:
        raise ValueError(f"{path}: expected exactly one metadata row, got {len(rows)}")
    return {clean_text(key): clean_text(value) for key, value in rows[0].items()}


def xlsx_headers(worksheet: Any) -> tuple[list[str], dict[str, int]]:
    first_row = next(
        worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None
    )
    if first_row is None:
        raise ValueError("empty worksheet")
    headers = [clean_text(value) for value in first_row]
    duplicates = [name for name, count in Counter(headers).items() if name and count > 1]
    if duplicates:
        raise ValueError(f"duplicate technical headers: {duplicates}")
    return headers, {name: index for index, name in enumerate(headers) if name}


def require_headers(path: Path, indexes: dict[str, int], names: Iterable[str]) -> None:
    missing = [name for name in names if name not in indexes]
    if missing:
        raise ValueError(f"{path}: missing XLSX columns {missing}")


def read_contract_file(path: Path, data_start_row: int) -> ContractFileSummary:
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        _, indexes = xlsx_headers(worksheet)
        require_headers(path, indexes, ["contract_id", "client_id"])
        counts: Counter[str] = Counter()
        client_ids: dict[str, str] = {}
        for row in worksheet.iter_rows(min_row=data_start_row, values_only=True):
            contract_id = normalize_contract_id(row[indexes["contract_id"]])
            if contract_id:
                counts[contract_id] += 1
                client_ids.setdefault(
                    contract_id, normalize_client_id(row[indexes["client_id"]])
                )
    finally:
        workbook.close()
    return ContractFileSummary(
        contract_ids=set(counts),
        row_count=sum(counts.values()),
        duplicates={key for key, count in counts.items() if count > 1},
        client_ids=client_ids,
    )


def repair_membership_fact_row(raw: list[str], path: Path) -> tuple[list[str], bool]:
    if len(raw) == len(MEMBERSHIP_FACT_FIELDS):
        return raw, False

    def is_hex_ref(value: str) -> bool:
        return bool(re.fullmatch(r"[0-9A-F]{32}", value or ""))

    effective_ref_idx = next(
        (index for index in range(5, len(raw)) if is_hex_ref(raw[index])), -1
    )
    if effective_ref_idx < 0:
        raise ValueError(f"{path}: cannot repair TSV row with {len(raw)} columns")
    effective_fio_start = effective_ref_idx + 2
    effective_fio_index = MEMBERSHIP_FACT_FIELDS.index("effective_client_fio")
    tail_after_effective_fio = len(MEMBERSHIP_FACT_FIELDS) - effective_fio_index - 1
    effective_fio_end = len(raw) - tail_after_effective_fio
    fixed = (
        raw[:4]
        + [" ".join(part.strip() for part in raw[4:effective_ref_idx] if part.strip())]
        + [raw[effective_ref_idx], raw[effective_ref_idx + 1]]
        + [
            " ".join(
                part.strip()
                for part in raw[effective_fio_start:effective_fio_end]
                if part.strip()
            )
        ]
        + raw[effective_fio_end:]
    )
    if len(fixed) != len(MEMBERSHIP_FACT_FIELDS):
        raise ValueError(
            f"{path}: cannot repair TSV row with {len(raw)} columns; "
            f"repaired to {len(fixed)}"
        )
    return fixed, True


def read_membership_facts(path: Path) -> FactsSummary:
    if not path.is_file():
        raise FileNotFoundError(path)
    result = FactsSummary(path=path)
    field_indexes = {name: index for index, name in enumerate(MEMBERSHIP_FACT_FIELDS)}
    duplicate_counts: Counter[str] = Counter()
    variants: dict[str, set[tuple[str, str]]] = defaultdict(set)
    with path.open("r", encoding="utf-16", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for raw in reader:
            if not raw:
                continue
            values, repaired = repair_membership_fact_row(raw, path)
            result.row_count += 1
            result.repaired_rows += int(repaired)
            contract_id = normalize_contract_id(
                values[field_indexes["document_number"]]
            )
            sale_date_text = clean_text(values[field_indexes["sale_date"]])
            active = clean_text(values[field_indexes["is_active_on_cutoff"]])
            cutoff = clean_text(values[field_indexes["cutoff_at"]])
            result.cutoff_values[cutoff] += 1
            if not contract_id:
                continue
            fact = Fact(
                client_id=normalize_client_id(values[field_indexes["client_id"]]),
                original_client_id=normalize_client_id(
                    values[field_indexes["original_client_id"]]
                ),
                effective_client_id=normalize_client_id(
                    values[field_indexes["effective_client_id"]]
                ),
                subscription_ref=clean_text(
                    values[field_indexes["subscription_ref"]]
                ),
                owner_change_ref=clean_text(
                    values[field_indexes["owner_change_ref"]]
                ),
                sale_date=as_date(sale_date_text),
                is_active_on_cutoff=active,
            )
            if contract_id in result.by_contract:
                duplicate_counts[contract_id] += 1
            else:
                result.by_contract[contract_id] = fact
            variants[contract_id].add((sale_date_text, active))
    result.duplicate_contracts = {
        key: count + 1 for key, count in duplicate_counts.items()
    }
    result.conflicting_contracts = {
        key: sorted(value) for key, value in variants.items() if len(value) > 1
    }
    return result


def read_service_cutoffs(path: Path) -> tuple[int, Counter[str], Counter[int]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    row_count = 0
    cutoffs: Counter[str] = Counter()
    column_counts: Counter[int] = Counter()
    with path.open("r", encoding="utf-16", newline="") as handle:
        reader = csv.reader(handle, delimiter="\t")
        for raw in reader:
            if not raw:
                continue
            row_count += 1
            column_counts[len(raw)] += 1
            if len(raw) < 2:
                cutoffs[""] += 1
            else:
                # cutoff_at is the penultimate field; using the tail keeps this
                # check robust if a free-text field contains an embedded tab.
                cutoffs[clean_text(raw[-2])] += 1
    return row_count, cutoffs, column_counts


def read_source_client_ids(applications_path: Path, refusers_path: Path) -> set[str]:
    """Reconstruct the exact client population consumed by the membership build."""

    if not applications_path.is_file():
        raise FileNotFoundError(applications_path)
    if not refusers_path.is_file():
        raise FileNotFoundError(refusers_path)
    client_ids: set[str] = set()
    workbook = load_workbook(applications_path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        _, indexes = xlsx_headers(worksheet)
        require_headers(applications_path, indexes, ["client_id"])
        for values in worksheet.iter_rows(min_row=3, values_only=True):
            client_id = normalize_client_id(values[indexes["client_id"]])
            if client_id:
                client_ids.add(client_id)
    finally:
        workbook.close()
    with refusers_path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if "client_id" not in (reader.fieldnames or []):
            raise ValueError(f"{refusers_path}: missing client_id")
        for row in reader:
            client_id = normalize_client_id(row.get("client_id"))
            if client_id:
                client_ids.add(client_id)
    return client_ids


def read_csv_contracts(path: Path) -> ContractFileSummary:
    if not path.is_file():
        raise FileNotFoundError(path)
    counts: Counter[str] = Counter()
    client_ids: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"contract_id", "client_id"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise ValueError(f"{path}: missing CSV columns {missing}")
        for row in reader:
            contract_id = normalize_contract_id(row.get("contract_id"))
            if not contract_id:
                continue
            counts[contract_id] += 1
            client_ids.setdefault(
                contract_id, normalize_client_id(row.get("client_id"))
            )
    return ContractFileSummary(
        contract_ids=set(counts),
        row_count=sum(counts.values()),
        duplicates={key for key, count in counts.items() if count > 1},
        client_ids=client_ids,
    )


def read_exclusion_rules(path: Path) -> dict[str, str]:
    if not path.is_file():
        raise FileNotFoundError(path)
    result: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"contract_id", "rule"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise ValueError(f"{path}: missing CSV columns {missing}")
        for row in reader:
            contract_id = normalize_contract_id(row.get("contract_id"))
            rule = clean_text(row.get("rule"))
            if not contract_id:
                continue
            if contract_id in result:
                raise ValueError(f"{path}: duplicate contract_id={contract_id}")
            result[contract_id] = rule
    return result


def read_dedupe_evidence(
    path: Path,
) -> tuple[dict[str, set[str]], set[str], int]:
    if not path.is_file():
        raise FileNotFoundError(path)
    loser_to_winners: dict[str, set[str]] = defaultdict(set)
    winner_ids: set[str] = set()
    row_count = 0
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        required = {"client_id", "winner_client_id", "dedupe_reason"}
        missing = sorted(required - set(reader.fieldnames or []))
        if missing:
            raise ValueError(f"{path}: missing CSV columns {missing}")
        for row in reader:
            row_count += 1
            loser = normalize_client_id(row.get("client_id"))
            winner = normalize_client_id(row.get("winner_client_id"))
            reason = normalized_text(row.get("dedupe_reason"))
            if "normalized phone component" not in reason:
                raise ValueError(
                    f"{path}: unsupported dedupe_reason at row {row_count + 1}: "
                    f"{row.get('dedupe_reason')!r}"
                )
            if loser and winner:
                loser_to_winners[loser].add(winner)
                winner_ids.add(winner)
    return dict(loser_to_winners), winner_ids, row_count


class ValidationContext:
    def __init__(self, output_dir: Path, work_dir: Path, baseline_dir: Path):
        self.output_dir = output_dir
        self.work_dir = work_dir
        self.baseline_dir = baseline_dir
        self.membership_facts_path = locate_work_file(
            work_dir,
            [
                "imports/staging/membership_import_facts.tsv",
                "staging/membership_import_facts.tsv",
                "membership_import_facts.tsv",
            ],
        )
        self.service_facts_path = locate_work_file(
            work_dir,
            [
                "imports/staging/services_import_facts.tsv",
                "staging/services_import_facts.tsv",
                "services_import_facts.tsv",
            ],
        )
        self.owner_metadata_path = locate_work_file(
            work_dir,
            [
                "owner/staging/staging_run_metadata.csv",
                "owner/staging_run_metadata.csv",
            ],
        )
        self.raw_metadata_path = locate_work_file(
            work_dir,
            [
                "raw/staging/staging_run_metadata.csv",
                "raw/staging_run_metadata.csv",
            ],
        )
        self.owner_applications_path = locate_work_file(
            work_dir,
            [f"owner/{APPLICATIONS_FILE}", APPLICATIONS_FILE],
        )
        self.refusers_path = locate_work_file(
            work_dir,
            ["owner/csv/new_application_refusers.csv", "csv/new_application_refusers.csv"],
        )
        self.dedupe_report_path = locate_work_file(
            work_dir,
            [
                "owner/reports/phone_deduplication_removed_clients.csv",
                "reports/phone_deduplication_removed_clients.csv",
            ],
        )
        self.membership_rows_path = locate_work_file(
            work_dir,
            [
                "imports/staging/membership_import_rows.csv",
                "staging/membership_import_rows.csv",
                "membership_import_rows.csv",
            ],
        )
        self.membership_exclusions_path = locate_work_file(
            work_dir,
            [
                "imports/staging/membership_import_excluded_rows.csv",
                "staging/membership_import_excluded_rows.csv",
                "membership_import_excluded_rows.csv",
            ],
        )
        self._facts: FactsSummary | None = None
        self._delivery: MembershipDeliverySummary | None = None
        self._baseline_contracts: dict[str, ContractFileSummary] | None = None
        self._source_client_ids: set[str] | None = None
        self._membership_rows: ContractFileSummary | None = None
        self._exclusions: dict[str, str] | None = None
        self._dedupe_evidence: tuple[dict[str, set[str]], set[str], int] | None = None

    def facts(self) -> FactsSummary:
        if self._facts is None:
            self._facts = read_membership_facts(self.membership_facts_path)
        return self._facts

    def delivery(self) -> MembershipDeliverySummary:
        if self._delivery is None:
            self._delivery = self._scan_delivery_memberships()
        return self._delivery

    def baseline_contracts(self) -> dict[str, ContractFileSummary]:
        if self._baseline_contracts is None:
            files = {
                MEMBERSHIP_FILE: 3,
                **{PROBLEM_FILES[number]: 2 for number in (1, 2, 3)},
            }
            self._baseline_contracts = {
                name: read_contract_file(self.baseline_dir / name, start_row)
                for name, start_row in files.items()
            }
        return self._baseline_contracts

    def source_client_ids(self) -> set[str]:
        if self._source_client_ids is None:
            self._source_client_ids = read_source_client_ids(
                self.owner_applications_path, self.refusers_path
            )
        return self._source_client_ids

    def membership_rows(self) -> ContractFileSummary:
        if self._membership_rows is None:
            self._membership_rows = read_csv_contracts(self.membership_rows_path)
        return self._membership_rows

    def exclusions(self) -> dict[str, str]:
        if self._exclusions is None:
            self._exclusions = read_exclusion_rules(self.membership_exclusions_path)
        return self._exclusions

    def dedupe_evidence(self) -> tuple[dict[str, set[str]], set[str], int]:
        if self._dedupe_evidence is None:
            self._dedupe_evidence = read_dedupe_evidence(self.dedupe_report_path)
        return self._dedupe_evidence

    def _scan_delivery_memberships(self) -> MembershipDeliverySummary:
        facts = self.facts()
        result = MembershipDeliverySummary()
        selected_contracts = (
            set(OWNER_CASES)
            | set(OLD_OWNER_CONTROL_CONTRACTS)
            | set(CYCLE_BALANCE_CONTROLS)
            | {SHULEYKO_CONTRACT_ID}
        )
        files = {
            MEMBERSHIP_FILE: 3,
            **{name: 2 for name in PROBLEM_FILES.values()},
        }
        seen_by_file: dict[str, Counter[str]] = {}
        memberships_by_partition: dict[str, set[str]] = {}

        for name, data_start_row in files.items():
            path = self.output_dir / name
            if not path.is_file():
                result.missing_files.append(name)
                continue
            workbook = load_workbook(path, read_only=True, data_only=True)
            try:
                worksheet = workbook.active
                _, indexes = xlsx_headers(worksheet)
                required = [
                    "contract_id",
                    "client_id",
                    "client_fio",
                    "contract_name",
                    "create_date",
                    "duration",
                    "visits_left",
                    "price",
                ]
                require_headers(path, indexes, required)
                counts: Counter[str] = Counter()
                for row_number, values in enumerate(
                    worksheet.iter_rows(min_row=data_start_row, values_only=True),
                    start=data_start_row,
                ):
                    contract_id = normalize_contract_id(
                        values[indexes["contract_id"]]
                    )
                    if not contract_id:
                        continue
                    counts[contract_id] += 1
                    result.real_row_count += 1
                    membership = MembershipRow(
                        contract_id=contract_id,
                        client_id=normalize_client_id(values[indexes["client_id"]]),
                        client_fio=clean_text(values[indexes["client_fio"]]),
                        contract_name=clean_text(values[indexes["contract_name"]]),
                        create_date=values[indexes["create_date"]],
                        duration=values[indexes["duration"]],
                        visits_left=values[indexes["visits_left"]],
                        price=values[indexes["price"]],
                        source_file=name,
                        row_number=row_number,
                    )
                    result.contract_clients.setdefault(
                        contract_id, membership.client_id
                    )
                    if contract_id in selected_contracts:
                        result.selected_rows[contract_id].append(membership)
                    if is_cycle_name(membership.contract_name):
                        result.cycle_rows.append(membership)

                    fact = facts.by_contract.get(contract_id)
                    if fact is None:
                        result.missing_staging_facts.append(
                            f"{contract_id} ({name}:{row_number})"
                        )
                        continue
                    if membership.client_id != fact.client_id:
                        result.contract_client_mismatches.append(
                            f"{contract_id} ({name}:{row_number}): "
                            f"xlsx.client_id={membership.client_id}, "
                            f"staging.client_id={fact.client_id}"
                        )
                    actual_create_date = as_date(membership.create_date)
                    if actual_create_date is None:
                        result.invalid_create_dates.append(
                            f"{contract_id} ({name}:{row_number})={membership.create_date!r}"
                        )
                    elif fact.sale_date is None or actual_create_date != fact.sale_date:
                        result.mismatched_create_dates.append(
                            f"{contract_id} ({name}:{row_number}): "
                            f"create_date={actual_create_date}, sale_date={fact.sale_date}"
                        )
                    else:
                        result.matched_create_dates += 1
            finally:
                workbook.close()
            seen_by_file[name] = counts
            memberships_by_partition[name] = set(counts)
            result.file_sets[name] = set(counts)
            result.file_counts[name] = sum(counts.values())
            duplicates = sorted(key for key, count in counts.items() if count > 1)
            if duplicates:
                result.duplicate_contracts[name] = duplicates

        for name, values in memberships_by_partition.items():
            result.union.update(values)
            for other_name, other_values in memberships_by_partition.items():
                if name >= other_name:
                    continue
                overlap = sorted(values & other_values)
                if overlap:
                    result.partition_overlaps[f"{name} <-> {other_name}"] = overlap
        return result


def is_cycle_name(value: Any) -> bool:
    normalized = normalized_text(value)
    return "сайкл" in normalized or bool(re.search(r"\bcycle\b", normalized))


def make_result(
    check_id: str,
    title: str,
    errors: list[str],
    details: list[str],
    metrics: dict[str, Any] | None = None,
    pass_summary: str = "бизнес-инвариант соблюдён",
) -> CheckResult:
    status = "PASS" if not errors else "FAIL"
    summary = pass_summary if not errors else f"ошибок: {len(errors)}"
    return CheckResult(
        check_id=check_id,
        title=title,
        status=status,
        summary=summary,
        details=details,
        errors=errors,
        metrics=metrics or {},
    )


def check_cutoff(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    metrics: dict[str, Any] = {"expected_cutoff": EXPECTED_CUTOFF_TEXT}
    metadata_paths = {
        "owner": ctx.owner_metadata_path,
        "raw": ctx.raw_metadata_path,
    }
    for layer, path in metadata_paths.items():
        if not path.is_file():
            errors.append(f"{layer}: нет metadata CSV: {path}")
            continue
        try:
            record = read_csv_record(path)
        except Exception as exc:  # input corruption must become a report failure
            errors.append(f"{layer}: не удалось прочитать metadata: {exc}")
            continue
        observed = {
            "cutoff_date": record.get("cutoff_date", ""),
            "cutoff_at": record.get("cutoff_at", ""),
            "backup_finish_at": record.get("backup_finish_at", ""),
        }
        metrics[f"{layer}_metadata"] = observed
        details.append(f"{layer} metadata: `{observed}`")
        if as_date(observed["cutoff_date"]) != EXPECTED_CUTOFF_DATE:
            errors.append(
                f"{layer}.cutoff_date={observed['cutoff_date']!r}, "
                f"ожидается {EXPECTED_CUTOFF_DATE}"
            )
        for field_name in ("cutoff_at", "backup_finish_at"):
            if as_cutoff_datetime(observed[field_name]) != EXPECTED_CUTOFF:
                errors.append(
                    f"{layer}.{field_name}={observed[field_name]!r}, "
                    f"ожидается {EXPECTED_CUTOFF_TEXT}"
                )
        if as_cutoff_datetime(observed["cutoff_at"]) != as_cutoff_datetime(
            observed["backup_finish_at"]
        ):
            errors.append(f"{layer}: cutoff_at не равен backup_finish_at")

    try:
        membership_facts = ctx.facts()
        metrics["membership_fact_rows"] = membership_facts.row_count
        metrics["membership_cutoff_values"] = dict(membership_facts.cutoff_values)
        details.append(
            f"membership staging: {membership_facts.row_count} строк, "
            f"cutoff values={dict(membership_facts.cutoff_values)}"
        )
        if membership_facts.row_count == 0:
            errors.append("membership staging пуст")
        bad = {
            value: count
            for value, count in membership_facts.cutoff_values.items()
            if as_cutoff_datetime(value) != EXPECTED_CUTOFF
        }
        if bad:
            errors.append(f"membership staging: неверные cutoff_at={bad}")
    except Exception as exc:
        errors.append(f"membership staging: {exc}")

    try:
        service_rows, service_cutoffs, column_counts = read_service_cutoffs(
            ctx.service_facts_path
        )
        metrics["service_fact_rows"] = service_rows
        metrics["service_cutoff_values"] = dict(service_cutoffs)
        metrics["service_column_counts"] = dict(column_counts)
        details.append(
            f"service staging: {service_rows} строк, "
            f"cutoff values={dict(service_cutoffs)}"
        )
        if service_rows == 0:
            errors.append("service staging пуст")
        bad = {
            value: count
            for value, count in service_cutoffs.items()
            if as_cutoff_datetime(value) != EXPECTED_CUTOFF
        }
        if bad:
            errors.append(f"service staging: неверные cutoff_at={bad}")
    except Exception as exc:
        errors.append(f"service staging: {exc}")

    return make_result(
        "MF-01",
        "Единый cutoff_at из backup",
        errors,
        details,
        metrics,
        pass_summary=f"все слои используют {EXPECTED_CUTOFF_TEXT}",
    )


def check_problem4_routing(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    clean = delivery.file_sets.get(MEMBERSHIP_FILE)
    problem4_name = PROBLEM_FILES[4]
    problem4 = delivery.file_sets.get(problem4_name)
    if clean is None:
        errors.append(f"нет clean membership: {MEMBERSHIP_FILE}")
    if problem4 is None:
        errors.append(f"нет problem4: {problem4_name}")
    else:
        if problem4 != {PROBLEM4_CONTRACT_ID}:
            errors.append(
                f"problem4 contract_id={sorted(problem4)}, "
                f"ожидается [{PROBLEM4_CONTRACT_ID}]"
            )
        if delivery.file_counts.get(problem4_name) != 1:
            errors.append(
                f"problem4 содержит "
                f"{delivery.file_counts.get(problem4_name, 0)} строк, ожидается 1"
            )
    if clean is not None and PROBLEM4_CONTRACT_ID in clean:
        errors.append(f"{PROBLEM4_CONTRACT_ID} остался в clean membership")
    details.append(
        f"problem4 rows={delivery.file_counts.get(problem4_name, 0)}, "
        f"contract_id={sorted(problem4 or set())}"
    )
    details.append(
        f"{PROBLEM4_CONTRACT_ID} in clean={bool(clean and PROBLEM4_CONTRACT_ID in clean)}"
    )
    return make_result(
        "MF-02",
        "Маршрутизация договора 151350 в problem4",
        errors,
        details,
        {
            "problem4_rows": delivery.file_counts.get(problem4_name, 0),
            "problem4_contract_ids": sorted(problem4 or set()),
            "present_in_clean": bool(clean and PROBLEM4_CONTRACT_ID in clean),
        },
        pass_summary="151350 единожды в problem4 и отсутствует в clean",
    )


def check_problem123_baseline(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    facts = ctx.facts()
    baseline = ctx.baseline_contracts()
    metrics: dict[str, Any] = {}
    for number in (1, 2, 3):
        name = PROBLEM_FILES[number]
        actual = delivery.file_sets.get(name)
        expected_summary = baseline[name]
        if actual is None:
            errors.append(f"problem{number}: нет {name}")
            continue
        missing = sorted(expected_summary.contract_ids - actual)
        unexpected = sorted(actual - expected_summary.contract_ids)
        duplicates = delivery.duplicate_contracts.get(name, [])
        if missing:
            errors.append(
                f"problem{number}: потеряны contract_id: {format_sample(missing)}"
            )
        if unexpected:
            errors.append(
                f"problem{number}: добавлены contract_id: {format_sample(unexpected)}"
            )
        if duplicates:
            errors.append(
                f"problem{number}: дубли contract_id: {format_sample(duplicates)}"
            )

        baseline_path = ctx.baseline_dir / name
        actual_path = ctx.output_dir / name
        (
            baseline_headers,
            baseline_header_rows,
            baseline_rows,
            baseline_row_duplicates,
        ) = read_keyed_xlsx(baseline_path, "contract_id", 2)
        (
            actual_headers,
            actual_header_rows,
            actual_rows,
            actual_row_duplicates,
        ) = read_keyed_xlsx(actual_path, "contract_id", 2)
        if baseline_headers != actual_headers or baseline_header_rows != actual_header_rows:
            errors.append(f"problem{number}: headers differ from baseline")
        if baseline_row_duplicates or actual_row_duplicates:
            errors.append(
                f"problem{number}: keyed row duplicates; "
                f"baseline={sorted(baseline_row_duplicates)}, "
                f"actual={sorted(actual_row_duplicates)}"
            )
        indexes = {header: index for index, header in enumerate(actual_headers)}
        require_headers(
            actual_path,
            indexes,
            [
                "contract_id",
                "create_date",
                "client_id",
                "phone",
                "client_fio",
                "card",
                "manager",
                *sorted(FINANCIAL_MEMBERSHIP_FIELDS),
            ],
        )

        changes: list[dict[str, Any]] = []
        for contract_id in sorted(set(baseline_rows) & set(actual_rows)):
            old_row = baseline_rows[contract_id]
            new_row = actual_rows[contract_id]
            for index, (old_value, new_value) in enumerate(zip(old_row, new_row)):
                if old_value == new_value:
                    continue
                changes.append(
                    {
                        "contract_id": contract_id,
                        "field": actual_headers[index],
                        "old": old_value,
                        "new": new_value,
                    }
                )

        # The expected date changes are derived independently from baseline
        # and the new SQL facts.  This pins both the exact contracts and their
        # target values without trusting the newly generated problem XLSX.
        expected_date_contracts: set[str] = set()
        for contract_id, old_row in baseline_rows.items():
            fact = facts.by_contract.get(normalize_contract_id(contract_id))
            if fact is None:
                errors.append(
                    f"problem{number} {contract_id}: no membership SQL fact"
                )
                continue
            old_create_date = as_date(old_row[indexes["create_date"]])
            if old_create_date != fact.sale_date:
                expected_date_contracts.add(normalize_contract_id(contract_id))
        expected_date_count = EXPECTED_PROBLEM_CREATE_DATE_CHANGES[number]
        if len(expected_date_contracts) != expected_date_count:
            errors.append(
                f"problem{number}: baseline/staging create_date delta="
                f"{len(expected_date_contracts)}, ожидается {expected_date_count}"
            )

        expected_change_keys = {
            (contract_id, "create_date")
            for contract_id in expected_date_contracts
        }
        expected_owner_values = EXPECTED_PROBLEM_OWNER_VALUES.get(number, {})
        for contract_id, expected_fields in expected_owner_values.items():
            old_row = baseline_rows.get(contract_id)
            if old_row is None:
                errors.append(
                    f"problem{number}: owner-control {contract_id} absent in baseline"
                )
                continue
            new_row = actual_rows.get(contract_id)
            if new_row is None:
                errors.append(
                    f"problem{number}: owner-control {contract_id} absent in actual"
                )
                continue
            for field_name, expected_value in expected_fields.items():
                observed_value = clean_text(new_row[indexes[field_name]])
                if observed_value != expected_value:
                    errors.append(
                        f"problem{number} {contract_id}.{field_name}="
                        f"{observed_value!r}, expected={expected_value!r}"
                    )
                if clean_text(old_row[indexes[field_name]]) != expected_value:
                    expected_change_keys.add((contract_id, field_name))

        actual_change_keys = {
            (change["contract_id"], change["field"]) for change in changes
        }
        missing_changes = sorted(expected_change_keys - actual_change_keys)
        forbidden_changes = sorted(actual_change_keys - expected_change_keys)
        if missing_changes:
            errors.append(
                f"problem{number}: expected semantic changes absent: "
                f"{format_sample(missing_changes)}"
            )
        if forbidden_changes:
            errors.append(
                f"problem{number}: forbidden non-owner/date changes: "
                f"{format_sample(forbidden_changes)}"
            )

        bad_create_values: list[str] = []
        for contract_id in sorted(expected_date_contracts):
            actual_row = actual_rows.get(contract_id)
            fact = facts.by_contract.get(contract_id)
            if actual_row is None or fact is None:
                continue
            observed_date = as_date(actual_row[indexes["create_date"]])
            if observed_date != fact.sale_date:
                bad_create_values.append(
                    f"{contract_id}: actual={observed_date}, staging={fact.sale_date}"
                )
        if bad_create_values:
            errors.append(
                f"problem{number}: create_date does not equal staging sale_date: "
                f"{format_sample(bad_create_values)}"
            )

        financial_changes = [
            change
            for change in changes
            if change["field"] in FINANCIAL_MEMBERSHIP_FIELDS
        ]
        if financial_changes:
            errors.append(
                f"problem{number}: financial semantics changed: "
                f"{format_sample(financial_changes)}"
            )
        field_counts = Counter(change["field"] for change in changes)
        changed_contracts = {change["contract_id"] for change in changes}
        metrics[f"problem{number}"] = {
            "actual_unique": len(actual),
            "baseline_unique": len(expected_summary.contract_ids),
            "missing": missing,
            "unexpected": unexpected,
            "duplicates": duplicates,
            "changed_cells": len(changes),
            "changed_contracts": len(changed_contracts),
            "field_change_counts": dict(field_counts),
            "expected_create_date_contracts": sorted(expected_date_contracts),
            "owner_expected_values": expected_owner_values,
            "missing_expected_changes": missing_changes,
            "forbidden_changes": forbidden_changes,
            "financial_changes": financial_changes,
            "all_changes": changes,
        }
        details.append(
            f"problem{number}: contracts actual/baseline={len(actual)}/"
            f"{len(expected_summary.contract_ids)}, changed cells={len(changes)}, "
            f"changed contracts={len(changed_contracts)}, fields={dict(field_counts)}"
        )
        details.append(
            f"problem{number}: financial semantics unchanged="
            f"{not financial_changes}; forbidden changes={len(forbidden_changes)}"
        )
    return make_result(
        "MF-03",
        "Состав и financial semantics problem1–problem3",
        errors,
        details,
        metrics,
        pass_summary=(
            "contract_id совпадают; financial semantics неизменна; "
            "дельта ровно create_date 9/27/147 + owner fields 150337/137855"
        ),
    )


def validate_selected_membership_row(
    delivery: MembershipDeliverySummary,
    contract_id: str,
    expected: dict[str, str],
    errors: list[str],
    details: list[str],
    *,
    label: str,
) -> None:
    rows = [
        row
        for row in delivery.selected_rows.get(contract_id, [])
        if row.source_file == MEMBERSHIP_FILE
    ]
    if len(rows) != 1:
        errors.append(
            f"{label} {contract_id}: в clean найдено {len(rows)} строк, ожидается 1"
        )
        return
    row = rows[0]
    details.append(
        f"{contract_id}: client_id={row.client_id}, client_fio={row.client_fio!r}, "
        f"row={row.source_file}:{row.row_number}"
    )
    if row.client_id != expected["client_id"]:
        errors.append(
            f"{label} {contract_id}: client_id={row.client_id}, "
            f"ожидается {expected['client_id']}"
        )
    if normalized_text(row.client_fio) != normalized_text(expected["client_fio"]):
        errors.append(
            f"{label} {contract_id}: client_fio={row.client_fio!r}, "
            f"ожидается {expected['client_fio']!r}"
        )


def check_owner_contracts(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    for contract_id, expected in OWNER_CASES.items():
        validate_selected_membership_row(
            delivery,
            contract_id,
            expected,
            errors,
            details,
            label="перенос владельца",
        )
    for contract_id, expected in OLD_OWNER_CONTROL_CONTRACTS.items():
        validate_selected_membership_row(
            delivery,
            contract_id,
            expected,
            errors,
            details,
            label="собственный договор старого владельца",
        )
    return make_result(
        "MF-04",
        "Владельцы договоров 133547 и 144947",
        errors,
        details,
        pass_summary="оба переноса применены, собственные договоры старых владельцев сохранены",
    )


def read_application_controls(path: Path) -> dict[str, list[dict[str, str]]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    wanted = set(APPLICATION_CARD_CONTROLS)
    result: dict[str, list[dict[str, str]]] = defaultdict(list)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers, indexes = xlsx_headers(worksheet)
        require_headers(path, indexes, ["client_id", "client_fio", "funnel", "funnel_step"])
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            client_id = normalize_client_id(values[indexes["client_id"]])
            if client_id not in wanted:
                continue
            record = {header: clean_text(values[index]) for index, header in enumerate(headers)}
            record["_row"] = str(row_number)
            result[client_id].append(record)
    finally:
        workbook.close()
    return result


def read_card_controls(path: Path) -> list[dict[str, str]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    result: list[dict[str, str]] = []
    wanted_names = {
        normalized_text(spec["client_fio"]) for spec in APPLICATION_CARD_CONTROLS.values()
    }
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        _, indexes = xlsx_headers(worksheet)
        require_headers(path, indexes, ["фио", "номер пластиковой карты"])
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            fio = clean_text(values[indexes["фио"]])
            if normalized_text(fio) not in wanted_names:
                continue
            result.append(
                {
                    "client_fio": fio,
                    "card": clean_text(values[indexes["номер пластиковой карты"]]),
                    "row": str(row_number),
                }
            )
    finally:
        workbook.close()
    return result


def check_owner_applications_cards(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    applications = read_application_controls(ctx.output_dir / APPLICATIONS_FILE)
    cards = read_card_controls(ctx.output_dir / CARDS_FILE)
    metrics: dict[str, Any] = {"applications": {}, "cards": {}}
    active_funnel = normalized_text("Действующие абонементы")
    active_step = normalized_text("Все действующие абонементы")
    for client_id, expected in APPLICATION_CARD_CONTROLS.items():
        rows = applications.get(client_id, [])
        good_application_rows = [
            row
            for row in rows
            if normalized_text(row.get("client_fio"))
            == normalized_text(expected["client_fio"])
            and normalized_text(row.get("funnel")) == active_funnel
            and normalized_text(row.get("funnel_step")) == active_step
        ]
        matching_cards = [
            row
            for row in cards
            if normalized_text(row["client_fio"])
            == normalized_text(expected["client_fio"])
            and row["card"] == expected["card"]
        ]
        metrics["applications"][client_id] = rows
        metrics["cards"][client_id] = matching_cards
        details.append(
            f"{client_id} ({expected['role']}): active applications={len(good_application_rows)}, "
            f"cards={len(matching_cards)}"
        )
        if len(good_application_rows) != 1:
            observed = [
                f"row={row.get('_row')}, fio={row.get('client_fio')!r}, "
                f"funnel={row.get('funnel')!r}, step={row.get('funnel_step')!r}"
                for row in rows
            ]
            errors.append(
                f"{client_id} ({expected['role']}): ожидается ровно одна "
                f"активная заявка; найдено {len(good_application_rows)}; "
                f"все строки={observed}"
            )
        if len(matching_cards) != 1:
            observed_cards = [
                row
                for row in cards
                if normalized_text(row["client_fio"])
                == normalized_text(expected["client_fio"])
            ]
            errors.append(
                f"{client_id} ({expected['role']}): карта {expected['card']} "
                f"строго один раз не найдена; строки={observed_cards}"
            )
    return make_result(
        "MF-05",
        "Заявки и карты после смены владельцев",
        errors,
        details,
        metrics,
        pass_summary="новые и прежние владельцы есть в активных заявках и картах",
    )


def check_membership_create_dates(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    facts = ctx.facts()
    delivery = ctx.delivery()
    if facts.duplicate_contracts:
        errors.append(
            f"staging содержит дубли document_number: "
            f"{format_sample(sorted(facts.duplicate_contracts))}"
        )
    if facts.conflicting_contracts:
        errors.append(
            f"staging содержит противоречивые sale_date: "
            f"{format_sample(sorted(facts.conflicting_contracts))}"
        )
    if delivery.missing_staging_facts:
        errors.append(
            f"для {len(delivery.missing_staging_facts)} реальных XLSX-строк нет "
            f"staging fact: {format_sample(delivery.missing_staging_facts)}"
        )
    if delivery.invalid_create_dates:
        errors.append(
            f"невалидный create_date в {len(delivery.invalid_create_dates)} строках: "
            f"{format_sample(delivery.invalid_create_dates)}"
        )
    if delivery.mismatched_create_dates:
        errors.append(
            f"create_date != staging.sale_date в "
            f"{len(delivery.mismatched_create_dates)} строках: "
            f"{format_sample(delivery.mismatched_create_dates)}"
        )
    if delivery.contract_client_mismatches:
        errors.append(
            f"xlsx.client_id != staging.client_id в "
            f"{len(delivery.contract_client_mismatches)} строках: "
            f"{format_sample(delivery.contract_client_mismatches)}"
        )
    details.extend(
        [
            f"real membership rows={delivery.real_row_count}",
            f"matched create_date/sale_date={delivery.matched_create_dates}",
            f"missing staging facts={len(delivery.missing_staging_facts)}",
            f"invalid create_date={len(delivery.invalid_create_dates)}",
            f"date mismatches={len(delivery.mismatched_create_dates)}",
            f"client_id mismatches={len(delivery.contract_client_mismatches)}",
        ]
    )
    return make_result(
        "MF-06",
        "create_date реальных абонементов",
        errors,
        details,
        {
            "real_rows": delivery.real_row_count,
            "matched": delivery.matched_create_dates,
            "missing_staging_fact_count": len(delivery.missing_staging_facts),
            "invalid_create_date_count": len(delivery.invalid_create_dates),
            "mismatch_count": len(delivery.mismatched_create_dates),
            "client_id_mismatch_count": len(delivery.contract_client_mismatches),
            "missing_staging_fact_sample": sample(delivery.missing_staging_facts),
            "invalid_create_date_sample": sample(delivery.invalid_create_dates),
            "mismatch_sample": sample(delivery.mismatched_create_dates),
            "client_id_mismatch_sample": sample(
                delivery.contract_client_mismatches
            ),
        },
        pass_summary="create_date всех реальных строк равен sale_date из нового staging",
    )


def read_membership_templates(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    result: list[dict[str, Any]] = []
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers, indexes = xlsx_headers(worksheet)
        require_headers(path, indexes, ["name", "duration", "visits"])
        for row_number, values in enumerate(
            worksheet.iter_rows(min_row=3, values_only=True), start=3
        ):
            record = {header: values[index] for index, header in enumerate(headers)}
            record["_row"] = row_number
            result.append(record)
    finally:
        workbook.close()
    return result


def expected_cycle_visits(name: str) -> int | None:
    match = re.search(r"(?<!\d)(8|12)\s*(?:пос|посещ)", normalized_text(name))
    return int(match.group(1)) if match else None


def check_cycle_templates(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    rows = [
        row
        for row in read_membership_templates(
            ctx.output_dir / MEMBERSHIP_TEMPLATES_FILE
        )
        if is_cycle_name(row.get("name"))
    ]
    if len(rows) != 4:
        errors.append(f"шаблонов САЙКЛ={len(rows)}, ожидается 4")
    observed: list[dict[str, Any]] = []
    for row in rows:
        name = clean_text(row.get("name"))
        expected = expected_cycle_visits(name)
        actual = as_decimal(row.get("visits"))
        observed.append(
            {
                "row": row["_row"],
                "name": name,
                "expected_visits": expected,
                "actual_visits": clean_text(row.get("visits")),
            }
        )
        details.append(
            f"row {row['_row']}: {name!r}, visits={row.get('visits')!r}, expected={expected}"
        )
        if expected not in (8, 12):
            errors.append(f"не удалось вывести 8/12 из имени шаблона {name!r}")
        elif actual != Decimal(expected):
            errors.append(
                f"шаблон {name!r}: visits={row.get('visits')!r}, "
                f"ожидается {expected}"
            )
    return make_result(
        "MF-07",
        "Посещения в шаблонах САЙКЛ",
        errors,
        details,
        {"cycle_template_count": len(rows), "templates": observed},
        pass_summary="все 4 шаблона САЙКЛ содержат visits=8/12 по названию",
    )


def check_cycle_memberships(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    blank_rows = [
        f"{row.contract_id} ({row.source_file}:{row.row_number})"
        for row in delivery.cycle_rows
        if is_blank(row.visits_left)
    ]
    if not delivery.cycle_rows:
        errors.append("не найдено ни одной реальной строки САЙКЛ")
    if blank_rows:
        errors.append(
            f"visits_left пуст в {len(blank_rows)} строках САЙКЛ: "
            f"{format_sample(blank_rows)}"
        )
    details.append(f"строк САЙКЛ={len(delivery.cycle_rows)}")
    details.append(f"пустой visits_left={len(blank_rows)}")
    return make_result(
        "MF-08",
        "visits_left всех абонементов САЙКЛ",
        errors,
        details,
        {
            "cycle_rows": len(delivery.cycle_rows),
            "blank_visits_left_count": len(blank_rows),
            "blank_visits_left_sample": sample(blank_rows),
        },
        pass_summary="visits_left заполнен во всех реальных строках САЙКЛ",
    )


def check_cycle_balances(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    metrics: dict[str, Any] = {}
    delivery = ctx.delivery()
    facts = ctx.facts()
    for contract_id, expected_balance in CYCLE_BALANCE_CONTROLS.items():
        rows = [
            row
            for row in delivery.selected_rows.get(contract_id, [])
            if row.source_file == MEMBERSHIP_FILE
        ]
        fact = facts.by_contract.get(contract_id)
        observed_balance = None
        if len(rows) != 1:
            errors.append(
                f"{contract_id}: в clean найдено {len(rows)} строк, ожидается 1"
            )
        else:
            observed_balance = as_decimal(rows[0].visits_left)
            if observed_balance != Decimal(expected_balance):
                errors.append(
                    f"{contract_id}: visits_left={rows[0].visits_left!r}, "
                    f"ожидается {expected_balance}"
                )
            if not is_cycle_name(rows[0].contract_name):
                errors.append(
                    f"{contract_id}: контрольная строка не распознана как САЙКЛ: "
                    f"{rows[0].contract_name!r}"
                )
        if fact is None:
            errors.append(f"{contract_id}: нет в membership staging")
            active = None
        else:
            active = fact.is_active_on_cutoff
            if clean_text(active) != "1":
                errors.append(
                    f"{contract_id}: staging.is_active_on_cutoff={active!r}, ожидается '1'"
                )
        metrics[contract_id] = {
            "expected_visits_left": expected_balance,
            "actual_visits_left": None if observed_balance is None else str(observed_balance),
            "staging_is_active_on_cutoff": active,
        }
        details.append(
            f"{contract_id}: visits_left={observed_balance}, expected={expected_balance}, active={active}"
        )
    return make_result(
        "MF-09",
        "Контрольные остатки восьми активных САЙКЛ",
        errors,
        details,
        metrics,
        pass_summary="8/8 контрольных остатков точно совпали",
    )


def check_shuleyko(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    rows = [
        row
        for row in delivery.selected_rows.get(SHULEYKO_CONTRACT_ID, [])
        if row.source_file == MEMBERSHIP_FILE
    ]
    membership: MembershipRow | None = rows[0] if len(rows) == 1 else None
    if len(rows) != 1:
        errors.append(
            f"{SHULEYKO_CONTRACT_ID}: в clean найдено {len(rows)} строк, ожидается 1"
        )
    if membership is not None:
        observations = {
            "client_fio": membership.client_fio,
            "create_date": str(as_date(membership.create_date)),
            "duration": clean_text(membership.duration),
            "price": clean_text(membership.price),
            "contract_name": membership.contract_name,
        }
        details.append(f"membership: {observations}")
        if normalized_text(membership.client_fio) != normalized_text(SHULEYKO_NAME):
            errors.append(
                f"{SHULEYKO_CONTRACT_ID}: client_fio={membership.client_fio!r}, "
                f"ожидается {SHULEYKO_NAME!r}"
            )
        if as_date(membership.create_date) != SHULEYKO_SALE_DATE:
            errors.append(
                f"{SHULEYKO_CONTRACT_ID}: create_date={membership.create_date!r}, "
                f"ожидается {SHULEYKO_SALE_DATE}"
            )
        if as_decimal(membership.duration) != SHULEYKO_DURATION:
            errors.append(
                f"{SHULEYKO_CONTRACT_ID}: duration={membership.duration!r}, ожидается 12"
            )
        if as_decimal(membership.price) != SHULEYKO_PRICE:
            errors.append(
                f"{SHULEYKO_CONTRACT_ID}: price={membership.price!r}, ожидается 12990"
            )

    template_matches: list[dict[str, Any]] = []
    if membership is not None and membership.contract_name:
        template_matches = [
            row
            for row in read_membership_templates(
                ctx.output_dir / MEMBERSHIP_TEMPLATES_FILE
            )
            if normalized_text(row.get("name"))
            == normalized_text(membership.contract_name)
        ]
        if len(template_matches) != 1:
            errors.append(
                f"шаблон {membership.contract_name!r}: найдено "
                f"{len(template_matches)} строк, ожидается 1"
            )
        elif as_decimal(template_matches[0].get("duration")) != SHULEYKO_DURATION:
            errors.append(
                f"шаблон {membership.contract_name!r}: "
                f"duration={template_matches[0].get('duration')!r}, ожидается 12"
            )
        if template_matches:
            details.append(
                f"template row={template_matches[0].get('_row')}, "
                f"duration={template_matches[0].get('duration')!r}"
            )
    return make_result(
        "MF-10",
        "Абонемент Шулейко 140663 и его шаблон",
        errors,
        details,
        {
            "membership_rows": len(rows),
            "template_rows": len(template_matches),
        },
        pass_summary="140663: date=2025-09-23, duration=12, price=12990; template duration=12",
    )


def zip_member_differences(left: Path, right: Path) -> tuple[list[str], list[str], list[str]]:
    with zipfile.ZipFile(left) as left_zip, zipfile.ZipFile(right) as right_zip:
        left_names = set(left_zip.namelist())
        right_names = set(right_zip.namelist())
        changed = sorted(
            name
            for name in left_names & right_names
            if left_zip.read(name) != right_zip.read(name)
        )
    return sorted(left_names - right_names), sorted(right_names - left_names), changed


def read_keyed_xlsx(
    path: Path, key_field: str, data_start_row: int
) -> tuple[list[str], list[tuple[Any, ...]], dict[str, tuple[Any, ...]], set[str]]:
    if not path.is_file():
        raise FileNotFoundError(path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        worksheet = workbook.active
        headers, indexes = xlsx_headers(worksheet)
        require_headers(path, indexes, [key_field])
        header_rows = list(
            worksheet.iter_rows(
                min_row=1, max_row=data_start_row - 1, values_only=True
            )
        )
        counts: Counter[str] = Counter()
        keyed: dict[str, tuple[Any, ...]] = {}
        for values in worksheet.iter_rows(min_row=data_start_row, values_only=True):
            key = clean_text(values[indexes[key_field]])
            if not key:
                continue
            counts[key] += 1
            keyed.setdefault(key, tuple(values))
    finally:
        workbook.close()
    return headers, header_rows, keyed, {
        key for key, count in counts.items() if count > 1
    }


def check_service_semantics(ctx: ValidationContext) -> CheckResult:
    """Allow only the owner/final-funnel cascade proven for the new rebuild."""

    errors: list[str] = []
    details: list[str] = []
    metrics: dict[str, Any] = {}

    template_actual = ctx.output_dir / SERVICE_TEMPLATES_FILE
    template_baseline = ctx.baseline_dir / SERVICE_TEMPLATES_FILE
    missing, unexpected, changed = zip_member_differences(
        template_baseline, template_actual
    )
    template_non_core_changes = [
        name for name in changed if name != "docProps/core.xml"
    ]
    if missing or unexpected or template_non_core_changes:
        errors.append(
            f"service templates: изменения за пределами core timestamp; "
            f"missing ZIP members={missing}, unexpected={unexpected}, "
            f"changed_non_core={template_non_core_changes}"
        )
    metrics[SERVICE_TEMPLATES_FILE] = {
        "actual_sha256": sha256(template_actual),
        "baseline_sha256": sha256(template_baseline),
        "changed_zip_members": changed,
        "missing_zip_members": missing,
        "unexpected_zip_members": unexpected,
    }
    details.append(
        f"{SERVICE_TEMPLATES_FILE}: changed ZIP members={changed}; "
        "разрешён только docProps/core.xml (время пересборки)"
    )

    clients_actual = ctx.output_dir / CLIENT_SERVICES_FILE
    clients_baseline = ctx.baseline_dir / CLIENT_SERVICES_FILE
    actual_headers, actual_header_rows, actual_rows, actual_duplicates = read_keyed_xlsx(
        clients_actual, "service_id", 3
    )
    baseline_headers, baseline_header_rows, baseline_rows, baseline_duplicates = (
        read_keyed_xlsx(clients_baseline, "service_id", 3)
    )
    if actual_headers != baseline_headers or actual_header_rows != baseline_header_rows:
        errors.append("service clients: technical/Russian headers differ from baseline")
    if actual_duplicates or baseline_duplicates:
        errors.append(
            f"service clients: duplicate service_id; actual={sorted(actual_duplicates)}, "
            f"baseline={sorted(baseline_duplicates)}"
        )
    removed = set(baseline_rows) - set(actual_rows)
    added = set(actual_rows) - set(baseline_rows)
    if removed != EXPECTED_SERVICE_REMOVED:
        errors.append(
            f"service clients: removed service_id={sorted(removed)}, "
            f"expected={sorted(EXPECTED_SERVICE_REMOVED)}"
        )
    if added != set(EXPECTED_SERVICE_ADDED):
        errors.append(
            f"service clients: added service_id={sorted(added)}, "
            f"expected={sorted(EXPECTED_SERVICE_ADDED)}"
        )

    header_index = {name: index for index, name in enumerate(actual_headers)}
    shared_changes: list[dict[str, Any]] = []
    for service_id in sorted(set(actual_rows) & set(baseline_rows)):
        old_row = baseline_rows[service_id]
        new_row = actual_rows[service_id]
        for index, (old_value, new_value) in enumerate(zip(old_row, new_row)):
            if old_value == new_value:
                continue
            shared_changes.append(
                {
                    "service_id": service_id,
                    "field": actual_headers[index],
                    "old": old_value,
                    "new": new_value,
                }
            )
    expected_shared_changes = [
        {
            "service_id": service_id,
            "field": "manager",
            "old": old_manager,
            "new": new_manager,
        }
        for service_id, (old_manager, new_manager) in sorted(
            EXPECTED_SERVICE_MANAGER_CHANGES.items()
        )
    ]
    if shared_changes != expected_shared_changes:
        errors.append(
            f"service clients: shared-row changes={shared_changes}, "
            f"expected={expected_shared_changes}"
        )

    added_observations: dict[str, dict[str, Any]] = {}
    for service_id, expected in EXPECTED_SERVICE_ADDED.items():
        values = actual_rows.get(service_id)
        if values is None:
            continue
        observed = {
            "client_id": normalize_client_id(values[header_index["client_id"]]),
            "client_fio": clean_text(values[header_index["client_fio"]]),
            "service_name": clean_text(values[header_index["service_name"]]),
            "create_date": as_date(values[header_index["create_date"]]),
            "price": as_decimal(values[header_index["price"]]),
            "manager": clean_text(values[header_index["manager"]]),
            "филиал": clean_text(values[header_index["филиал"]]),
        }
        added_observations[service_id] = observed
        if observed != expected:
            errors.append(
                f"service clients: added {service_id}={observed}, expected={expected}"
            )

    _, _, client_zip_changes = zip_member_differences(
        clients_baseline, clients_actual
    )
    allowed_client_members = {"docProps/core.xml", "xl/worksheets/sheet1.xml"}
    if set(client_zip_changes) - allowed_client_members:
        errors.append(
            f"service clients: unexpected changed ZIP members={client_zip_changes}"
        )
    metrics[CLIENT_SERVICES_FILE] = {
        "actual_sha256": sha256(clients_actual),
        "baseline_sha256": sha256(clients_baseline),
        "baseline_rows": len(baseline_rows),
        "actual_rows": len(actual_rows),
        "removed_service_ids": sorted(removed),
        "added_service_ids": sorted(added),
        "shared_changes": shared_changes,
        "added_row_controls": added_observations,
        "changed_zip_members": client_zip_changes,
    }
    details.extend(
        [
            f"service clients rows: baseline={len(baseline_rows)}, actual={len(actual_rows)}",
            f"expected key replacements: removed={sorted(removed)}, added={sorted(added)}",
            f"shared-row field changes={shared_changes}",
            f"client workbook changed ZIP members={client_zip_changes}",
        ]
    )
    return make_result(
        "MF-11",
        "Семантический diff service XLSX",
        errors,
        details,
        metrics,
        pass_summary=(
            "service templates отличаются только core timestamp; service clients "
            "имеют ровно 2 ожидаемые замены ключей и 3 замены manager"
        ),
    )


def check_contract_preservation(ctx: ValidationContext) -> CheckResult:
    errors: list[str] = []
    details: list[str] = []
    delivery = ctx.delivery()
    facts = ctx.facts()
    source_client_ids = ctx.source_client_ids()
    full_membership_rows = ctx.membership_rows()
    exclusions = ctx.exclusions()
    dedupe_losers, dedupe_winners, dedupe_row_count = ctx.dedupe_evidence()
    baseline = ctx.baseline_contracts()
    baseline_clean = baseline[MEMBERSHIP_FILE].contract_ids
    baseline_union = set(baseline_clean)
    baseline_clients = dict(baseline[MEMBERSHIP_FILE].client_ids)
    for number in (1, 2, 3):
        baseline_union.update(baseline[PROBLEM_FILES[number]].contract_ids)
        baseline_clients.update(baseline[PROBLEM_FILES[number]].client_ids)
    actual_clean = delivery.file_sets.get(MEMBERSHIP_FILE, set())

    # First prove that the delivery is an exact partition of the newly built
    # full membership CSV.  Baseline is not the source of truth after owner
    # changes because phone-dedupe winners can legitimately change.
    problem_union = set()
    for name in PROBLEM_FILES.values():
        problem_union.update(delivery.file_sets.get(name, set()))
    expected_clean = full_membership_rows.contract_ids - problem_union
    partition_missing = sorted(full_membership_rows.contract_ids - delivery.union)
    partition_unexpected = sorted(delivery.union - full_membership_rows.contract_ids)
    clean_missing = sorted(expected_clean - actual_clean)
    clean_unexpected = sorted(actual_clean - expected_clean)
    staging_delivery_client_mismatches = sorted(
        contract_id
        for contract_id in full_membership_rows.contract_ids & delivery.union
        if full_membership_rows.client_ids.get(contract_id, "")
        != delivery.contract_clients.get(contract_id, "")
    )
    delivered_applications_path = ctx.output_dir / APPLICATIONS_FILE
    owner_applications_same_bytes = (
        delivered_applications_path.is_file()
        and ctx.owner_applications_path.is_file()
        and sha256(delivered_applications_path) == sha256(ctx.owner_applications_path)
    )

    if partition_missing:
        errors.append(
            f"из full-new membership потеряны contract_id: "
            f"{format_sample(partition_missing)}"
        )
    if partition_unexpected:
        errors.append(
            f"в delivery есть contract_id, которых нет в full-new membership: "
            f"{format_sample(partition_unexpected)}"
        )
    if clean_missing or clean_unexpected:
        errors.append(
            "clean не равен full-new membership минус problem1–problem4: "
            f"missing={format_sample(clean_missing)}, "
            f"unexpected={format_sample(clean_unexpected)}"
        )
    if full_membership_rows.duplicates:
        errors.append(
            f"membership_import_rows.csv: дубли contract_id="
            f"{format_sample(sorted(full_membership_rows.duplicates))}"
        )
    if staging_delivery_client_mismatches:
        errors.append(
            "membership_import_rows.csv и delivery расходятся по client_id: "
            f"{format_sample(staging_delivery_client_mismatches)}"
        )
    if not owner_applications_same_bytes:
        errors.append(
            "owner applications XLSX, использованный для отбора membership, "
            "не совпадает побайтно с applications XLSX в delivery"
        )

    # Account for every SQL fact.  A fact may be absent only when the explicit
    # business-exclusion CSV says so, or when its effective client is outside
    # the exact applications+refusers population consumed by the builder.
    fact_ids = set(facts.by_contract)
    delivery_without_fact = sorted(delivery.union - fact_ids)
    facts_not_delivered = fact_ids - delivery.union
    allowed_exclusion_rules = {"exclude_active_later_contact_full"}
    invalid_exclusion_rules = {
        contract_id: rule
        for contract_id, rule in exclusions.items()
        if rule not in allowed_exclusion_rules
    }
    explicit_exclusions = {
        contract_id
        for contract_id in facts_not_delivered
        if contract_id in exclusions
        and exclusions[contract_id] in allowed_exclusion_rules
    }
    not_in_final_client_xlsx = {
        contract_id
        for contract_id in facts_not_delivered - explicit_exclusions
        if facts.by_contract[contract_id].client_id not in source_client_ids
    }
    unexplained_facts = sorted(
        facts_not_delivered - explicit_exclusions - not_in_final_client_xlsx
    )
    exclusions_still_delivered = sorted(set(exclusions) & delivery.union)
    exclusions_without_fact = sorted(set(exclusions) - fact_ids)

    if facts.row_count != EXPECTED_MEMBERSHIP_FACTS:
        errors.append(
            f"membership SQL facts rows={facts.row_count}, "
            f"ожидается {EXPECTED_MEMBERSHIP_FACTS}"
        )
    if len(fact_ids) != EXPECTED_MEMBERSHIP_FACTS:
        errors.append(
            f"unique membership SQL fact document_number={len(fact_ids)}, "
            f"ожидается {EXPECTED_MEMBERSHIP_FACTS}"
        )
    if delivery_without_fact:
        errors.append(
            f"в delivery есть contract_id без SQL fact: "
            f"{format_sample(delivery_without_fact)}"
        )
    if invalid_exclusion_rules:
        errors.append(f"неизвестные membership exclusion rules={invalid_exclusion_rules}")
    if exclusions_still_delivered:
        errors.append(
            f"исключённые contract_id всё ещё в delivery: "
            f"{format_sample(exclusions_still_delivered)}"
        )
    if exclusions_without_fact:
        errors.append(
            f"exclusion CSV ссылается на отсутствующие facts: "
            f"{format_sample(exclusions_without_fact)}"
        )
    if unexplained_facts:
        errors.append(
            f"SQL facts без строки и без документированной причины: "
            f"{format_sample(unexplained_facts)}"
        )

    # Baseline delta is evidence, not a forbidden change.  The owner fix
    # changes active status and therefore the winner in shared-phone
    # components.  Every old-only/new-only ID must be tied to that cascade or
    # to a real owner-change row in the new staging.
    old_only = sorted(baseline_union - delivery.union)
    new_only = sorted(delivery.union - baseline_union)
    removed_explanations: list[dict[str, Any]] = []
    unexplained_old_only: list[str] = []
    for contract_id in old_only:
        fact = facts.by_contract.get(contract_id)
        if (
            fact is not None
            and fact.client_id not in source_client_ids
            and fact.client_id in dedupe_losers
            and contract_id in not_in_final_client_xlsx
        ):
            removed_explanations.append(
                {
                    "contract_id": contract_id,
                    "baseline_client_id": baseline_clients.get(contract_id, ""),
                    "new_staging_client_id": fact.client_id,
                    "reason": "not_in_final_client_xlsx_after_phone_dedupe",
                    "dedupe_winner_client_ids": sorted(
                        dedupe_losers[fact.client_id]
                    ),
                }
            )
        else:
            unexplained_old_only.append(contract_id)

    added_explanations: list[dict[str, Any]] = []
    unexplained_new_only: list[str] = []
    for contract_id in new_only:
        fact = facts.by_contract.get(contract_id)
        if fact is None or fact.client_id not in source_client_ids:
            unexplained_new_only.append(contract_id)
            continue
        reasons: list[str] = []
        if fact.client_id in dedupe_winners:
            reasons.append("selected_phone_dedupe_winner")
        if (
            fact.owner_change_ref
            and fact.original_client_id
            and fact.original_client_id != fact.effective_client_id
        ):
            reasons.append("effective_client_changed_by_owner_document")
        if not reasons:
            unexplained_new_only.append(contract_id)
            continue
        added_explanations.append(
            {
                "contract_id": contract_id,
                "new_client_id": fact.client_id,
                "original_client_id": fact.original_client_id,
                "reason": reasons,
            }
        )

    if len(old_only) != EXPECTED_BASELINE_OLD_ONLY:
        errors.append(
            f"baseline old-only contracts={len(old_only)}, "
            f"ожидается {EXPECTED_BASELINE_OLD_ONLY}"
        )
    if len(new_only) != EXPECTED_BASELINE_NEW_ONLY:
        errors.append(
            f"baseline new-only contracts={len(new_only)}, "
            f"ожидается {EXPECTED_BASELINE_NEW_ONLY}"
        )
    if unexplained_old_only:
        errors.append(
            f"baseline old-only без phone-dedupe объяснения: "
            f"{format_sample(unexplained_old_only)}"
        )
    if unexplained_new_only:
        errors.append(
            f"baseline new-only без owner/phone-dedupe объяснения: "
            f"{format_sample(unexplained_new_only)}"
        )
    if PROBLEM4_CONTRACT_ID not in baseline_clean:
        errors.append(f"baseline clean не содержит {PROBLEM4_CONTRACT_ID}")
    if delivery.duplicate_contracts:
        errors.append(
            f"дубли contract_id внутри файлов: "
            f"{ {name: sample(values) for name, values in delivery.duplicate_contracts.items()} }"
        )
    if delivery.partition_overlaps:
        errors.append(
            f"contract_id одновременно в нескольких частях delivery: "
            f"{ {name: sample(values) for name, values in delivery.partition_overlaps.items()} }"
        )
    details.extend(
        [
            f"SQL facts={facts.row_count}; delivered={len(delivery.union)}; "
            f"not_in_final_client_xlsx={len(not_in_final_client_xlsx)}; "
            f"explicit exclusions={len(explicit_exclusions)}; unexplained={len(unexplained_facts)}",
            f"full-new CSV={len(full_membership_rows.contract_ids)}, "
            f"delivery union={len(delivery.union)}, clean={len(actual_clean)}, "
            f"problem union={len(problem_union)}",
            f"owner applications work/output same_bytes={owner_applications_same_bytes}",
            f"baseline union={len(baseline_union)}, new union={len(delivery.union)}",
            f"baseline old-only ({len(old_only)}): {', '.join(old_only)}",
            f"baseline new-only ({len(new_only)}): {', '.join(new_only)}",
            f"old-only explained by phone-dedupe={len(removed_explanations)}/{len(old_only)}",
            f"new-only explained by owner/phone-dedupe={len(added_explanations)}/{len(new_only)}",
            f"phone-dedupe evidence rows={dedupe_row_count}",
        ]
    )
    return make_result(
        "MF-12",
        "Полная трассируемость SQL facts и baseline contract delta",
        errors,
        details,
        {
            "sql_fact_rows": facts.row_count,
            "sql_fact_unique_contracts": len(fact_ids),
            "delivered_contracts": len(delivery.union),
            "documented_reasons": {
                "not_in_final_client_xlsx": sorted(not_in_final_client_xlsx),
                "explicit_exclusions": {
                    contract_id: exclusions[contract_id]
                    for contract_id in sorted(explicit_exclusions)
                },
                "unexplained": unexplained_facts,
            },
            "source_final_client_ids": len(source_client_ids),
            "full_new_membership_contracts": len(
                full_membership_rows.contract_ids
            ),
            "partition_missing": partition_missing,
            "partition_unexpected": partition_unexpected,
            "clean_missing": clean_missing,
            "clean_unexpected": clean_unexpected,
            "staging_delivery_client_mismatches": (
                staging_delivery_client_mismatches
            ),
            "owner_applications_work_output_same_bytes": (
                owner_applications_same_bytes
            ),
            "baseline_clean": len(baseline_clean),
            "actual_new_clean": len(actual_clean),
            "baseline_union": len(baseline_union),
            "actual_union": len(delivery.union),
            "baseline_old_only": old_only,
            "baseline_new_only": new_only,
            "baseline_old_only_explanations": removed_explanations,
            "baseline_new_only_explanations": added_explanations,
            "unexplained_old_only": unexplained_old_only,
            "unexplained_new_only": unexplained_new_only,
            "partition_overlaps": {
                name: values for name, values in delivery.partition_overlaps.items()
            },
        },
        pass_summary=(
            "101436/101436 SQL facts имеют строку или причину; "
            "baseline delta 45/41 полностью объяснён owner/phone-dedupe"
        ),
    )


CHECKS: list[Callable[[ValidationContext], CheckResult]] = [
    check_cutoff,
    check_problem4_routing,
    check_problem123_baseline,
    check_owner_contracts,
    check_owner_applications_cards,
    check_membership_create_dates,
    check_cycle_templates,
    check_cycle_memberships,
    check_cycle_balances,
    check_shuleyko,
    check_service_semantics,
    check_contract_preservation,
]


def safe_check(
    function: Callable[[ValidationContext], CheckResult],
    context: ValidationContext,
) -> CheckResult:
    try:
        return function(context)
    except Exception as exc:
        check_number = CHECKS.index(function) + 1
        return CheckResult(
            check_id=f"MF-{check_number:02d}",
            title=function.__name__,
            status="FAIL",
            summary="проверка не выполнена из-за ошибки входных данных",
            errors=[f"{type(exc).__name__}: {exc}"],
        )


def markdown_escape(value: Any) -> str:
    return clean_text(value).replace("|", "\\|").replace("\n", " ")


def render_markdown(
    verdict: str,
    generated_at: str,
    context: ValidationContext,
    checks: list[CheckResult],
) -> str:
    pass_count = sum(check.status == "PASS" for check in checks)
    fail_count = len(checks) - pass_count
    lines = [
        "# Бизнес-валидация manager fixes на срез 2026-06-30",
        "",
        f"- Вердикт: **{verdict}**",
        f"- Пройдено: `{pass_count}`; не пройдено: `{fail_count}`",
        f"- Ожидаемый cutoff_at: `{EXPECTED_CUTOFF_TEXT}`",
        f"- Новая delivery: `{context.output_dir}`",
        f"- Work/staging: `{context.work_dir}`",
        f"- Baseline: `{context.baseline_dir}`",
        f"- Сформировано UTC: `{generated_at}`",
        "",
        "## Сводка",
        "",
        "| ID | Статус | Проверка | Итог |",
        "| --- | --- | --- | --- |",
    ]
    for check in checks:
        lines.append(
            f"| {check.check_id} | **{check.status}** | "
            f"{markdown_escape(check.title)} | {markdown_escape(check.summary)} |"
        )
    lines.extend(["", "## Подробности", ""])
    for check in checks:
        lines.extend(
            [
                f"### {check.check_id} — {check.title}: {check.status}",
                "",
                f"{check.summary}.",
                "",
            ]
        )
        if check.errors:
            lines.append("Ошибки:")
            lines.append("")
            lines.extend(f"- {error}" for error in check.errors)
            lines.append("")
        if check.details:
            lines.append("Наблюдения:")
            lines.append("")
            lines.extend(f"- {detail}" for detail in check.details)
            lines.append("")
    return "\n".join(lines).rstrip() + "\n"


def resolve_cli_path(value: str) -> Path:
    path = Path(value).expanduser()
    return path.resolve() if not path.is_absolute() else path


def validate(args: argparse.Namespace) -> int:
    output_dir = resolve_cli_path(args.output_dir)
    work_dir = resolve_cli_path(args.work_dir)
    baseline_dir = resolve_cli_path(args.baseline_dir)
    report_path = resolve_cli_path(args.report)
    json_report_path = resolve_cli_path(args.json_report)
    context = ValidationContext(output_dir, work_dir, baseline_dir)
    checks = [safe_check(function, context) for function in CHECKS]
    verdict = "PASS" if all(check.status == "PASS" for check in checks) else "FAIL"
    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()

    payload = {
        "verdict": verdict,
        "generated_at_utc": generated_at,
        "expected": {
            "cutoff_at": EXPECTED_CUTOFF_TEXT,
            "cutoff_date": str(EXPECTED_CUTOFF_DATE),
            "date_stamp": DATE_STAMP,
        },
        "inputs": {
            "output_dir": str(output_dir),
            "work_dir": str(work_dir),
            "baseline_dir": str(baseline_dir),
        },
        "summary": {
            "checks": len(checks),
            "passed": sum(check.status == "PASS" for check in checks),
            "failed": sum(check.status == "FAIL" for check in checks),
        },
        "checks": [asdict(check) for check in checks],
    }

    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text(
        render_markdown(verdict, generated_at, context, checks), encoding="utf-8"
    )
    json_report_path.parent.mkdir(parents=True, exist_ok=True)
    json_report_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )

    print(f"verdict={verdict}")
    print(f"checks_passed={payload['summary']['passed']}")
    print(f"checks_failed={payload['summary']['failed']}")
    print(f"report={report_path}")
    print(f"json_report={json_report_path}")
    for check in checks:
        print(f"{check.check_id}={check.status}: {check.summary}")
    return 0 if verdict == "PASS" else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", required=True, help="new final delivery directory")
    parser.add_argument("--work-dir", required=True, help="new run work/staging root")
    parser.add_argument("--baseline-dir", required=True, help="frozen prior delivery")
    parser.add_argument("--report", required=True, help="Markdown report path")
    parser.add_argument("--json-report", required=True, help="JSON report path")
    return parser.parse_args()


if __name__ == "__main__":
    raise SystemExit(validate(parse_args()))
