#!/usr/bin/env python3
"""Build a business review workbook for active full overlap no-payment cash rows."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"


ACTION_BY_GROUP = {
    "A_status_refusal_or_booking_no_payment_no_visits": (
        "candidate_auto_exclude_if_confirmed",
        "Подтвердить удаление: статус Отказ/Бронь, платежа нет, посещений нет, есть другое active full.",
    ),
    "B_duplicate_same_name_dates_no_payment": (
        "manual_dedupe_decision",
        "Дубль с тем же названием и датами: какую строку оставить или удалить весь дубль?",
    ),
    "C_duplicate_same_dates_different_name_no_payment": (
        "manual_dedupe_decision",
        "Дубль на те же даты, но другое название: какая строка правильная?",
    ),
    "D_future_start_no_payment_overlap": (
        "manual_future_sale_payment_check",
        "Будущий старт после cutoff: это реальная будущая продажа без платежа или лишняя строка?",
    ),
    "E_started_no_payment_no_visits_other_has_basis": (
        "manual_possible_extra_sale",
        "Похоже на Попову: старт уже был, платежа/посещений нет, другое членство дает основание входа. Удалять?",
    ),
    "F_candidate_has_visits_do_not_auto_delete": (
        "keep_do_not_auto_delete",
        "По спорному абонементу есть посещения. Автоматически не удалять.",
    ),
    "G_needs_manual_context": (
        "manual_context_needed",
        "Платежей/посещений нет и у спорной, и у выбранной другой строки. Нужен контекст 1С.",
    ),
}


SUMMARY_ORDER = [
    "A_status_refusal_or_booking_no_payment_no_visits",
    "B_duplicate_same_name_dates_no_payment",
    "C_duplicate_same_dates_different_name_no_payment",
    "D_future_start_no_payment_overlap",
    "E_started_no_payment_no_visits_other_has_basis",
    "F_candidate_has_visits_do_not_auto_delete",
    "G_needs_manual_context",
]


REVIEW_HEADERS = [
    "risk_group",
    "recommended_action",
    "question_for_business",
    "business_decision",
    "business_comment",
    "contract_id",
    "client_id",
    "client_fio",
    "contract_name",
    "status",
    "sale_date",
    "start_date",
    "end_date",
    "price",
    "paid_candidate",
    "payment_count_candidate",
    "matched_payment_ref",
    "nearby_payment_docs_14d",
    "nearby_payment_total_14d",
    "linked_sale_docs",
    "linked_sale_sum_fld1140",
    "linked_sale_sum_fld1154",
    "linked_sale_sum_fld1160",
    "candidate_visit_docs",
    "candidate_first_visit",
    "candidate_last_visit",
    "other_active_full_count_in_final",
    "best_other_contract_id",
    "best_other_contract_name",
    "best_other_status",
    "best_other_start_date",
    "best_other_end_date",
    "best_other_has_payment",
    "best_other_payment_amount",
    "best_other_payment_method",
    "best_other_visit_docs",
    "best_other_first_visit",
    "best_other_last_visit",
    "funnel",
    "funnel_step",
    "final_funnel_selected_contract_id",
    "final_funnel_selected_subscription_name",
    "final_funnel_selected_start_date",
    "final_funnel_selected_end_date",
    "final_funnel_active_full_subscription_count",
    "final_funnel_validation_status",
    "target_is_final_funnel_selected",
    "best_other_is_final_funnel_selected",
    "selected_card_number",
    "target_card",
    "best_other_card",
    "relation_to_best_other",
    "same_card_as_best_other",
    "sibling_no_payment_cash_active_count",
    "same_name_dates_sibling_count",
    "same_dates_sibling_count",
    "doc_posted",
    "doc_marked",
    "normalized_club",
    "manual_comment_if_any",
]


def as_abs(path: str | Path) -> Path:
    path = Path(path)
    return path if path.is_absolute() else ROOT / path


def read_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def action_for(row: dict[str, str]) -> tuple[str, str]:
    return ACTION_BY_GROUP.get(row["risk_group"], ("manual_context_needed", "Нужно ручное решение."))


def autosize(ws) -> None:
    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 55)


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def append_table(ws, headers: list[str], rows: list[list[Any]]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    autosize(ws)


def build_workbook(rows: list[dict[str, str]]) -> Workbook:
    wb = Workbook()
    summary = wb.active
    summary.title = "summary"

    group_counts = Counter(row["risk_group"] for row in rows)
    visit_count = sum(1 for row in rows if int(row.get("candidate_visit_docs") or 0) > 0)
    manual_comment_count = sum(1 for row in rows if row.get("manual_comment_if_any"))
    nearby_payment_count = sum(1 for row in rows if int(row.get("nearby_payment_docs_14d") or 0) > 0)
    target_selected_count = sum(1 for row in rows if int(row.get("target_is_final_funnel_selected") or 0) == 1)
    best_other_selected_count = sum(
        1 for row in rows if int(row.get("best_other_is_final_funnel_selected") or 0) == 1
    )

    summary_rows = [
        ["metric", "value"],
        ["total_rows", len(rows)],
        ["clients", len({row["client_id"] for row in rows})],
        ["candidate_visit_docs_positive", visit_count],
        ["candidate_visit_docs_zero", len(rows) - visit_count],
        ["manual_comments_present", manual_comment_count],
        ["rows_with_nearby_payment_docs_14d", nearby_payment_count],
        ["target_is_final_funnel_selected", target_selected_count],
        ["best_other_is_final_funnel_selected", best_other_selected_count],
    ]
    for row in summary_rows:
        summary.append(row)
    style_header(summary)
    autosize(summary)

    groups = wb.create_sheet("groups")
    group_rows = []
    for group in SUMMARY_ORDER:
        action, question = ACTION_BY_GROUP[group]
        group_rows.append([group, group_counts.get(group, 0), action, question])
    append_table(groups, ["risk_group", "rows", "recommended_action", "question_for_business"], group_rows)

    review = wb.create_sheet("review_63")
    review_rows: list[list[Any]] = []
    for row in rows:
        action, question = action_for(row)
        review_rows.append([
            row.get("risk_group", ""),
            action,
            question,
            "",
            "",
            row.get("contract_id", ""),
            row.get("client_id", ""),
            row.get("client_fio", ""),
            row.get("contract_name", ""),
            row.get("status", ""),
            row.get("sale_date", ""),
            row.get("start_date", ""),
            row.get("end_date", ""),
            row.get("price", ""),
            row.get("paid_candidate", ""),
            row.get("payment_count_candidate", ""),
            row.get("matched_payment_ref", ""),
            int(row.get("nearby_payment_docs_14d") or 0),
            row.get("nearby_payment_total_14d", ""),
            int(row.get("linked_sale_docs") or 0),
            row.get("linked_sale_sum_fld1140", ""),
            row.get("linked_sale_sum_fld1154", ""),
            row.get("linked_sale_sum_fld1160", ""),
            int(row.get("candidate_visit_docs") or 0),
            row.get("candidate_first_visit", ""),
            row.get("candidate_last_visit", ""),
            int(row.get("other_active_full_count_in_final") or 0),
            row.get("best_other_contract_id", ""),
            row.get("best_other_contract_name", ""),
            row.get("best_other_status", ""),
            row.get("best_other_start_date", ""),
            row.get("best_other_end_date", ""),
            int(row.get("best_other_has_payment") or 0),
            row.get("best_other_payment_amount", ""),
            row.get("best_other_payment_method", ""),
            int(row.get("best_other_visit_docs") or 0),
            row.get("best_other_first_visit", ""),
            row.get("best_other_last_visit", ""),
            row.get("funnel", ""),
            row.get("funnel_step", ""),
            row.get("final_funnel_selected_contract_id", ""),
            row.get("final_funnel_selected_subscription_name", ""),
            row.get("final_funnel_selected_start_date", ""),
            row.get("final_funnel_selected_end_date", ""),
            row.get("final_funnel_active_full_subscription_count", ""),
            row.get("final_funnel_validation_status", ""),
            int(row.get("target_is_final_funnel_selected") or 0),
            int(row.get("best_other_is_final_funnel_selected") or 0),
            row.get("selected_card_number", ""),
            row.get("target_card", ""),
            row.get("best_other_card", ""),
            row.get("relation_to_best_other", ""),
            int(row.get("same_card_as_best_other") or 0),
            int(row.get("sibling_no_payment_cash_active_count") or 0),
            int(row.get("same_name_dates_sibling_count") or 0),
            int(row.get("same_dates_sibling_count") or 0),
            row.get("doc_posted", ""),
            row.get("doc_marked", ""),
            row.get("normalized_club", ""),
            row.get("manual_comment_if_any", ""),
        ])
    append_table(review, REVIEW_HEADERS, review_rows)

    legend = wb.create_sheet("legend")
    legend_rows = [
        ["field", "meaning"],
        ["business_decision", "Колонка для ручного ответа: delete / keep / dedupe_keep_this / dedupe_delete_this / unclear."],
        ["business_comment", "Свободный комментарий бизнеса."],
        ["nearby_payment_docs_14d", "Posted платежи клиента в окне sale_datetime +/- 14 дней. Для всех 63 строк SQL дал 0."],
        ["linked_sale_docs", "Связанные документы продажи Document154 по строке абонемента."],
        ["linked_sale_sum_fld1140/1154/1160", "Суммы из строки продажи Document154; они не являются доказательством оплаты."],
        ["candidate_visit_docs", "Посещения по спорному абонементу из dbo._Document150."],
        ["best_other_*", "Лучшее другое active/not-finished full-членство клиента для сравнения."],
        ["final_funnel_selected_*", "Абонемент, который выбран в staging final_funnel_clients как основание клиента."],
        ["target_is_final_funnel_selected", "1, если спорная строка выбрана как selected_subscription в воронке клиентов."],
    ]
    for row in legend_rows:
        legend.append(row)
    style_header(legend)
    autosize(legend)

    return wb


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--audit-csv",
        default=f"output/20251115_0800_fix_owner_new_import/reports/no_payment_cash_active_full_overlap_deep_audit.csv",
    )
    parser.add_argument(
        "--output-xlsx",
        default=f"output/20251115_0800_fix_owner_new_import/no_payment_cash_active_full_overlap_review_{DATE_STAMP}.xlsx",
    )
    args = parser.parse_args()

    rows = read_rows(as_abs(args.audit_csv))
    wb = build_workbook(rows)
    output = as_abs(args.output_xlsx)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
