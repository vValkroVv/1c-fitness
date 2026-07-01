#!/usr/bin/env python3
"""Build the no-payment cash active-full overlap audit from reproducible inputs."""

from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"
CUTOFF_DATE = date(2026, 5, 25)

FACT_FIELDS = [
    "client_ref",
    "client_id",
    "original_client_ref",
    "original_client_id",
    "original_client_fio",
    "effective_client_ref",
    "effective_client_id",
    "effective_client_fio",
    "owner_change_ref",
    "owner_change_number",
    "owner_change_datetime",
    "owner_change_old_client_ref",
    "owner_change_new_client_ref",
    "owner_change_modifier_name",
    "owner_change_count_for_membership",
    "subscription_ref",
    "document_number",
    "holder_client_ref",
    "payer_client_ref",
    "client_role_source",
    "product_ref",
    "product_code",
    "subscription_name",
    "product_class",
    "is_full_subscription",
    "is_trial_or_guest",
    "is_subrent",
    "is_limited_subrent",
    "sale_date",
    "sale_datetime",
    "start_date",
    "end_date",
    "duration_days",
    "status",
    "booking_status_ref",
    "booking_status_name",
    "doc_posted",
    "doc_marked",
    "register_duration_days",
    "is_active_on_cutoff",
    "is_finished_before_cutoff",
    "days_to_end",
    "days_since_end",
    "raw_club",
    "normalized_club",
    "club_source",
    "raw_source",
    "doc_duration_value",
    "rg_duration_days",
    "rg_freeze_days",
    "rg_guests",
    "rg_price",
    "rg_paid_candidate",
    "rg_payment_count_candidate",
    "rg_visits_candidate_8007",
    "rg_visits_candidate_8008",
    "rg_visits_candidate_8009",
    "subrent_visit_limit",
    "subrent_active_by_dates_on_cutoff",
    "subrent_finished_by_dates_before_cutoff",
    "subrent_rg3336_receipt_qty",
    "subrent_rg3336_expense_qty",
    "subrent_rg3336_signed_balance",
    "subrent_rg3336_visit_doc_expense_qty",
    "subrent_rg3336_receipt_rows",
    "subrent_rg3336_expense_rows",
    "subrent_rg3336_case_group",
    "matched_payment_ref",
    "matched_payment_datetime",
    "matched_payment_amount",
    "matched_payment_method",
    "matched_payment_operation",
    "matched_payment_match_source",
    "cutoff_at",
]

DEEP_AUDIT_FIELDS = [
    "risk_group",
    "contract_id",
    "client_id",
    "client_fio",
    "contract_name",
    "status",
    "sale_date",
    "sale_datetime",
    "start_date",
    "end_date",
    "started_by_cutoff",
    "price",
    "paid_candidate",
    "payment_count_candidate",
    "matched_payment_ref",
    "nearby_payment_docs_14d",
    "nearby_payment_total_14d",
    "linked_sale_docs",
    "linked_sale_sum_fld1140",
    "linked_sale_sum_fld1154",
    "linked_sale_sum_fld1160",
    "candidate_visit_docs",
    "candidate_first_visit",
    "candidate_last_visit",
    "other_active_full_count_in_final",
    "best_other_contract_id",
    "best_other_contract_name",
    "best_other_status",
    "best_other_start_date",
    "best_other_end_date",
    "best_other_has_payment",
    "best_other_payment_amount",
    "best_other_payment_method",
    "best_other_visit_docs",
    "best_other_first_visit",
    "best_other_last_visit",
    "funnel",
    "funnel_step",
    "final_funnel_selected_contract_id",
    "final_funnel_selected_subscription_name",
    "final_funnel_selected_start_date",
    "final_funnel_selected_end_date",
    "final_funnel_active_full_subscription_count",
    "final_funnel_validation_status",
    "target_is_final_funnel_selected",
    "best_other_is_final_funnel_selected",
    "selected_card_number",
    "target_card",
    "best_other_card",
    "relation_to_best_other",
    "same_card_as_best_other",
    "sibling_no_payment_cash_active_count",
    "same_name_dates_sibling_count",
    "same_dates_sibling_count",
    "doc_posted",
    "doc_marked",
    "normalized_club",
    "manual_comment_if_any",
]


def as_abs(path: str | Path) -> Path:
    path = Path(path)
    return path if path.is_absolute() else ROOT / path


def decimal_value(value: object) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def parse_date(value: object) -> date | None:
    if value in (None, ""):
        return None
    try:
        return datetime.strptime(str(value)[:10], "%Y-%m-%d").date()
    except ValueError:
        return None


def normalized_status(value: str) -> str:
    return (value or "").strip() or "blank"


def read_facts(path: Path) -> list[dict[str, str]]:
    facts: list[dict[str, str]] = []
    with path.open(encoding="utf-16", newline="") as f:
        for row in csv.DictReader(f, fieldnames=FACT_FIELDS, delimiter="\t"):
            if row["document_number"] and row["document_number"] != "document_number":
                facts.append(row)
    return facts


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def read_final_funnel(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    rows = read_csv(path)
    return {row["client_id"]: row for row in rows}


def read_manual_comments(path: Path) -> dict[str, str]:
    comments: dict[str, str] = {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    for values in ws.iter_rows(min_row=3, values_only=True):
        contract_id = str(values[0] or "").strip()
        if not contract_id.startswith("000"):
            continue
        tail = [str(value) for value in values[20:] if value not in (None, "")]
        if tail:
            comments[contract_id] = " | ".join(tail)
    wb.close()
    return comments


def read_visit_counts(path: Path) -> dict[str, dict[str, str]]:
    visits: dict[str, dict[str, str]] = {}
    with path.open(encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f, delimiter="|")
        for row in reader:
            if len(row) < 5 or row[0] in ("subscription_ref", "----------------"):
                continue
            try:
                int(row[2])
            except ValueError:
                continue
            visits[row[1]] = {
                "visit_docs": row[2],
                "first_visit": "" if row[3] == "NULL" else row[3],
                "last_visit": "" if row[4] == "NULL" else row[4],
            }
    return visits


def read_probe_details(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    lines = path.read_text(encoding="utf-8", errors="replace").replace("\x00", "").splitlines()
    try:
        start = lines.index("02 target details") + 1
    except ValueError:
        return {}
    headers = lines[start].split("|")
    details: dict[str, dict[str, str]] = {}
    for line in lines[start + 2 :]:
        if line.startswith("03 nearby payment details"):
            break
        if not line.strip() or set(line.replace("|", "")) <= {"-"}:
            continue
        values = line.split("|")
        if len(values) != len(headers):
            continue
        row = {key: ("" if value == "NULL" else value) for key, value in zip(headers, values)}
        details[row["document_number"]] = row
    return details


def fact_has_payment(fact: dict[str, str]) -> bool:
    return bool(fact.get("matched_payment_ref")) or decimal_value(fact.get("matched_payment_amount")) > 0


def visit_count(visits: dict[str, dict[str, str]], contract_id: str) -> int:
    return int(visits.get(contract_id, {}).get("visit_docs") or 0)


def relation_to_other(candidate: dict[str, str], other: dict[str, str]) -> str:
    start = parse_date(candidate.get("start_date"))
    end = parse_date(candidate.get("end_date"))
    other_start = parse_date(other.get("start_date"))
    other_end = parse_date(other.get("end_date"))
    if not start or not end or not other_start or not other_end:
        return "date_unknown"
    if start >= other_start and end <= other_end:
        return "candidate_inside_other"
    if start <= other_start and end >= other_end:
        return "candidate_covers_other"
    if other_start <= start <= other_end:
        return "candidate_starts_inside_other"
    if start <= other_start <= end:
        return "other_starts_inside_candidate"
    return "no_date_overlap_but_both_not_finished"


def classify(
    final_row: dict[str, str],
    fact: dict[str, str],
    siblings: list[dict[str, str]],
    visits: dict[str, dict[str, str]],
    best_other: dict[str, str],
) -> str:
    contract_id = fact["document_number"]
    status = normalized_status(fact.get("status", ""))
    start = parse_date(fact.get("start_date"))
    started_by_cutoff = bool(start and start <= CUTOFF_DATE)
    has_visits = visit_count(visits, contract_id) > 0
    other_has_basis = fact_has_payment(best_other) or visit_count(visits, best_other["document_number"]) > 0
    same_dates = [
        item
        for item in siblings
        if item["start_date"] == fact["start_date"] and item["end_date"] == fact["end_date"]
    ]
    same_name_dates = [item for item in same_dates if item["subscription_name"] == fact["subscription_name"]]

    if status in ("Отказ", "Бронь абонемента") and not has_visits:
        return "A_status_refusal_or_booking_no_payment_no_visits"
    if same_name_dates:
        return "B_duplicate_same_name_dates_no_payment"
    if same_dates:
        return "C_duplicate_same_dates_different_name_no_payment"
    if not started_by_cutoff:
        return "D_future_start_no_payment_overlap"
    if not has_visits and other_has_basis:
        return "E_started_no_payment_no_visits_other_has_basis"
    if has_visits:
        return "F_candidate_has_visits_do_not_auto_delete"
    return "G_needs_manual_context"


def build_audit_rows(
    facts: list[dict[str, str]],
    final_rows: list[dict[str, str]],
    final_funnel_by_client: dict[str, dict[str, str]],
    visits: dict[str, dict[str, str]],
    probe_details: dict[str, dict[str, str]],
    manual_comments: dict[str, str],
) -> list[dict[str, object]]:
    facts_by_doc = {row["document_number"]: row for row in facts}
    facts_by_ref = {row["subscription_ref"]: row for row in facts}
    final_by_doc = {row["contract_id"]: row for row in final_rows}

    active_full_by_client: dict[str, list[dict[str, str]]] = defaultdict(list)
    for contract_id in final_by_doc:
        fact = facts_by_doc.get(contract_id)
        if fact and fact["is_full_subscription"] == "1" and fact["is_active_on_cutoff"] == "1":
            active_full_by_client[fact["client_id"]].append(fact)

    candidate_pairs: list[tuple[dict[str, str], dict[str, str], list[dict[str, str]]]] = []
    for contract_id, final_row in final_by_doc.items():
        fact = facts_by_doc.get(contract_id)
        if not fact:
            continue
        is_candidate = (
            decimal_value(final_row.get("price")) > 0
            and final_row.get("type_of_payment") == "наличные"
            and not final_row.get("_payment_match_source")
            and fact.get("is_full_subscription") == "1"
            and fact.get("is_active_on_cutoff") == "1"
        )
        if not is_candidate:
            continue
        others = [item for item in active_full_by_client[fact["client_id"]] if item["document_number"] != contract_id]
        if others:
            candidate_pairs.append((final_row, fact, others))

    rows: list[dict[str, object]] = []
    for final_row, fact, others in candidate_pairs:
        contract_id = fact["document_number"]

        def other_key(other: dict[str, str]) -> tuple[int, int, str, str, str]:
            return (
                1 if fact_has_payment(other) else 0,
                visit_count(visits, other["document_number"]),
                other.get("end_date", ""),
                other.get("start_date", ""),
                other.get("document_number", ""),
            )

        best_other = max(others, key=other_key)
        no_payment_siblings = [
            item
            for item in active_full_by_client[fact["client_id"]]
            if item["document_number"] != contract_id
            and item["document_number"] in final_by_doc
            and decimal_value(final_by_doc[item["document_number"]].get("price")) > 0
            and final_by_doc[item["document_number"]].get("type_of_payment") == "наличные"
            and not final_by_doc[item["document_number"]].get("_payment_match_source")
        ]
        same_dates_siblings = [
            item
            for item in no_payment_siblings
            if item["start_date"] == fact["start_date"] and item["end_date"] == fact["end_date"]
        ]
        same_name_dates_siblings = [
            item for item in same_dates_siblings if item["subscription_name"] == fact["subscription_name"]
        ]
        risk_group = classify(final_row, fact, no_payment_siblings, visits, best_other)
        detail = probe_details.get(contract_id, {})
        candidate_visits = visits.get(contract_id, {})
        other_visits = visits.get(best_other["document_number"], {})
        best_other_final = final_by_doc.get(best_other["document_number"], {})
        funnel_row = final_funnel_by_client.get(fact["client_id"], {})
        selected_ref = funnel_row.get("selected_subscription_ref", "")
        selected_fact = facts_by_ref.get(selected_ref, {})
        selected_contract_id = selected_fact.get("document_number", "")
        same_card = bool(
            final_row.get("card")
            and best_other_final.get("card")
            and final_row.get("card") == best_other_final.get("card")
        )
        start = parse_date(fact.get("start_date"))

        rows.append(
            {
                "risk_group": risk_group,
                "contract_id": contract_id,
                "client_id": fact["client_id"],
                "client_fio": fact["effective_client_fio"],
                "contract_name": fact["subscription_name"],
                "status": normalized_status(fact.get("status", "")),
                "sale_date": fact["sale_date"],
                "sale_datetime": fact["sale_datetime"],
                "start_date": fact["start_date"],
                "end_date": fact["end_date"],
                "started_by_cutoff": int(bool(start and start <= CUTOFF_DATE)),
                "price": decimal_value(final_row.get("price")),
                "paid_candidate": fact["rg_paid_candidate"],
                "payment_count_candidate": fact["rg_payment_count_candidate"],
                "matched_payment_ref": fact["matched_payment_ref"],
                "nearby_payment_docs_14d": detail.get("nearby_payment_docs_14d", "0"),
                "nearby_payment_total_14d": detail.get("nearby_payment_total_14d", ".00"),
                "linked_sale_docs": detail.get("linked_sale_docs", ""),
                "linked_sale_sum_fld1140": detail.get("linked_sale_sum_fld1140", ""),
                "linked_sale_sum_fld1154": detail.get("linked_sale_sum_fld1154", ""),
                "linked_sale_sum_fld1160": detail.get("linked_sale_sum_fld1160", ""),
                "candidate_visit_docs": candidate_visits.get("visit_docs", "0"),
                "candidate_first_visit": candidate_visits.get("first_visit", ""),
                "candidate_last_visit": candidate_visits.get("last_visit", ""),
                "other_active_full_count_in_final": len(others),
                "best_other_contract_id": best_other["document_number"],
                "best_other_contract_name": best_other["subscription_name"],
                "best_other_status": normalized_status(best_other.get("status", "")),
                "best_other_start_date": best_other["start_date"],
                "best_other_end_date": best_other["end_date"],
                "best_other_has_payment": int(fact_has_payment(best_other)),
                "best_other_payment_amount": best_other["matched_payment_amount"],
                "best_other_payment_method": best_other["matched_payment_method"],
                "best_other_visit_docs": other_visits.get("visit_docs", "0"),
                "best_other_first_visit": other_visits.get("first_visit", ""),
                "best_other_last_visit": other_visits.get("last_visit", ""),
                "funnel": funnel_row.get("funnel", ""),
                "funnel_step": funnel_row.get("funnel_step", ""),
                "final_funnel_selected_contract_id": selected_contract_id,
                "final_funnel_selected_subscription_name": funnel_row.get("selected_subscription_name", ""),
                "final_funnel_selected_start_date": funnel_row.get("selected_subscription_start_date", ""),
                "final_funnel_selected_end_date": funnel_row.get("selected_subscription_end_date", ""),
                "final_funnel_active_full_subscription_count": funnel_row.get("active_full_subscription_count", ""),
                "final_funnel_validation_status": funnel_row.get("validation_status", ""),
                "target_is_final_funnel_selected": int(selected_contract_id == contract_id),
                "best_other_is_final_funnel_selected": int(selected_contract_id == best_other["document_number"]),
                "selected_card_number": funnel_row.get("selected_card_number", ""),
                "target_card": final_row.get("card", ""),
                "best_other_card": best_other_final.get("card", ""),
                "relation_to_best_other": relation_to_other(fact, best_other),
                "same_card_as_best_other": int(same_card),
                "sibling_no_payment_cash_active_count": len(no_payment_siblings),
                "same_name_dates_sibling_count": len(same_name_dates_siblings),
                "same_dates_sibling_count": len(same_dates_siblings),
                "doc_posted": detail.get("doc_posted", fact.get("doc_posted", "")),
                "doc_marked": detail.get("doc_marked", fact.get("doc_marked", "")),
                "normalized_club": detail.get("normalized_club", fact.get("normalized_club", "")),
                "manual_comment_if_any": manual_comments.get(contract_id, ""),
            }
        )

    rows.sort(key=lambda row: (str(row["risk_group"]), str(row["client_id"]), str(row["contract_id"])))
    return rows


def write_csv(path: Path, rows: list[dict[str, object]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--facts-tsv",
        default="output/20251115_0800_fix_owner_new_import/staging/membership_import_facts.tsv",
    )
    parser.add_argument(
        "--final-rows-csv",
        default="output/20251115_0800_fix_owner_new_import/staging/membership_import_rows.csv",
    )
    parser.add_argument(
        "--final-funnel-csv",
        default="output/20251115_0800_fix_owner/staging/final_funnel_clients.csv",
    )
    parser.add_argument(
        "--visits-log",
        default="logs/new-changes/prolem_2/64_membership_visit_counts_by_subscription.txt",
    )
    parser.add_argument(
        "--probe-log",
        default="logs/new-changes/prolem_2/65_no_payment_cash_active_overlap_probe.txt",
    )
    parser.add_argument(
        "--manual-xlsx",
        default="output/20251115_0800_fix_owner_new_import/membership_import_representative_30_examples_20260525_0800-with-answers.xlsx",
    )
    parser.add_argument(
        "--deep-audit-csv",
        default="output/20251115_0800_fix_owner_new_import/reports/no_payment_cash_active_full_overlap_deep_audit.csv",
    )
    parser.add_argument(
        "--current-final-csv",
        default="output/20251115_0800_fix_owner_new_import/reports/no_payment_cash_active_full_overlap_current_final.csv",
    )
    args = parser.parse_args()

    facts = read_facts(as_abs(args.facts_tsv))
    final_rows = read_csv(as_abs(args.final_rows_csv))
    final_funnel_by_client = read_final_funnel(as_abs(args.final_funnel_csv))
    visits = read_visit_counts(as_abs(args.visits_log))
    probe_details = read_probe_details(as_abs(args.probe_log))
    manual_comments = read_manual_comments(as_abs(args.manual_xlsx))
    audit_rows = build_audit_rows(
        facts,
        final_rows,
        final_funnel_by_client,
        visits,
        probe_details,
        manual_comments,
    )

    write_csv(as_abs(args.deep_audit_csv), audit_rows, DEEP_AUDIT_FIELDS)
    current_fields = [
        "contract_id",
        "client_id",
        "client_fio",
        "contract_name",
        "status",
        "sale_date",
        "start_date",
        "end_date",
        "price",
        "paid_candidate",
        "payment_count_candidate",
        "other_active_full_count_in_final",
        "best_other_contract_id",
        "best_other_contract_name",
        "best_other_status",
        "best_other_start_date",
        "best_other_end_date",
        "best_other_has_payment",
        "best_other_payment_amount",
        "best_other_payment_method",
        "best_other_visit_docs",
        "funnel",
        "funnel_step",
        "final_funnel_selected_contract_id",
        "target_is_final_funnel_selected",
        "best_other_is_final_funnel_selected",
        "final_funnel_validation_status",
        "relation_to_best_other",
    ]
    write_csv(as_abs(args.current_final_csv), audit_rows, current_fields)
    print(f"wrote {len(audit_rows)} rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
