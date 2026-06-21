#!/usr/bin/env python3
"""Check Fitbase import_заявки XLSX for implicit phone/name duplicates.

The main case is the same normalized phone with names that differ only by a
trailing duplicate suffix, for example `Иванов Иван` and `Иванов Иван {2}`.
The script also reports all same-phone groups where FIO differs, because those
are the broader implicit duplicate candidates. It is read-only for the XLSX and
writes a CSV report next to it by default.
"""

from __future__ import annotations

import argparse
import csv
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = (
    ROOT
    / "output"
    / "part2_20260525_0800_final_combined"
    / "fitbase_active_clients_import_zayavki_20260525_0800__all_funnels.xlsx"
)
EXPECTED_HEADERS = [
    "client_id",
    "phone",
    "client_fio",
    "email",
    "funnel",
    "funnel_step",
    "budget",
    "create_date",
    "manager",
    "филиал",
]
PHONE_SPLIT_RE = re.compile(r"[,;]\s*")
TRAILING_DUPLICATE_SUFFIX_RE = re.compile(r"(?:\s*\{\d+\})+\s*$")
SPACE_RE = re.compile(r"\s+")


@dataclass(frozen=True)
class ImportRow:
    row_number: int
    client_id: str
    phone_raw: str
    client_fio: str
    funnel: str
    funnel_step: str
    create_date: str
    branch: str


def as_abs(path: str | Path) -> Path:
    p = Path(path)
    return p if p.is_absolute() else ROOT / p


def normalize_phone_token(value: str) -> str:
    digits = re.sub(r"\D+", "", value or "")
    if len(digits) == 11 and digits[0] in {"7", "8"}:
        return "7" + digits[1:]
    if len(digits) == 10:
        return "7" + digits
    return ""


def normalize_phones(value: str) -> list[str]:
    phones: list[str] = []
    for part in PHONE_SPLIT_RE.split(value or ""):
        phone = normalize_phone_token(part)
        if phone and phone not in phones:
            phones.append(phone)
    return phones


def normalize_fio(value: str) -> str:
    value = TRAILING_DUPLICATE_SUFFIX_RE.sub("", value or "")
    return SPACE_RE.sub(" ", value.casefold().strip())


def has_duplicate_suffix(value: str) -> bool:
    return bool(TRAILING_DUPLICATE_SUFFIX_RE.search(value or ""))


def read_import_rows(path: Path) -> list[ImportRow]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(1, col).value for col in range(1, len(EXPECTED_HEADERS) + 1)]
    if headers != EXPECTED_HEADERS:
        wb.close()
        raise ValueError(f"Unexpected XLSX headers: expected {EXPECTED_HEADERS}, got {headers}")

    rows: list[ImportRow] = []
    for row_number, values in enumerate(
        ws.iter_rows(min_row=3, max_col=len(EXPECTED_HEADERS), values_only=True),
        start=3,
    ):
        if not any(value not in (None, "") for value in values):
            continue
        rows.append(
            ImportRow(
                row_number=row_number,
                client_id=str(values[0] or ""),
                phone_raw=str(values[1] or ""),
                client_fio=str(values[2] or ""),
                funnel=str(values[4] or ""),
                funnel_step=str(values[5] or ""),
                create_date=str(values[7] or ""),
                branch=str(values[9] or ""),
            )
        )
    wb.close()
    return rows


def build_phone_index(rows: Iterable[ImportRow]) -> dict[str, list[ImportRow]]:
    by_phone: dict[str, list[ImportRow]] = defaultdict(list)
    for row in rows:
        for phone in normalize_phones(row.phone_raw):
            by_phone[phone].append(row)
    return by_phone


def duplicate_type(group_rows: list[ImportRow]) -> str:
    raw_names = {row.client_fio for row in group_rows}
    base_names = {normalize_fio(row.client_fio) for row in group_rows}
    if len(base_names) == 1 and any(has_duplicate_suffix(row.client_fio) for row in group_rows):
        return "suffix_duplicate_name"
    if len(base_names) == 1 and len(raw_names) > 1:
        return "same_base_name_different_format"
    return "same_phone_different_fio"


def find_implicit_duplicates(rows: list[ImportRow]) -> list[dict[str, str]]:
    by_phone = build_phone_index(rows)
    report_rows: list[dict[str, str]] = []
    group_id = 0

    for phone, phone_rows in sorted(by_phone.items()):
        if len(phone_rows) < 2:
            continue

        raw_names = {row.client_fio for row in phone_rows}
        if len(raw_names) < 2:
            continue

        group_id += 1
        base_names = sorted({normalize_fio(row.client_fio) for row in phone_rows})
        group_type = duplicate_type(phone_rows)
        row_numbers = ";".join(str(row.row_number) for row in phone_rows)
        client_ids = ";".join(row.client_id for row in phone_rows)
        raw_names_joined = " | ".join(sorted(raw_names))
        base_names_joined = " | ".join(base_names)
        funnels = " | ".join(sorted({f"{row.funnel}/{row.funnel_step}" for row in phone_rows}))
        for row in phone_rows:
            report_rows.append(
                {
                    "group_id": str(group_id),
                    "duplicate_type": group_type,
                    "normalized_phone": phone,
                    "normalized_base_fio": normalize_fio(row.client_fio),
                    "distinct_normalized_base_fio_count": str(len(base_names)),
                    "group_size": str(len(phone_rows)),
                    "row_number": str(row.row_number),
                    "client_id": row.client_id,
                    "phone_raw": row.phone_raw,
                    "client_fio": row.client_fio,
                    "has_trailing_duplicate_suffix": "1" if has_duplicate_suffix(row.client_fio) else "0",
                    "funnel": row.funnel,
                    "funnel_step": row.funnel_step,
                        "create_date": row.create_date,
                        "branch": row.branch,
                        "group_row_numbers": row_numbers,
                    "group_client_ids": client_ids,
                    "group_raw_names": raw_names_joined,
                    "group_normalized_base_names": base_names_joined,
                    "group_funnels": funnels,
                }
            )

    return report_rows


def write_report(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "group_id",
        "duplicate_type",
        "normalized_phone",
        "normalized_base_fio",
        "distinct_normalized_base_fio_count",
        "group_size",
        "row_number",
        "client_id",
        "phone_raw",
        "client_fio",
        "has_trailing_duplicate_suffix",
        "funnel",
        "funnel_step",
        "create_date",
        "branch",
        "group_row_numbers",
        "group_client_ids",
        "group_raw_names",
        "group_normalized_base_names",
        "group_funnels",
    ]
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, lineterminator="\n")
        writer.writeheader()
        writer.writerows(rows)


def print_summary(rows: list[ImportRow], report_rows: list[dict[str, str]], report_path: Path) -> None:
    by_phone = build_phone_index(rows)
    duplicate_phone_groups = {phone: items for phone, items in by_phone.items() if len(items) > 1}
    duplicate_phone_extra_rows = sum(len(items) - 1 for items in duplicate_phone_groups.values())
    implicit_group_ids = {row["group_id"] for row in report_rows}
    implicit_row_numbers = {row["row_number"] for row in report_rows}
    implicit_phone_groups = {row["normalized_phone"] for row in report_rows}
    row_occurrence_type_counts = Counter(row["duplicate_type"] for row in report_rows)
    group_type_by_id: dict[str, str] = {}
    for row in report_rows:
        group_type_by_id.setdefault(row["group_id"], row["duplicate_type"])
    group_type_counts = Counter(group_type_by_id.values())
    suffix_rows_total = sum(1 for row in rows if has_duplicate_suffix(row.client_fio))
    invalid_or_empty_phone_rows = sum(1 for row in rows if not normalize_phones(row.phone_raw))

    print(f"xlsx_rows={len(rows)}")
    print(f"unique_normalized_phones={len(by_phone)}")
    print(f"rows_without_parseable_phone={invalid_or_empty_phone_rows}")
    print(f"phone_groups_with_multiple_rows={len(duplicate_phone_groups)}")
    print(f"duplicate_phone_extra_rows={duplicate_phone_extra_rows}")
    print(f"rows_with_trailing_duplicate_suffix={{n}}={suffix_rows_total}")
    print(f"same_phone_different_fio_candidate_groups={len(implicit_group_ids)}")
    print(f"same_phone_different_fio_candidate_phone_groups={len(implicit_phone_groups)}")
    print(f"same_phone_different_fio_candidate_distinct_rows={len(implicit_row_numbers)}")
    print(f"same_phone_different_fio_candidate_report_row_occurrences={len(report_rows)}")
    for key in ["suffix_duplicate_name", "same_base_name_different_format", "same_phone_different_fio"]:
        print(f"{key}_groups={group_type_counts.get(key, 0)}")
        print(f"{key}_report_row_occurrences={row_occurrence_type_counts.get(key, 0)}")
    print(f"report={report_path}")

    if not report_rows:
        print("examples=None")
        return

    print("examples:")
    rows_by_group: dict[str, list[dict[str, str]]] = defaultdict(list)
    for row in report_rows:
        rows_by_group[row["group_id"]].append(row)
    for group_id in sorted(rows_by_group, key=lambda value: int(value))[:10]:
        group_rows = rows_by_group[group_id]
        first = group_rows[0]
        print(
            f"- group={group_id} type={first['duplicate_type']} phone={first['normalized_phone']} "
            f"base_fio={first['normalized_base_fio']!r} size={first['group_size']}"
        )
        for row in group_rows:
            print(
                "  "
                f"row={row['row_number']} client_id={row['client_id']} "
                f"fio={row['client_fio']!r} funnel={row['funnel']!r}"
            )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Path to import_заявки XLSX.")
    parser.add_argument(
        "--report",
        default="",
        help="CSV report path. Defaults to <xlsx stem>__implicit_duplicate_check.csv next to the XLSX.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path = as_abs(args.xlsx)
    report_path = (
        as_abs(args.report)
        if args.report
        else xlsx_path.with_name(f"{xlsx_path.stem}__implicit_duplicate_check.csv")
    )

    rows = read_import_rows(xlsx_path)
    report_rows = find_implicit_duplicates(rows)
    write_report(report_path, report_rows)
    print_summary(rows, report_rows, report_path)


if __name__ == "__main__":
    main()
