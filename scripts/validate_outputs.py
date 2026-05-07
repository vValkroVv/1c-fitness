#!/usr/bin/env python3
"""Validate final Fitbase CSV/XLSX outputs against the migration plan."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from build_fitbase_xlsx import CARD_HEADERS, MAIN_HEADERS, MAIN_RUS_HEADERS, ROOT, normalized_fio, normalized_phones_set


DEFAULT_CUTOFF_DATE = "2026-04-29"
DEFAULT_DATE_STAMP = "20260429"
DEFAULT_OUTPUT_DIR = ROOT / "output"
ALLOWED_STEPS = {
    "Бронь",
    "60-31 день до окончания",
    "30-8 дней до окончания",
    "7-0 день до окончания",
    "Действующие клиенты",
}


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def int_value(value: object, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        return int(str(value))
    except ValueError:
        return default


def split_values(value: str | None) -> list[str]:
    result: list[str] = []
    for part in (value or "").split(","):
        item = part.strip()
        if item and item not in result:
            result.append(item)
    return result


def workbook_headers(path: Path, row: int, width: int) -> list[object]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    headers = [ws.cell(row, col).value for col in range(1, width + 1)]
    wb.close()
    return headers


def workbook_data_row_count(path: Path, first_data_row: int, width: int) -> int:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    count = 0
    for row in ws.iter_rows(min_row=first_data_row, max_col=width, values_only=True):
        if any(value not in (None, "") for value in row):
            count += 1
    wb.close()
    return count


def add_error(errors: list[str], condition: bool, message: str) -> None:
    if not condition:
        errors.append(message)


def validate(args: argparse.Namespace) -> int:
    output_dir = Path(args.output_dir)
    date_stamp = args.date_stamp
    final_csv = output_dir / f"final_active_clients_{date_stamp}.csv"
    main_xlsx = output_dir / f"fitbase_active_clients_import_zayavki_{date_stamp}.xlsx"
    cards_xlsx = output_dir / f"fitbase_active_clients_plastic_cards_{date_stamp}.xlsx"
    mini_main_xlsx = output_dir / f"mini_fitbase_active_clients_import_zayavki_{date_stamp}.xlsx"
    mini_cards_xlsx = output_dir / f"mini_fitbase_active_clients_plastic_cards_{date_stamp}.xlsx"

    required_files = [
        final_csv,
        main_xlsx,
        cards_xlsx,
        output_dir / "validation_report.md",
        output_dir / "stage_distribution.csv",
        output_dir / "duplicates_report.csv",
        output_dir / "missing_required_fields.csv",
        output_dir / "missing_sales_report.csv",
        output_dir / "missing_cards_report.csv",
        output_dir / "multiple_active_subscriptions_report.csv",
        output_dir / "multiple_cards_report.csv",
        output_dir / "booking_without_active_subscription_report.csv",
        output_dir / "table_mapping_report.md",
        output_dir / "schema_inventory.csv",
        output_dir / "splits" / "split_summary.csv",
        mini_main_xlsx,
        mini_cards_xlsx,
    ]

    errors: list[str] = []
    warnings: list[str] = []
    for path in required_files:
        add_error(errors, path.exists(), f"missing required file: {path.relative_to(ROOT)}")

    if errors:
        rows: list[dict[str, str]] = []
    else:
        rows = read_csv(final_csv)

    add_error(errors, bool(rows), "final active-client CSV is empty")

    if rows:
        add_error(errors, workbook_headers(main_xlsx, 1, len(MAIN_HEADERS)) == MAIN_HEADERS, "main XLSX technical headers mismatch")
        add_error(errors, workbook_headers(main_xlsx, 2, len(MAIN_RUS_HEADERS)) == MAIN_RUS_HEADERS, "main XLSX Russian headers mismatch")
        add_error(errors, workbook_headers(cards_xlsx, 1, len(CARD_HEADERS)) == CARD_HEADERS, "plastic cards XLSX headers mismatch")
        add_error(errors, workbook_data_row_count(main_xlsx, 3, len(MAIN_HEADERS)) == len(rows), "main XLSX row count mismatch")
        add_error(errors, workbook_data_row_count(cards_xlsx, 2, len(CARD_HEADERS)) == len(rows), "plastic cards XLSX row count mismatch")

        client_ids = [row.get("client_id", "") for row in rows]
        duplicated_client_ids = [key for key, count in Counter(client_ids).items() if key and count > 1]
        add_error(errors, not duplicated_client_ids, f"duplicate client_id values remain: {len(duplicated_client_ids)}")

        exact_keys = Counter(
            (normalized_fio(row.get("client_fio")), normalized_phones_set(row.get("phones")))
            for row in rows
            if normalized_fio(row.get("client_fio")) and normalized_phones_set(row.get("phones"))
        )
        exact_dups = [key for key, count in exact_keys.items() if count > 1]
        add_error(errors, not exact_dups, f"exact FIO+phone duplicates remain: {len(exact_dups)}")

        invalid_steps = sorted({row.get("funnel_step", "") for row in rows if row.get("funnel_step", "") not in ALLOWED_STEPS})
        add_error(errors, not invalid_steps, f"invalid funnel steps: {invalid_steps}")
        add_error(errors, all(row.get("funnel") == "Действующие клиенты" for row in rows), "unexpected funnel value found")
        add_error(errors, all(int_value(row.get("budget")) == 0 for row in rows), "non-zero budget found")
        add_error(errors, all(row.get("create_date") for row in rows), "empty create_date found")

        for row in rows:
            step = row.get("funnel_step", "")
            days = int_value(row.get("days_to_end"), -999999)
            has_booking = row.get("has_active_booking") == "1"
            if has_booking and step != "Бронь":
                errors.append(f"booking priority violation: client_id={row.get('client_id')}")
                break
            if step == "60-31 день до окончания" and not (31 <= days <= 60):
                errors.append(f"60-31 boundary violation: client_id={row.get('client_id')}, days={days}")
                break
            if step == "30-8 дней до окончания" and not (8 <= days <= 30):
                errors.append(f"30-8 boundary violation: client_id={row.get('client_id')}, days={days}")
                break
            if step == "7-0 день до окончания" and not (0 <= days <= 7):
                errors.append(f"7-0 boundary violation: client_id={row.get('client_id')}, days={days}")
                break

        missing_required = read_csv(output_dir / "missing_required_fields.csv")
        missing_sales = read_csv(output_dir / "missing_sales_report.csv")
        missing_cards = read_csv(output_dir / "missing_cards_report.csv")
        multiple_subs = read_csv(output_dir / "multiple_active_subscriptions_report.csv")
        multiple_cards = read_csv(output_dir / "multiple_cards_report.csv")
        booking_without_active = read_csv(output_dir / "booking_without_active_subscription_report.csv")

        add_error(errors, len(missing_required) == sum(1 for row in rows if not row.get("phones") or not row.get("client_fio") or not row.get("create_date")), "missing_required_fields.csv count mismatch")
        add_error(errors, len(missing_sales) == sum(1 for row in rows if not row.get("first_sale_date") or not row.get("create_date")), "missing_sales_report.csv count mismatch")
        add_error(errors, len(missing_cards) == sum(1 for row in rows if not row.get("plastic_card_number") or int_value(row.get("active_card_count")) == 0), "missing_cards_report.csv count mismatch")
        add_error(errors, len(multiple_subs) == sum(1 for row in rows if int_value(row.get("active_subscription_count")) > 1), "multiple_active_subscriptions_report.csv count mismatch")
        add_error(errors, len(multiple_cards) == sum(1 for row in rows if int_value(row.get("active_card_count")) > 1), "multiple_cards_report.csv count mismatch")
        add_error(errors, len(booking_without_active) == 0, "booking_without_active_subscription_report.csv is not empty")

        card_numbers = [
            card
            for row in rows
            for card in split_values(row.get("plastic_card_number"))
        ]
        duplicated_cards = [key for key, count in Counter(card_numbers).items() if key and count > 1]
        add_error(errors, not duplicated_cards, f"duplicate individual plastic-card numbers remain: {len(duplicated_cards)}")
        add_error(
            errors,
            all("," in row.get("plastic_card_number", "") for row in rows if int_value(row.get("active_card_count")) > 1),
            "some multiple-card clients do not have comma-separated card numbers",
        )

        split_summary = read_csv(output_dir / "splits" / "split_summary.csv")
        add_error(errors, len(split_summary) == 9, "split_summary.csv should contain 9 groups")
        add_error(errors, sum(int_value(row.get("actual_rows")) for row in split_summary) == len(rows), "split row total mismatch")
        for split in split_summary:
            expected = int_value(split.get("actual_rows"))
            main_split = ROOT / split.get("main_xlsx_file", "")
            cards_split = ROOT / split.get("cards_xlsx_file", "")
            add_error(errors, main_split.exists(), f"missing split main XLSX: {split.get('main_xlsx_file')}")
            add_error(errors, cards_split.exists(), f"missing split cards XLSX: {split.get('cards_xlsx_file')}")
            if main_split.exists():
                add_error(errors, workbook_data_row_count(main_split, 3, len(MAIN_HEADERS)) == expected, f"split main row count mismatch: {split.get('group_key')}")
            if cards_split.exists():
                add_error(errors, workbook_data_row_count(cards_split, 2, len(CARD_HEADERS)) == expected, f"split cards row count mismatch: {split.get('group_key')}")

        if mini_main_xlsx.exists() and mini_cards_xlsx.exists():
            mini_main_rows = workbook_data_row_count(mini_main_xlsx, 3, len(MAIN_HEADERS))
            mini_cards_rows = workbook_data_row_count(mini_cards_xlsx, 2, len(CARD_HEADERS))
            add_error(errors, mini_main_rows == mini_cards_rows and mini_main_rows > 0, "mini-test XLSX row counts mismatch or empty")
        else:
            warnings.append("mini-test XLSX files are missing")

    verdict = "PASS" if not errors else "FAIL"
    report_lines = [
        "# Final Fitbase Output Audit",
        "",
        f"Run date: {datetime.now().isoformat(timespec='seconds')}",
        f"cutoff_date: `{args.cutoff_date}`",
        f"date_stamp: `{date_stamp}`",
        f"final_rows: `{len(rows)}`",
        "",
        "## Verdict",
        "",
        f"`{verdict}`",
        "",
    ]
    if errors:
        report_lines.extend(["## Errors", ""])
        report_lines.extend(f"- {error}" for error in errors)
        report_lines.append("")
    if warnings:
        report_lines.extend(["## Warnings", ""])
        report_lines.extend(f"- {warning}" for warning in warnings)
        report_lines.append("")
    report_lines.extend(
        [
            "## Residual External Checks",
            "",
            "- Actual Fitbase mini-test import still has to be run in Fitbase UI/API.",
            "- If Fitbase rejects comma-separated card numbers or Excel date cells, only the output formatting should be changed after confirmation.",
            "",
        ]
    )
    report_path = output_dir / "final_audit_report.md"
    report_path.write_text("\n".join(report_lines), encoding="utf-8")

    print(f"verdict={verdict}")
    print(f"errors={len(errors)}")
    print(f"warnings={len(warnings)}")
    print(f"report={report_path.relative_to(ROOT)}")
    return 0 if not errors else 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--cutoff-date", default=DEFAULT_CUTOFF_DATE)
    parser.add_argument("--date-stamp", default=DEFAULT_DATE_STAMP)
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR))
    return parser.parse_args()


def main() -> None:
    raise SystemExit(validate(parse_args()))


if __name__ == "__main__":
    main()
