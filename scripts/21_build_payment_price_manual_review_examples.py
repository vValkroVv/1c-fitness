#!/usr/bin/env python3
"""Build a compact XLSX with membership rows for manual business review."""

from __future__ import annotations

import argparse
import csv
from collections import Counter
from copy import copy
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"

CLIENT_HEADERS = [
    "contract_id",
    "client_id",
    "phone",
    "client_fio",
    "contract_name",
    "card",
    "duration",
    "duration_type",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "freeze",
    "guests",
    "visits_left",
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "manager",
]

EXAMPLE_HEADERS = [
    "question_for_manual_check",
    "correct_price",
    "correct_payment_type",
    "comment",
    *CLIENT_HEADERS,
]


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def text(row: dict[str, str], key: str) -> str:
    return (row.get(key) or "").strip()


def price(row: dict[str, str]) -> Decimal:
    return decimal_value(row.get("price"))


def payment_type(row: dict[str, str]) -> str:
    return text(row, "type_of_payment")


def business_override(row: dict[str, str]) -> str:
    return text(row, "_business_override")


def payment_match_source(row: dict[str, str]) -> str:
    return text(row, "_payment_match_source")


def payment_method_raw(row: dict[str, str]) -> str:
    return text(row, "_payment_method_raw")


def contract_name(row: dict[str, str]) -> str:
    return text(row, "contract_name")


def payment_year(row: dict[str, str]) -> str:
    return text(row, "payment_date")[:4]


def row_sort_key(row: dict[str, str]) -> tuple[str, str, str]:
    return (text(row, "payment_date"), text(row, "client_id"), text(row, "contract_id"))


def rows_by_preference(
    rows: list[dict[str, str]],
    count: int,
    used_contract_ids: set[str],
    prefer_years: list[str] | None = None,
    prefer_prices: list[Decimal] | None = None,
) -> list[dict[str, str]]:
    pool = [row for row in rows if text(row, "contract_id") not in used_contract_ids]
    picked: list[dict[str, str]] = []

    for year in prefer_years or []:
        for row in sorted(pool, key=row_sort_key):
            if payment_year(row) == year and row not in picked:
                picked.append(row)
                break

    for preferred_price in prefer_prices or []:
        for row in sorted(pool, key=row_sort_key):
            if price(row) == preferred_price and row not in picked:
                picked.append(row)
                break

    for row in sorted(pool, key=row_sort_key):
        if len(picked) >= count:
            break
        if row not in picked:
            picked.append(row)

    for row in picked[:count]:
        used_contract_ids.add(text(row, "contract_id"))
    return picked[:count]


def build_examples(rows: list[dict[str, str]]) -> list[tuple[str, dict[str, str]]]:
    non_business = [row for row in rows if not business_override(row)]
    used_contract_ids: set[str] = set()
    examples: list[tuple[str, dict[str, str]]] = []

    def add(
        question: str,
        candidates: list[dict[str, str]],
        count: int,
        prefer_years: list[str] | None = None,
        prefer_prices: list[Decimal] | None = None,
    ) -> None:
        for row in rows_by_preference(candidates, count, used_contract_ids, prefer_years, prefer_prices):
            examples.append((question, row))

    def add_named(
        question: str,
        candidates: list[dict[str, str]],
        names: list[str],
        fallback_count: int = 0,
    ) -> None:
        for name in names:
            add(question, [row for row in candidates if contract_name(row) == name], 1)
        if fallback_count:
            add(question, candidates, fallback_count)

    zero_no_payment_blank = [
        row for row in rows if business_override(row) == "business_zero_no_payment_blank_payment_type"
    ]
    add(
        "Проверить правило: price=0, платеж не найден, type_of_payment оставлен пустым. Это корректно?",
        [row for row in zero_no_payment_blank if contract_name(row) == "Абонемент Неделя Фитнес"],
        2,
        prefer_years=["2023", "2024"],
    )
    add_named(
        "Проверить правило: price=0, платеж не найден, type_of_payment оставлен пустым. Это корректно?",
        zero_no_payment_blank,
        ["Абонемент Неделя сайт 2023", "СУБАРЕНДА безлимит"],
    )

    zero_fallback_blank = [
        row for row in rows if business_override(row) == "business_zero_fallback_payment_type_blank"
    ]
    add_named(
        "Проверить правило: price=0, платеж найден только fallback по клиенту/date, type_of_payment очищен. Это корректно?",
        zero_fallback_blank,
        [
            "Абонемент УЛЬТРА 12 месяцев",
            "Абонемент МУЛЬТИКАРТА 12 месяцев",
            "Абонемент Неделя Фитнес",
            "Абонемент Неделя сайт 2023",
            "СУБАРЕНДА безлимит",
        ],
    )

    zero_direct_kept = [
        row
        for row in non_business
        if price(row) == 0
        and payment_type(row)
        and payment_match_source(row) == "direct_doc152_vt1083_doc154_vt1137_doc163"
    ]
    add_named(
        "Проверить правило: price=0, direct-платеж найден, type_of_payment оставлен как в платеже. Это корректно?",
        zero_direct_kept,
        [
            "Абонемент УЛЬТРА 12 месяцев",
            "Абонемент МУЛЬТИКАРТА 12 месяцев",
            "Абонемент НЕДЕЛЯ САЙТ",
            "Абонемент УЛЬТРА 12 месяцев СПЕЦПРЕДЛОЖЕНИЕ",
        ],
    )

    positive_no_payment_cash = [
        row
        for row in non_business
        if price(row) > 0 and payment_type(row) == "наличные" and not payment_match_source(row)
    ]
    add_named(
        "Проверить правило: price>0, платеж не найден, type_of_payment поставлен наличные. Это корректно?",
        positive_no_payment_cash,
        [
            "Абонемент УЛЬТРА 12 месяцев",
            "Абонемент МУЛЬТИКАРТА 15 месяцев (подарок) спецпредложение",
            "Абонемент МУЛЬТИКАРТА 12 месяцев",
            "Абонемент МУЛЬТИКАРТА 12 месяцев + подарок (СПЕЦПРЕДЛОЖЕНИЕ)",
            "Абонемент УЛЬТРА 1 МЕСЯЦ БЕЗЛИМИТ",
        ],
    )

    zero_raw_blank = [
        row for row in rows if business_override(row) == "business_zero_raw_blank_payment_type_blank"
    ]
    add_named(
        "Проверить правило: price=0, платеж найден, raw method пустой, type_of_payment оставлен пустым. Это корректно?",
        zero_raw_blank,
        [
            "Абонемент УЛЬТРА 12 месяцев",
            "Абонемент МУЛЬТИКАРТА 12 месяцев",
            "Абонемент УЛЬТРА 12+2 месяца в подарок",
        ],
    )

    add_named(
        "Контроль: подтвержденная бесплатная неделя, price=0, type_of_payment пустой. Это корректно?",
        [row for row in rows if business_override(row) == "business_confirmed_free_trial_zero_price_blank_payment"],
        ["Абонемент НЕДЕЛЯ САЙТ", "Абонемент НЕДЕЛЯ ФИТНЕСА БЕСПЛАТНО"],
    )
    add_named(
        "Контроль: full price=0 без платежа оставлен как ввод остатков/корпоративный/модификатор. Это корректно?",
        [row for row in rows if business_override(row) == "business_full_zero_no_payment_initial_balance_corporate_or_modifier"],
        ["Абонемент МУЛЬТИКАРТА 12 месяцев", "Абонемент УЛЬТРА 12 месяцев"],
    )
    add_named(
        "Контроль: заранее подтвержденный бесплатный/пробный абонемент, price=0, type_of_payment пустой. Это корректно?",
        [row for row in rows if business_override(row) == "business_free_trial_zero_price_blank_payment"],
        ["Абонемент НЕДЕЛЯ ДРУГ", "Абонемент 10 ДНЕЙ"],
    )

    add(
        "Контроль: технический raw method замаплен в безналичные. Это корректно?",
        [
            row
            for row in non_business
            if payment_type(row) == "безналичные"
            and ("для ошибок" in payment_method_raw(row).lower() or "бар ип иконников" in payment_method_raw(row).lower())
        ],
        1,
    )
    add_named(
        "Контроль: price>0, direct-платеж есть, raw method пустой, type_of_payment поставлен наличные. Это корректно?",
        [
            row
            for row in non_business
            if price(row) > 0
            and payment_type(row) == "наличные"
            and not payment_method_raw(row)
            and payment_match_source(row) == "direct_doc152_vt1083_doc154_vt1137_doc163"
        ],
        ["Абонемент УЛЬТРА 12 месяцев", "Абонемент НЕДЕЛЯ САЙТ"],
    )

    if len(examples) != 30:
        raise RuntimeError(f"Expected 30 examples, got {len(examples)}")
    return examples


def write_examples(output_path: Path, examples: list[tuple[str, dict[str, str]]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "manual_review_30"
    ws.append(EXAMPLE_HEADERS)

    for question, row in examples:
        ws.append([question, "", "", "", *[row.get(header, "") for header in CLIENT_HEADERS]])

    header_fill = PatternFill("solid", fgColor="1F4E78")
    manual_fill = PatternFill("solid", fgColor="FFF2CC")
    thin = Side(style="thin", color="D9D9D9")

    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(2, 5):
        ws.cell(row=1, column=col_idx).font = Font(color="000000", bold=True)
        ws.cell(row=1, column=col_idx).fill = manual_fill

    widths = {
        "A": 58,
        "B": 16,
        "C": 20,
        "D": 34,
        "E": 16,
        "F": 14,
        "G": 18,
        "H": 30,
        "I": 42,
        "J": 18,
        "K": 10,
        "L": 16,
        "M": 12,
        "N": 12,
        "O": 12,
        "P": 12,
        "Q": 10,
        "R": 10,
        "S": 12,
        "T": 12,
        "U": 12,
        "V": 14,
        "W": 16,
        "X": 28,
    }
    for column, width in widths.items():
        ws.column_dimensions[column].width = width

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = ws.dimensions
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", default="output/20251115_0800_fix_owner_new_import")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = ROOT / output_dir

    staging_csv = output_dir / "staging" / "membership_import_rows.csv"
    with staging_csv.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))

    examples = build_examples(rows)
    output_path = output_dir / f"payment_price_manual_review_examples_30_after_rules_{args.date_stamp}.xlsx"
    write_examples(output_path, examples)

    print(f"manual review examples: {output_path}")
    print(f"rows: {len(examples)}")
    print("contract names:")
    for key, value in Counter(contract_name(row) for _, row in examples).most_common():
        print(f"- {key}: {value}")


if __name__ == "__main__":
    main()
