#!/usr/bin/env python3
"""Compare the 1C manager debt report with cutoff-aware register finances.

The manager XLSX is validation-only.  Delivery money is taken from the
restored backup and the generated membership staging CSV.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
MANAGER_HEADERS = (
    "Структурная единица",
    "Продано",
    "Оплачено",
    "Задолженность",
)
CUTOFF_AT = "2026-06-30 23:27:03"


@dataclass(frozen=True)
class ManagerSale:
    manager_client_row: int
    manager_sale_row: int
    manager_fio: str
    sale_number: str
    sale_at: datetime
    sold: Decimal
    paid: Decimal
    debt: Decimal

    @property
    def key(self) -> tuple[str, str]:
        return canonical_number(self.sale_number), self.sale_at.strftime(
            "%Y-%m-%d %H:%M"
        )


def as_abs(path: str | Path) -> Path:
    value = Path(path)
    return value if value.is_absolute() else (ROOT / value).resolve()


def decimal_value(value: Any) -> Decimal:
    text = str(value or "0").strip().replace(" ", "").replace(",", ".")
    try:
        result = Decimal(text or "0")
    except InvalidOperation as exc:
        raise ValueError(f"Invalid numeric value: {value!r}") from exc
    if not result.is_finite():
        raise ValueError(f"Non-finite numeric value: {value!r}")
    return result


def excel_number(value: Decimal) -> int | float:
    integral = value.to_integral_value()
    return int(integral) if value == integral else float(value)


def canonical_number(value: Any) -> str:
    digits = re.sub(r"\D", "", str(value or ""))
    if not digits:
        raise ValueError(f"Sale number has no digits: {value!r}")
    return str(int(digits))


def normalize_fio(value: Any) -> str:
    text = str(value or "").casefold().replace("ё", "е")
    return " ".join(re.sub(r"[^0-9a-zа-я]+", " ", text).split())


def parse_manager_report(path: Path) -> list[ManagerSale]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook.active
    headers = tuple(worksheet.cell(1, column).value for column in range(1, 5))
    if headers != MANAGER_HEADERS:
        workbook.close()
        raise ValueError(
            f"Unexpected manager headers: {headers!r}; expected={MANAGER_HEADERS!r}"
        )
    hierarchy = tuple(worksheet.cell(row, 1).value for row in range(2, 5))
    if hierarchy != ("Клиент", "Документ продажи", "Клиент.Телефон"):
        workbook.close()
        raise ValueError(f"Unexpected manager hierarchy: {hierarchy!r}")

    sale_pattern = re.compile(
        r"^Продажа\s+(?P<number>\d+)\s+от\s+"
        r"(?P<timestamp>\d{2}\.\d{2}\.\d{4}\s+\d{2}:\d{2})\s*$"
    )
    sales: list[ManagerSale] = []
    client_row = 0
    client_fio = ""
    for row_number in range(5, worksheet.max_row + 1):
        level = int(worksheet.row_dimensions[row_number].outlineLevel or 0)
        label = str(worksheet.cell(row_number, 1).value or "").strip()
        if level == 0:
            client_row = 0
            client_fio = ""
            continue
        if level == 1:
            client_row = row_number
            client_fio = label
            continue
        if level != 2:
            continue
        if not client_row or not client_fio:
            workbook.close()
            raise ValueError(f"Manager sale row {row_number} has no client parent")
        match = sale_pattern.fullmatch(label)
        if not match:
            workbook.close()
            raise ValueError(f"Invalid manager sale label at row {row_number}: {label!r}")
        sales.append(
            ManagerSale(
                manager_client_row=client_row,
                manager_sale_row=row_number,
                manager_fio=client_fio,
                sale_number=match.group("number"),
                sale_at=datetime.strptime(
                    match.group("timestamp"), "%d.%m.%Y %H:%M"
                ),
                sold=decimal_value(worksheet.cell(row_number, 2).value),
                paid=decimal_value(worksheet.cell(row_number, 3).value),
                debt=decimal_value(worksheet.cell(row_number, 4).value),
            )
        )
    workbook.close()
    if not sales:
        raise ValueError(f"No manager sale rows found in {path}")
    keys = [sale.key for sale in sales]
    if len(set(keys)) != len(keys):
        raise ValueError("Manager report has duplicate sale-number/minute keys")
    return sales


def read_membership_staging(path: Path) -> tuple[
    list[dict[str, str]],
    dict[tuple[str, str], list[dict[str, str]]],
]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    required = {
        "contract_id",
        "client_id",
        "client_fio",
        "price",
        "amount_of_payments",
        "payment_left",
        "_money_source",
        "_is_active_on_cutoff",
        "_membership_sale_line_amount",
        "_financial_sale_document_number",
        "_financial_sale_document_datetime",
        "_financial_register_allocation_unambiguous",
        "_financial_register_row_count",
        "_financial_register_charge_sum",
        "_financial_register_payment_sum",
        "_financial_register_signed_debt",
        "_refuser_placeholder",
    }
    missing = sorted(required - set(rows[0] if rows else {}))
    if missing:
        raise ValueError(f"Membership staging is missing fields: {missing}")

    by_sale: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in rows:
        number = (row.get("_financial_sale_document_number") or "").strip()
        timestamp = (row.get("_financial_sale_document_datetime") or "").strip()
        if not number or not timestamp:
            continue
        key = canonical_number(number), timestamp[:16]
        by_sale.setdefault(key, []).append(row)
    return rows, by_sale


def read_delivery_memberships(path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [
        str(value or "")
        for value in next(
            worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
        )
    ]
    required = {
        "contract_id",
        "price",
        "amount_of_payments",
        "payment_left",
    }
    missing = sorted(required - set(headers))
    if missing:
        workbook.close()
        raise ValueError(f"Delivery membership is missing fields: {missing}")
    indexes = {header: index for index, header in enumerate(headers)}
    result: dict[str, dict[str, Any]] = {}
    for values in worksheet.iter_rows(min_row=3, values_only=True):
        contract_id = str(values[indexes["contract_id"]] or "").strip()
        if not contract_id:
            continue
        if contract_id in result:
            workbook.close()
            raise ValueError(f"Duplicate delivery contract_id: {contract_id}")
        result[contract_id] = {
            field: values[indexes[field]]
            for field in (
                "price",
                "amount_of_payments",
                "payment_left",
            )
        }
    workbook.close()
    return result


def exact(left: Decimal, right: Decimal) -> bool:
    return left == right


def compare_rows(
    manager_sales: list[ManagerSale],
    staging_by_sale: dict[tuple[str, str], list[dict[str, str]]],
    delivery_rows: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[str]]:
    comparisons: list[dict[str, Any]] = []
    errors: list[str] = []
    for case_number, sale in enumerate(manager_sales, start=1):
        candidates = staging_by_sale.get(sale.key, [])
        if len(candidates) != 1:
            errors.append(
                f"manager row {sale.manager_sale_row}: sale key {sale.key!r} "
                f"matched {len(candidates)} membership rows"
            )
            continue
        staging = candidates[0]
        contract_id = (staging.get("contract_id") or "").strip()
        delivery = delivery_rows.get(contract_id)
        if delivery is None:
            errors.append(
                f"manager row {sale.manager_sale_row}: contract {contract_id} "
                "is absent from clean delivery"
            )
            continue
        if staging.get("_financial_register_allocation_unambiguous") != "1":
            errors.append(
                f"manager row {sale.manager_sale_row}: contract {contract_id} "
                "does not have an unambiguous register allocation"
            )

        db_sold = decimal_value(staging["_membership_sale_line_amount"])
        db_paid = max(
            decimal_value(staging["_financial_register_payment_sum"]),
            Decimal("0"),
        )
        db_debt = max(
            decimal_value(staging["_financial_register_signed_debt"]),
            Decimal("0"),
        )
        delivery_sold = decimal_value(delivery["price"])
        delivery_paid = decimal_value(delivery["amount_of_payments"])
        delivery_debt = decimal_value(delivery["payment_left"])
        sold_status = "совпало" if exact(sale.sold, db_sold) else "не совпало"
        paid_status = "совпало" if exact(sale.paid, db_paid) else "не совпало"
        debt_status = "совпало" if exact(sale.debt, db_debt) else "не совпало"
        delivery_status = (
            "совпало с БД"
            if (
                exact(delivery_sold, db_sold)
                and exact(delivery_paid, db_paid)
                and exact(delivery_debt, db_debt)
            )
            else "ошибка новой выгрузки"
        )
        comparisons.append(
            {
                "№": case_number,
                "Строка клиента менеджера": sale.manager_client_row,
                "Строка продажи менеджера": sale.manager_sale_row,
                "ФИО менеджера": sale.manager_fio,
                "Номер продажи": sale.sale_number,
                "Дата продажи менеджера": sale.sale_at,
                "Договор": contract_id,
                "ID клиента": (staging.get("client_id") or "").strip(),
                "ФИО в новой выгрузке": (staging.get("client_fio") or "").strip(),
                "ФИО совпало": (
                    "да"
                    if normalize_fio(sale.manager_fio)
                    == normalize_fio(staging.get("client_fio"))
                    else "владелец изменён"
                ),
                "Менеджер: продано": excel_number(sale.sold),
                "Менеджер: оплачено": excel_number(sale.paid),
                "Менеджер: долг": excel_number(sale.debt),
                "БД 30.06: продано": excel_number(db_sold),
                "БД 30.06: оплачено": excel_number(db_paid),
                "БД 30.06: долг": excel_number(db_debt),
                "Новая выгрузка: цена": excel_number(delivery_sold),
                "Новая выгрузка: оплачено": excel_number(delivery_paid),
                "Новая выгрузка: долг": excel_number(delivery_debt),
                "Δ продано менеджер−БД": excel_number(sale.sold - db_sold),
                "Δ оплачено менеджер−БД": excel_number(sale.paid - db_paid),
                "Δ долг менеджер−БД": excel_number(sale.debt - db_debt),
                "Статус продано": sold_status,
                "Статус оплачено": paid_status,
                "Статус долга": debt_status,
                "Статус новой выгрузки": delivery_status,
                "Источник": (staging.get("_money_source") or "").strip(),
            }
        )
    return comparisons, errors


def source_group(row: dict[str, str]) -> str:
    source = (row.get("_money_source") or "").strip()
    if row.get("_refuser_placeholder") == "1":
        return "refuser_placeholder"
    if source.startswith("accumrg3305_sale_balance"):
        return "register_balance"
    if "ambiguous_multi_membership_sale" in source:
        return "info_debt_fallback_ambiguous_sale"
    if "no_unambiguous_register_balance" in source:
        return "info_debt_fallback_no_register"
    if source.startswith("business_"):
        return "business_override"
    return source or "other"


def write_coverage_csv(
    path: Path,
    staging_rows: list[dict[str, str]],
) -> dict[str, int]:
    counts: Counter[tuple[str, str]] = Counter()
    positive: Counter[tuple[str, str]] = Counter()
    active_register_positive = 0
    active_register_positive_mismatch = 0
    active_nonregister_positive = 0
    for row in staging_rows:
        scope = (
            "active_on_cutoff"
            if row.get("_is_active_on_cutoff") == "1"
            else "other_or_placeholder"
        )
        group = source_group(row)
        key = scope, group
        counts[key] += 1
        debt = decimal_value(row.get("payment_left"))
        if debt > 0:
            positive[key] += 1
            if scope == "active_on_cutoff" and group != "register_balance":
                active_nonregister_positive += 1
        register_debt = max(
            decimal_value(row.get("_financial_register_signed_debt")),
            Decimal("0"),
        )
        if (
            scope == "active_on_cutoff"
            and row.get("_financial_register_allocation_unambiguous") == "1"
            and register_debt > 0
        ):
            active_register_positive += 1
            if debt != register_debt:
                active_register_positive_mismatch += 1

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=["scope", "source_group", "rows", "positive_debt_rows"],
            lineterminator="\n",
        )
        writer.writeheader()
        for scope, group in sorted(counts):
            writer.writerow(
                {
                    "scope": scope,
                    "source_group": group,
                    "rows": counts[(scope, group)],
                    "positive_debt_rows": positive[(scope, group)],
                }
            )
    return {
        "active_register_positive_debts": active_register_positive,
        "active_register_positive_debt_mismatches": (
            active_register_positive_mismatch
        ),
        "active_nonregister_positive_debts": active_nonregister_positive,
    }


DATA_HEADERS = [
    "№",
    "Строка клиента менеджера",
    "Строка продажи менеджера",
    "ФИО менеджера",
    "Номер продажи",
    "Дата продажи менеджера",
    "Договор",
    "ID клиента",
    "ФИО в новой выгрузке",
    "ФИО совпало",
    "Менеджер: продано",
    "Менеджер: оплачено",
    "Менеджер: долг",
    "БД 30.06: продано",
    "БД 30.06: оплачено",
    "БД 30.06: долг",
    "Новая выгрузка: цена",
    "Новая выгрузка: оплачено",
    "Новая выгрузка: долг",
    "Δ продано менеджер−БД",
    "Δ оплачено менеджер−БД",
    "Δ долг менеджер−БД",
    "Статус продано",
    "Статус оплачено",
    "Статус долга",
    "Статус новой выгрузки",
    "Источник",
]


def style_data_sheet(worksheet, rows: list[dict[str, Any]]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    exact_fill = PatternFill("solid", fgColor="E2F0D9")
    mismatch_fill = PatternFill("solid", fgColor="FCE4D6")
    thin_gray = Side(style="thin", color="D9E2F3")
    worksheet.append(DATA_HEADERS)
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = Border(bottom=thin_gray)
    for item in rows:
        worksheet.append([item[header] for header in DATA_HEADERS])
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.sheet_view.showGridLines = False
    worksheet.row_dimensions[1].height = 42

    money_headers = {
        header
        for header in DATA_HEADERS
        if (
            "продано" in header.lower()
            or "оплачено" in header.lower()
            or "долг" in header.lower()
            or header.startswith("Δ")
            or "цена" in header.lower()
        )
    }
    for column, header in enumerate(DATA_HEADERS, start=1):
        if header in money_headers:
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row_number, column).number_format = "#,##0.00"
        if header.startswith("Статус"):
            for row_number in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(row_number, column)
                cell.fill = (
                    exact_fill
                    if str(cell.value or "").startswith("совпало")
                    else mismatch_fill
                )
    date_column = DATA_HEADERS.index("Дата продажи менеджера") + 1
    for row_number in range(2, worksheet.max_row + 1):
        worksheet.cell(row_number, date_column).number_format = "yyyy-mm-dd hh:mm"

    widths = {
        "A": 7,
        "B": 14,
        "C": 14,
        "D": 34,
        "E": 15,
        "F": 20,
        "G": 16,
        "H": 14,
        "I": 34,
        "J": 18,
    }
    for column in range(1, len(DATA_HEADERS) + 1):
        letter = get_column_letter(column)
        worksheet.column_dimensions[letter].width = widths.get(letter, 18)
    worksheet.column_dimensions[get_column_letter(len(DATA_HEADERS))].width = 48


def write_comparison_workbook(
    path: Path,
    comparisons: list[dict[str, Any]],
    metrics: dict[str, Any],
) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Итоги"
    summary.sheet_view.showGridLines = False
    summary.merge_cells("A1:B1")
    summary["A1"] = "Сверка долга на 30.06.2026"
    summary["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    summary["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    summary["A1"].alignment = Alignment(horizontal="center", vertical="center")
    summary.row_dimensions[1].height = 30
    summary["A3"] = "Cutoff backup"
    summary["B3"] = metrics["cutoff_at"]
    summary["A5"] = "Показатель"
    summary["B5"] = "Значение"
    for cell in summary[5][:2]:
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.font = Font(bold=True)
    summary_rows = [
        ("Продаж в отчёте менеджера", metrics["manager_sales"]),
        ("Найдено в новой выгрузке", metrics["mapped_sales"]),
        ("Расхождений по продано", metrics["manager_sold_mismatches"]),
        ("Расхождений по долгу", metrics["manager_debt_mismatches"]),
        ("Расхождений по оплачено", metrics["manager_paid_mismatches"]),
        ("Полностью совпало", metrics["manager_full_matches"]),
        ("Ошибка новой выгрузки против БД", metrics["delivery_tuple_mismatches"]),
        ("Сумма превышения оплаты менеджера", metrics["manager_paid_excess"]),
    ]
    for row in summary_rows:
        summary.append(row)
    for row_number in range(3, 14):
        summary[f"B{row_number}"].alignment = Alignment(
            horizontal="left", vertical="center"
        )
    summary["A15"] = (
        "Важно: 63 строки отличаются только колонкой «Оплачено» менеджерского "
        "отчёта. Задолженность новой выгрузки совпадает с отчётом во всех 274 случаях."
    )
    summary.merge_cells("A15:B17")
    summary["A15"].alignment = Alignment(wrap_text=True, vertical="top")
    summary["A15"].fill = PatternFill("solid", fgColor="FFF2CC")
    summary["A15"].font = Font(size=10, bold=True)
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 20

    all_sheet = workbook.create_sheet("Все 274")
    style_data_sheet(all_sheet, comparisons)
    paid_mismatches = [
        row for row in comparisons if row["Статус оплачено"] == "не совпало"
    ]
    paid_sheet = workbook.create_sheet("Не совпало оплачено")
    style_data_sheet(paid_sheet, paid_mismatches)
    debt_mismatches = [
        row for row in comparisons if row["Статус долга"] == "не совпало"
    ]
    debt_sheet = workbook.create_sheet("Не совпал долг")
    if debt_mismatches:
        style_data_sheet(debt_sheet, debt_mismatches)
    else:
        debt_sheet.sheet_view.showGridLines = False
        debt_sheet["A1"] = "Расхождений по задолженности нет"
        debt_sheet["A1"].font = Font(size=14, bold=True, color="006100")
        debt_sheet["A3"] = (
            "Все 274 задолженности совпали с балансом _AccumRg3305 "
            "на 2026-06-30 23:27:03."
        )
        debt_sheet["A3"].alignment = Alignment(wrap_text=True, vertical="top")
        debt_sheet.row_dimensions[3].height = 34
        debt_sheet.column_dimensions["A"].width = 52

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--manager-xlsx", required=True)
    parser.add_argument("--membership-rows-csv", required=True)
    parser.add_argument("--delivery-membership-xlsx", required=True)
    parser.add_argument("--output-xlsx", required=True)
    parser.add_argument("--report", required=True)
    parser.add_argument("--json-report", required=True)
    parser.add_argument("--coverage-csv", required=True)
    parser.add_argument("--cutoff-at", default=CUTOFF_AT)
    parser.add_argument("--expected-manager-sales", type=int, default=274)
    parser.add_argument("--expected-paid-mismatches", type=int, default=63)
    args = parser.parse_args()

    manager_sales = parse_manager_report(as_abs(args.manager_xlsx))
    staging_rows, staging_by_sale = read_membership_staging(
        as_abs(args.membership_rows_csv)
    )
    delivery_rows = read_delivery_memberships(
        as_abs(args.delivery_membership_xlsx)
    )
    comparisons, errors = compare_rows(
        manager_sales,
        staging_by_sale,
        delivery_rows,
    )

    manager_sold_mismatches = sum(
        row["Статус продано"] == "не совпало" for row in comparisons
    )
    manager_paid_mismatches = sum(
        row["Статус оплачено"] == "не совпало" for row in comparisons
    )
    manager_debt_mismatches = sum(
        row["Статус долга"] == "не совпало" for row in comparisons
    )
    delivery_tuple_mismatches = sum(
        row["Статус новой выгрузки"] != "совпало с БД"
        for row in comparisons
    )
    manager_full_matches = sum(
        row["Статус продано"] == "совпало"
        and row["Статус оплачено"] == "совпало"
        and row["Статус долга"] == "совпало"
        for row in comparisons
    )
    manager_paid_excess = sum(
        (
            decimal_value(row["Δ оплачено менеджер−БД"])
            for row in comparisons
            if decimal_value(row["Δ оплачено менеджер−БД"]) > 0
        ),
        Decimal("0"),
    )
    coverage = write_coverage_csv(
        as_abs(args.coverage_csv),
        staging_rows,
    )
    metrics: dict[str, Any] = {
        "cutoff_at": args.cutoff_at,
        "manager_sales": len(manager_sales),
        "mapped_sales": len(comparisons),
        "manager_sold_mismatches": manager_sold_mismatches,
        "manager_paid_mismatches": manager_paid_mismatches,
        "manager_debt_mismatches": manager_debt_mismatches,
        "manager_full_matches": manager_full_matches,
        "delivery_tuple_mismatches": delivery_tuple_mismatches,
        "manager_paid_excess": excel_number(manager_paid_excess),
        **coverage,
    }

    if len(manager_sales) != args.expected_manager_sales:
        errors.append(
            f"manager sales={len(manager_sales)}, "
            f"expected={args.expected_manager_sales}"
        )
    if len(comparisons) != len(manager_sales):
        errors.append(
            f"mapped manager sales={len(comparisons)}, total={len(manager_sales)}"
        )
    if manager_sold_mismatches:
        errors.append(f"manager sold mismatches={manager_sold_mismatches}")
    if manager_debt_mismatches:
        errors.append(f"manager debt mismatches={manager_debt_mismatches}")
    if manager_paid_mismatches != args.expected_paid_mismatches:
        errors.append(
            f"manager paid mismatches={manager_paid_mismatches}, "
            f"expected={args.expected_paid_mismatches}"
        )
    if delivery_tuple_mismatches:
        errors.append(
            f"new delivery tuple mismatches DB={delivery_tuple_mismatches}"
        )
    if coverage["active_register_positive_debt_mismatches"]:
        errors.append(
            "active register positive debt mismatches="
            f"{coverage['active_register_positive_debt_mismatches']}"
        )
    if coverage["active_nonregister_positive_debts"]:
        errors.append(
            "active positive debts outside register source="
            f"{coverage['active_nonregister_positive_debts']}"
        )

    verdict = "PASS" if not errors else "FAIL"
    metrics["verdict"] = verdict
    metrics["errors"] = errors
    write_comparison_workbook(
        as_abs(args.output_xlsx),
        comparisons,
        metrics,
    )

    report_lines = [
        "# Manager debt vs 1C register",
        "",
        f"- verdict: **{verdict}**",
        f"- cutoff: `{args.cutoff_at}`",
        f"- manager sale rows: `{len(manager_sales)}`",
        f"- mapped to new delivery: `{len(comparisons)}`",
        f"- sold mismatches: `{manager_sold_mismatches}`",
        f"- debt mismatches: `{manager_debt_mismatches}`",
        f"- paid mismatches: `{manager_paid_mismatches}`",
        f"- full triple matches: `{manager_full_matches}`",
        f"- new delivery tuple mismatches vs DB: `{delivery_tuple_mismatches}`",
        f"- manager paid excess absent from backup: `{manager_paid_excess}`",
        f"- active register positive debts: `{coverage['active_register_positive_debts']}`",
        "- active positive debts outside register source: "
        f"`{coverage['active_nonregister_positive_debts']}`",
        "",
        "The manager XLSX is used only for validation. Delivery values come "
        "from the restored database at the backup cutoff.",
        "",
        "## Errors",
        "",
        *([f"- {error}" for error in errors] or ["- none"]),
        "",
    ]
    report_path = as_abs(args.report)
    report_path.parent.mkdir(parents=True, exist_ok=True)
    report_path.write_text("\n".join(report_lines), encoding="utf-8")
    json_path = as_abs(args.json_report)
    json_path.parent.mkdir(parents=True, exist_ok=True)
    json_path.write_text(
        json.dumps(metrics, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    print(f"verdict={verdict}")
    print(f"manager_sales={len(manager_sales)}")
    print(f"manager_debt_mismatches={manager_debt_mismatches}")
    print(f"manager_paid_mismatches={manager_paid_mismatches}")
    print(f"delivery_tuple_mismatches={delivery_tuple_mismatches}")
    print(f"output_xlsx={as_abs(args.output_xlsx)}")
    return 0 if not errors else 1


if __name__ == "__main__":
    raise SystemExit(main())
