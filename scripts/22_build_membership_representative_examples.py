#!/usr/bin/env python3
"""Build a representative 30-row sample from the final membership import."""

from __future__ import annotations

import argparse
import csv
from copy import copy
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Callable, Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DATE_STAMP = "20260525_0800"

CLIENT_HEADERS = [
    "tag",
    "contract_id",
    "client_id",
    "phone",
    "client_fio",
    "contract_name",
    "card",
    "duration",
    "duration_type",
    "create_date",
    "payment_date",
    "activation_date",
    "end_date",
    "freeze",
    "guests",
    "visits_left",
    "price",
    "amount_of_payments",
    "payment_left",
    "type_of_payment",
    "manager",
]


@dataclass(frozen=True)
class ExampleSpec:
    key: str
    description: str
    predicate: Callable[[dict[str, str]], bool]
    prefer_names: tuple[str, ...] = ()
    prefer_years: tuple[str, ...] = ()
    prefer_payment_types: tuple[str, ...] = ()


def decimal_value(value: Any) -> Decimal:
    if value in (None, ""):
        return Decimal("0")
    try:
        return Decimal(str(value).strip().replace(",", "."))
    except (InvalidOperation, AttributeError):
        return Decimal("0")


def text(row: dict[str, str], key: str) -> str:
    return (row.get(key) or "").strip()


def year(row: dict[str, str]) -> str:
    return text(row, "payment_date")[:4]


def sort_key(row: dict[str, str]) -> tuple[int, str, str, str]:
    has_phone = 0 if text(row, "phone") else 1
    return has_phone, text(row, "payment_date"), text(row, "client_id"), text(row, "contract_id")


def score(row: dict[str, str], spec: ExampleSpec) -> tuple[int, int, int, tuple[int, str, str, str]]:
    name_score = spec.prefer_names.index(text(row, "contract_name")) if text(row, "contract_name") in spec.prefer_names else 99
    year_score = spec.prefer_years.index(year(row)) if year(row) in spec.prefer_years else 99
    payment_type_score = (
        spec.prefer_payment_types.index(text(row, "type_of_payment"))
        if text(row, "type_of_payment") in spec.prefer_payment_types
        else 99
    )
    return name_score, year_score, payment_type_score, sort_key(row)


def select_examples(rows: list[dict[str, str]]) -> tuple[list[tuple[ExampleSpec, dict[str, str]]], list[ExampleSpec]]:
    def price(row: dict[str, str]) -> Decimal:
        return decimal_value(row.get("price"))

    def payment_left(row: dict[str, str]) -> Decimal:
        return decimal_value(row.get("payment_left"))

    specs = [
        ExampleSpec(
            "ordinary_ultra_cashless",
            "Обычный массовый full-абонемент УЛЬТРА, цена > 0, direct-платеж, безналичные.",
            lambda r: text(r, "contract_name") == "Абонемент УЛЬТРА 12 месяцев"
            and price(r) > 0
            and text(r, "type_of_payment") == "безналичные"
            and text(r, "_money_source") == "assume_paid_full_when_no_installment_marker"
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163",
            prefer_years=("2024", "2025", "2026"),
        ),
        ExampleSpec(
            "ordinary_multicard_cashless",
            "Обычный массовый full-абонемент МУЛЬТИКАРТА, цена > 0, direct-платеж, безналичные.",
            lambda r: text(r, "contract_name") == "Абонемент МУЛЬТИКАРТА 12 месяцев"
            and price(r) > 0
            and text(r, "type_of_payment") == "безналичные"
            and text(r, "_money_source") == "assume_paid_full_when_no_installment_marker"
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163",
            prefer_years=("2024", "2025", "2026"),
        ),
        ExampleSpec(
            "ordinary_special_offer_cashless",
            "Платный full-абонемент спецпредложение/подарок, direct-платеж, безналичные.",
            lambda r: price(r) > 0
            and text(r, "type_of_payment") == "безналичные"
            and text(r, "_money_source") == "assume_paid_full_when_no_installment_marker"
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163"
            and ("СПЕЦПРЕДЛОЖЕНИЕ" in text(r, "contract_name") or "подар" in text(r, "contract_name").lower()),
            prefer_years=("2025", "2026", "2024"),
        ),
        ExampleSpec(
            "ordinary_sbp",
            "Обычный платный абонемент с типом оплаты СБП.",
            lambda r: price(r) > 0 and text(r, "type_of_payment") == "сбп",
            prefer_years=("2025", "2026", "2024"),
        ),
        ExampleSpec(
            "cash_direct_blank_raw_ultra",
            "Платный direct-платеж с пустым raw method, поставлен тип оплаты наличные.",
            lambda r: price(r) > 0
            and text(r, "type_of_payment") == "наличные"
            and not text(r, "_payment_method_raw")
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163",
            prefer_names=("Абонемент УЛЬТРА 12 месяцев",),
            prefer_years=("2019", "2020", "2021"),
        ),
        ExampleSpec(
            "cash_direct_blank_raw_week",
            "Платная НЕДЕЛЯ САЙТ с direct-платежом и пустым raw method, поставлены наличные.",
            lambda r: text(r, "contract_name") == "Абонемент НЕДЕЛЯ САЙТ"
            and price(r) > 0
            and text(r, "type_of_payment") == "наличные"
            and not text(r, "_payment_method_raw")
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163",
        ),
        ExampleSpec(
            "cash_no_payment_default",
            "Платная строка без найденного платежа, по бизнес-решению поставлены наличные.",
            lambda r: price(r) > 0
            and text(r, "type_of_payment") == "наличные"
            and not text(r, "_payment_match_source"),
            prefer_years=("2026", "2025", "2024"),
        ),
        ExampleSpec(
            "technical_raw_cashless",
            "Технический raw method из группы ошибок, замаплен в безналичные.",
            lambda r: price(r) > 0
            and text(r, "type_of_payment") == "безналичные"
            and ("для ошибок" in text(r, "_payment_method_raw").lower() or "бар ип иконников" in text(r, "_payment_method_raw").lower()),
        ),
        ExampleSpec(
            "partial_paid_candidate",
            "Частичная оплата из InfoRg3060._Fld3072, payment_left > 0.",
            lambda r: text(r, "_money_source") == "rg_fld3072_paid_candidate" and payment_left(r) > 0,
            prefer_names=("Абонемент УЛЬТРА 12 месяцев", "Абонемент МУЛЬТИКАРТА 12 месяцев"),
            prefer_years=("2026", "2025", "2024"),
        ),
        ExampleSpec(
            "installment_named",
            "Именная рассрочка, где оплата восстановлена через matched payment amount.",
            lambda r: "рассроч" in text(r, "contract_name").lower()
            and text(r, "_money_source") == "matched_payment_amount_for_installment",
        ),
        ExampleSpec(
            "owner_change_paid",
            "Строка с переоформлением/сменой владельца, платный full-абонемент.",
            lambda r: bool(text(r, "_owner_change_ref")) and price(r) > 0 and text(r, "_product_class") == "full_subscription",
            prefer_years=("2025", "2026", "2024"),
        ),
        ExampleSpec(
            "owner_change_zero",
            "Строка со сменой владельца и price=0, закрыта бизнес-правилом full zero no payment.",
            lambda r: bool(text(r, "_owner_change_ref"))
            and text(r, "_business_override") == "business_full_zero_no_payment_initial_balance_corporate_or_modifier",
        ),
        ExampleSpec(
            "legacy_2018_full_zero",
            "Старый 2018 full/gift пласт, price=0 и пустой type_of_payment.",
            lambda r: text(r, "_business_override") == "business_legacy_2018_full_subscription_zero_price_blank_payment",
            prefer_names=("Абонемент УЛЬТРА 12 месяцев", "Абонемент МУЛЬТИКАРТА 12 месяцев"),
        ),
        ExampleSpec(
            "free_week_friend",
            "Заранее подтвержденный бесплатный НЕДЕЛЯ ДРУГ.",
            lambda r: text(r, "contract_name") == "Абонемент НЕДЕЛЯ ДРУГ"
            and text(r, "_business_override") == "business_free_trial_zero_price_blank_payment",
        ),
        ExampleSpec(
            "free_10_days",
            "Заранее подтвержденный бесплатный/пробный Абонемент 10 ДНЕЙ.",
            lambda r: text(r, "contract_name") == "Абонемент 10 ДНЕЙ"
            and text(r, "_business_override") == "business_free_trial_zero_price_blank_payment",
            prefer_years=("2026", "2025"),
        ),
        ExampleSpec(
            "confirmed_week_site_zero",
            "Подтвержденная бесплатная НЕДЕЛЯ САЙТ при price=0 и не-direct платеже.",
            lambda r: text(r, "contract_name") == "Абонемент НЕДЕЛЯ САЙТ"
            and text(r, "_business_override") == "business_confirmed_free_trial_zero_price_blank_payment",
        ),
        ExampleSpec(
            "confirmed_week_fitness_free_zero",
            "Подтвержденная бесплатная НЕДЕЛЯ ФИТНЕСА БЕСПЛАТНО.",
            lambda r: text(r, "contract_name") == "Абонемент НЕДЕЛЯ ФИТНЕСА БЕСПЛАТНО"
            and text(r, "_business_override") == "business_confirmed_free_trial_zero_price_blank_payment",
        ),
        ExampleSpec(
            "zero_no_payment_blank_week",
            "Остаточный zero-price без платежа: оставляем price=0 и пустой type_of_payment.",
            lambda r: text(r, "_business_override") == "business_zero_no_payment_blank_payment_type"
            and text(r, "contract_name") == "Абонемент Неделя Фитнес",
        ),
        ExampleSpec(
            "zero_no_payment_blank_subrent",
            "Остаточный zero-price без платежа на СУБАРЕНДА безлимит: оставляем пустой type_of_payment.",
            lambda r: text(r, "_business_override") == "business_zero_no_payment_blank_payment_type"
            and text(r, "contract_name") == "СУБАРЕНДА безлимит",
        ),
        ExampleSpec(
            "zero_fallback_blank_full",
            "Price=0 с fallback-платежом у full-абонемента: type_of_payment очищен.",
            lambda r: text(r, "_business_override") == "business_zero_fallback_payment_type_blank"
            and text(r, "_product_class") == "full_subscription",
            prefer_names=("Абонемент УЛЬТРА 12 месяцев", "Абонемент МУЛЬТИКАРТА 12 месяцев"),
        ),
        ExampleSpec(
            "zero_fallback_blank_trial",
            "Price=0 с fallback-платежом у недельного/пробного слоя: type_of_payment очищен.",
            lambda r: text(r, "_business_override") == "business_zero_fallback_payment_type_blank"
            and text(r, "_product_class") == "trial_or_guest",
            prefer_names=("Абонемент Неделя сайт 2023", "Абонемент Неделя Фитнес"),
        ),
        ExampleSpec(
            "zero_direct_kept_full",
            "Price=0 с direct-платежом у full-абонемента: type_of_payment оставлен как есть.",
            lambda r: price(r) == 0
            and not text(r, "_business_override")
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163"
            and text(r, "_product_class") == "full_subscription"
            and text(r, "type_of_payment"),
            prefer_names=("Абонемент УЛЬТРА 12 месяцев", "Абонемент МУЛЬТИКАРТА 12 месяцев"),
        ),
        ExampleSpec(
            "zero_direct_kept_week_site",
            "Price=0 с direct-платежом у НЕДЕЛЯ САЙТ: type_of_payment оставлен как есть.",
            lambda r: price(r) == 0
            and not text(r, "_business_override")
            and text(r, "contract_name") == "Абонемент НЕДЕЛЯ САЙТ"
            and text(r, "_payment_match_source") == "direct_doc152_vt1083_doc154_vt1137_doc163"
            and text(r, "type_of_payment"),
        ),
        ExampleSpec(
            "zero_direct_raw_blank",
            "Price=0, direct-платеж найден, raw method пустой: type_of_payment оставлен пустым.",
            lambda r: text(r, "_business_override") == "business_zero_raw_blank_payment_type_blank",
            prefer_names=("Абонемент УЛЬТРА 12 месяцев", "Абонемент МУЛЬТИКАРТА 12 месяцев"),
        ),
        ExampleSpec(
            "subrent_unlimited",
            "Безлимитная субаренда: visits_left пустой.",
            lambda r: text(r, "_is_subrent") == "1"
            and text(r, "_is_limited_subrent") == "0"
            and price(r) > 0,
        ),
        ExampleSpec(
            "subrent_limited_active_nonzero",
            "Активная ограниченная субаренда: visits_left из регистра, остаток > 0.",
            lambda r: text(r, "_visits_left_source") == "rg3336_correct_dimension_balance"
            and decimal_value(r.get("visits_left")) > 0,
            prefer_names=("СУБАРЕНДА 15 посещений", "СУБАРЕНДА 12 посещений", "СУБАРЕНДА 8 посещений"),
        ),
        ExampleSpec(
            "subrent_limited_active_zero",
            "Активная ограниченная субаренда: реальный остаток visits_left = 0.",
            lambda r: text(r, "_visits_left_source") == "rg3336_correct_dimension_balance"
            and decimal_value(r.get("visits_left")) == 0,
        ),
        ExampleSpec(
            "subrent_limited_expired_zero",
            "Просроченная ограниченная субаренда: visits_left = 0 по бизнес-правилу.",
            lambda r: text(r, "_visits_left_source") == "business_expired_limited_subrent_zero_visits_left",
        ),
        ExampleSpec(
            "non_membership_solarium",
            "Нестандартный full-продукт солярия в общей выгрузке.",
            lambda r: "соляр" in text(r, "contract_name").lower() and price(r) > 0,
        ),
        ExampleSpec(
            "short_paid_trial",
            "Короткий платный trial/guest слой, не zero-price.",
            lambda r: text(r, "_product_class") == "trial_or_guest"
            and price(r) > 0
            and text(r, "type_of_payment"),
            prefer_names=("Абонемент НЕДЕЛЯ САЙТ", "Абонемент 2 недели Фитнес"),
        ),
    ]

    selected: list[tuple[ExampleSpec, dict[str, str]]] = []
    missing: list[ExampleSpec] = []
    used_contract_ids: set[str] = set()
    for spec in specs:
        candidates = [
            row
            for row in rows
            if text(row, "contract_id") not in used_contract_ids and spec.predicate(row)
        ]
        if not candidates:
            missing.append(spec)
            continue
        picked = sorted(candidates, key=lambda row: score(row, spec))[0]
        selected.append((spec, picked))
        used_contract_ids.add(text(picked, "contract_id"))

    return selected, missing


def read_staging_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def read_main_rows(path: Path) -> tuple[list[Any], list[Any], dict[str, list[Any]], list[dict[str, Any]]]:
    wb = load_workbook(path)
    ws = wb.active
    header_row = [ws.cell(1, col).value for col in range(1, len(CLIENT_HEADERS) + 1)]
    russian_row = [ws.cell(2, col).value for col in range(1, len(CLIENT_HEADERS) + 1)]
    row_styles: list[dict[str, Any]] = []
    for col in range(1, len(CLIENT_HEADERS) + 1):
        cell = ws.cell(3, col)
        row_styles.append(
            {
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }
        )
    data_by_contract_id: dict[str, list[Any]] = {}
    contract_idx = header_row.index("contract_id")
    for row_idx in range(3, ws.max_row + 1):
        values = [ws.cell(row_idx, col).value for col in range(1, len(CLIENT_HEADERS) + 1)]
        contract_id = str(values[contract_idx] or "").strip()
        if contract_id:
            data_by_contract_id[contract_id] = values
    wb.close()
    return header_row, russian_row, data_by_contract_id, row_styles


def write_xlsx(
    source_xlsx: Path,
    output_xlsx: Path,
    selected: list[tuple[ExampleSpec, dict[str, str]]],
) -> None:
    header_row, russian_row, data_by_contract_id, row_styles = read_main_rows(source_xlsx)
    wb = load_workbook(source_xlsx)
    ws = wb.active
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)
    for col, value in enumerate(header_row, start=1):
        ws.cell(1, col).value = value
    for col, value in enumerate(russian_row, start=1):
        ws.cell(2, col).value = value

    for _, staging_row in selected:
        contract_id = text(staging_row, "contract_id")
        values = data_by_contract_id.get(contract_id)
        if values is None:
            raise RuntimeError(f"Selected contract_id is missing in main XLSX: {contract_id}")
        ws.append(values)

    for row_idx in range(3, 3 + len(selected)):
        for col_idx, style in enumerate(row_styles, start=1):
            cell = ws.cell(row_idx, col_idx)
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.number_format = style["number_format"]
            cell.protection = copy(style["protection"])

    ws.freeze_panes = "A3"
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_xlsx)
    wb.close()


def write_markdown(
    path: Path,
    xlsx_path: Path,
    selected: list[tuple[ExampleSpec, dict[str, str]]],
    missing: list[ExampleSpec],
    date_stamp: str,
) -> None:
    lines = [
        f"# Representative {len(selected)} membership examples",
        "",
        f"date_stamp: `{date_stamp}`.",
        "",
        "Назначение: это не файл для правки правил `price/payment_type`, а",
        "репрезентативная выборка из всей финальной выгрузки абонементов.",
        "XLSX содержит только те же `20` колонок, что основной файл импорта",
        f"`fitbase_import_abonementy_clientov_{date_stamp}.xlsx`: без",
        "`question_for_manual_check`, `correct_price`, `correct_payment_type`,",
        "`comment` и без технических staging-полей.",
        "",
        f"Файл: `{xlsx_path}`",
        "",
        "## Состав 30 строк",
        "",
        "| # | category | contract_id | client_id | contract_name | payment_date | price | type_of_payment | зачем включено |",
        "|---:|---|---|---|---|---|---:|---|---|",
    ]
    for idx, (spec, row) in enumerate(selected, start=1):
        lines.append(
            "| "
            + " | ".join(
                [
                    str(idx),
                    spec.key,
                    text(row, "contract_id"),
                    text(row, "client_id"),
                    text(row, "contract_name").replace("|", "\\|"),
                    text(row, "payment_date"),
                    str(decimal_value(row.get("price"))),
                    text(row, "type_of_payment") or "blank",
                    spec.description.replace("|", "\\|"),
                ]
            )
            + " |"
        )
    lines.extend(
        [
            "",
            "## Missing categories",
            "",
        ]
    )
    if missing:
        for spec in missing:
            lines.append(f"- `{spec.key}`: {spec.description}")
    else:
        lines.append("- none")
    lines.extend(
        [
            "",
            "## Покрытие",
            "",
            "- массовые full-абонементы `УЛЬТРА` и `МУЛЬТИКАРТА`;",
            "- разные типы оплаты: `безналичные`, `наличные`, `сбп`, blank;",
            "- direct-платежи, fallback-платежи и отсутствие найденного платежа;",
            "- частичная оплата и рассрочка;",
            "- owner-change/переоформление;",
            "- все основные zero-price бизнес-правила;",
            "- активная/просроченная/безлимитная субаренда;",
            "- нестандартный full-продукт солярия;",
            "- короткий платный trial/guest слой.",
            "",
        ]
    )
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", default="output/20251115_0800_fix_owner_new_import")
    parser.add_argument("--date-stamp", default=DATE_STAMP)
    parser.add_argument("--report-md", default="")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    output_dir = Path(args.output_dir)
    if not output_dir.is_absolute():
        output_dir = ROOT / output_dir
    source_xlsx = output_dir / f"fitbase_import_abonementy_clientov_{args.date_stamp}.xlsx"
    staging_csv = output_dir / "staging" / "membership_import_rows.csv"

    staging_rows = read_staging_rows(staging_csv)
    selected, missing = select_examples(staging_rows)
    sample_xlsx = output_dir / f"membership_import_representative_{len(selected)}_examples_{args.date_stamp}.xlsx"
    if args.report_md:
        sample_md = Path(args.report_md)
        if not sample_md.is_absolute():
            sample_md = ROOT / sample_md
    else:
        sample_md = output_dir / "reports" / f"representative_{len(selected)}_examples_{args.date_stamp}.md"
    write_xlsx(source_xlsx, sample_xlsx, selected)
    write_markdown(sample_md, sample_xlsx, selected, missing, args.date_stamp)

    print(f"representative examples xlsx: {sample_xlsx}")
    print(f"representative examples md: {sample_md}")
    print(f"rows: {len(selected)}")
    print(f"missing_categories: {len(missing)}")
    for idx, (spec, row) in enumerate(selected, start=1):
        print(f"{idx:02d}. {spec.key}: {text(row, 'contract_id')} | {text(row, 'contract_name')}")


if __name__ == "__main__":
    main()
